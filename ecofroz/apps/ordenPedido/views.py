import logging
import io
import os
from io import BytesIO
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseNotFound, Http404, FileResponse, JsonResponse
from django.urls import reverse_lazy
from django.db.models import Q
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.template import context, loader
from django.contrib.sites.shortcuts import get_current_site
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from django.db import connection
from django.db.models import Subquery, OuterRef
from datetime import date, timedelta
from django.core import serializers
from django.core.files.storage import FileSystemStorage

# from wkhtmltopdf.views import PDFTemplateView

from django.conf import settings
from .models import OrdenesPedidos, DetallePedido, CotizaPedido, RechazaOrdenesPedidos, \
    RechazaDetallePedido, RechazaCotizaPedido, SecuencialCodifica, OrdenesPago, ConsultaExperto, Consultados, PreCotiza
from .forms import OrdenForm, DetForm, EditaForm, DetEditaForm, ApruebaForm, \
    DetApruebaForm, CotizaForm, DetCotizaFormSet, CodificaForm, CodificaFormDet, \
    EditaCodificaForm, EditaCodificaFormDet, IngresoPagosForm, nombreProyectoForm, selectAsignadoa, Predocumentos
from apps.usuarios.models import User, Cotizador, Autorizador, Generador, Email
from apps.activos.models import activo_tipo, desc_activo, activo_nomenclatura, activo_grupo, \
    activo_depar, activo_areas, detalle_desc_activo, salida_activos, activo_ubica
from apps.proveedores.models import proveedor
from apps.parametrosGlobales.models import *
from django.core.paginator import Paginator
from apps.parametrosGlobales.models import proyectos_contabilidad


# Create your views here.

logger = logging.getLogger(__name__)

### FUNCIONES INTERNAS ###

def envioMail(subject, email, template, queryset, queryset2):
    html_message = loader.render_to_string(
        'ordenPedido/pedidotmp/%s' %template,
            {
                'aprob':queryset,
                'aprob2':queryset2,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except Exception as err:
        logger.error(err)
        # logger.error("Unable to send mail.")

### VISTAS ###

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def AnularOrdenes(request):
    busqueda = request.GET.get("buscar")
    anuladas = DetallePedido.objects.filter(numpedido__genera_compra=3)[:25]
    if busqueda:
        consulta = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = busqueda) | Q(descripcion__icontains = busqueda))

        return render(request,'ordenPedido/pedidotmp/anula_ordenes.html', {'consulta':consulta, 'anuladas':anuladas})

    return render(request,'ordenPedido/pedidotmp/anula_ordenes.html', {'anuladas':anuladas})

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def AnulaOrdenesConfirma(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        idpedido = numpedido
        aprobado = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        aprob = aprobado.numpedido.genera_compra
        print(aprob)
        return render(request, 'ordenPedido/pedidotmp/anula_ordenes_confirma.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})
        
def confirma_anulacion(request, numpedido):
    print('estoy aqui')
    print(numpedido)
    observa = request.GET.get('text')
    orden = OrdenesPedidos.objects.get(numpedido = numpedido)
    orden.genera_compra = 3
    orden.observa_compra = observa
    orden.aprobado = 2
    orden.fchgenerac = datetime.now()
    orden.estado_cotiza = None
    orden.save()

    return redirect('ordenpedido:anular_orden')

def ubicaAreaAjax(request):
    data = []
    action = request.GET['action']
    area = request.GET['id']
    print('hola' + ' ' + area)
    try:
        for i in activo_areas.objects.filter(area_ubica=area):
            data.append({
                'area_codigo': i.area_codigo,
                'area_nombre': i.area_nombre,
            })

        return JsonResponse(data, safe=False)
    except Exception as e:
        print(e)

class IngresoPedido(PermissionRequiredMixin, CreateView):
    permission_required = 'ordenPedido.add_ordenespedidos'
    model = OrdenesPedidos
    template_name = 'ordenPedido/pedidotmp/ingreso_pedido.html'
    form_class = DetForm
    second_form_class = OrdenForm
    proyecto = nombreProyectoForm
    success_url = reverse_lazy('ordenpedido:listar_ordenes')

    def get_context_data(self, **kwargs):
        context = super(IngresoPedido, self).get_context_data(**kwargs)
        grupo = activo_grupo.objects.all()
        ubica = activo_ubica.objects.all()
        area = activo_areas.objects.none()
        activo = desc_activo.objects.all().order_by('activo_codigo')
        

        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'proyecto' not in context:
            context['proyecto'] = self.proyecto(self.request.GET)
        if 'activo' not in context:
            context['activo'] = activo
        
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        form2 = self.second_form_class(request.POST)
        proyecto = request.POST.get('nombre_proyecto') 
        activo = request.POST.get('activo_reemplazo')
        vida = request.POST.get('tiempo_vida')
        
        try:
            consulta = proyectos_contabilidad.objects.get(id=int(proyecto))
        
        except:
            consulta = None
        
        

        if form.is_valid() and form2.is_valid():
            print("Si entra forms correctos")
            
            
            solicitud = form.save(commit=False)
            cab = form2.save()
            solicitud.numpedido = cab
            solicitud.tiempo_vida = int(vida) 
            # cab.usuario_solicita = request.user
            if consulta:
                cab.numproyecto = consulta
            else:
                cab.numproyecto = None
            if activo:
                cab.codigo_activo_reemplazado = activo
            else:
                cab.codigo_activo_reemplazado = None
            cab.save()
            solicitud.save()

            queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__usuario_solicita=request.user.id).last()
            
            emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)


            ### SECCION PARA GRABAR NOTICACIONES 2-JUN-2022 - DM.
            
            id_usuario = queryset.numpedido.usuario_solicita_id

            id_solicita = User.objects.get(id=id_usuario)
            aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
            id_aprobador = User.objects.get(id=aprobador.user_id.id)

            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=7) # 7. PEDIDOS ACTIVOS AUTORIZACION GERENTE AREA
            app_number = modelo_App.objects.get(id=2) # 2. Activos

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = queryset.numpedido.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_aprobador,
                )

            r.save()

            ### FIN DE SECCION DE NOTIFICACIONES ##



            html_message = loader.render_to_string(
                'ordenPedido/pedidotmp/email.html',
                {
                    'mail':queryset
                }
            )
            email_subject = 'Solicitud de orden de pedido'
            to_list = emailaprob.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            if queryset.numpedido.reemplazo_accion == 1:
                emailaprob = emailaprob.e_mail
                html_message = loader.render_to_string(
                    'ordenPedido/pedidotmp/email_dd_baja.html',
                    {
                        'mail':queryset
                    }
                )
                email_subject = 'Solicitud de orden de pedido'
                to_list = emailaprob
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")
            
        
            return HttpResponseRedirect(self.get_success_url())
        else:
            print(form.errors)
            print(form2.errors)
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class DuplicaPedido(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.add_ordenespedidos'
    model = OrdenesPedidos
    second_model = DetallePedido
    template_name = 'ordenPedido/pedidotmp/duplica_pedido.html'
    form_class = OrdenForm
    second_form_class = DetForm
    success_url = reverse_lazy('ordenpedido:listar_ordenes')

    def get_context_data(self, **kwargs):
        context = super(DuplicaPedido, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numpedido=pk)
        detalle = self.second_model.objects.get(numpedido=orden.numpedido)
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        area = orden.area.area_codigo
        gruposelect = orden.grupo
        grupo = activo_grupo.objects.all()
        ubica = activo_ubica.objects.all()
        areas = activo_areas.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'area' not in context:
            context['area'] = area
        if 'areas' not in context:
            context['areas'] = areas
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'gruposelect' not in context:
            context['gruposelect'] = gruposelect
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        form2 = self.second_form_class(request.POST, request.FILES)

        proyecto = request.POST.get('numproyecto') 
        

        try:
            consulta = proyectos_contabilidad.objects.get(id=int(proyecto))
        except:
            consulta = None
        


        
        if form.is_valid() and form2.is_valid():
            solicitud = form2.save(commit=False)
            cab = form.save()
            solicitud.numpedido = cab
            cab.usuario_solicita = request.user
            if consulta:
                cab.numproyecto = consulta
            else:
                cab.numproyecto = None
            solicitud.save()
            cab.save()

            queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__usuario_solicita=request.user.id).last()
            
            emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
            html_message = loader.render_to_string(
                'ordenPedido/pedidotmp/email.html',
                {
                    'mail':queryset
                }
            )
            email_subject = 'Solicitud de orden de pedido'
            to_list = emailaprob.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            if queryset.numpedido.reemplazo_accion == 1:
                emailaprob = emailaprob.e_mail
                html_message = loader.render_to_string(
                    'ordenPedido/pedidotmp/email_dd_baja.html',
                    {
                        'mail':queryset
                    }
                )
                email_subject = 'Solicitud de orden de pedido'
                to_list = emailaprob
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class EditaPedido(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.change_ordenespedidos'
    model = OrdenesPedidos
    second_model = DetallePedido
    template_name = 'ordenPedido/pedidotmp/edita_pedido.html'
    form_class = EditaForm
    second_form_class = DetEditaForm
    success_url = reverse_lazy('ordenpedido:listar_ordenes')

    def get_context_data(self, **kwargs):
        context = super(EditaPedido, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numpedido=pk)
        detalle = self.second_model.objects.get(numpedido=orden.numpedido)
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        area = orden.area.area_codigo
        gruposelect = orden.grupo
        grupo = activo_grupo.objects.all()
        ubica = activo_ubica.objects.all()
        areas = activo_areas.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'areas' not in context:
            context['areas'] = areas
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'gruposelect' not in context:
            context['gruposelect'] = gruposelect
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numpedido=id_orden)
        detalle = self.second_model.objects.get(numpedido=orden.numpedido)
        form = self.form_class(request.POST, instance=orden)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        estado = OrdenesPedidos.objects.get(numpedido=id_orden)
        try:
            proy = proyectos_contabilidad.objects.get(id=request.POST.get('numproyecto'))
            # proyn = proy.nombre_proyecto
        except Exception as e:
            proy = None

        orden.numproyecto = proy

        if form.is_valid() and form2.is_valid():
            form.save()
            form2.save()
            orden.save()
    
            if orden.aprobado == 0:
                queryset = DetallePedido.objects.get(numpedido__numpedido=id_orden)
                print(queryset.numpedido)
                emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
                html_message = loader.render_to_string(
                    'ordenPedido/pedidotmp/email.html',
                    {
                        'mail':queryset
                    }
                )
                email_subject = 'Solicitud de orden de pedido'
                to_list = emailaprob.e_mail
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())

class EditaPedidoAutoriza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.change_ordenespedidos'
    model = OrdenesPedidos
    second_model = DetallePedido
    template_name = 'ordenPedido/pedidotmp/edita_pedido_autori.html'
    form_class = EditaForm
    second_form_class = DetEditaForm
    success_url = reverse_lazy('ordenpedido:aprueba_ordenes')

    def get_context_data(self, **kwargs):
        context = super(EditaPedidoAutoriza, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numpedido=pk)
        detalle = self.second_model.objects.get(numpedido=orden.numpedido)
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None
        
        ubicacion = orden.ubica
        gruposelect = orden.grupo
        grupo = activo_grupo.objects.all()
        ubica = activo_ubica.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'gruposelect' not in context:
            context['gruposelect'] = gruposelect
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        fch = datetime.now()
        orden = self.model.objects.get(numpedido=id_orden)
        detalle = self.second_model.objects.get(numpedido=orden.numpedido)
        form = self.form_class(request.POST, instance=orden)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        try:
            proy = proyectos_contabilidad.objects.get(id=request.POST.get('numproyecto'))
            # proyn = proy.nombre_proyecto
        except Exception as e:
            proy = None
        
        orden.aprobado = 1
        orden.usuario_aprueba = request.user.id
        orden.fchaprueba = fch
        orden.numproyecto = proy
        
        
        if form.is_valid() and form2.is_valid():
            form.save()
            form2.save()
            orden.save()

            email_usuario = User.objects.get(id=orden.usuario_solicita.id)

            queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=id_orden)
            
            for i in queryset:
                id_usuario = i.numpedido.usuario_solicita_id 

       
            ##Inicio de seccion de notificaciones Globales
            id_solicita = User.objects.get(id=id_usuario)
            # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
            
            id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
            id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
            id_compra3 = User.objects.get(id=22) # EVELY ROBLES
            id_compra4 = User.objects.get(id=118) # SOL RAMOS


            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=8) # 2. PEDIDOS ACTIVOS COTIZACION

            app_number = modelo_App.objects.get(id=2) # 2. ordenPedido

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = orden.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra1,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = orden.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra2,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = orden.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra3,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = orden.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra4,
                )
            r.save()

            ##Fin Seccion Notificaciones Globales


            html_message = loader.render_to_string(
                'ordenPedido/pedidotmp/email_aprobado.html',
                {
                    'aprob':queryset
                }
            )

            email_subject = 'Aprobación de orden de pedido Gerente de Area'
            to_list = 'adquisiciones@ecofroz.com'
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            envioMail('Aprobación de Orden de Pedido Gerente de Area', email_usuario.email, 'email_aprobado_gerente.html', queryset, '')

            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())
            # print(form.errors)
            # return HttpResponse('Error de algo')

@login_required
@permission_required('ordenPedido.view_ordenespedidos', raise_exception=True)
def ordenesPedido(request):
    queryset = request.GET.get("buscar")
    if queryset:
        orden = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            # Q(numpedido__numproyecto__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numpedido__usuario_solicita_id = request.user.id)
    else:
        orden2 = DetallePedido.objects.filter(numpedido__usuario_solicita_id = request.user.id).order_by('-numpedido__numpedido')
    
        paginator = Paginator(orden2, 100)
        page = request.GET.get('page')
        orden = paginator.get_page(page)
    
    return render(request, 'ordenPedido/pedidotmp/listar_ordenes.html', {'form':orden, 'busqueda':queryset})

def recibeProducto(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset.entrega_solicita = True
    queryset.save()

    return redirect('ordenpedido:listar_ordenes')

class EliminaPedido(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'ordenPedido.delete_ordenespedidos'
    model = OrdenesPedidos
    success_url = reverse_lazy('ordenpedido:listar_ordenes')

# -----------APROBACIÓN DE PEDIDOS GERENTE AREA---------------

@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def ordenesAprueba(request):
    queryset = request.GET.get("buscar")
    usu_depend = Autorizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(autorizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)
    print(id_usu)
    ordenx = DetallePedido.objects.filter(Q(numpedido__usuario_solicita__in=id_usu) & ~Q(numpedido__aprobado=0)).order_by('-numpedido__numpedido')
    # orden = DetallePedido.objects.filter(numpedido__aprobado__isnull=True).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
    # aprobados = DetallePedido.objects.filter(Q(numpedido__aprobado = 1) | Q(numpedido__aprobado = 0)).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
    if queryset:
        ordenx = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset) 
        ).filter(Q(numpedido__usuario_solicita__in=id_usu) & ~Q(numpedido__aprobado=0))

    paginator = Paginator(ordenx, 100)
    page = request.GET.get('page')
    orden = paginator.get_page(page)



        # orden = DetallePedido.objects.filter(
        #     Q(numpedido__numpedido__icontains = queryset) |
        #     Q(descripcion__icontains = queryset) 
        # ).filter(numpedido__aprobado__isnull=True).filter(numpedido__usuario_solicita__in=id_usu)

        # aprobados = DetallePedido.objects.filter(
        #     Q(numpedido__numpedido__icontains = queryset) |
        #     Q(descripcion__icontains = queryset) 
        # ).filter(Q(numpedido__aprobado=1) | 
        # Q(numpedido__aprobado=0)
        # ).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
            
    # else:
    #     aprobados = DetallePedido.objects.filter(Q(numpedido__aprobado = 1) | Q(numpedido__aprobado = 0)).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
        # orden = DetallePedido.objects.filter(numpedido__aprobado__isnull=True).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
    # return render(request, 'ordenPedido/pedidotmp/para_aprobar.html', {'form':orden, 'form2': aprobados, 'busqueda':queryset})
    return render(request, 'ordenPedido/pedidotmp/para_aprobar.html', {'form':orden, 'busqueda':queryset})

@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def apruebaDirect(request, numpedido):
    aprueba = OrdenesPedidos.objects.get(numpedido=numpedido)
    fch = datetime.now()
    if request.method == 'GET':
        aprueba.aprobado = 1
        aprueba.usuario_aprueba = request.user.id
        aprueba.fchaprueba = fch
        aprueba.save()

        email_usuario = User.objects.get(id=aprueba.usuario_solicita.id)

        queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        for i in queryset:
            id_tipo = i.numpedido.tipoactivo_id
            id_usuario = i.numpedido.usuario_solicita_id
        
        
            ##Inicio de seccion de notificaciones Globales
            id_solicita = User.objects.get(id=id_usuario)
            # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
            
            id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
            id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
            id_compra3 = User.objects.get(id=22) # EVELY ROBLES
            id_compra4 = User.objects.get(id=118) # SOL RAMOS


            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=8) # 2. PEDIDOS ACTIVOS COTIZACION

            app_number = modelo_App.objects.get(id=2) # 2. ordenPedido

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra1,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra2,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra3,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numpedido,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra4,
                )
            r.save()

            ##Fin Seccion Notificaciones Globales

        html_message = loader.render_to_string(
            'ordenPedido/pedidotmp/email_aprobado.html',
            {
                'aprob':queryset
            }
        )

        email_subject = 'Aprobación de orden de pedido'
        to_list = 'adquisiciones@ecofroz.com'
        mail = EmailMultiAlternatives(
                email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        mail.attach_alternative(html_message, "text/html")
        try:
            mail.send()
        except:
            logger.error("Unable to send mail.")

        envioMail('Aprobaciòn de Orden de Pedido Gerente de Area', email_usuario.email, 'email_aprobado_gerente.html', queryset, '')

    return redirect('ordenpedido:aprueba_ordenes')

@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def rechazaDirect(request, numpedido):
    pedido = request.GET.get("pedido")
    rechaza = OrdenesPedidos.objects.get(numpedido=pedido)
    if request.method == 'GET':
        rechaza.aprobado = 0
        rechaza.motRechaza = request.GET.get("text")
        rechaza.save()

        email_usuario = User.objects.get(id=rechaza.usuario_solicita.id)
        queryset = DetallePedido.objects.get(numpedido__numpedido=pedido)

        envioMail('Rechazo de Orden de Pedido', email_usuario.email, 'email_rechaza_pedido.html', queryset, '')

    return redirect('ordenpedido:aprueba_ordenes')

###########------------COTIZACIONES DE PEDIDOS-----------###########

@login_required
@permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def cotizaPedidos(request):
    queryset = request.GET.get("buscar")
    startdate = date.today() + timedelta(days=1) 
    enddate = startdate - timedelta(days=360)
   

    if queryset:
        orden = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset) 
        ).filter(numpedido__estado_cotiza__isnull=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')

        aprobados = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset) 
        ).filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
    
    
    else:
        # orden = OrdenesPedidos.objects.filter(estado_cotiza__isnull=True).filter(aprobado=1)
        orden = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__estado_cotiza__isnull=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
        
        
        aprobados2 = DetallePedido.objects.select_related('numpedido').filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).filter(numpedido__fchsolicita__range=[enddate,startdate]).order_by('-numpedido__ord')
        paginator = Paginator(aprobados2, 50)
        page = request.GET.get('page')
        aprobados = paginator.get_page(page)
    
    return render(request, 'ordenPedido/pedidotmp/para_cotizar.html', {'form':orden, 'form2': aprobados, 'busqueda':queryset})

class IngresoCotizaciones(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'ordenPedido.add_cotizapedido'
    model = OrdenesPedidos
    template_name = 'ordenPedido/pedidotmp/cotiza_ordenes.html'
    form_class = CotizaForm
    success_url = reverse_lazy('ordenpedido:cotiza_pedidos')

    def get(self, request, *args, **kwargs):
        self.object = None
        pk = self.kwargs.get('pk',0)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=pk)
        detalle = DetCotizaFormSet()
        return self.render_to_response(self.get_context_data(detalle = detalle, cot = cotiza))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numpedido=id_orden)
        aprob = None
        prove = None
        seleccion = None
        cd = False
        usuario = request.user
        orden.estado_cotiza = True
        orden.cotiza_observa = request.POST.get('observa_cotiza')
        orden.compra_directa = cd
        orden.genera_compra = aprob
        orden.select_cotiza = seleccion
        orden.usuario_cotiza = usuario.username
        orden.fchcotiza = datetime.now()

        form2 = DetCotizaFormSet(request.POST, request.FILES, instance=orden)
        if form2.is_valid():
            for form in form2.forms:
                if form.cleaned_data.get('valor'):
                    form.save()
            
            try:
                otros_docs_adqui = request.FILES['otros_docs_adqui']
                storage = FileSystemStorage()
                file_path = storage.save('otros_doc/'+otros_docs_adqui.name, otros_docs_adqui)
                modelo = DetallePedido.objects.get(numpedido=id_orden)
                modelo.otros_doc_adqui = file_path
                modelo.save()
            except Exception as e:
                print(e)


            if cd == True:
                cotiza = CotizaPedido.objects.get(numpedido_id=id_orden)
                cotiza.cotiza_seleccion = True
                cotiza.save()
            orden.save()

            if cd != True:
                queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=id_orden)
                for i in queryset:
                    usuario_id = i.numpedido.usuario_solicita_id
                # email_selectcotiza = User.objects.get(id=usuario_id)
                email_selectcotiza = User.objects.filter(id=usuario_id)
                for i in email_selectcotiza:
                    for a in i.autorizador.filter(status=True):
                        email_autoriza = a.e_mail
                
                email_selectcotiza = User.objects.get(id=usuario_id)

                # print('aqui email:' + email_selectcotiza.cotizador.e_mail) 
                        
                html_message = loader.render_to_string(
                    'ordenPedido/pedidotmp/email_cotizado.html',
                    {
                        'aprob':queryset
                    }
                )

                email_subject = 'Cotización de orden de pedido'
                to_list = email_selectcotiza.cotizador.e_mail
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

                return HttpResponseRedirect(self.get_success_url())
            else:
                queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=id_orden)
                for i in queryset:
                    usuario_id = i.numpedido.usuario_solicita_id
                # email_selectcotiza = User.objects.get(id=usuario_id)
                email_selectcotiza = User.objects.get(id=usuario_id)
                # for i in email_selectcotiza:
                #     for a in i.autorizador.filter(status=True):
                #         email_autoriza = a.e_mail

                html_message = loader.render_to_string(
                    'ordenPedido/pedidotmp/email_cotizado.html',
                    {
                        'aprob':queryset
                    }
                )

                email_subject = 'Cotización de orden de pedido'
                to_list = email_selectcotiza.email
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

                return HttpResponseRedirect(self.get_success_url())
            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())

@login_required
def observaReferencial(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        idpedido = numpedido
        aprobado = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        aprob = aprobado.numpedido.genera_compra
        print(aprob)
        return render(request, 'ordenPedido/pedidotmp/referencial_cotiza.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})

@login_required
def lookCotiza(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        idpedido = numpedido
        aprobado = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        aprob = aprobado.numpedido.genera_compra
        print(aprob)
        return render(request, 'ordenPedido/pedidotmp/look_cotiza.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})


@login_required
@permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def compraPedidos(request):
    queryset = request.GET.get("buscar")
    
    if queryset:
        orden2 = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(numpedido__proveedor__nombre_empresa__icontains = queryset)
        ).filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
    else:
       
        orden2 = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
    
    paginator = Paginator(orden2, 50)
    page = request.GET.get('page')
    orden = paginator.get_page(page)
    
    return render(request, 'ordenPedido/pedidotmp/para_comprar.html', {'form2': orden, 'busqueda':queryset})

@login_required
@permission_required('ordenPedido.selecciona_cotizapedido', raise_exception=True)
def pedidosCotizados(request):
    queryset = request.GET.get("buscar")
    usu_depend = Cotizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(cotizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)
    print(usu_depend2)
    cotizados = DetallePedido.objects.filter(numpedido__select_cotiza=True).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
    # cotizados = DetallePedido.objects.filter(numpedido__select_cotiza=True).filter(numpedido__usuario_solicita_id = request.user.id).order_by('-numpedido__numpedido')
    if queryset:
        orden = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion = queryset) 
        ).filter(numpedido__select_cotiza__isnull=True).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')
    else:
        orden = DetallePedido.objects.filter(numpedido__estado_cotiza__isnull=False).filter(numpedido__select_cotiza__isnull=True).filter(numpedido__usuario_solicita__in=id_usu).order_by('-numpedido__numpedido')

    return render(request, 'ordenPedido/pedidotmp/selec_cotizacion.html', {'form':orden, 'form2':cotizados, 'busqueda':queryset})


class ApruebaCotiza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.selecciona_cotizapedido'
    model = OrdenesPedidos
    template_name = 'ordenPedido/pedidotmp/selec_cotiza.html'
    form_class = CotizaForm
    success_url = reverse_lazy('ordenpedido:selec_cotiza')

    def get_context_data(self, **kwargs):
        context = super(ApruebaCotiza, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        detalle = CotizaPedido.objects.filter(numpedido=pk)
        detalle_data = []
        for det in detalle:
            d={
               'id':det.id,
               'valor':det.valor,
               'empresa_cotiza':det.empresa_cotiza,
               'pdf_cotiza':det.pdf_cotiza,
               'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=pk)
        prove = proveedor.objects.all()
        if 'detalle' not in context:
            context['detalle'] = detalle_cotiza
        if 'cot' not in context:
            context['cot'] = cotiza
        if 'prov' not in context:
            context['prov'] = prove
        if 'pk' not in context:
            context['pk'] = pk
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        id_orden = kwargs['pk']
        fch = datetime.now()
        orden = self.model.objects.get(numpedido=id_orden)
        cotiza = CotizaPedido.objects.filter(numpedido=orden.numpedido)
        form2 = DetCotizaFormSet(request.POST, queryset=cotiza)
        orden.observa_selec_cot = request.POST.get('observa_aut_cot')
        orden.proveedor = proveedor.objects.get(id=request.POST.get('prove'))
        orden.select_cotiza = True
        orden.fchselectcotiza = fch
        if form2.is_valid():
            for form in form2.forms:
                clave = form.cleaned_data.get('id')
                if clave:
                    if form.cleaned_data.get('cotiza_seleccion'):
                        queryset = CotizaPedido.objects.get(id=clave.id)
                        queryset.cotiza_seleccion = True
                        queryset.save()
                    else:
                        queryset = CotizaPedido.objects.get(id=clave.id)
                        queryset.cotiza_seleccion = False
                        queryset.save()
                
            orden.save()

            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())


class DevuelveCotiza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.selecciona_cotizapedido'
    model = OrdenesPedidos
    template_name = 'ordenPedido/selec_cotiza.html'
    form_class = CotizaForm
    success_url = reverse_lazy('ordenpedido:selec_cotiza')

    def get_context_data(self, **kwargs):
        pass
        # print("Hola mundo",pk)

        # context = super(DevuelveCotiza, self).get_context_data(**kwargs)
        # pk = self.kwargs.get('pk',0)
        # detalle = CotizaPedido.objects.filter(numpedido=pk)
        # detalle_data = []
        # for det in detalle:
        #     d={
        #        'id':det.id,
        #        'valor':det.valor,
        #        'empresa_cotiza':det.empresa_cotiza,
        #        'pdf_cotiza':det.pdf_cotiza,
        #        'cotiza_seleccion':det.cotiza_seleccion, 
        #     }
        #     detalle_data.append(d)
       
        # detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        # cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=pk)
        # if 'detalle' not in context:
        #     context['detalle'] = detalle_cotiza
        # if 'cot' not in context:
        #     context['cot'] = cotiza
        # return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numpedido=id_orden)
        cotiza = CotizaPedido.objects.filter(numpedido=orden.numpedido).delete()
        #form2 = DetCotizaFormSet(request.POST, queryset=cotiza)
        orden.observa_selec_cot = request.POST.get('observa_aut_cot')
        #orden.proveedor = proveedor.objects.get(id=request.POST.get('prove'))
        orden.select_cotiza = None
        orden.estado_cotiza = None
        orden.compra_directa = None
        orden.usuario_cotiza = None
                
        orden.save()

        #id_genera = User.objects.get(id=id_usuario)
        queryset = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=id_orden)
       
        envioMail('Devolución de Cotización Gerente de Area', 'adquisiciones@ecofroz.com,dmencias@ecofroz.com', 'email_cotizacion_devuelta_gerente.html', queryset, '')

            
        return HttpResponseRedirect(self.get_success_url())





################ORDENES DE COMPRA##########################

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def ordenesCompra(request):
    queryset = request.GET.get("buscar")
    form2 = selectAsignadoa()
    if queryset:
        query_cotiza_seleccion = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True)

        q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(numpedido__select_cotiza=True).annotate(
                valor=Subquery(query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]
                ))

        q2 = q1.annotate(cotiza_seleccion=Subquery(
            query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]
        ))
           
        q3 = q2.annotate(empresa_cotiza=Subquery(
            query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]
        ))
            
        val2 = q3.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(empresa_cotiza__icontains = queryset) |
            Q(numpedido__usuario_cotiza__icontains = queryset)).filter(cotiza_seleccion=True).order_by('-numpedido__genera_compra', '-numpedido__fchsolicita')
    
        preguntas_experto = ConsultaExperto.objects.all()
            

    else:

        query_cotiza_seleccion = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True)

        q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(numpedido__select_cotiza=True).annotate(
                valor=Subquery(query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]
                ))


        q2 = q1.annotate(cotiza_seleccion=Subquery(
            query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]
        ))

                    
        q3 = q2.annotate(empresa_cotiza=Subquery(
            query_cotiza_seleccion.filter(numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]
        ))


        val2 = q3.filter(cotiza_seleccion=True).order_by('-numpedido__genera_compra', '-numpedido__fchsolicita')

        preguntas_experto = ConsultaExperto.objects.all()

    paginator = Paginator(val2, 80)
    page = request.GET.get('page')
    orden = paginator.get_page(page)

    # datos = []
    # for i in orden:
    #     valor = i.numpedido_id
    #     datos.append(valor)

    # print(datos)


        # orden = OrdenesPedidos.objects.filter(estado_cotiza__isnull=  False).filter(aprobado=1)
        #q1 = CotizaPedido.objects.filter(numpedido__estado_cotiza__isnull=False).filter(numpedido__aprobado=1).filter(cotiza_seleccion=True).order_by('-numpedido__genera_compra', '-numpedido__fchsolicita')
        # q2 = RechazaCotizaPedido.objects.filter(numpedido__genera_compra=2).filter(numpedido__aprobado=1).filter(cotiza_seleccion=True).order_by('-numpedido__ord')
        # orden = q1.union(q2)
        # orden = CotizaPedido.objects.filter(numpedido__estado_cotiza__isnull=False).filter(numpedido__aprobado=1).filter(cotiza_seleccion=True).select_related('numpedido').order_by('-numpedido__ord')
    
    # orden = []
    # for data in q1:
    #     q2 = DetallePedido.objects.get(numpedido__numpedido = data.numpedido.numpedido)
    #     d={
    #         'pedido':data.numpedido.numpedido,
    #         'usuario':data.numpedido.usuario_solicita.first_name + ' ' + data.numpedido.usuario_solicita.last_name,
    #         'departamento':data.numpedido.departamento,
    #         'proyecto':data.numpedido.numproyecto,
    #         'fecha':data.numpedido.fchsolicita,
    #         'tipo':data.numpedido.tipoactivo,
    #         'descripcion':q2.descripcion,
    #         'valor':data.valor,
    #         'estado':data.numpedido.genera_compra,
    #         'compradora':data.numpedido.usuario_cotiza,
    #         'genera_compra':data.numpedido.genera_compra,
    #         # 'proveedor':data.empresa_cotiza.nombre_empresa,
    #     }
    #     orden.append(d)

    return render(request, 'ordenPedido/pedidotmp/listar_compras.html', {'form':orden, 'busqueda':queryset, 'form2':form2, 'preguntas':preguntas_experto})

def autComprasExcel(request):
    try:          
        # q1 = CotizaPedido.objects.filter(numpedido__estado_cotiza__isnull=False).filter(numpedido__aprobado=1).filter(cotiza_seleccion=True).order_by('-numpedido__genera_compra', '-numpedido__fchsolicita')
        #orden = []
        #for data in q1:
        #    q2 = DetallePedido.objects.get(numpedido__numpedido = data.numpedido.numpedido)
        #    d={
        #        'pedido':data.numpedido.numpedido,
        #        'usuario':data.numpedido.usuario_solicita.first_name + data.numpedido.usuario_solicita.last_name,
        #        'departamento':data.numpedido.departamento,
        #        'proyecto':data.numpedido.numproyecto,
        #        'fecha':data.numpedido.fchsolicita,
        #        'tipo':data.numpedido.tipoactivo,
        #        'descripcion':q2.descripcion,
        #        'valor':data.valor,
        #        'estado':data.numpedido.genera_compra,
        #        'compradora':data.numpedido.usuario_cotiza,
        #        'genera_compra':data.numpedido.genera_compra,
        #    }   
        #    orden.append(d)
        # print(q1)

        cursor = connection.cursor()
        # cursor.execute("""select o.numpedido, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
        #     (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto, o.fchsolicita,
        #     d.descripcion, c.valor, o.usuario_cotiza,
        #     	case
		#             when o.genera_compra = 1 then 'APROBADO'
		#             else 'PENDIENTE'
	    #         end as estado, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
        #     o.fchaprueba, o.fchgenerac,
        #         case
        #             when o.activado_activo = True then 'SI'
        #             when o.activado_activo = False then 'NO'
        #             else 'PENDIENTE'
        #         end as activo
        #     from pedidos.ordenespedidos o inner join pedidos.detallepedido d on (o.numpedido = d.numpedido_id)
        #     inner join pedidos.cotizapedido c on (o.numpedido = c.numpedido_id)
        #     where c.cotiza_seleccion = true order by o.fchsolicita desc""")

        cursor.execute("""select o.numpedido, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
            (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, (select e.nombre_proyecto from parametros.proyectos_contabilidad e where id = numproyecto_id) as proyecto, o.fchsolicita,
            d.descripcion, c.valor, o.usuario_cotiza,
            	case
		            when o.genera_compra = 1 then 'APROBADO'
		            else 'PENDIENTE'
	            end as estado, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
            o.fchaprueba, o.fchgenerac,
                case
                    when o.activado_activo = True then 'SI'
                    when o.activado_activo = False then 'NO'
                    else 'PENDIENTE'
                end as activo, o.observa_compra
            from pedidos.ordenespedidos o inner join pedidos.detallepedido d on (o.numpedido = d.numpedido_id)
            inner join pedidos.cotizapedido c on (o.numpedido = c.numpedido_id)
            where c.cotiza_seleccion = true order by o.fchsolicita desc""")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE COMPRAS DE ACTIVOS'

        ws.merge_cells('A1:M1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'USUARIO SOLICITANTE'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'PROYECTO'
        ws['E3'] = 'FECHA DE SOLICITUD'
        ws['F3'] = 'FECHA DE APROBACIÓN'
        ws['G3'] = 'DESCRIPCIÒN'
        ws['H3'] = 'VALOR'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'COMPRADORA'
        ws['K3'] = 'ESTADO GENERACIÒN COMPRA'
        ws['L3'] = 'FECHA APROBACIÓN ADMINISTRATIVA'
        ws['M3'] = 'OBSERVACIÓN ADMINISTRATIVA'
        ws['N3'] = 'ES ACTIVO'
              
        count = 4
        rowcount = 1   

        for i in cursor:
            print(i)
            #for (k, v) in i.items():
            #    print()
            #ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 1).value =  str(i[0])
            ws.cell(row = count, column = 2).value =  str(i[1])
            ws.cell(row = count, column = 3).value =  str(i[2])
            ws.cell(row = count, column = 4).value =  str(i[3])
            ws.cell(row = count, column = 5).value =  str(i[4])
            ws.cell(row = count, column = 6).value =  str(i[10])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[9])
            ws.cell(row = count, column = 10).value =  str(i[7])
            ws.cell(row = count, column = 11).value =  str(i[8])
            ws.cell(row = count, column = 12).value =  str(i[11])
            ws.cell(row = count, column = 13).value =  str(i[13])
            ws.cell(row = count, column = 14).value =  str(i[12])
            # ws.cell(row = count, column = 11).value =  str(i.compradora)
            # ws.cell(row = count, column = 12).value =  str(i.genera_compra)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE COMPRAS DE ACTIVOS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        #ahora = datetime.now()
        #ayer = ahora - timedelta(days=1)
        #fecha = ahora.strftime('%Y-%m-%d')
        #fechatxt = ahora.strftime('%Y%m%d')
        #fechaacttxt = ayer.strftime('%Y%m%d')
        #queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        #paginator = Paginator(queryset.order_by('-ord'), 50)
        #page = request.GET.get('page')
        #ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'
        print(e)

        return redirect('ordenpedido:ordenescompra')

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def generaCompra(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)

        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        idpedido = numpedido
        aprobado = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        aprob = aprobado.numpedido.genera_compra
        observa =  aprobado.numpedido.observa_compra
        
        return render(request, 'ordenPedido/pedidotmp/genera_compras.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob, 'observa':observa})

@login_required 
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def apruebaCompra(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        compra = OrdenesPedidos.objects.get(numpedido=numpedido)
        compra.genera_compra = 1
        compra.observa_compra = observa
        compra.fchgenerac = datetime.now()
        # print('Aprueba Compra Con observación: ' observa)
        compra.save()

        queryset = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido).filter(cotiza_seleccion=True)
        queryset2 = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)

        for i in queryset2:
            id_tipo = i.numpedido.tipoactivo_id
            id_usuario = i.numpedido.usuario_solicita_id


        id_solicita = User.objects.get(id=id_usuario)
        # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
        
        id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
        id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
        id_compra3 = User.objects.get(id=22) # EVELY ROBLES
        id_compra4 = User.objects.get(id=118) # SOL RAMOS


        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=6) # 6. ACTIVOS AUTORIZA COMPRA GER. ADMIN

        app_number = modelo_App.objects.get(id=2) # 1. INSUMOS / TRABAJOS

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numpedido,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra1,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numpedido,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra2,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numpedido,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra3,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numpedido,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra4,
            )
        r.save()

        # emailcompra = Cotizador.objects.get(tipo=id_tipo)

        # html_message = loader.render_to_string(
        # 'ordenPedido/pedidotmp/email_orden_compra.html',
        #     {
        #         'aprob':queryset,
        #         'aprob2':queryset2,
        #     }
        # )

        # html_message_bodega = loader.render_to_string(
        # 'ordenPedido/pedidotmp/email_orden_compra_bod.html',
        #     {
        #         'aprob':queryset,
        #         'aprob2':queryset2,
        #     }
        # )

        # html_message_contabilidad = loader.render_to_string(
        # 'ordenPedido/pedidotmp/email_orden_compra_cont.html',
        #     {
        #         'aprob':queryset,
        #         'aprob2':queryset2,
        #     }
        # )

####LINEAS FUNCIONALES CAMBIO POR PRUEBAS######
        envioMail('Generación de Orden de Compra', 'adquisiciones@ecofroz.com', 'email_orden_compra.html', queryset, queryset2)
        envioMail('Adquisisción de Activo', 'bodega@ecofroz.com', 'email_orden_compra_bod.html', queryset, queryset2)
        envioMail('Activacón de Activo', 'contabilidad@ecofroz.com', 'email_orden_compra_cont.html', queryset, queryset2)


        # envioMail('Generación de Orden de Compra', 'desarrollo@ecofroz.com', 'email_orden_compra.html', queryset, queryset2)
        # envioMail('Adquisisción de Activo', 'desarrollo@ecofroz.com', 'email_orden_compra_bod.html', queryset, queryset2)
        # envioMail('Activacón de Activo', 'desarrollo@ecofroz.com', 'email_orden_compra_cont.html', queryset, queryset2)

        # email_subject = 'Generación de Orden de Compra'
        # to_list = emailcompra.e_mail
        # mail = EmailMultiAlternatives(
        #         email_subject, '', '', [to_list])
        # mail.attach_alternative(html_message, "text/html")
        # try:
        #     mail.send()
        # except:
        #     logger.error("Unable to send mail.")

        return redirect('ordenpedido:ordenescompra')
    else:
        return redirect('ordenpedido:ordenescompra')

@login_required 
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def actualizaObserva(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        compra = OrdenesPedidos.objects.get(numpedido=numpedido)
        compra.observa_compra = observa
        # print('Actualiza Con observación: ' + observa)
        compra.save()

        return redirect('ordenpedido:ordenescompra')
    else:
        return redirect('ordenpedido:ordenescompra')

def cancelaCompra(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        compra = OrdenesPedidos.objects.get(numpedido=numpedido)
        compra.genera_compra = 2
        compra.observa_compra = observa
        compra.aprobado = None
        compra.estado_cotiza = None
        compra.select_cotiza = None
        # print('Prueba de canceacion de compra:' + observa)
        compra.save()

        ord_historic = OrdenesPedidos.objects.get(numpedido=numpedido)
        det_historic = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        cot_historic = CotizaPedido.objects.filter(numpedido__numpedido=numpedido)

        rop = RechazaOrdenesPedidos(
            numpedido = ord_historic.numpedido,
            departamento = ord_historic.departamento,
            area = ord_historic.area,
            usuario_solicita = ord_historic.usuario_solicita,
            usuario_aprueba = ord_historic.usuario_aprueba,
            aprobado = ord_historic.aprobado,
            observa_selec_cot = ord_historic.observa_selec_cot,
            fchsolicita = ord_historic.fchsolicita,
            fchsolicitatxt = ord_historic.fchsolicitatxt,
            tipoactivo = ord_historic.tipoactivo,
            tiempo_vida = ord_historic.tiempo_vida,
            tiempo_tipo = ord_historic.tiempo_tipo,
            numproyecto = ord_historic.numproyecto,
            regimenespecial = ord_historic.regimenespecial,
            proveedor = str(ord_historic.proveedor),
            # files = ord_historic.files,
            fchaprueba = ord_historic.fchaprueba,
            fchapruebatxt = ord_historic.fchapruebatxt,
            motRechaza = ord_historic.motRechaza,
            estado_cotiza = ord_historic.estado_cotiza,
            cotiza_observa = ord_historic.cotiza_observa,
            select_cotiza = ord_historic.select_cotiza,
            genera_compra = ord_historic.genera_compra,
            observa_compra = ord_historic.observa_compra,
            motivo_compra = ord_historic.motivo_compra,
            justificacion_compra = ord_historic.justificacion_compra,
            reemplazo_accion = ord_historic.reemplazo_accion,
            reemplazo_observa = ord_historic.reemplazo_observa,
            activado_activo = ord_historic.activado_activo,
            fecha_activacion = ord_historic.fecha_activacion,
            recibido_bodega = ord_historic.recibido_bodega,
            fecha_recepcion = ord_historic.fecha_recepcion,
            codigo_genera = ord_historic.codigo_genera,
            entregado_bodega =  ord_historic.entregado_bodega,
            fecha_entrega = ord_historic.fecha_entrega,
            entrega_observa = ord_historic.entrega_observa,
            ord = ord_historic.ord,
            cod_activo = ord_historic.cod_activo,
        )
        rop.save()

        rdp = RechazaDetallePedido(
            numpedido = rop,
            cantidad = det_historic.cantidad,
            descripcion = det_historic.descripcion,
            unimedida = det_historic.unimedida,
            img_1 = det_historic.img_1,
            img_2 = det_historic.img_2,
            cotiza_Ref = det_historic.cotiza_Ref,
        )
        rdp.save()

        for i in cot_historic:
            rcp = RechazaCotizaPedido(
                numpedido = rop,
                valor = i.valor,
                empresa_cotiza = i.empresa_cotiza,
                pdf_cotiza = i.pdf_cotiza,
                cotiza_seleccion = i.cotiza_seleccion,
            )
            rcp.save()

        cotiza = CotizaPedido.objects.filter(numpedido=numpedido).delete()

        emailaprob = User.objects.get(id=compra.usuario_solicita.id).autorizador.get(status=True)
        emailaprob_auto = emailaprob.e_mail
        print(emailaprob_auto)
        queryset = observa
        queryset2 = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)

##ARREGLAR CORREO QUE SEA AUTOMATICO###

        envioMail('Rechazo de Compra', emailaprob_auto, 'email_rechaza_cotizacion.html', queryset, queryset2)

        return redirect('ordenpedido:ordenescompra')
    else:
        return redirect('ordenpedido:ordenescompra')

###### ACTIVAR ACTIVOS ######

@login_required
@permission_required('ordenPedido.facturacion', raise_exception=True)
def listarComprasConta(request):
    queryset = request.GET.get("buscar")
    # aprobados = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
    # aprobados = OrdenesPedidos.objects.filter(estado_cotiza = True).filter(aprobado=1)
    if queryset:
        orden2 = CotizaPedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset)
        ).filter(numpedido__genera_compra=1, cotiza_seleccion=True).order_by('-numpedido__ord')
    else:
        # orden = OrdenesPedidos.objects.filter(estado_cotiza__isnull=True).filter(aprobado=1)
        orden2 = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido__genera_compra=1, cotiza_seleccion=True).order_by('-numpedido__ord')

    paginator = Paginator(orden2, 100)
    page = request.GET.get('page')
    orden = paginator.get_page(page)

    return render(request, 'ordenPedido/pedidotmp/listar_compras_conta.html', {'form':orden, 'busqueda':queryset})

def datosCompraConta(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        valor = 0
        detalle_data = []
        for det in detalle:
            if det.cotiza_seleccion == True:
                valor = det.valor
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        idpedido = numpedido
        aprobado = DetallePedido.objects.get(numpedido__numpedido=numpedido)
        activado = aprobado.numpedido.activado_activo
        aprob = aprobado.numpedido.genera_compra
        fact = aprobado.numpedido.aceptacion_factura
        print(aprob)
        return render(request, 'ordenPedido/pedidotmp/compras_conta.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob, 'valor':valor, 'fact':fact, 'activo':activado})

@login_required
@permission_required('ordenPedido.activa_activos', raise_exception=True)
def listarActivaciones(request):
    queryset = request.GET.get("buscar")
    startdate = date.today() + timedelta(days=1) 
    enddate = startdate - timedelta(days=360)
   
    orden_activada = OrdenesPedidos.objects.filter(genera_compra=1).filter(activado_activo__isnull=False).filter(fchsolicita__range=[enddate,startdate]).order_by('-ord')
    if queryset:
        # orden = CotizaPedido.objects.filter(
        #     Q(numpedido__numpedido__icontains = queryset) |
        #     Q(numpedido__numproyecto__icontains = queryset) 
        # ).filter(numpedido__estado_cotiza__isnull=False).filter(cotiza_seleccion=True).order_by('-numpedido__ord')
        orden = CotizaPedido.objects.filter(numpedido__numpedido__icontains = queryset).filter(numpedido__estado_cotiza__isnull=False).filter(cotiza_seleccion=True).order_by('-numpedido__ord')
    else:
        orden = CotizaPedido.objects.filter(numpedido__genera_compra=1).filter(numpedido__activado_activo__isnull=True).filter(cotiza_seleccion=True).order_by('-numpedido__ord')
        # orden = OrdenesPedidos.objects.filter(genera_compra=1).filter(activado_activo__isnull=True)
        # orden_activada = OrdenesPedidos.objects.filter(genera_compra=1).filter(activado_activo__isnull=False)

    return render(request, 'ordenPedido/pedidotmp/listar_activaciones.html', {'form':orden, 'form2':orden_activada, 'busqueda':queryset})

@login_required
@permission_required('ordenPedido.activa_activos', raise_exception=True)
def activaActivo(request, numpedido):
    pk = numpedido
    activa = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=pk)
    cotiza = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=pk).filter(cotiza_seleccion=True)

    return render(request, 'ordenPedido/pedidotmp/activaciones.html', {'act':activa, 'cot':cotiza})

@login_required
@permission_required('ordenPedido.activa_activos', raise_exception=True)
def activacionSi(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    observa = request.GET.get('observa_activa')
    if request.method == 'GET':
        queryset.activado_activo = True
        queryset.fecha_activacion = datetime.now()
        queryset.activado_observa = observa
        queryset.save()

        queryset = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido).filter(cotiza_seleccion=True)
        queryset2 = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)
        ##CORREOS A JEFE BODEGA y contabilidad
        envioMail('Activacion de Pedido', 'jefe.bodega@ecofroz.com', 'email_activacion_activo.html', queryset, queryset2)
        envioMail('Activacion de Pedido', 'contabilidad_sis@ecofroz.com', 'email_activacion_activo.html', queryset, queryset2)
    
    return redirect('ordenpedido:activaciones')

def activacionNo(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    observa = request.GET.get('observa_activa')
    if request.method == 'GET':
        queryset.activado_activo = False
        queryset.fecha_activacion = datetime.now()
        queryset.activado_observa = observa
        queryset.save()

        queryset = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=numpedido).filter(cotiza_seleccion=True)
        queryset2 = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=numpedido)

        envioMail('Activacion de Pedido', 'jefe.bodega@ecofroz.com', 'email_activacion_activo_no.html', queryset, queryset2)
        envioMail('Activacion de Pedido', 'contabilidad_sis@ecofroz.com', 'email_activacion_activo_no.html', queryset, queryset2)

    return redirect('ordenpedido:activaciones')

###### RECEPCION ACTIVOS Y GENERACION DE CODIGOS ######

@login_required
def listarRecepciones(request):
    queryset = request.GET.get("buscar")
    activate = 0
    if queryset:
        recepcion = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numpedido__genera_compra=1).filter(Q(numpedido__codigo_genera=None) | Q(numpedido__codigo_genera=False)).order_by('numpedido')
        generado = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numpedido__genera_compra=1).filter(numpedido__recibido_bodega=True) & DetallePedido.objects.filter(numpedido__codigo_genera=True) & DetallePedido.objects.filter(numpedido__entregado_bodega__isnull=True).order_by('numpedido')
        entregado = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numpedido__genera_compra=1).filter(numpedido__entregado_bodega=True).order_by('numpedido')

        q1 = recepcion.count()
        q2 = generado.count()
        q3 = entregado.count()

        if q1 > q2 and q1 > q3:
            activate = 'pendientes'
        elif q2 > q1 and q2 > q3:
            activate = 'generadas'
        elif q3 > q1 and q3 > q2:
            activate = 'entregadas'
        else:
            activate = 0
    else:
        recepcion = DetallePedido.objects.filter(numpedido__genera_compra=1).filter(Q(numpedido__codigo_genera=None) | Q(numpedido__codigo_genera=False)).exclude(numpedido__no_codificable=True).order_by('numpedido')
        generado = DetallePedido.objects.filter(numpedido__genera_compra=1).filter(numpedido__recibido_bodega=True) & DetallePedido.objects.filter(numpedido__codigo_genera=True) & DetallePedido.objects.filter(numpedido__entregado_bodega__isnull=True).order_by('numpedido')
        entregado = DetallePedido.objects.filter(numpedido__genera_compra=1).filter(numpedido__entregado_bodega=True).order_by('-numpedido__fecha_entrega')

    paginator_recep = Paginator(recepcion, 25) # Show 25 contacts per page.
    page_number_recep = request.GET.get('page_recep')
    page_obj_recep = paginator_recep.get_page(page_number_recep)

    paginator_gen = Paginator(generado, 25) # Show 25 contacts per page.
    page_number_gen = request.GET.get('page_gen')
    page_obj_gen = paginator_gen.get_page(page_number_gen)

    paginator_ent = Paginator(entregado, 25) # Show 25 contacts per page.
    page_number_ent = request.GET.get('page_ent')
    page_obj_ent = paginator_ent.get_page(page_number_ent)

    if page_number_recep:
        activate = 'pendientes'
    elif page_number_gen:
        activate = 'generadas'
    elif page_number_ent:
        activate = 'entregadas'

    return render(request, 'ordenPedido/pedidotmp/listar_recepciones.html', {'form':page_obj_recep, 'form2':page_obj_gen, 'form3':page_obj_ent, 'activate':activate})

@login_required  
def recibidoProd(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset2 = DetallePedido.objects.filter(numpedido__numpedido=numpedido)
    if request.method == 'GET':
        queryset.recibido_bodega = True
        queryset.fecha_recepcion = datetime.now()
        queryset.save()

        email_usuario = User.objects.get(id=queryset.usuario_solicita.id) + ',' + 'jefe.bodega@ecofroz.com'
        print(email_usuario.email)

        envioMail('Recepción de Pedido', email_usuario.email, 'email_recepcion.html', queryset, queryset2)

    return redirect('ordenpedido:recepcion')

def grabadoAct(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset2 = desc_activo.objects.get(activo_codigo=queryset.cod_activo)
    if request.method == 'GET':
        estado = request.GET.get("grabado")
        queryset2.grabado = estado
        queryset2.save()
        queryset.grabado_activo = True
        queryset.save()

    return redirect('ordenpedido:recepcion')


def listarPreguntasExperto(request):

    usuario = request.user.id
    us = User.objects.get(id=usuario)
    print(usuario, us)

    form = ConsultaExperto.objects.filter(consul__responsable=us).filter(numpedido__estado_consulta_experto=1).order_by('-fecha_pregunta')
    respondidas = ConsultaExperto.objects.filter(consul__responsable=us).filter(numpedido__estado_consulta_experto=2).order_by('-fecha_respuesta')
    print(respondidas)

    return render(request, 'ordenPedido/pedidotmp/listar_preguntas_experto.html', {'form':form,'form2':respondidas})



class CodificaIngreso(CreateView):
    model = desc_activo
    template_name = 'ordenPedido/pedidotmp/ingreso_codificacion.html'
    form_class = CodificaFormDet
    second_form_class = CodificaForm
    success_url = reverse_lazy('ordenpedido:recepcion')

    def get_context_data(self, **kwargs):
        context = super(CodificaIngreso, self).get_context_data(**kwargs)
        pk =  self.kwargs.get('pk',0)
        queryset = DetallePedido.objects.all().select_related('numpedido').get(numpedido__numpedido=pk)
        proveedor = CotizaPedido.objects.get(numpedido=pk, cotiza_seleccion=True)
        # tipo = queryset.numpedido.tipoactivo_id
        # nomenclatura = activo_tipo.objects.all().select_related('tipo_grupo').get(id=tipo)
        # numcontinua = desc_activo.objects.filter(activo_codigo__contains=nomenclatura.tipo_nomenclatura).order_by('activo_codigo').last()
        if queryset.numpedido.parte_activo == True:
            numcontinua = SecuencialCodifica.objects.filter(codigo='S').order_by('numeracion').last()
            codigo = 'S'
        else:
            numcontinua = SecuencialCodifica.objects.filter(codigo='P').order_by('numeracion').last()
            codigo = 'P'

        # nomenclatura_cod = activo_nomenclatura.objects.get(nomenclatura_codigo=nomenclatura.tipo_nomenclatura)
        
        # if nomenclatura_cod:
        #     nom_cod = nomenclatura_cod.id
        # else:
        #     nom_cod = 99
        
        num = ''
        if numcontinua:
            num = numcontinua.numeracion
            # for i in numcontinua.activo_codigo:
            #     if i.isnumeric():
            #         num = num + i
        else:
            num = 999

        numero = int(num) + 1
        cod = len(str(numero))
        if cod == 1: 
            # nomen = nomenclatura.tipo_nomenclatura + '00' + str(numero)
            nomen = codigo + '000' + str(numero)
        elif cod == 2:
            # nomen = nomenclatura.tipo_nomenclatura + '0' + str(numero)
            nomen = codigo + '00' + str(numero)
        elif cod == 3:
            nomen = codigo + '0' + str(numero)
        else:
            # nomen = nomenclatura.tipo_nomenclatura + str(numero)
            nomen = codigo + str(numero)

        initial_data = {
            # 'activo_nomenclatura': nom_cod,
            'activo_tipo': queryset.numpedido.tipoactivo.id,
            'activo_depar': queryset.numpedido.departamento.id,
            'activo_area': queryset.numpedido.area,
            'activo_descripcion': queryset.descripcion,
            'activo_grupo': queryset.numpedido.grupo, #nomenclatura.tipo_grupo.id,
            'activo_ubica': queryset.numpedido.ubica,
            'numero_factura': queryset.numpedido.numero_factura,
            'activo_valor_compra': queryset.numpedido.valor_factura,
            'orden_de_pedido':queryset.numpedido.numpedido,
        }

        inicio = {
            'desc_activo_proveedor': 4,
        }
        print(queryset.numpedido.fchgenerac)
        # print(queryset.numpedido.codigo_mba)

        activa = DetallePedido.objects.all().select_related('numpedido').filter(numpedido=pk)
        cotiza = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=pk).filter(cotiza_seleccion=True)

        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET or None, initial=inicio)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET or None, initial=initial_data)
        if 'nomen' not in context:
            context['nomen'] = nomen
        if 'proveedor' not in context:
            context['proveedor'] = proveedor.empresa_cotiza
        if 'act' not in context:
            context['act'] = activa
        if 'cot' not in context:
            context['cot'] = cotiza
        if 'codigo' not in context:
            context['codigo'] = codigo
        if 'numero' not in context:
            context['numero'] = numero
        if 'fchcompra' not in context:
            context['fchcompra'] = queryset.numpedido.fchgenerac
        if 'valor' not in context:
            context['valor'] = queryset.numpedido.valor_factura
        if 'codmba' not in context:
            context['codmba'] = queryset.numpedido.codigo_mba
        return context

    def post(self, request, *args, **kwargs):
        pk =  self.kwargs.get('pk',0)
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)
        actualiza = OrdenesPedidos.objects.get(numpedido=pk)
        # estado = request.POST.get('Activaciones')
        
        # activa = 0
        # if estado == '1':
        #     activa = True
        # elif estado == '2':
        #     activa = False

        # print(estado)
        # print(activa)
            
        if form.is_valid() and form2.is_valid():
            solicitud = form.save(commit=False)
            solicitud.desc_activo_codigo = form2.save()
            solicitud.save()
            actualiza.codigo_genera = True
            actualiza.cod_activo = request.POST.get('activo_codigo')
            # actualiza.fecha_activacion = datetime.now()
            # actualiza.activado_observa = request.POST.get('observa_activa')
            # actualiza.activado_activo = activa
            actualiza.save()
            codigo2 = request.POST.get('codigo')
            numero2 = request.POST.get('numero')
            r = SecuencialCodifica(
                codigo=codigo2,
                numeracion=numero2,
            )
            r.save()
            
            queryset = CotizaPedido.objects.all().select_related('numpedido').filter(numpedido=pk).filter(cotiza_seleccion=True)
            queryset2 = DetallePedido.objects.filter(numpedido=pk)

            email_usuario = User.objects.get(id=actualiza.usuario_solicita.id)
###coreeo debe ser automatico
            envioMail('Codificación de producto', email_usuario.email, 'email_codificacion_entrega.html', queryset, queryset2)
            envioMail('Codificación de producto', 'jefe.bodega@ecofroz.com', 'email_codificacion_entrega_pb.html', queryset, queryset2)

            return HttpResponseRedirect(self.get_success_url())
        else:
            print(form.errors)
            print(form2.errors)
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

def editaCodifica(request, numpedido):
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    if request.method == 'GET':
        activo = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=queryset.cod_activo)
        data = {
            'activo_nomenclatura':activo.desc_activo_codigo.activo_nomenclatura,
			'activo_codigo':activo.desc_activo_codigo.activo_codigo,
			'activo_grupo':activo.desc_activo_codigo.activo_grupo,
			'activo_tipo':activo.desc_activo_codigo.activo_tipo,
			'activo_descripcion':activo.desc_activo_codigo.activo_descripcion,
			'activo_ubica':activo.desc_activo_codigo.activo_ubica,
			'activo_depar':activo.desc_activo_codigo.activo_depar,
			'activo_area':activo.desc_activo_codigo.activo_area,
            'activo_valor':activo.desc_activo_codigo.activo_valor,
            'orden_de_pedido':activo.desc_activo_codigo.orden_de_pedido,
			'numero_factura':activo.desc_activo_codigo.numero_factura,
			'cod_activo_padre':activo.desc_activo_codigo.cod_activo_padre,

            'desc_activo_marca':activo.desc_activo_marca,
            'desc_activo_num_serie':activo.desc_activo_num_serie,
            'desc_activo_modelo':activo.desc_activo_modelo,
            'desc_activo_proveedor':activo.desc_activo_proveedor,
			'desc_activo_usuario_registra':activo.desc_activo_usuario_registra,
        }
        form = EditaCodificaForm(initial=data)
        form2 = EditaCodificaFormDet(initial=data)
        return render(request, 'ordenPedido/pedidotmp/edita_codificacion.html', {'form':form, 'form2':form2})

    if request.method == 'POST':
        marca = request.POST.get('desc_activo_marca')
        serie = request.POST.get('desc_activo_num_serie')
        modelo = request.POST.get('desc_activo_modelo')

        actualiza = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=queryset.cod_activo)
        actualiza.desc_activo_marca = marca
        actualiza.desc_activo_num_serie = serie
        actualiza.desc_activo_modelo = modelo
        actualiza.save()

        queryset.codigo_genera = True   
        queryset.save()

        return redirect('ordenpedido:recepcion')

def entregaPedido(request, numpedido):
    pedido = request.GET.get('pedido')
    queryset = OrdenesPedidos.objects.get(numpedido=pedido)
    if request.method == 'GET':
        queryset.entregado_bodega = True
        queryset.entrega_observa = request.GET.get("text")
        queryset.fecha_entrega = datetime.now()
        queryset.save()

    return redirect('ordenpedido:recepcion')

@login_required
def entrega_de_activos(request,numpedido):
    #pedido = request.GET.get('pedido')
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset_activo = desc_activo.objects.get(activo_codigo = queryset.cod_activo)
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    

    if request.method == 'POST':
        nuevo_depar = activo_depar.objects.get(id=request.POST.get("activo_depar"))
        nueva_area = activo_areas.objects.get(area_codigo=request.POST.get("activo_area"))

        queryset.entregado_bodega = True
        queryset.entrega_observa = request.POST.get("observaciones")
        queryset.fecha_entrega = datetime.now()
        queryset.persona_entrega_en_bodega = request.POST.get("usuario_registra")
        queryset.persona_retira_de_bodega = request.POST.get("usuario_retira")
        queryset.save()

        actualiza = desc_activo.objects.filter(activo_codigo=queryset.cod_activo).update(activo_depar=nuevo_depar,activo_area=nueva_area)

        return redirect('ordenpedido:recepcion')
    
    user = request.user

    return render(request, 'ordenPedido/pedidotmp/entrega_de_activos.html',{'queryset':queryset,'queryset_activo':queryset_activo,'depar':depar,'area':area,'user':user})


def confirma_custodio(request,numpedido):
    #pedido = request.GET.get('pedido')
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset_activo = desc_activo.objects.get(activo_codigo = queryset.cod_activo)
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    

    if request.method == 'POST':
        nuevo_depar = activo_depar.objects.get(id=request.POST.get("activo_depar"))
        nueva_area = activo_areas.objects.get(area_codigo=request.POST.get("activo_area"))
        nuevo_custodio = request.POST.get("custodio")

        actualiza = desc_activo.objects.filter(activo_codigo=queryset.cod_activo).update(activo_depar=nuevo_depar,activo_area=nueva_area)
        actualiza2 = detalle_desc_activo.objects.filter(desc_activo_codigo=queryset_activo.id).update(desc_activo_custodio=nuevo_custodio)

        return redirect('ordenpedido:recepcion')
    

    return render(request, 'ordenPedido/pedidotmp/confirma_custodio.html',{'queryset':queryset,'queryset_activo':queryset_activo,'depar':depar,'area':area})



def pdfCodigo(request, nomen):
    codigo = nomen
    print(codigo)
    fuentes = str(settings.STATICFILES_DIRS)

    pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
    pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer)
    p.setPageSize((144, 72))
    # text = p.beginText(5, 5)
    # text.setFont('3 of 9 Barcode')
    p.setFont('3 of 9 Barcode', 32)

    # Draw things on the PDF. Here's where the PDF generation happens.
    # See the ReportLab documentation for the full list of functionality.
    p.drawString(7, 25, '*' + codigo + '*')

    p.setFont('Arial', 16)
    p.drawString(50, 10, codigo)

    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=False, filename='hello.pdf')

def muestra(request):
    departamentos = activo_depar.objects.all()
    areas = activo_areas.objects.all()
    tipo = activo_tipo.objects.all()

    return render(request, 'ordenPedido/pedidotmp/multi_ingreso.html', {'depar':departamentos, 'area':areas, 'tipo':tipo})

def apruebaFactura(request):
    pedido = request.GET.get('pedido')
    valor = request.GET.get('valor')
    factura = request.GET.get('factura')
    if request.GET.get('mba'):
        mba = request.GET.get('mba')
    else:
        mba = 'REGISTRO'
    fecha = datetime.now()
    ordenes = OrdenesPedidos.objects.get(numpedido=pedido)
    ordenes.aceptacion_factura = True
    ordenes.valor_factura = valor
    ordenes.numero_factura = factura
    ordenes.usuario_factura = request.user.username
    ordenes.fch_factura = fecha
    ordenes.fch_factura_txt = fecha.strftime('%Y%m%d')
    codigo_mba = mba
    ordenes.save()

    if ordenes.cod_activo is not None:
        activos = desc_activo.objects.get(activo_codigo=ordenes.cod_activo)
        activos.activo_valor_compra = valor
        activos.numero_factura = factura
        activos.save()
    
    return redirect('ordenpedido:compras_conta')

# ''''''''''ORDENES DE PAGO'''''''''''''''''''#

# '''''''''''ADQUISICIONES''''''''''''''''''''#

def muestraOrdenes(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = OrdenesPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = OrdenesPago.objects.all()

    return render(request, 'ordenPedido/pedidotmp/listar_pagos_sol.html', {'ordenes':ordenes,})

class GeneraPagos(CreateView):
    model = OrdenesPago
    template_name = 'ordenPedido/pedidotmp/ingreso_pagos.html'
    form_class = IngresoPagosForm
    success_url = reverse_lazy('ordenpedido:ingreso_pagos')

    def get_context_data(self, **kwargs):
        context = super(GeneraPagos, self).get_context_data(**kwargs)
        numero = OrdenesPedidos.objects.all()
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'numero' not in context:
            context['numero'] = numero
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            form.save()

            queryset = OrdenesPago.objects.all().filter(usuario_solicita=request.user.id).last()
            queryset2 = ""

            envioMail('Generación de Orden de Pago', 'gerencia.administrativa@ecofroz.com', 'email_orden_pago_sol.html', queryset, queryset2)

        return HttpResponseRedirect(self.get_success_url())

# '''''''''''GERENCIA ADMINISTRATIVA''''''''''''''''''''#

def muestraOrdenesGa(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = OrdenesPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = OrdenesPago.objects.all().order_by('fch_genera')

    return render(request, 'ordenPedido/pedidotmp/listar_pagos_ga.html', {'ordenes':ordenes,})

def detallePagos(request, orden):
    orden = OrdenesPago.objects.get(id=orden)
    
    return render(request, 'ordenPedido/pedidotmp/aprueba_pagos.html', {'orden':orden,})

def apruebaPagos(request, orden):
    orden =  OrdenesPago.objects.get(id=orden)
    orden.observaciones_ga = request.GET.get("text")
    orden.estado = True
    orden.save()

    queryset = orden
    queryset2 = ""

    envioMail('Generación de Orden de Pago', 'adquisiciones@ecofroz.com', 'email_orden_pago_ga_ap.html', queryset, queryset2)
    envioMail('Generación de Orden de Pago', 'contabilidad_sis@ecofroz.com', 'email_orden_pago_ga_ap.html', queryset, queryset2)

    return redirect('ordenpedido:aprueba_pagos')

def rechazaPagos(request, orden):
    orden =  OrdenesPago.objects.get(id=orden)
    orden.observaciones_ga = request.GET.get("text")
    orden.estado = False
    orden.save()

    queryset = orden
    queryset2 = ""

    envioMail('Generación de Orden de Pago', 'adquisiciones@ecofroz.com', 'email_orden_pago_ga_re.html', queryset, queryset2)
    envioMail('Generación de Orden de Pago', 'contabilidad_sis@ecofroz.com', 'email_orden_pago_ga_re.html', queryset, queryset2)

    return redirect('ordenpedido:aprueba_pagos')

def consulta_expertos(request):
    
    pregunta = request.GET.get("pregunta")
    numpedido = request.GET.get("numpe")
    
    if request.method == 'GET':
        
        fec = datetime.today()
        form = selectAsignadoa(request.GET or None)

        if form.is_valid():
            data = form.cleaned_data['asignacion']
            
            ids = []
            for i in data:
                id = i.id
                ids.append(id)
            
            print(ids)
                        
             
            fecha=datetime.today()
            pregunta_compuesta = str(datetime.strftime(fecha, '%Y-%m-%d %H:%M')) + " " + pregunta 

            numero = OrdenesPedidos.objects.get(numpedido=numpedido)
            
            usuarios = User.objects.filter(id__in=ids)

            r = ConsultaExperto(
                numpedido = numero,
                fecha_pregunta = fecha,
                pregunta = pregunta_compuesta,
                persona_pregunta = User.objects.get(genera_consultas_a_expertos=True),
                )
            r.save()

            # Actualiza cabecera de pedidos con el estado de la consulta
            actualiza_estado = OrdenesPedidos.objects.filter(numpedido=numpedido).update(estado_consulta_experto=1)
            
            queryset = ConsultaExperto.objects.filter(numpedido=numpedido).last()
            
            for i in usuarios:
                responsable = Consultados(responsable=i)
                responsable.save()
                responsable.consultadoa.add(queryset)

            
            queryset2 = Consultados.objects.filter(consultadoa=queryset.id)
            
            d= []
            emailaprob=""
            for i in usuarios:
                emailaprob = emailaprob + i.email + ','
                d.append(emailaprob)
            
            print("Lista de correo, emailaprob")

            id_solicita = User.objects.get(genera_consultas_a_expertos=True) # MONICA SEVILLA
            
            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=19) # 19. PEDIDOS ACTIVOS CONSULTA A ROL EXPERTO

            app_number = modelo_App.objects.get(id=4) # 2. ordenPedido - ConsultaExperto

            for i in usuarios:

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = queryset.numpedido.numpedido,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = User.objects.get(id=i.id),
                    )
                r.save()

            envioMail('Consulta para Experto Técnico en orden de pedido', emailaprob, 'email_consulta_experto.html', queryset, queryset2)
            
        else:
            return HttpResponse("Errores")
        
        return redirect('ordenpedido:ordenescompra')

def verPreguntasExperto(request,id):
    resp = request.POST.get('respuesta')
    try:
        query = ConsultaExperto.objects.get(id=id)
        numpe = query.numpedido.numpedido
    except: 
        query = ConsultaExperto.objects.get(numpedido=id)
        numpe = query.numpedido.numpedido
        id = query.id


    if request.method == 'GET':
        pregunta = query.pregunta
        query = DetallePedido.objects.get(numpedido__numpedido=numpe)
        return render(request, 'ordenPedido/pedidotmp/ver_preguntas_rol_experto.html', {'solic':query, 'pregunta':pregunta}) 

    else:
        fecha = datetime.now()
        respuesta_compuesta = str(datetime.strftime(fecha, '%Y-%m-%d %H:%M')) + " " + resp 


        actualiza_respuesta = ConsultaExperto.objects.filter(id=id).update(respuesta=respuesta_compuesta,
        fecha_respuesta=fecha)

        actualiza_estado = OrdenesPedidos.objects.filter(numpedido=numpe).update(estado_consulta_experto=2)

        mail = User.objects.get(genera_consultas_a_expertos=True) # MONICA SEVILLA
        #mail = User.objects.get(id=10) # DM
        recibemail = mail.email

        queryset = ConsultaExperto.objects.get(id=id)
        
        queryset2 = Consultados.objects.filter(consultadoa=queryset.id)
        

        envioMail('Ha recibido una respuesta de Consulta a Rol Experto en orden de pedido de Activos', recibemail, 'email_responde_experto.html', queryset, queryset2)
        #envioMail('Ha recibido una respuesta de Consulta a Rol Experto en orden de pedido de insumos', recibemail, 'email_responde_experto.html', queryset, "")


        return redirect('ordenpedido:listar_preguntas_experto')

def recuperaPreguntas(request):
    numpe = request.GET['numpe']
    data = []
    
    print(numpe)
    try:
        query = ConsultaExperto.objects.filter(numpedido=numpe)
        data = []
    
        for queryset in query:
            d={
                'id':queryset.numpedido.numpedido,
                'pregunta':queryset.pregunta,
                'respuesta':queryset.respuesta,
                'estado':queryset.numpedido.estado_consulta_experto,

            }   
            
            data.append(d)
        
        # initial_data = Consultados.objects.filter(asignadoa__numpedido=numpe).values_list(
        #         'responsable', flat=True
        #     )
        
        # print(initial_data)
        
        # asignacion = selectAsignadoa(initial={"asignacion":[cat for cat in initial_data]})
        
    except:
        data=None
        
    
    return JsonResponse(data, safe=False)





# '''''''''''CONTABILIDAD''''''''''''''''''''#

def muestraOrdenesConta(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = OrdenesPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = OrdenesPago.objects.filter(~Q(estado=False)).order_by('fch_genera')

    return render(request, 'ordenPedido/pedidotmp/listar_pagos_conta.html', {'ordenes':ordenes,})


# '''''''''''''''ACTA DE ENTREGA DE ACTIVOS''''''''''''''''
@login_required
def imprimeActaEntregaActivos(request,numpedido):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        reg = DetallePedido.objects.filter(numpedido=numpedido).select_related("numpedido")
        
        for i in reg:
            num = str(i.numpedido.numpedido)
            fec = str(i.numpedido.fecha_entrega)
            desc = i.descripcion
            codigo_genera = str(i.numpedido.cod_activo)
            tipo = str(i.numpedido.tipoactivo.tipo_nombre)
            depar = str(i.numpedido.departamento.dep_nombre)
            sec = str(i.numpedido.area.area_nombre)
            solicita = str(i.numpedido.usuario_solicita.first_name+' '+i.numpedido.usuario_solicita.last_name)
            retira = str(i.numpedido.persona_retira_de_bodega)
            entrega_obs = str(i.numpedido.entrega_observa)
            entrega = str(i.numpedido.persona_entrega_en_bodega)

        desc_string = ''
        for i in desc:
            desc_string = desc_string + i
        
        desc_string = (desc_string[:150] + '..') if len(desc_string) > 200 else desc_string
            
        #w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, codigo_genera)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Acta de Entrega-Recepción de Activos Nuevos")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Requisición #:")
        p.drawString(50,710, "Fecha de Entrega:")
        p.drawString(50,690, "Código del Activo:")
        p.drawString(50,670, "Descripción del Activo:")

        p.drawString(50,510, "Tipo Activo:")
        #p.drawString(50,580, "Código de Activo:")
        #p.drawString(50,560, "Empresa de Mantenimiento:")
        p.drawString(50,490, "Departamento:")
        p.drawString(50,470, "Sector:")
        p.drawString(50,550, "Observaciones de entrega:")
        p.drawString(30,194, "Solicitante:")
        p.drawString(220,194, "Retira:")
        p.drawString(440,194, "Entrega:")
        p.drawString(190,100, "---------------------------------------------------")
        p.drawString(215,90, "Firma Persona Retira ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in desc_string:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        texto2 = ''
        e = 1
        for i in entrega_obs:
            if e == 46:
                texto2 = texto2 + i
                texto2 = texto2 + '\n'
                e = 0
            else:
                texto2 = texto2 + i
                e = e + 1
        texto2 = texto2.split('\n')

        p.drawString(250,730, num)
        p.drawString(250,710, fec)
        p.drawString(250,690, codigo_genera)
        p.drawString(250,510, tipo)
        p.drawString(250,490, depar)
        p.drawString(250,470, sec)
        #p.drawString(250,550, entrega_obs)
        
        a = 0
        b = 670
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
       
        a = 0
        b = 550
        for i in texto2:
            p.drawString(250, b, texto2[a])
            a = a + 1
            b = b - 15
       
       
       
        p.drawString(30,180, solicita)
        p.drawString(220,180, retira)
        p.drawString(440,180, entrega)
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="entrega.pdf")


##REPORTE DE COMPRAS CON CRITERIO DE ACTIVACIONES (BAJO DEMANDA - NO BORRAR)

def activacionesExcel(request):
    try:          
        
        cursor = connection.cursor()

        cursor.execute("""select o.numpedido, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
            (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto, o.fchsolicita,
            d.descripcion, c.valor, o.usuario_cotiza,
            	case
		            when o.genera_compra = 1 then 'APROBADO'
		            else 'PENDIENTE'
	            end as estado, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
            o.fchaprueba, o.fchgenerac,
                case
                    when o.activado_activo = True then 'SI'
                    when o.activado_activo = False then 'NO'
                    else 'PENDIENTE'
                end as activo, o.observa_compra, o.cod_activo
            from pedidos.ordenespedidos o inner join pedidos.detallepedido d on (o.numpedido = d.numpedido_id)
            inner join pedidos.cotizapedido c on (o.numpedido = c.numpedido_id)
            where c.cotiza_seleccion = true order by o.fchsolicita desc""")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE COMPRAS DE ACTIVOS CON CRITERIO DE ACTIVACION'

        ws.merge_cells('A1:M1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'USUARIO SOLICITANTE'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'PROYECTO'
        ws['E3'] = 'FECHA DE SOLICITUD'
        ws['F3'] = 'FECHA DE APROBACIÓN'
        ws['G3'] = 'DESCRIPCIÒN'
        ws['H3'] = 'VALOR'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'COMPRADORA'
        ws['K3'] = 'ESTADO GENERACIÒN COMPRA'
        ws['L3'] = 'FECHA APROBACIÓN ADMINISTRATIVA'
        ws['M3'] = 'OBSERVACIÓN ADMINISTRATIVA'
        ws['N3'] = 'ACTIVADO'
        ws['O3'] = 'CODIGO ACTIVO'
              
        count = 4
        rowcount = 1   

        for i in cursor:
            print(i)
            #for (k, v) in i.items():
            #    print()
            #ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 1).value =  str(i[0])
            ws.cell(row = count, column = 2).value =  str(i[1])
            ws.cell(row = count, column = 3).value =  str(i[2])
            ws.cell(row = count, column = 4).value =  str(i[3])
            ws.cell(row = count, column = 5).value =  str(i[4])
            ws.cell(row = count, column = 6).value =  str(i[10])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[9])
            ws.cell(row = count, column = 10).value =  str(i[7])
            ws.cell(row = count, column = 11).value =  str(i[8])
            ws.cell(row = count, column = 12).value =  str(i[11])
            ws.cell(row = count, column = 13).value =  str(i[13])
            ws.cell(row = count, column = 14).value =  str(i[12])
            ws.cell(row = count, column = 15).value =  str(i[14])
            # ws.cell(row = count, column = 11).value =  str(i.compradora)
            # ws.cell(row = count, column = 12).value =  str(i.genera_compra)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE COMPRAS DE ACTIVOS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
       

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'
        print(e)

        return redirect('ordenpedido:ordenescompra')

def eliminar_documentos(request, pk, pedido):
    docu = PreCotiza.objects.get(id=pk)
    if request.method == 'POST':
        docu.delete()
        return redirect('ordenpedido:valida_cotizacion',pedido)
    return render(request,'ordenPedido/pedidotmp/eliminar_documentos.html',{'documento':docu})


class validaCotizacion(View):
    model = PreCotiza
    template_name = 'ordenPedido/pre_cotiza.html'
    success_url = reverse_lazy('ordenpedido:cotiza_trabajos')
    
    def get(self,request,**kwargs):
        pk = self.kwargs.get('pk',0)
        archivos_list = PreCotiza.objects.filter(numpedido = pk)
        for i in archivos_list:
            print(i.archivos.url)
        q = DetallePedido.objects.filter(numpedido = pk).select_related('numpedido')
        
        return render(self.request, 'ordenPedido/pedidotmp/pre_cotiza.html', {'pk':pk,'archivos_list':archivos_list, 'query':q})


    def post(self, request, **kwargs):

        doc_id = kwargs['pk']  
        form = Predocumentos(self.request.POST, self.request.FILES)
        
        print(form)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.numpedido_id = doc_id
            
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()
        
        else:
            data = {'is_valid': False}
        return JsonResponse(data)


def save_estado_precotiza(request):    #AJAX PARA GURDAR EL ESTADO DE LA SELECCION INDIVIDUAL
    id = request.GET.get("id")
    valor = request.GET.get("valor")
    print(valor)
    if valor == 'true':
        valor = True
    else:
        valor = False

    actualiza = PreCotiza.objects.filter(id = id).update(seleccion=valor)

    data = 'OK'

    return JsonResponse(data,safe=False)




def save_record_pre_cotiza(request, pk):
    pedido = OrdenesPedidos.objects.filter(numpedido=pk)
    detpedido = DetallePedido.objects.filter(numpedido=pk)
    
    observaciones_adqui = request.POST.get("observaciones")
    success_url = reverse_lazy('ordenpedido:cotiza_pedidos')
    compradora = request.user.username
   
    if request.method == 'POST':

        try:
            guarda_obs_adqui = detpedido.update(observa_envia_precotiza=observaciones_adqui)
            guarda_fecha_precotiza = PreCotiza.objects.filter(numpedido=pk).update(escrito=True)
            guarda_precotiza_cab = pedido.update(fchprecotiza = date.today(),usuario_precotiza=compradora,estado_precotiza = True)
            
        except:
            guarda_fecha_precotiza = PreCotiza.objects.filter(numpedido=pk).update(escrito=True)
            guarda_precotiza_cab = pedido.update(fchprecotiza = date.today(),usuario_precotiza=compradora,estado_precotiza = True)
            

    #     #Obtiene dirección de correo de solicitante

        queryset = DetallePedido.objects.get(numpedido=pk)
        
        id_usuario = queryset.numpedido.usuario_solicita.id
        id_solicita = User.objects.get(id=id_usuario)
        id_precotiza =  User.objects.get(username=request.user.username)
     
            
        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=23) # 1. ACTIVOS PRE COTIZACION
        app_number = modelo_App.objects.get(id=2) # 1. Activos

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = queryset.numpedido.numpedido,
            tipo = noti, 
            usuario_activa = id_precotiza,
            autorizador_id = id_solicita,
            )

        r.save()

         ### FIN DE SECCION DE NOTIFICACIONES ##

        emailaprob = id_solicita
        html_message = loader.render_to_string(
            'ordenPedido/pedidotmp/email_envia_precotiza.html',
            {
                'aprob':queryset
            
            }
        )
        email_subject = 'Envío de Pre Cotización Activos N°'+' '+str(pk)
        to_list = emailaprob.email
        mail = EmailMultiAlternatives(
                email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        mail.attach_alternative(html_message, "text/html")
        try:
            mail.send()
        except:
            logger.error("Unable to send mail.")

        return HttpResponseRedirect(success_url)

class revisaCotizacion(View):
    model = PreCotiza
    template_name = 'ordenPedido/pedidotmp/rev_pre_cotiza.html'
    
    def get(self,request,**kwargs):
        pk = self.kwargs.get('pk',0)
        archivos_list = PreCotiza.objects.filter(numpedido = pk)
        queryset = DetallePedido.objects.get(numpedido=pk)
        
        return render(self.request, 'ordenPedido/pedidotmp/rev_pre_cotiza.html', {'queryset':queryset,'archivos_list':archivos_list})

    def post(self, request, **kwargs):
        success_url = reverse_lazy('ordenpedido:listar_ordenes')
        pk = self.kwargs.get('pk',0)
        observa_responde = request.POST.get('observaciones')

        pedido = OrdenesPedidos.objects.filter(numpedido=pk)
        ped = OrdenesPedidos.objects.get(numpedido=pk)
        detpedido = DetallePedido.objects.filter(numpedido=pk)

        try:
            guarda_obs_responde = detpedido.update(observa_responde_precotiza=observa_responde)
            guarda_precotiza_cab = pedido.update(fchpreresponde = date.today(),estado_responde_precotiza = True)
            
        except:
            
            guarda_precotiza_cab = pedido.update(fchpreresponde = date.today(),estado_responde_precotiza = True)
            

    #     #Obtiene dirección de correo de compradora

        
        queryset = DetallePedido.objects.get(numpedido=pk)
        queryset2 = PreCotiza.objects.filter(numpedido=pk)
        
        id_usuario = queryset.numpedido.usuario_solicita.id
        id_solicita = User.objects.get(id=id_usuario)

        id_precotiza =  User.objects.get(username=ped.usuario_precotiza)
     
            
        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=24) # 1. ACTIVOS RESPONDE PRE COTIZACION
        app_number = modelo_App.objects.get(id=2) # 2. Activos

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = queryset.numpedido.numpedido,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_precotiza,
            )

        r.save()

    #     ### FIN DE SECCION DE NOTIFICACIONES ##

        emailaprob = id_precotiza
        html_message = loader.render_to_string(
            'ordenPedido/pedidotmp/email_responde_precotiza.html',
            {
                'aprob':queryset,
                'aprob2':queryset2
            
            }
        )
        email_subject = 'Respuesta de Pre Cotización Activos N°'+' '+str(pk)
        to_list = emailaprob.email
        mail = EmailMultiAlternatives(
                email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        mail.attach_alternative(html_message, "text/html")
        try:
            mail.send()
        except:
            logger.error("Unable to send mail.")

        return HttpResponseRedirect(success_url)
