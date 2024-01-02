import logging
from rest_framework.generics import ListAPIView
from django.shortcuts import render, redirect, get_object_or_404
from apps.activos.models import activo_depar, activo_areas
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseNotFound, Http404, FileResponse, JsonResponse
from django.contrib.auth.decorators import login_required, permission_required
from django.urls import reverse
from datetime import datetime, date, timedelta
from django.shortcuts import render
from django.core.paginator import Paginator
from django.utils import timezone, formats
from django.db.models import Q, F
from .models import *
from .forms import *
from django.template import context, loader
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from apps.usuarios.models import *
from apps.parametrosGlobales.models import tipo_notificacion, modelo_App, notificaciones_globales, proyectos_contabilidad
from openpyxl import Workbook
import pymongo
import pytz
from django.utils import timezone, formats
import openai
from apps.usuarios.models import User
from apps.proveedores.models import proveedor_det
from apps.activos.models import activo_depar, activo_ubica
from .serializers import SectoresSerializer2
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4 
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from io import BytesIO
import io
import os
from django.conf import settings
from django.db import transaction, IntegrityError
from django.db import connection
from itertools import chain
import pytz
from django.utils import timezone
from datetime import datetime, timedelta
from django.utils import timezone, formats
#import pandas as pd
from .resources import *
from django.db import connection, connections
from tablib import Dataset
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View


logger = logging.getLogger(__name__)

openai.api_key = "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"

def envioMail(subject, email, template, queryset, queryset2):
    html_message = loader.render_to_string(
        'procesosAdministrativos/email/%s' %template,
            {
                'aprob':queryset,
                'aprob2':queryset2,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except:
        logger.error("Unable to send mail.")

@login_required
def listarSolicitudes(request):
    queryset = request.GET.get("buscar")
    # if queryset:
    #     query = DetSolicitudTrabajoInterno.objects.filter(
    #         Q(numtrabajo__numtrabajo = queryset) |
    #         Q(numtrabajo__descripcion__icontains = queryset)).filter(numtrabajo__usuario_solicita = request.user.id).exclude(numtrabajo__anula=True)
    # else:
    #     query = DetSolicitudTrabajoInterno.objects.select_related('numtrabajo').filter(numtrabajo__usuario_solicita = request.user.id).exclude(
    #         numtrabajo__anula=True).order_by('-numtrabajo__fchsolicita')
    

    # paginator = Paginator(query, 50)
    # page = request.GET.get('page')
    # solic = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/listar_solicitudes.html', {})



@login_required
def ingresoSolicitudes(request):
    if request.method == "GET":
        # form = ingresoTrabajos(request.GET or None)
        
        # form = cabIngresoTrabajos(request.GET or None)
        # form2 = detIngresoTrabajos(request.GET or None)

        return render(request, 'procesosAdministrativos/ingreso.html', {})
    # else:
    #     pass:
        # form = cabIngresoTrabajos(request.POST)
        # form2 = detIngresoTrabajos(request.POST, request.FILES)
        # sector = request.POST.get("nsector")
        # print("nsector")

        # if form.is_valid() and form2.is_valid():
        #     cab = form.save(commit=False)
        #     cab.area = sector
        #     cab.usuario_solicita_id = request.user.id
        #     cab.save()
        #     fecha = datetime.today()
        #     # fec = datetime.strftime(fecha, '%Y-%m-%d %H:%M:%S.%fZ') 
            
        #     actualiza = SolicitudTrabajoInterno.objects.filter(numtrabajo=cab.numtrabajo).update(fchsolicita=fecha,estado1n=1)
            
        #     det = form2.save(commit=False)
        #     det.numtrabajo = cab
            
        #     form2.save()
            
        #     queryset = DetSolicitudTrabajoInterno.objects.filter(numtrabajo=cab.numtrabajo)
        #     emailaprob = User.objects.get(id=request.user.id)
        #     aut_oti = emailaprob.autoriza_oti.e_mail
        #     autoriza_instancia = Autorizador.objects.get(id=emailaprob.autoriza_oti_id)
        #     autoriza_user_id = autoriza_instancia.user_id.id
        #     autoriza_user = User.objects.get(id=autoriza_user_id)
            

        #     solicitud = DetSolicitudTrabajoInterno.objects.get(numtrabajo=cab.numtrabajo)
        #     ## SECCION DE NOTIFICACIONES GLOBALES

        #     ##Guarda cambio de estado en tabla de mensajeria
        #     noti = tipo_notificacion.objects.get(id=14) # 14. TRABAJOS INTERNOS SOLICITUD

        #     app_number = modelo_App.objects.get(id=3) # 3. trabajosInternos

        #     r = notificaciones_globales(
        #         app_origen = app_number,
        #         estado = True,
        #         identificador = solicitud.numtrabajo.numtrabajo,
        #         tipo = noti, 
        #         usuario_activa = emailaprob,
        #         autorizador_id = autoriza_user,
        #         )

        #     r.save()

        #     print("Helouuuu")
        #     envioMail('Solicitud Autorización de Orden de Trabajo Interna', aut_oti, 'email_aprobado_gerente.html', queryset, '')

        #     return redirect('trabajosinternos:listar_trabajos_internos')
        # else:
        #     print(form.errors)
        #     print(form2.errors)
        #     return HttpResponse("errores")




@login_required
def reporteRaytec(request):
    ahora = datetime.now()
    # ayer = ahora - timedelta(days=1)
    # fecha = ahora.strftime('%Y-%m-%d')
    # fechatxt = ahora.strftime('%Y%m%d')
    # fechaacttxt = ayer.strftime('%Y%m%d')
    #queryset = eventos_raytec.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
    queryset = EventosRaytec.objects.all()

    paginator = Paginator(queryset.order_by('-fecha_evento'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    l1 = request.GET.get("l1")
    l2 = request.GET.get("l2")
    l3 = request.GET.get("l3")

    mensaje = ''

    if inicio != 'None':
        print("Entra DM")
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        if inicio and fin:
            if inicio and fin and l2:
                print(inicio)
        
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea='L2')
                paginator = Paginator(filtro.order_by('-fecha_evento'), 500)
                page = request.GET.get('page')
                reg = paginator.get_page(page)

            elif inicio and fin and l3:
                
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea='L3')
                paginator = Paginator(filtro.order_by('-fecha_evento'), 50)
                page = request.GET.get('page')
                reg = paginator.get_page(page)

            elif inicio and fin and l1:
                
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea='L1')
                paginator = Paginator(filtro.order_by('-fecha_evento'), 50)
                page = request.GET.get('page')
                reg = paginator.get_page(page)
            
            elif inicio and fin and l1 and l2:
                
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea__in=('L1','L2'))
                paginator = Paginator(filtro.order_by('-fecha_evento'), 50)
                page = request.GET.get('page')
                reg = paginator.get_page(page)
                
            elif inicio and fin and l1 and l3:
                
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea__in=('L1','L3'))
                paginator = Paginator(filtro.order_by('-fecha_evento'), 50)
                page = request.GET.get('page')
                reg = paginator.get_page(page)

            elif inicio and fin and l2 and l3:
                
                # iniciotxt = iniciodate.strftime('%Y%m%d')
                # fintxt = findate.strftime('%Y%m%d')
                filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).filter(linea__in=('L2','L3'))
                paginator = Paginator(filtro.order_by('-fecha_evento'), 50)
                page = request.GET.get('page')
                reg = paginator.get_page(page) 

        else:
            mensaje = 'No seleccionó las fechas de busqueda'

    print(inicio,type(inicio))

    return render(request, 'procesosAdministrativos/reporte_raytec.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin})


@login_required
def reportePareceticoL3(request):
    ahora = datetime.now()
    queryset = EventosPareceticoL3.objects.all()

    paginator = Paginator(queryset.order_by('-fecha','ubicacion','tipo'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    

    mensaje = ''
    umedida = ' PPMs'

    if inicio != 'None':
        print("Si")
        
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        print(iniciodate,findate)
        if inicio and fin:
            print("llegamos")
             
            filtro = EventosPareceticoL3.objects.filter(fecha__date__gte=iniciodate,fecha__date__lt=findate)
            paginator = Paginator(filtro.order_by('-fecha'), 500)
            page = request.GET.get('page')
            reg = paginator.get_page(page)

        else:
            mensaje = 'No seleccionó las fechas de busqueda'


    return render(request, 'procesosAdministrativos/reporte_parecetico_l3.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin,'umedida':umedida})


@login_required
def reporteCloroL2(request):
    ahora = datetime.now()
   
    queryset = EventosCloroL2.objects.all()

    paginator = Paginator(queryset.order_by('-fecha','ubicacion','tipo'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    
    mensaje = ''
    umedida = ' mg/l'

    if inicio != 'None':
        print("Si")
        
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        print(iniciodate,findate)
        if inicio and fin:
            print("llegamos")
             
            filtro = EventosCloroL2.objects.filter(fecha__date__gte=iniciodate,fecha__date__lt=findate)
            paginator = Paginator(filtro.order_by('-fecha'), 100)
            page = request.GET.get('page')
            reg = paginator.get_page(page)

        else:
            mensaje = 'No seleccionó las fechas de busqueda'


    return render(request, 'procesosAdministrativos/reporte_cloro_l2.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin,'umedida':umedida})


@login_required
def reporteIshidaR1(request):
    ahora = datetime.now()
   
    queryset = EventosIshidaR1.objects.all()

    paginator = Paginator(queryset.order_by('-fecha'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    
    mensaje = ''
    umedida = ' g'

    if inicio != 'None':
        print("Si")
        
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        print(iniciodate,findate)
        if inicio and fin:
            print("llegamos")
             
            filtro = EventosIshidaR1.objects.filter(fecha__date__gte=iniciodate,fecha__date__lt=findate)
            paginator = Paginator(filtro.order_by('-fecha'), 100)
            page = request.GET.get('page')
            reg = paginator.get_page(page)

        else:
            mensaje = 'No seleccionó las fechas de busqueda'


    return render(request, 'procesosAdministrativos/reporte_ishida_r1.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin,'umedida':umedida})




@login_required
def reporteIshidaR2(request):
    ahora = datetime.now()
   
    queryset = EventosIshidaR2.objects.all()

    paginator = Paginator(queryset.order_by('-fecha'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    
    mensaje = ''
    umedida = ' g'

    if inicio != 'None':
        print("Si")
        
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        print(iniciodate,findate)
        if inicio and fin:
            print("llegamos")
             
            filtro = EventosIshidaR2.objects.filter(fecha__date__gte=iniciodate,fecha__date__lt=findate)
            paginator = Paginator(filtro.order_by('-fecha'), 100)
            page = request.GET.get('page')
            reg = paginator.get_page(page)

        else:
            mensaje = 'No seleccionó las fechas de busqueda'


    return render(request, 'procesosAdministrativos/reporte_ishida_r2.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin,'umedida':umedida})


@login_required
def reporteIshidaR3(request):
    ahora = datetime.now()
   
    queryset = EventosIshidaR3.objects.all()

    paginator = Paginator(queryset.order_by('-fecha'), 50)
    page = request.GET.get('page')
    reg = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    
    mensaje = ''
    umedida = ' g'

    if inicio != 'None':
        print("Si")
        
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        print(iniciodate,findate)
        if inicio and fin:
            print("llegamos")
             
            filtro = EventosIshidaR3.objects.filter(fecha__date__gte=iniciodate,fecha__date__lt=findate)
            paginator = Paginator(filtro.order_by('-fecha'), 100)
            page = request.GET.get('page')
            reg = paginator.get_page(page)

        else:
            mensaje = 'No seleccionó las fechas de busqueda'


    return render(request, 'procesosAdministrativos/reporte_ishida_r3.html',{'reg':reg,'msj':mensaje,'inicio':inicio,'fin':fin,'umedida':umedida})





def exportExcelEventosRaytec(request, inicio, fin):
    #try:
    inicio = str(inicio)
    fin = str(fin)

    iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
    findate = datetime.strptime(fin, '%Y-%m-%d')
    iniciotxt = iniciodate.strftime('%Y%m%d')
    fintxt = findate.strftime('%Y%m%d')
            
    filtro = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin).order_by('-fecha_evento')  

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE DE EVENTOS RAYTEC' + '  -  ' + inicio + '  -  ' + fin + '  -  ' + 'RECHAZOS SUMARIZADOS CADA 5 MINUTOS'

    ws.merge_cells('A1:M1')
    ws['A3'] = 'N°'
    ws['B3'] = 'PROCESO'
    ws['C3'] = 'RECHAZOS'
    ws['D3'] = 'DEFECTO'
    ws['E3'] = 'RECETA'
    ws['F3'] = 'LINEA'
    ws['G3'] = 'FECHA'
    ws['H3'] = 'INICIA PROCESO'
    ws['I3'] = 'FINALIZA PROCESO'

    count = 4
    rowcount = 1  

    for i in filtro:
        ws.cell(row = count, column = 1).value =  rowcount
        if i.proceso is None:
            ws.cell(row = count, column = 2).value = "--"
        else:
            ws.cell(row = count, column = 2).value =  str(i.proceso)
        ws.cell(row = count, column = 3).value =  str(i.counter)
        ws.cell(row = count, column = 4).value =  str(i.defecto)
        ws.cell(row = count, column = 5).value =  str(i.receta)
        ws.cell(row = count, column = 6).value =  str(i.linea)
        ws.cell(row = count, column = 7).value =  str(i.fecha_evento)
        if i.usuini is None:
            ws.cell(row = count, column = 8).value =  "--"
        else:
             ws.cell(row = count, column = 8).value =  str(i.usuini)
        if i.usufin is None:
            ws.cell(row = count, column = 9).value =  "--"
        else:
            ws.cell(row = count, column = 9).value =  str(i.usufin)

        count+=1
        rowcount+=1

    nombre_archivo = "REPORTE DE EVENTOS RAYTEC" + inicio + "_" + fin + ".xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    #except Exception as e:
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = EventosRaytec.objects.filter(time__gte=inicio, time__lte=fin)
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    registros = paginator.get_page(page)

    # mensaje = 'Error'
    # mensaje_e = 'Error en los valores, por favor verifique las fechas'
    #'msje':mensaje_e,

    return render(request, 'procesosAdministrativos/reporte_raytec.html', {'reg':registros,'inicio':inicio, 'fin':fin})

@login_required
def exportExcelEventosPeraceticoL3(request, inicio, fin):
    #try:
    inicio = str(inicio)
    fin = str(fin)

    iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
    findate = datetime.strptime(fin, '%Y-%m-%d')
    iniciotxt = iniciodate.strftime('%Y%m%d')
    fintxt = findate.strftime('%Y%m%d')
            
    filtro = EventosPareceticoL3.objects.filter(fecha__date__gte=iniciodate,fecha__date__lte=findate).order_by('-fecha')  

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE DE EVENTOS APA L3' + '  -  ' + inicio + '  -  ' + fin 

    ws.merge_cells('A1:M1')
    ws['A3'] = 'ROW ID'
    ws['B3'] = 'PROCESO'
    ws['C3'] = 'FECHA'
    ws['D3'] = 'UBICACION'
    ws['E3'] = 'TIPO'
    ws['F3'] = 'VALOR'
    ws['G3'] = 'LINEA'
    ws['H3'] = 'INICIA PROCESO'
    ws['I3'] = 'FINALIZA PROCESO'

    count = 4
    rowcount = 1  

    for i in filtro:
        # ws.cell(row = count, column = 1).value =  rowcount
        ws.cell(row = count, column = 1).value =  i.id
        if i.proceso is None:
            ws.cell(row = count, column = 2).value = "--"
        else:
            ws.cell(row = count, column = 2).value =  str(i.proceso)
        ws.cell(row = count, column = 3).value =  i.fecha + timedelta(hours=5)
        ws.cell(row = count, column = 4).value =  str(i.ubicacion)
        ws.cell(row = count, column = 5).value =  str(i.tipo)
        ws.cell(row = count, column = 6).value =  str(i.valor)
        ws.cell(row = count, column = 7).value =  str(i.linea)
        if i.usuini is None:
            ws.cell(row = count, column = 8).value =  "--"
        else:
             ws.cell(row = count, column = 8).value =  str(i.usuini)
        if i.usufin is None:
            ws.cell(row = count, column = 9).value =  "--"
        else:
            ws.cell(row = count, column = 9).value =  str(i.usufin)

        count+=1
        rowcount+=1

    nombre_archivo = "REPORTE DE EVENTOS APA L3" + inicio + "_" + fin + ".xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    #except Exception as e:
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = EventosPareceticoL3.filter(time__gte=inicio, time__lte=fin)
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    registros = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/reporte_parecetico_l3.html', {'reg':registros,'inicio':inicio, 'fin':fin})


@login_required
def exportExcelCloroL2(request, inicio, fin):
    #try:
    inicio = str(inicio)
    fin = str(fin)

    iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
    findate = datetime.strptime(fin, '%Y-%m-%d')
    iniciotxt = iniciodate.strftime('%Y%m%d')
    fintxt = findate.strftime('%Y%m%d')

    filtro = EventosCloroL2.objects.filter(fecha__date__gte=iniciodate,fecha__date__lte=findate).order_by('-fecha','ubicacion','tipo') 

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE DE EVENTOS DOSIFICACION CLORO L2' + '  -  ' + inicio + '  -  ' + fin 

    ws.merge_cells('A1:M1')
    ws['A3'] = 'Id ROW'
    ws['B3'] = 'PROCESO'
    ws['C3'] = 'FECHA'
    ws['D3'] = 'UBICACION'
    ws['E3'] = 'TIPO'
    ws['F3'] = 'VALOR'
    ws['G3'] = 'FLUJO'
    ws['H3'] = 'LINEA'
    ws['I3'] = 'INICIA PROCESO'
    ws['J3'] = 'FINALIZA PROCESO'

    count = 4
    rowcount = 1  

    for i in filtro:
        # ws.cell(row = count, column = 1).value =  rowcount
        ws.cell(row = count, column = 1).value =  i.id
        if i.proceso is None:
            ws.cell(row = count, column = 2).value = "--"
        else:
            ws.cell(row = count, column = 2).value =  str(i.proceso)

        ws.cell(row = count, column = 3).value =  i.fecha + timedelta(hours=5)  #Para corregir el deficit de 5 horas respecto al tiempo local y lo que se muestra en el reporte
        ws.cell(row = count, column = 4).value =  str(i.ubicacion)
        ws.cell(row = count, column = 5).value =  str(i.tipo)
        ws.cell(row = count, column = 6).value =  float(i.valor)
        ws.cell(row = count, column = 7).value =  str(i.aiEstado)
        
        ws.cell(row = count, column = 8).value =  str(i.linea)
        if i.usuini is None:
            ws.cell(row = count, column = 9).value =  "--"
        else:
             ws.cell(row = count, column = 10).value =  str(i.usuini)
        if i.usufin is None:
            ws.cell(row = count, column = 11).value =  "--"
        else:
            ws.cell(row = count, column = 12).value =  str(i.usufin)

        count+=1
        rowcount+=1

    nombre_archivo = "REPORTE DE EVENTOS DOSIFICACION CLORO L2" + inicio + "_" + fin + ".xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    #except Exception as e:
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = EventosCloroL2.filter(time__gte=inicio, time__lte=fin)
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    registros = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/reporte_cloro_l2.html', {'reg':registros,'inicio':inicio, 'fin':fin})


# CONSULTA DE EVENTOS EN VIVO
@login_required
def consultaEventosL1(request):
    MONGO_URL='mongodb://admin:Eco2021desa@192.168.27.62/bigdata?authSource=admin'
    connection = pymongo.MongoClient(MONGO_URL)
    mongo_db = connection.bigdata
    #Busca e imprime los elementos de una colección

    collection = mongo_db['LINEA1_LIVE'].find().sort("time", -1)
    
    data=[]
    for i in collection:
        data.append(i)

    return render(request,'procesosAdministrativos/listar_eventos_l1.html',{'datos':data})

@login_required
def consultaEventosL2(request):
    MONGO_URL='mongodb://admin:Eco2021desa@192.168.27.62/bigdata?authSource=admin'
    connection = pymongo.MongoClient(MONGO_URL)
    mongo_db = connection.bigdata
    #Busca e imprime los elementos de una colección

    collection = mongo_db['LINEA2_LIVE'].find().sort("time", -1)
    
    data=[]
    for i in collection:
        data.append(i)

    return render(request,'procesosAdministrativos/listar_eventos_l2.html',{'datos':data})


@login_required
def administra_eventos(request):
    all_events = Eventos.objects.all()
    form = cargaDoc(files=request.FILES )

    query = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado1n=2).exclude(
            numpedido__anula=True).exclude(Q(numpedido__estado2n=2) | Q(numpedido__estado2n=3) | Q(numpedido__estado2n=4)).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
        

    context = {
        "events":all_events,
    }
    return render(request,'procesosAdministrativos/administrar_eventos_transporte.html',{'context':context,'form':form,'nuevas_solicitudes':query})

@login_required
def ver_rutas_fijas(request):
    hoy = datetime.today()
    events = Eventos.objects.exclude(name__in = ['Camioneta Administrativa']).filter(start__gte=hoy)
    navega = request.user.username

    context = {
        "events":events,
        "navega":navega,
    }
    return render(request,'procesosAdministrativos/ver_rutas_fijas.html',context)

@login_required
def all_events(request):
    all_events = Eventos.objects.all()
    out = []
    for event in all_events:
        out.append({
            'title': event.name,
            'id': event.id,
            'start':event.start.strftime("%m-%d-%Y, %H:%M:%S"),
            'end':event.end.strftime("%m-%d-%Y, %H:%M:%S"),
            'desc':event.desc_evento,
            'backgroundColor':event.color,
            'allDay':event.allday,
            'extendedProps': {
                'descripcion':event.desc_evento,
                'asientos':event.conteo_disponible,
                'ruta':event.ruta.origen+' => '+event.ruta.destino,
                'ruta_id':event.ruta.id,
                'transportista':event.transportista,
                'allday':event.allday,
                'archivo_nombre':str(event.docs),
                }
        })
        
    
    return JsonResponse(out,safe=False)


@login_required
def all_events_wchis(request):
    hoy = datetime.today()
    estasemana = datetime.strftime(hoy - timedelta(days=5),'%Y-%m-%d')
    all_events = Eventos.objects.exclude(name__in = ['Camioneta Administrativa']).filter(start__gte=estasemana)
    out = []
    for event in all_events:
        out.append({
            'title': event.name,
            'id': event.id,
            'start':event.start.strftime("%m-%d-%Y, %H:%M:%S"),
            'end':event.end.strftime("%m-%d-%Y, %H:%M:%S"),
            'desc':event.desc_evento,
            'backgroundColor':event.color,
            'extendedProps': {
                'descripcion':event.desc_evento,
                'asientos':event.conteo_disponible,
                'ruta':event.ruta.origen+" => "+event.ruta.destino,
                'transportista':event.transportista,
                'allday':event.allday,
                }
        })
    
    return JsonResponse(out,safe=False)


@login_required
def update_event_tarjeta(request):
    id = request.GET.get("id",None)
    title = request.GET.get("title",None)
    inicio = request.GET.get("start",None)
    fin = request.GET.get("end", None)
    desc = request.GET.get("desc_evento",None)
    transportista = request.GET.get("transportista",None)
    ruta = request.GET.get("ruta",None)
    asientos = request.GET.get("asientos",None)

    
    c = ','
    pos = transportista.find(c)

    nombre_empre = ''
    for i in transportista:
        if i == c:
            break
        else:
            nombre_empre = nombre_empre + i
            
    try:

        prove = proveedor_det.objects.get(codigo_id__nombre_empresa=nombre_empre)
        asientos_d = prove.asientos
        color = prove.color

        if color == "AZUL":
            color = "#0080FF"
        elif color == "VERDE":
            color = "#74DF00"
        elif color == "ROJO":
            color = "#FF0040"
        elif color == "NARANJA":
            color = "#FE9A2E"
        else:
            color = "#CCFB5D"   

    except:
        color = "#CCFB5D"
    
    actualiza = Eventos.objects.filter(id=id).update(name=title,start=inicio,end=fin,desc_evento=desc,ruta=ruta,transportista=transportista,conteo_disponible=asientos,color=color)

    data = {}
    return JsonResponse(data)


@login_required
def add_event(request):
    
    if request.method == 'POST':
        print ("Si entra")
        
        form = cargaDoc(request.POST, request.FILES)
    
        inicio = request.POST.get("start", None)
        fin = request.POST.get("end", None)
        iniciotd = request.POST.get("startsf", None)
        title = request.POST.get("title",None)
        desc = request.POST.get("desc_evento",None)
        color = request.POST.get("color",None)
        transportista = request.POST.get("transportista",None)
        transportista_id = request.POST.get("transportista_id",None)
        ruta = request.POST.get("ruta",None)
        allday = request.POST.get("allday",None)
        tarifa = request.POST.get("tarifa",None)
        
        trans = str(transportista)
        print(trans)

        root = RutaTransporte.objects.get(id=ruta)

        
        if allday:
            print(inicio)
            # inicio = datetime.combine(inicioca, '00:00:00')
            # fin = datetime.combine(inicioca, '00:00:00')
            new_fecha = datetime.strptime(iniciotd, '%Y-%m-%d')
            fin = datetime.strftime(new_fecha + timedelta(days=1,hours=5),'%Y-%m-%d')
            inicio = datetime.strftime(new_fecha + timedelta(hours=5),'%Y-%m-%d')


            c = ','
            pos = trans.find(c)

            nombre_empre = ''
            for i in trans:
                if i == c:
                    break
                else:
                    nombre_empre = nombre_empre + i
            
            try:

                prove = proveedor_det.objects.get(codigo_id__nombre_empresa=nombre_empre)
                asientos_d = prove.asientos
                color = prove.color

                if color == "AZUL":
                    color = "#0080FF"
                elif color == "VERDE":
                    color = "#74DF00"
                elif color == "ROJO":
                    color = "#FF0040"
                elif color == "NARANJA":
                    color = "#FE9A2E"
                else:
                     color = "#CCFB5D"   

                
                form = cargaDoc(request.POST,request.FILES)
                
                if form.is_valid():
                
                    cab = form.save(commit=False)
                    cab.name = str(title)
                    cab.start = inicio
                    cab.end = fin
                    cab.desc_evento = str("Camioneta Administración")
                    cab.color = str(color)
                    # cab.ruta = root
                    cab.transportista = str(transportista)
                    cab.conteo_disponible=asientos_d
                    cab.color = color
                    cab.allday = True
                    # cab.tarifa_evento = tarifa
                    cab.save()

            except:
                asientos_d = 0
                if form.is_valid():
                    cab = form.save(commit=False)
                    cab.name = str("Camioneta Administración")
                    cab.start = inicio
                    cab.end = fin
                    cab.desc_evento = str(desc)
                    cab.color = str(color)
                    # cab.ruta = root
                    cab.transportista = str(transportista)
                    cab.conteo_disponible=asientos_d
                    cab.color = color
                    cab.allday = True
                    # cab.tarifa_evento = tarifa
                    cab.save()
    


            # event = Eventos(name=str(title), start=inicio, end=fin, desc_evento=str(desc), color=str(color), allday=True)
            # event.save()
        
        else:

            # Rutina para obtener el nombre del proveedor de la cadena que se se recoge en el POST Y obtener los asientos parametrizados
            c = ','
            pos = trans.find(c)

            nombre_empre = ''
            for i in trans:
                if i == c:
                    break
                else:
                    nombre_empre = nombre_empre + i
            
            try:

                prove = proveedor_det.objects.get(codigo_id__nombre_empresa=nombre_empre)
                asientos_d = prove.asientos
                color = prove.color

                if color == "AZUL":
                    color = "#0080FF"
                elif color == "VERDE":
                    color = "#74DF00"
                elif color == "ROJO":
                    color = "#FF0040"
                elif color == "NARANJA":
                    color = "#FE9A2E"
                else:
                     color = "#CCFB5D"   

                
                form = cargaDoc(request.POST,request.FILES)
                
                if form.is_valid():
                
                    cab = form.save(commit=False)
                    cab.name = str(title)
                    cab.start = inicio
                    cab.end = fin
                    cab.desc_evento = str(desc)
                    cab.color = str(color)
                    cab.ruta = root
                    cab.transportista = str(transportista)
                    cab.conteo_disponible=asientos_d
                    cab.color = color
                    cab.tarifa_evento = tarifa
                    cab.save()

            except:
                asientos_d = 0
                if form.is_valid():
                    cab = form.save(commit=False)
                    cab.name = str(title)
                    cab.start = inicio
                    cab.end = fin
                    cab.desc_evento = str(desc)
                    cab.color = str(color)
                    cab.ruta = root
                    cab.transportista = str(transportista)
                    cab.conteo_disponible=asientos_d
                    cab.color = color
                    cab.tarifa_evento = tarifa
                    cab.save()

                            
        data = {
            'message': 'Salio por json response'
        }

        return JsonResponse(data)

@login_required
def update_event(request):
    start = request.GET.get("start", None)
    end = request.GET.get("end", None)
    title = request.GET.get("title",None)
    id = request.GET.get("id",None)
    allday = request.GET.get("allday",None)
    event = Eventos.objects.get(id=id)
    # if allday:
    #     print("Entra")
    #     new_start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
    #     new_end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
    #     # start = datetime.strftime(new_start + timedelta(hours=5),'%Y-%m-%d %H:%M:%S')
    #     # end = datetime.strftime(new_end + timedelta(hours=5),'%Y-%m-%d %H:%M:%S')

    event.start = start
    event.end = end
    event.name = title
    event.save()

    if allday:
        print("Entra")
        # new_start = datetime.strptime(start, '%Y-%m-%d %H:%M:%S')
        # new_end = datetime.strptime(end, '%Y-%m-%d %H:%M:%S')
        # start = datetime.strftime(new_start + timedelta(hours=5),'%Y-%m-%d %H:%M:%S')
        # end = datetime.strftime(new_end + timedelta(hours=5),'%Y-%m-%d %H:%M:%S')
        time_change = timedelta(hours=5)
        Eventos.objects.filter(id=id).update(start=F('start')+time_change)
        Eventos.objects.filter(id=id).update(end=F('end')+time_change)
        
    data = {}
    return JsonResponse(data)

@login_required
def remove_event(request):
    
    id = request.GET.get("id",None)
    event = Eventos.objects.get(id=id)
    event.delete()
    data = {}
    return JsonResponse(data)


@login_required
def listarSolicitudesTransporte(request):
    queryset = request.GET.get("buscar")
    if queryset:
        query = DetSolicitudTransporte.objects.all().filter(
            Q(numpedido__numpedido = queryset)).filter(numpedido__usuario_solicita = request.user.id).exclude(numpedido__anula=True)
    else:
        query = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__usuario_solicita = request.user.id).exclude(
            numpedido__anula=True).order_by('-numpedido__fchsolicita')
        
        # query = Booking.objects.filter(usuario_solicita=request.user.id).order_by('-fchsolicita')
    

    paginator = Paginator(query, 50)
    page = request.GET.get('page')
    solic = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/listar_solicitudes_transporte.html', {'form':solic,'busqueda':queryset})


@login_required
def listarParaAprobarTransporte(request):
    
    usu_depend = Autorizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(autorizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)

    solic = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__usuario_solicita__in=id_usu).filter(
        Q(numpedido__estado1n=1)).exclude(
            numpedido__anula=True).exclude(numpedido__evento_de_calendario=True).filter(numpedido__estado1n=1).order_by('-numpedido__fchsolicita')
    
    query = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__usuario_solicita__in=id_usu).filter(
        Q(numpedido__estado1n=2) | Q(numpedido__estado1n=3)).exclude(
            numpedido__anula=True).exclude(numpedido__evento_de_calendario=True).order_by('-numpedido__fchsolicita')

    paginator = Paginator(query, 50)
    page = request.GET.get('page')
    query2 = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/listar_solicitudes_transporte_1n.html', {'form':solic,'procesadas':query2})

@login_required
def autorizaTransporte1n(request,numpedido):
    if request.method == 'GET':
        pedido = DetSolicitudTransporte.objects.select_related('numpedido').get(numpedido__numpedido=numpedido)

        return render(request, 'procesosAdministrativos/ver_solicitudestransporte1n.html', {'solic':pedido})
    else:
    
        fecha = datetime.today()
        if request.POST.get('notas_rechazo'):
            actualiza_rechazo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                estado1n=3,fchaprueba1n=fecha, observaciones_rechazo1n=str(request.POST.get('notas_rechazo')))
        else:
            actualiza_campo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                estado1n=2,estado2n=1,fchaprueba1n=fecha, usuario_aprueba1n_id=request.user.id)

            queryset = SolicitudTransporte.objects.get(numpedido=numpedido)
            emailconta = 'contabilidad_sis@ecofroz.com'

            if queryset.tipo == 3:   #Vale de Combustible
                envioMail('Solicitud Autorizada de Vale de Combustible', emailconta, 'email_generado_vale.html', queryset, '')

            else:

                emailaprobb = User.objects.get(rol='GTR')  #Gestor de Transporte
                emailaprob = emailaprobb.email
                
                # copia = 'mantenimiento@ecofroz.com'

                ## Sección para Notificaciones Globales
                query = SolicitudTransporte.objects.get(numpedido=numpedido)

                id_usuario = query.usuario_solicita_id
                    
                id_genera = User.objects.get(id=id_usuario)
                # emailgenera =  Generador.objects.get(id=id_genera.generador_id)
                id_aprueba = User.objects.get(rol='GTR')  #ROL = GESTOR DE TRANSPORTE

                ##Guarda cambio de estado en tabla de mensajeria
                noti = tipo_notificacion.objects.get(id=25) # 25. TRANSPORTE AUTORIZACION GERENTE AREA

                app_number = modelo_App.objects.get(id=6) # 6. PROCESOS ADMINISTRATIVOS 

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = query.numpedido,
                    tipo = noti, 
                    usuario_activa = id_genera,
                    autorizador_id = id_aprueba,
                    )

                r.save()


                envioMail('Solicitud Autorizada de Transporte', emailaprob, 'email_generado_transporte.html', queryset, '')
            #     #envioMail('Solicitud Autorizada de Orden de Trabajo Interna', copia, 'email_generado.html', queryset, '')

        return redirect('procesosadministrativos:listar_aprueba_1n')


@login_required
def autorizaVales(request,numpedido):
    if request.method == 'GET':
        pedido = DetSolicitudTransporte.objects.select_related('numpedido').get(numpedido__numpedido=numpedido)

        return render(request, 'procesosAdministrativos/ver_solicitudesvales.html', {'solic':pedido})
    else:
        confirma = request.POST.get("confirmav")
        fecha = datetime.today()
        numero = request.POST.get("numvale")
        valor = request.POST.get("valorvale")

        if confirma == "registro":
            actualiza_campo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                estado_vale=1,fecha_gestiona_vale=fecha,
                 usu_gestiona_vale=request.user.username,observaciones_vale=request.POST.get("notas_registro"),
                 num_vale=numero,valor_vale=valor)
            query = SolicitudTransporte.objects.get(numpedido=numpedido)
            usuario_solic = query.usuario_solicita.email
            envioMail('Vale de Combustible Generado', usuario_solic, 'vale_generado.html', query, '')

        else:
        
            if request.POST.get('notas_rechazo'):
                actualiza_rechazo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                    estado_vale=2,fecha_gestiona_vale=fecha,usu_gestiona_vale=request.user.id, observaciones_vale=str(request.POST.get('notas_rechazo')))
            else:
                actualiza_campo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                    estado_vale=2,fecha_gestiona_vale=fecha, usu_gestiona_vale=request.user.id)

            # queryset = SolicitudTransporte.objects.get(numpedido=numpedido)
            # emailaprobb = User.objects.get(rol='GTR')
            # emailaprob = emailaprobb.email
            
            # copia = 'mantenimiento@ecofroz.com'

            ## Sección para Notificaciones Globales
            
                
            # id_genera = User.objects.get(id=id_usuario)
            # emailgenera =  Generador.objects.get(id=id_genera.generador_id)
            # id_aprueba = User.objects.get(rol='GTR')  #ROL = GESTOR DE TRANSPORTE

            ##Guarda cambio de estado en tabla de mensajeria
            # noti = tipo_notificacion.objects.get(id=25) # 25. TRANSPORTE AUTORIZACION GERENTE AREA

            # app_number = modelo_App.objects.get(id=6) # 6. PROCESOS ADMINISTRATIVOS 

            # r = notificaciones_globales(
            #     app_origen = app_number,
            #     estado = True,
            #     identificador = query.numpedido,
            #     tipo = noti, 
            #     usuario_activa = id_genera,
            #     autorizador_id = id_aprueba,
            #     )

            # r.save()


            
        #     #envioMail('Solicitud Autorizada de Orden de Trabajo Interna', copia, 'email_generado.html', queryset, '')

        return redirect('procesosadministrativos:gestiona_vales')





@login_required
def imprimeAutorizacionSalida(request,numpedido):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        reg = SolicitudTransporte.objects.get(numpedido=numpedido)
        autorizador = reg.usuario_aprueba1n
        pasajero = reg.pasajero
        otros = reg.otros_pasajeros
        # seguridad_fisica = User.objects.get(username=reg.pers_autoriza_seguridad)
        

        num = str(reg.numpedido)
        fec = str(reg.fchsolicita)
        dep = str(reg.departamento)
        ruta = str(reg.ruta.origen+" => "+reg.ruta.destino)
        otra_ruta = str(reg.otra_ruta)
        
        w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        # p.setFont('3 of 9 Barcode', 32)
        p.setFont('Arial', 18)
        p.drawString(540, 800, num)
        p.setFont('Arial', 16)
        p.drawString(150, 780, "SOLICITUD PERMISO DE SALIDA (ECRH-R-8-1)")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Nombre:")
        p.drawString(110,730, pasajero)
        p.drawString(270,730, "Código:")
        p.drawString(440,730, "Fecha:")
        p.drawString(485,730, fec[:10])
        p.drawString(50,700, "Otro(s) Solitante(s):")
        p.drawString(160,700, otros )
    
        p.drawString(50,670, "Código(s):")
        p.drawString(50,640, "Departamento:")
        p.drawString(135,640, dep)
        p.drawString(270,640, "Area:")
        p.drawString(440,640, "Turno:")
        p.drawString(50,610, "Ruta:")
        if otra_ruta:
            p.drawString(130,610, otra_ruta)
        else:
            p.drawString(130,610, ruta)
        
        p.drawString(50,580, "Horas:")
        p.drawString(180,580, "Hora de Salida:")
        p.drawString(350,580, "Hora de Entrada:")
        p.drawString(50,550, "Observaciones/Notas:")
        
        p.drawString(240,450, "Persona Autoriza Area")
        p.drawString(240,435, str(autorizador))

        # p.drawString(320,535, "Firma del Empleado")
        # p.drawString(310,550, "--------------------------------------")

        p.setFont('Arial', 8)
        p.drawString(0,385, "-----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------")
        
        
        p.setFont('Arial', 11)

        vacio = ""

        # p.drawString(110,730, pasajero)
        # p.drawString(250,710, 'codigo aqui')
        # p.drawString(485,730, fec[:10])
        # p.drawString(90,670, ruta)
        # p.drawString(135,700, dep)
        p.drawString(250,650, vacio)
        p.drawString(250,630, vacio)
        p.drawString(250,610, vacio)
        p.drawString(250,580, vacio)
        p.drawString(250,560, vacio)
    
        p.drawString(250,520, vacio)
        
        
        #DUPLICADO
        escala = 455

        p.setFont('Arial', 18)
        p.drawString(540, 800-escala+10, num)
        p.setFont('Arial', 11)
        p.drawString(35, 800-escala+10, "--copia--")
        p.setFont('Arial', 16)
        p.drawString(150, 780-escala, "SOLICITUD PERMISO DE SALIDA (ECRH-R-8-1)")
        p.setFont('Arial', 12)
        p.drawString(50,730-escala, "Nombre:")
        p.drawString(110,730-escala, pasajero)
        p.drawString(270,730-escala, "Código:")
        p.drawString(440,730-escala, "Fecha:")
        p.drawString(485,730-escala, fec[:10])
        p.drawString(50,700-escala, "Otro(s) Solitante(s):")
        p.drawString(160,700-escala, otros )
        p.drawString(50,670-escala, "Departamento:")
        p.drawString(135,670-escala, dep)
        p.drawString(270,670-escala, "Area:")
        p.drawString(440,670-escala, "Turno:")
        p.drawString(50,640-escala, "Código(s):")
        
        
        p.drawString(50,610-escala, "Ruta:")
        if otra_ruta:
            p.drawString(90,610-escala, otra_ruta)
        else:
            p.drawString(90,610-escala, ruta)

        p.drawString(50,580-escala, "Horas:")
        p.drawString(180,580-escala, "Hora de Salida:")
        p.drawString(350,580-escala, "Hora de Entrada:")
        p.drawString(50,550-escala, "Observaciones/Notas:")
        
        p.drawString(240,500-escala, "Persona Autoriza Area")
        p.drawString(240,475-escala, str(autorizador))

        # p.drawString(320,535-escala, "Firma del Empleado")
        # p.drawString(310,550-escala, "--------------------------------------")

       

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="salida.pdf")

@login_required
def verSolicitudes(request,numpedido):
    if request.method == 'GET':
        instancia = SolicitudTransporte.objects.get(numpedido=numpedido)
        detalle = DetSolicitudTransporte.objects.get(numpedido=instancia.numpedido)
        # sectordj = instancia.area
        form = cabIngresoSolicitudTransporte(instance=instancia)
        form2 = detIngresoSolicitudTransporte(instance=detalle, files=request.FILES )
        return render(request, 'procesosAdministrativos/ver_solicitudes0.html', {'form':form, 'form2':form2,'cab':instancia})


@login_required
def editaSolicitudes(request,numpedido):
    if request.method == 'GET':
        instancia = SolicitudTransporte.objects.get(numpedido=numpedido)
        detalle = DetSolicitudTransporte.objects.get(numpedido=instancia.numpedido)
        # sectordj = instancia.area
        form = cabIngresoSolicitudTransporte(instance=instancia)
        form2 = detIngresoSolicitudTransporte(instance=detalle, files=request.FILES )
        return render(request, 'procesosadministrativos/edita_solicitudes0.html', {'form':form, 'form2':form2})

    
    if request.method == 'POST':
        instancia = SolicitudTrabajoInterno.objects.get(numtrabajo=numtrabajo)
        detalle = DetSolicitudTrabajoInterno.objects.get(numtrabajo=instancia.numtrabajo)
    
        form = cabIngresoTrabajos(request.POST, instance=instancia)
        form2 = detIngresoTrab(request.POST, request.FILES, instance=detalle)

        sectorantiguo = instancia.area
        fecha = datetime.today()
        
        if request.POST.get('nsector'):
            print(request.POST.get('nsector'))
            if form.is_valid() and form2.is_valid():
                cab = form.save(commit=False)
                cab.fchsolicita = fecha
                cab.area = str(request.POST.get('nsector'))
                cab.estado1n = 1
                cab.observaciones_rechazo1n = None
                det = form2.save(commit=False)
                det.numtrabajo = cab
                det.save()
                cab.save()
                
            else:
                return HttpResponse(form.errors,form2.errors)
                
        else:
            print("Sector Nulo")
            if form2.is_valid():
                print("Llega")
                cab = form.save(commit=False)
                cab.area = sectorantiguo
                cab.fchsolicita = fecha
                cab.estado1n = 1
                cab.observaciones_rechazo1n = None
                det = form2.save(commit=False)
                det.numtrabajo = cab
                det.save()
                cab.save()
            else:
                return HttpResponse(form.errors, form2.errors)


        return redirect('trabajosinternos:listar_trabajos_internos')


@login_required
def anulaSolicitudes(request, numpedido):
    if request.method == 'GET':
        anular = SolicitudTransporte.objects.filter(numpedido=numpedido).update(anula=True)

        return redirect('procesosadministrativos:listar_solicitudes_transporte')


@login_required
def gestionaTransporte2n(request):
    queryset = request.GET.get("buscar")
    if queryset:
        
        q1 = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado1n=2).exclude(
                numpedido__anula=True).exclude(Q(numpedido__estado2n=2) | Q(numpedido__estado2n=3) | Q(numpedido__estado2n=4)).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')

        query_nuevas = q1.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(numpedido__usuario_solicita__last_name__icontains = queryset) |
            Q(numpedido__usuario_solicita__first_name__icontains = queryset) |
            Q(numpedido__transportista__icontains = queryset))
        
        q2 = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado2n=2).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
        
        query_proceso = q2.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(numpedido__usuario_solicita__last_name__icontains = queryset) |
            Q(numpedido__usuario_solicita__first_name__icontains = queryset) |
            Q(numpedido__transportista__icontains = queryset))
        
        q3 = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado2n=4).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
          
        query_confirmadas = q3.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(numpedido__usuario_solicita__last_name__icontains = queryset) |
            Q(numpedido__usuario_solicita__first_name__icontains = queryset) |
            Q(numpedido__transportista__icontains = queryset))
        
        countn = q1.count()
        countp = q2.count()
        countc = q3.count()
          
    else:
       
        # queryset = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado1n=2).filter(
        #     Q(numpedido__estado2n=2) |
        #     Q(numpedido__estado2n=3) |
        #     Q(numpedido__estado2n=4)).exclude(
        #     numpedido__anula=True).order_by('-numpedido__fchaprueba1n')
    
        queryset = None
        query_nuevas = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado1n=2).exclude(
                numpedido__anula=True).exclude(Q(numpedido__estado2n=2) | Q(numpedido__estado2n=3) | Q(numpedido__estado2n=4)).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
            
        query_proceso = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado2n=2).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
            
        query_confirmadas = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado2n=4).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')[:100]
        query_confirmadast = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__estado2n=4).exclude(numpedido__tipo=3).order_by('-numpedido__fchaprueba1n')
        
        countn = query_nuevas.count()
        countp = query_proceso.count()
        countc = query_confirmadast.count()  

        
    return render(request, 'procesosAdministrativos/gestionar_solicitudes_transporte_2n.html', {'form2':query_nuevas, 'form3':query_proceso,'form4':query_confirmadas,'busqueda':queryset,'countn':countn,'countp':countp,'countc':countc})


@login_required        
def gestionaSolicitudesTransporte(request, numpedido):
    if request.method == 'GET':

        print(numpedido)
        
        solicitud = DetSolicitudTransporte.objects.filter(numpedido=numpedido).select_related('numpedido')
        print(solicitud)
        solic = SolicitudTransporte.objects.get(numpedido=numpedido)
       
        num = SolicitudTransporte.objects.get(numpedido=numpedido)
        numt = num.numpedido
        notas = num.notas_del_gestor
        transportista = num.transportista
        valor = num.tarifa_real
        factura = num.num_factura
        fecha_factura = num.fecha_factura
        retorno = num.retorno

        return render(request, 'procesosAdministrativos/gestiona_solicitudes_transporte.html', { 'cot': solicitud,'solic':solic, 'numt':numt,'notas':notas,'valor_real':valor,'transportista':transportista,'factura':factura,'fecha_factura':fecha_factura,'retorno':retorno})


@login_required
def devuelve2n(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('notas')
        fecha = datetime.today()

        actualiza = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
            estado2n=3, estado1n=1, fecha_confirma=fecha, observaciones_rechazo2n=observa)

        queryset = DetSolicitudTransporte.objects.get(numpedido=numpedido)
        emailaprob = queryset.numpedido.usuario_solicita.email

        envioMail('Devolución de Solicitud de Transporte', emailaprob, 'email_devolucion.html', queryset, '')
        
        return redirect('procesosadministrativos:listar_gestiona_2n')




@login_required
def ajax_cambia_enproceso(request):
    numpe = request.GET.get("sol_id")
    notas = request.GET.get("notasg")

    print(numpe)
    print(notas)

    fecha_hoy = datetime.today()

    comprueba = SolicitudTransporte.objects.get(numpedido=numpe)
    if comprueba.fecha_confirma:
        actualiza = SolicitudTransporte.objects.filter(numpedido=numpe).update(estado2n=2,fecha_en_proceso=fecha_hoy,notas_del_gestor=notas,fecha_confirma=None)

    else:
        actualiza = SolicitudTransporte.objects.filter(numpedido=numpe).update(estado2n=2,fecha_en_proceso=fecha_hoy,notas_del_gestor=notas)

    return JsonResponse ("Exito", safe=False)

@login_required
def ajax_cambia_confirma(request):
    numpe = request.GET.get("sol_id")
    notas = request.GET.get("notasg")

    print(numpe)
    print(notas)

    fecha_hoy = datetime.today()

    actualiza = SolicitudTransporte.objects.filter(numpedido=numpe).update(estado2n=4,fecha_confirma=fecha_hoy,notas_del_gestor=notas)

    query = SolicitudTransporte.objects.get(numpedido=numpe)
    usuario_solic = query.usuario_solicita.email

    envioMail('Confirmación de Transporte', usuario_solic, 'email_confirmado_transporte.html', query, '')

    return JsonResponse ("Exito", safe=False)

@login_required
def ajax_save(request):
    numpe = request.GET.get("sol_id")
    notas = request.GET.get("notasg")
    transpo = request.GET.get("transportista")
    # valor = request.GET.get("valor")
    # factura = int(request.GET.get("factura"))
    # fec_factura = request.GET.get("fec_factura")

    try:
        valor = int(request.GET.get("valor"))
    except:
        valor = None
    try:
        factura = int(request.GET.get("factura"))
    except:
        factura = None
    try:
        fec_factura = request.GET.get("fec_factura")
    except:
        fec_factura = None
    try:
        retorno = request.GET.get("val_retorno")
        if retorno == 'SI':
            retorno = True
        elif retorno == 'NO':
            retorno = False
        else:
            retorno = None
    except:
        retorno = None

    # AQUI ME QUEDE 
    actualiza = SolicitudTransporte.objects.filter(numpedido=numpe).update(notas_del_gestor=notas,transportista=transpo,
    tarifa_real=valor, num_factura=factura,fecha_factura=fec_factura,retorno=retorno)

    return JsonResponse ("Exito", safe=False)

@login_required
def gestionaVales(request):
    queryset = request.GET.get("buscar")
    if queryset:
        query = DetSolicitudTransporte.objects.filter(
            Q(numpedido__numpedido = queryset) |
            Q(numpedido__descripcion__icontains = queryset)).filter(numpedido__usuario_solicita = request.user.id).exclude(numpedido__anula=True)
    else:
        #TIPO 3 = VALE DE COMBUSTIBLE
        query = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__tipo = 3).filter(
            numpedido__estado1n=2).exclude(numpedido__estado_vale__in=[1,2]).exclude(numpedido__anula=True).order_by('-numpedido__fchsolicita')
        
        query_confirmadas = DetSolicitudTransporte.objects.select_related('numpedido').filter(numpedido__tipo = 3).filter(
            numpedido__estado_vale__in=[1,2]).exclude(numpedido__anula=True).order_by('-numpedido__fchsolicita')
    

    paginator = Paginator(query_confirmadas, 50)
    page = request.GET.get('page')
    solic = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/gestionar_vales.html', {'form':query,'procesadas':solic,'busqueda':queryset})
   



@login_required
def nuevaSolicitud(request):
    if request.method == "GET":
        # form = ingresoTrabajos(request.GET or None)
        
        form = cabIngresoSolicitudTransporte(request.GET or None)
        form2 = detIngresoSolicitudTransporte(request.GET or None)

        return render(request, 'procesosAdministrativos/ingreso_solicitud_transporte.html', {'form':form, 'form2':form2})
    else:
        rutat = request.POST.get("rutatexto")
        fec = request.POST.get("fechahora")
        fecha_r = fec.replace("T"," ")
        if fecha_r:
            fecha_evento = datetime.strptime(fecha_r, '%Y-%m-%d %H:%M')
            fecha_evento = datetime.strftime(fecha_evento, '%Y-%m-%d %H:%M') 
        else:
            fecha_evento = date.today()
       
        otra_ruta = request.POST.get("input-otra-ruta")
        otros_passengers = request.POST.get("input-nombre-otros-pasajeros")
        beneficiario = request.POST.get("input-nombre-beneficio-vale")
        regreso = request.POST.get("retorno")
        if regreso == 'on':
            regreso = True
        else:
            regreso = False

        if otra_ruta:
            
            otra_ruta = request.POST.get("input-otra-ruta")

            form = cabIngresoSolicitudTransporte(request.POST)
            form2 = detIngresoSolicitudTransporte(request.POST, request.FILES)
      
            if form.is_valid() and form2.is_valid():  
                cab = form.save(commit=False)
                cab.usuario_solicita_id = request.user.id
                cab.otra_ruta = otra_ruta
                cab.pasajero = request.POST.get("input-nombre-pasajero")
                cab.otros_pasajeros = otros_passengers
                cab.beneficiario_vale = beneficiario
                cab.retorno = regreso
                cab.fchevento = fecha_evento

                cab.save()

                fecha = datetime.today()
                fec = datetime.strftime(fecha, '%Y-%m-%d %H:%M:%S.%fZ') 
                
                actualiza = SolicitudTransporte.objects.filter(numpedido=cab.numpedido).update(fchsolicita=fecha,estado1n=1)
                
                det = form2.save(commit=False)
                det.numpedido = cab
                
                form2.save()
                
                queryset = DetSolicitudTransporte.objects.filter(numpedido=cab.numpedido)
                id_usuario = User.objects.get(id=request.user.id)
                # aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
                # id_aprobador = User.objects.get(id=aprobador.user_id.id)

                emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
                # autoriza_instancia = Autorizador.objects.get(id=emailaprob.autoriza_oti_id)
                # autoriza_user_id = autoriza_instancia.user_id.id
                # autoriza_user = User.objects.get(id=autoriza_user_id)
                

                # solicitud = DetSolicitudTrabajoInterno.objects.get(numpedido=cab.numpedido)
                ## SECCION DE NOTIFICACIONES GLOBALES

            #     ##Guarda cambio de estado en tabla de mensajeria
            #     noti = tipo_notificacion.objects.get(id=14) # 14. TRABAJOS INTERNOS SOLICITUD

            #     app_number = modelo_App.objects.get(id=3) # 3. trabajosInternos

            #     r = notificaciones_globales(
            #         app_origen = app_number,
            #         estado = True,
            #         identificador = solicitud.numtrabajo.numtrabajo,
            #         tipo = noti, 
            #         usuario_activa = emailaprob,
            #         autorizador_id = autoriza_user,
            #         )

            #     r.save()

        
                envioMail('Solicitud Autorización de Transporte / Vales Combustible', emailaprob.e_mail, 'email_aprobacion_gerente.html', queryset, '')

                return redirect('procesosadministrativos:listar_solicitudes_transporte')
            else:
                print(form.errors)
                print(form2.errors)
                return redirect('procesosadministrativos:listar_solicitudes_transporte')
                
                # return HttpResponse("errores")

        else:

            form = cabIngresoSolicitudTransporte(request.POST)
            form2 = detIngresoSolicitudTransporte(request.POST, request.FILES)
        

            if form.is_valid() and form2.is_valid():
                cab = form.save(commit=False)
                cab.usuario_solicita_id = request.user.id
                cab.pasajero = request.POST.get("input-nombre-pasajero")
                cab.otros_pasajeros = otros_passengers
                cab.beneficiario_vale = beneficiario
                cab.retorno = regreso
                cab.fchevento = fecha_evento
                cab.save()

                fecha = datetime.today()
                fec = datetime.strftime(fecha, '%Y-%m-%d %H:%M:%S.%fZ') 
                
                actualiza = SolicitudTransporte.objects.filter(numpedido=cab.numpedido).update(fchsolicita=fecha,estado1n=1)
                
                det = form2.save(commit=False)
                det.numpedido = cab
                
                form2.save()
                
                queryset = DetSolicitudTransporte.objects.filter(numpedido=cab.numpedido)
                id_usuario = User.objects.get(id=request.user.id)
                # aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
                # id_aprobador = User.objects.get(id=aprobador.user_id.id)

                emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
                print(emailaprob.e_mail)
                # autoriza_instancia = Autorizador.objects.get(id=emailaprob.autoriza_oti_id)
                # autoriza_user_id = autoriza_instancia.user_id.id
                # autoriza_user = User.objects.get(id=autoriza_user_id)
                

                # solicitud = DetSolicitudTrabajoInterno.objects.get(numpedido=cab.numpedido)
                ## SECCION DE NOTIFICACIONES GLOBALES

            #     ##Guarda cambio de estado en tabla de mensajeria
            #     noti = tipo_notificacion.objects.get(id=14) # 14. TRABAJOS INTERNOS SOLICITUD

            #     app_number = modelo_App.objects.get(id=3) # 3. trabajosInternos

            #     r = notificaciones_globales(
            #         app_origen = app_number,
            #         estado = True,
            #         identificador = solicitud.numtrabajo.numtrabajo,
            #         tipo = noti, 
            #         usuario_activa = emailaprob,
            #         autorizador_id = autoriza_user,
            #         )

            #     r.save()

        
                envioMail('Solicitud Autorización de Transporte/Vales Combustible', emailaprob.e_mail, 'email_aprobacion_gerente.html', queryset, '')

                return redirect('procesosadministrativos:listar_solicitudes_transporte')
            else:
            
                print(form.errors)
                print(form2.errors)
                return redirect('procesosadministrativos:listar_solicitudes_transporte')
                # return HttpResponse("errores")

@login_required
def nuevaRuta(request):
    codigo = request.GET.get("codigo_ruta")
    ori = request.GET.get("origen")
    dest = request.GET.get("destino")
    tar = request.GET.get("tarifa")
    comentarios = request.GET.get("comentarios")

    try:

        for i in range(1):
            r = RutaTransporte(
                
                codigo_ruta = codigo,
                origen = ori,
                destino = dest,
                tarifa = tar,
                comentarios = comentarios,
                activo = True,
                fecha_creacion = datetime.today(),
                fecha_modifica = datetime.today(),
                persona_edita = request.user.username,
                )
            r.save()
    
    except:
        for i in range(1):
            r = RutaTransporte(
                
                codigo_ruta = codigo,
                origen = ori,
                destino = dest,
                tarifa = tar,
                activo = True,
                fecha_creacion = datetime.today(),
                fecha_modifica = datetime.today(),
                persona_edita = request.user.username,
                )
            r.save()
        
    return redirect ('parametrosglobales:parametros_rutas')

@login_required
def actualizaEstadoRuta(request):
    if request.method == 'POST':
        valor_radio = request.POST.get("radio")
        id = request.POST.get("rutaid")
        fecha = datetime.today()
        codigo = request.POST.get("codigo_ruta2")
        origen = request.POST.get("origen2")
        destino = request.POST.get("destino2")
        tarifa = request.POST.get("tarifa2")
        comentarios = request.POST.get("comentarios2")
        print(comentarios)


        if valor_radio == "SI":
            print("Ya!")
            actualiza = RutaTransporte.objects.filter(id=id).update(activo=True,fecha_modifica=fecha,persona_edita=request.user.username,
            codigo_ruta=codigo,origen=origen,destino=destino,tarifa=tarifa,comentarios=comentarios)
        else:
            print("No!")
            actualiza = RutaTransporte.objects.filter(id=id).update(activo=False,fecha_modifica=fecha,persona_edita=request.user.username,
            codigo_ruta=codigo,origen=origen,destino=destino,tarifa=tarifa,comentarios=comentarios)
        
    return redirect ('parametrosglobales:parametros_rutas')


@login_required
def verificaSalidaAutotizada(request):
    if request.method == 'GET':
        try:
            codigo = request.GET.get('numpedido')
            print(codigo)
            
            if codigo != 'None':
                print("Si entra")
                
                queryset = SolicitudTransporte.objects.get(numpedido=int(codigo))
                print(queryset)

                return render(request, 'procesosAdministrativos/verifica_salida_personal.html',{'form':queryset, 'codigo':codigo})
            else:
                consulta = 1
                return render(request, 'procesosAdministrativos/verifica_salida_personal.html',{'consulta':consulta})
        except:
            return render(request, 'procesosAdministrativos/verifica_salida_personal.html') 



@login_required
def ajax_call_ruta(request):
    id = request.GET["id"]
    ruta = RutaTransporte.objects.filter(id=id) 
    
    context = []
    for i in ruta:
        d = {
            'codigo_ruta':i.codigo_ruta,
            'origen':i.origen,
            'destino':i.destino,
            'tarifa':i.tarifa,
            'comentarios':i.comentarios,
            'estado':i.activo,
            'id':i.id,
        }
        context.append(d)

    return JsonResponse (context, safe=False)

@login_required
def ajax_query_passengers(request):
    id = request.GET["id"]
    print(id)
    pasajeros = Booking.objects.filter(evento_id=id).order_by('fchsolicita')
    
    context = []
    for i in pasajeros:
        d = {
            'nombre':i.nombre_pasajero,
            'departamento':i.departamento,
            'fecha_book':datetime.strftime(i.fchsolicita, '%Y-%m-%d %H:%M') 
        }
        context.append(d)

    return JsonResponse (context, safe=False)



@login_required
def ajax_save_booking(request):
    id_evento = request.GET.get("evento_id")
    event = Eventos.objects.get(id=id_evento)

    try:
        with transaction.atomic():

            actualiza = Eventos.objects.filter(id=id_evento).update(conteo_disponible=event.conteo_disponible-1)
        
            solicita = request.user.username
            solip = request.user.id
            depar = request.GET.get("depar")
            pasajero = request.GET.get("pasajero")
            justifica = request.GET.get("justifica")
            tipo = request.GET.get("tipo")
            
            usuario = User.objects.get(username=str(solicita))

            context = []
            
            fechas = datetime.today()
            ubi = activo_ubica.objects.get(id=4)
            dep = activo_depar.objects.get(dep_nombre=str(depar))
            solipe = User.objects.get(id=solip)

            for x in range(1):
                r = Booking(
                    evento = event,
                    departamento = depar,
                    nombre_pasajero = pasajero,
                    justificacion = justifica,
                    fchsolicita = fechas,
                    usuario_solicita = usuario,
                    estado1n = 5,  #Booked
                    tipo = tipo,
                    )
                r.save()
            
            if usuario.rol != 'GTR':  #Se usa para que no lleguen solicitudes de transporte de Bookings hechos por Gestor de Transporte al 15-11-2023: Lesly Perez: GTR es el rol Gestor de Transporte

                for x in range(1):
                    s = SolicitudTransporte(
                        tipo = 1,
                        estado1n = 1,
                        fchsolicita = fechas,
                        justificacion = justifica,
                        departamento = dep,
                        ubica = ubi,
                        usuario_solicita = solipe,
                        ruta = event.ruta,
                        fchevento = event.start,
                        pasajero = pasajero,
                        )
                    s.save()

                q = SolicitudTransporte.objects.latest('numpedido')

                for x in range(1):
                    d = DetSolicitudTransporte(
                        numpedido = q
                    )
                    d.save()
                    queryset = DetSolicitudTransporte.objects.filter(numpedido=q.numpedido)
                    id_usuario = User.objects.get(id=request.user.id)
                    # aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
                    # id_aprobador = User.objects.get(id=aprobador.user_id.id)

                    emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
                
                    # autoriza_instancia = Autorizador.objects.get(id=emailaprob.autoriza_oti_id)
                    # autoriza_user_id = autoriza_instancia.user_id.id
                    # autoriza_user = User.objects.get(id=autoriza_user_id)
                    

                    # solicitud = DetSolicitudTrabajoInterno.objects.get(numpedido=cab.numpedido)
                    ## SECCION DE NOTIFICACIONES GLOBALES

                #     ##Guarda cambio de estado en tabla de mensajeria
                #     noti = tipo_notificacion.objects.get(id=14) # 14. TRABAJOS INTERNOS SOLICITUD

                #     app_number = modelo_App.objects.get(id=3) # 3. trabajosInternos

                #     r = notificaciones_globales(
                #         app_origen = app_number,
                #         estado = True,
                #         identificador = solicitud.numtrabajo.numtrabajo,
                #         tipo = noti, 
                #         usuario_activa = emailaprob,
                #         autorizador_id = autoriza_user,
                #         )

                #     r.save()

            
                    envioMail('Solicitud Autorización de Transporte', emailaprob.e_mail, 'email_aprobacion_gerente.html', queryset, '')
                    
                    print("Va bien")
              
            return JsonResponse("Satisfactorio",safe=False)
    
    except IntegrityError:
        context = []
        handle_exception()

        print("Va por error")
    
    
        return JsonResponse (context, safe=False)





@login_required
def ajax_call_depar(request):
    if 'term' in request.GET:
        dp =  activo_depar.objects.filter(dep_nombre__istartswith=request.GET.get('term')) 
        depar = list()
        for opciones in dp:
            depar.append(opciones.dep_nombre) 
        
        return JsonResponse(depar,safe=False)
    
    return render('procesosadministrativos/ver_rutas_fijas.html')

@login_required
def ajax_call_pasajero(request):
    if 'term' in request.GET:
        us =  User.objects.filter(Q(first_name__istartswith=request.GET.get('term')) | Q(last_name__istartswith=request.GET.get('term')))
        usuario = list()
        for opciones in us:
            usuario.append(opciones.first_name+' '+opciones.last_name) 
        
        return JsonResponse(usuario,safe=False)
    
    return render('procesosadministrativos/ver_rutas_fijas.html')


@login_required
def ajax_call_nombreruta(request):
    if 'term' in request.GET:
        qs =  RutaTransporte.objects.filter(Q(origen__icontains=request.GET.get('term')) | Q(destino__icontains=request.GET.get('term')))
        ruta = list()
        
        for opciones in qs:
            if opciones.id != 42:  #DESCARTAR FORMATO DE LABEL PARA SELECCION DE TERMINO "OTRA RUTA"
                d={
                    'id': opciones.id,
                    'value':opciones.codigo_ruta,
                    'label':opciones.origen+' => '+opciones.destino,
                    'tarifa':opciones.tarifa,
                }
                ruta.append(d)
            else:
                d={
                    'id': opciones.id,
                    'value':opciones.codigo_ruta,
                    'label':opciones.origen,
                    'tarifa':opciones.tarifa,
                }
                ruta.append(d)
            
            print(ruta)

        
        return JsonResponse(ruta,safe=False)
    
    return render('procesosadministrativos/administrar_eventos_transporte.html')
    

@login_required
def ajax_call_nombrepasajero(request):
    if 'term' in request.GET:
        qs =  User.objects.filter(Q(first_name__istartswith=request.GET.get('term')) | Q(last_name__istartswith=request.GET.get('term')))
        pasajero = list()
        for opciones in qs:
            pasajero.append(opciones.first_name+' '+opciones.last_name) 
        
        return JsonResponse(pasajero,safe=False)
    
    return render('procesosadministrativos/administrar_eventos_transporte.html')


@login_required
def ajax_call_nombretransportista(request):
    if 'term' in request.GET:
        qs =  proveedor_det.objects.filter(proveedor_ruta_fija = True).filter(codigo_id__nombre_empresa__icontains=request.GET.get('term')) 
        chofer = list()
        for opciones in qs:
            chofer.append(opciones.codigo_id.nombre_empresa+',') 
        
        return JsonResponse(chofer,safe=False)
    
    return render('procesosadministrativos/administrar_eventos_transporte.html')


@login_required
def reporteTransporte(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
   
    queryset = SolicitudTransporte.objects.filter(estado1n=2).exclude(tipo=3).exclude(anula=True)

    #Trunca la tabla de Reporte Booking
    cursor = connection.cursor()
    cursor.execute("TRUNCATE TABLE procesos.booking_reporte")

    #Consulta la tabla de Booking con registros realizados por persona con rol de Gestor de Transporte
    queryset_booking = Booking.objects.filter(usuario_solicita__rol__cod='GTR')
    
    ubi = activo_ubica.objects.get(id=4) #MACHACHI PLANTA
   
    for x in queryset_booking:

        r = BookingReporte(
            numpedido = x.id,
            ubica = ubi,
            tipo = 1,
            estado1n=2,
            fchevento= x.evento.start,
            departamento = activo_depar.objects.get(dep_nombre=x.departamento),
            pasajero = x.nombre_pasajero,
            ruta = x.evento.ruta,
            justificacion = x.evento.desc_evento,
            evento = x.evento,
            usuario_solicita = x.usuario_solicita,
            fchsolicita = x.fchsolicita,
            num_factura = None,
            fecha_factura = None,
            )
        r.save()
    
    query_booking = BookingReporte.objects.all()

    query = queryset.union(query_booking).order_by("-fchsolicita")
 

    paginator = Paginator(query,50)
    page = request.GET.get('page')
    datos = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    print("Ini:",inicio)
    print("Fin:",fin)
    # ingresos = ''
    mensaje = ''
    consulta = ''

    if inicio != 'None':
        print("Si entra")
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        finmasuno = findate + timedelta(days=1)
        if inicio and fin:
            # ini = iniciodate.strftime('%Y%m%d')
            # fi = findate.strftime('%Y%m%d')
            queryset = SolicitudTransporte.objects.filter(estado1n=2).exclude(tipo=3).exclude(anula=True).filter(
                fchaprueba1n__gte=iniciodate, fchaprueba1n__lt=finmasuno).order_by("-fchaprueba1n")
            
            # queryset = SolicitudTransporte.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE procesos.booking_reporte")
            
            queryset_booking = Booking.objects.filter(usuario_solicita__rol__cod='GTR').filter(
                fchsolicita__gte=iniciodate,fchsolicita__lt=finmasuno)
            
            ubi = activo_ubica.objects.get(id=4) #MACHACHI PLANTA
   
            for x in queryset_booking:

                r = BookingReporte(
                    numpedido = x.id,
                    ubica = ubi,
                    tipo = 1,
                    estado1n=2,
                    fchevento= x.evento.start,
                    departamento = activo_depar.objects.get(dep_nombre=x.departamento),
                    pasajero = x.nombre_pasajero,
                    ruta = x.evento.ruta,
                    justificacion = x.evento.desc_evento,
                    evento = x.evento,
                    usuario_solicita = x.usuario_solicita,
                    fchsolicita = x.fchsolicita
                    )
                r.save()
            
            query_booking = BookingReporte.objects.all()

            query = queryset.union(query_booking).order_by("-fchsolicita")
        
            paginator = Paginator(query,50) 
            page = request.GET.get('page')
            datos = paginator.get_page(page)
            
            consulta = 'SI'

        else:
            mensaje = 'No selecciono las fechas de busqueda'

    return render(request, 'procesosAdministrativos/reporte_transporte.html', {'reg':datos, 'msj':mensaje, 'inicio':inicio, 'fin':fin,'consulta':consulta})


def exportExcelTransporte(request, inicio, fin):

    if inicio != 'None':
    
        inicio = str(inicio)
        fin = str(fin)
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        finmasuno = findate + timedelta(days=1)
        queryset = SolicitudTransporte.objects.filter(estado1n=2).exclude(tipo=3).exclude(anula=True).filter(
                    fchaprueba1n__gte=iniciodate, fchaprueba1n__lt=finmasuno).order_by("-fchaprueba1n")

        query_booking = BookingReporte.objects.all()

        query = queryset.union(query_booking).order_by("-fchsolicita")

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE TRANSPORTE' + '  -  ' + inicio + '  -  ' + fin

        ws.merge_cells('A1:H1')
        ws['A3'] = 'NUMPEDIDO'
        ws['B3'] = 'DEPARTAMENTO'
        ws['C3'] = 'FECHA SOLICITA'
        ws['D3'] = 'FECHA APROBADO'
        ws['E3'] = 'REGISTRA'
        ws['F3'] = 'TIPO'
        ws['G3'] = 'RUTA'
        ws['H3'] = 'RETORNO'
        ws['I3'] = 'APRUEBA'
        ws['J3'] = 'PASAJERO(S)'
        ws['K3'] = 'FECHA EVENTO'
        ws['L3'] = 'TRANSPORTISTA'
        ws['M3'] = 'JUSTIFICACION'
        ws['N3'] = 'TARIFA REFERENCIAL'
        ws['O3'] = 'TARIFA REAL'
        ws['P3'] = 'N° FACTURA'
        ws['Q3'] = 'FECHA FACTURA'


        count = 4
        rowcount = 1  

        for i in query:
            ws.cell(row = count, column = 1).value =  i.numpedido
            ws.cell(row = count, column = 2).value =  str(i.departamento.dep_nombre)
            ws.cell(row = count, column = 3).value =  i.fchsolicita
            ws.cell(row = count, column = 4).value =  i.fchaprueba1n
            ws.cell(row = count, column = 5).value =  str(i.usuario_solicita.username)
            if i.tipo == 1:
                ws.cell(row = count, column = 6).value =  "Transporte personas"
            elif i.tipo == 2:
                ws.cell(row = count, column = 6).value =  "Encomienda"
            try:
                if i.ruta.id == 42:
                    ws.cell(row = count, column = 7).value =  str("OR:"+i.otra_ruta)
                else:
                    ws.cell(row = count, column = 7).value =  str(i.ruta)
            except:
                ws.cell(row = count, column = 7).value =  'NID'

            if i.retorno == True:
                ws.cell(row = count, column = 8).value =  "SI"
            else:
                ws.cell(row = count, column = 8).value =  "NO"

            if i.usuario_aprueba1n:
                ws.cell(row = count, column = 9).value =  str(i.usuario_aprueba1n.username)
            else:
                ws.cell(row = count, column = 9).value =  "--"


            if i.pasajero == '':
                ws.cell(row = count, column = 10).value =  str(i.otros_pasajeros)
            else:
                if i.pasajero and i.otros_pasajeros:
                    ws.cell(row = count, column = 10).value =  str(i.pasajero)+', '+str(i.otros_pasajeros)
                else:
                    ws.cell(row = count, column = 10).value =  str(i.pasajero)

            
            ws.cell(row = count, column = 11).value =  i.fchevento
            
            if i.transportista:
                ws.cell(row = count, column = 12).value =  str(i.transportista)
            else:
                ws.cell(row = count, column = 12).value = "--"
            
            ws.cell(row = count, column = 13).value =  str(i.justificacion)

            try:
                if i.ruta.id == 42:
                    ws.cell(row = count, column = 14).value =  "--"
                else:
                    ws.cell(row = count, column = 14).value =  i.ruta.tarifa
            except:
                ws.cell(row = count, column = 14).value =  "NID"

            ws.cell(row = count, column = 15).value =  i.tarifa_real
            ws.cell(row = count, column = 16).value =  i.num_factura
            ws.cell(row = count, column = 17).value =  i.fecha_factura
              
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE TRANSPORTE" + inicio + "_" + fin + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
        

        mensaje = ''

        if inicio != 'None':
            iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
            findate = datetime.strptime(fin, '%Y-%m-%d')
            finmasuno = findate + timedelta(days=1)
            if inicio and fin:
            
                queryset = SolicitudTransporte.objects.filter(estado1n=2).exclude(tipo=3).exclude(anula=True).filter(
                    fchaprueba1n__gte=iniciodate, fchaprueba1n__lt=finmasuno).order_by("-fchaprueba1n")
                
                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE procesos.booking_reporte")
                
                queryset_booking = Booking.objects.filter(usuario_solicita__rol__cod='GTR').filter(
                    fchsolicita__gte=iniciodate,fchsolicita__lt=finmasuno)
                
                ubi = activo_ubica.objects.get(id=4) #MACHACHI PLANTA
    
                for x in queryset_booking:

                    r = BookingReporte(
                        numpedido = x.id,
                        ubica = ubi,
                        tipo = 1,
                        estado1n=2,
                        fchevento= x.evento.start,
                        departamento = activo_depar.objects.get(dep_nombre=x.departamento),
                        pasajero = x.nombre_pasajero,
                        ruta = x.evento.ruta,
                        justificacion = x.evento.desc_evento,
                        evento = x.evento,
                        usuario_solicita = x.usuario_solicita,
                        fchsolicita = x.fchsolicita
                        )
                    r.save()
                
                query_booking = BookingReporte.objects.all()

                query = queryset.union(query_booking).order_by("-fchsolicita")
            
                paginator = Paginator(query,50) 
                page = request.GET.get('page')
                datos = paginator.get_page(page)
            else:
                mensaje = 'No selecciono las fechas de busqueda'

        return render(request, 'procesosAdministrativos/reporte_transporte.html', {'reg':datos, 'msj':mensaje, 'inicio':inicio, 'fin':fin})

    else:
        msj2 = 'Error'
        mensaje = 'No selecciono las fechas de busqueda'
        return render(request, 'procesosAdministrativos/reporte_transporte.html', {'msj':mensaje, 'msj2':msj2, 'inicio':inicio, 'fin':fin})




@login_required
def chat(request):
    if request.method == 'POST':
        user_input = request.POST['message']
        fechab = datetime.today()
        
        # p = User.objects.get(username=str(request.user.username))

        for i in range(1):
            r = RecGPT(
                pc = str(request.user.username),
                pr = str(user_input),
                fecha = fechab,
                )
            r.save() 
        
        response = openai.Completion.create(
            model="text-davinci-003",
            prompt=user_input,
            temperature=0.6,
            max_tokens=3000,
        )
        
        bot_response = response.choices[0].text.strip()
        return render(request, 'procesosAdministrativos/chat.html', {'bot_response': bot_response})
    else:
        return render(request, 'procesosAdministrativos/chat.html')


@login_required
def cargaPlanAgri(request): 
    if request.method == 'POST':

        semana = request.POST.get('semana')
        mes = request.POST.get('mes')
        print(mes,semana)
        inv_resource = InvResource()  
        dataset = Dataset()  
        # print(dataset)    
        inventario = request.FILES['xlsfile'] 
        imported_data = dataset.load(inventario.read())
        
        imported_data.headers = ['producto', 'tipo','peso','unimed']

        print(imported_data)
        result = inv_resource.import_data(imported_data, dry_run=True, raise_errors=True) # Test the data import 
        print(result) 
        print(result.row_errors())  
        if not result.has_errors():
            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE procesos.inventario_semanal_agri")  #carga_respuestas_prod_serv_varios se utiliza para cargar las respuestas de todas las categorias  
            # borra = InventarioSemanal.objects.filter(mes=mes,semana=semana).delete()
            inv_resource.import_data(imported_data, dry_run=False,raise_errors=True) # Actually import now  
            mensaje = 1
            borra = InventarioConsolidado.objects.filter(mes=mes,semana=semana).delete()
            consulta = InventarioSemanal.objects.all()
            for i in consulta:
                actualiza = InventarioConsolidado(
                    mes=mes,
                    semana=semana,
                    producto=i.producto,
                    tipo=i.tipo,
                    peso=i.peso,
                    unimed=i.unimed
                    )
                actualiza.save()

    else:
        mensaje=None
    
    return render(request,'procesosAdministrativos/cargaplanagri.html',{'mensaje':mensaje})

@login_required
def gestionaInventarioAgri(request):
    busqueda = request.GET.get('buscar')
    select_semana = request.GET.get('semana')
    if busqueda:
        queryset = InventarioConsolidado.objects.filter(producto__icontains=busqueda) 
    elif select_semana:
        queryset = InventarioConsolidado.objects.filter(semana=select_semana).order_by('producto')

    else:
        # query = InventarioConsolidado.objects.all().order_by('producto')  
        queryset = None

        # paginator = Paginator(query, 10)
        # page = request.GET.get('page')
        # queryset = paginator.get_page(page)

    return render(request, 'procesosAdministrativos/gestiona_inventario.html',{'queryset':queryset,'busqueda':busqueda})

@login_required
def registraEgreso(request):
    
    if request.method == 'GET':
        usuario = request.user
    
        form = solicitudMultiForm(request.GET or None)
        formdet = DetMultiFormSet(request.GET or None, queryset=InventarioConsolidado.objects.none())
        #Filtra Proveedores en la categoría Gestores Ambientales de Desecho
    #     prove1 = proveedor.objects.filter(categoria=5)
    #     #Aumentar proveedores que no son prte de la categoría de Gestores Ambientales pero que cumplen el rol
    #     prove2 = proveedor.objects.filter(id=397)
    #     proved = prove1 | prove2
    #     prove = proved.order_by('nombre_empresa')
    #     # tipo_des = tipo_desecho.objects.all().order_by('tipo')
    #     # tipo_so = tipo_solicitud.objects.all().order_by('tipo')

    # elif request.method == 'POST':
    #     form = solicitudMultiForm(request.POST)
    #     formdet = DetMultiFormSet(request.POST)

    #     if form.is_valid() and formdet.is_valid():
    #         print(form)
    #         print(formdet)
    #         solic = form.save(commit=False)
    #         solic.fecha_registro = datetime.now()
    #         solic = form.save()
    #         for i in formdet:
    #             detalle = i.save(commit=False)
    #             detalle.numsolicitud = solic
    #             detalle.save()
    #         solic.save()

    #         query = solicitud.objects.all().order_by('id').last()
    #         # query.fecha_registro = datetime.now()
    #         # query.save()

    #         if query.tipo_solicitud.id == 3:
    #             # pass
    #             envioMail('Solicitud de Aprobación de Donación de Desecho', 'dmencias@ecofroz.com,gerencia.administrativa@ecofroz.com', 'email_solicita_aprobacion_1n_dona.html', query)

    #         else:
    #             # pass
    #             #envioMail('Solicitud de Aprobación de Desecho', 'dmencias@ecofroz.com', 'email_solicita_aprobacion_1n.html', query)
    #             envioMail('Solicitud de Aprobación de Desecho', 'dmencias@ecofroz.com,auditoria@ecofroz.com,ambiente@ecofroz.com,contabilidad_sis@ecofroz.com', 'email_solicita_aprobacion_1n.html', query)


    #         return redirect('desecho:listar_solicitudes')

    # return render(request, 'procesosAdministrativos/registra_egreso.html',{'form':form, 'formdet':formdet,'proveedor':prove,'usuario':usuario})
    return render(request, 'procesosAdministrativos/registra_egreso.html',{'form':form, 'formdet':formdet,'usuario':usuario})

#Funcion AJA> para filtrar listas dinámicamente en el ingreso de solicitudes
def load_items(request):
    tipo_id = request.GET.get('tipo')

    itemsq = items.objects.filter(tipo_desecho=tipo_id).order_by('item')
    return render(request, 'desecho/items_dropdown_select.html', {'itemsq':itemsq})



####### AQUI VAN LOS SERVICIOS #########

class SectoresListApiView(ListAPIView):
    serializer_class = SectoresSerializer2

    def get_queryset(self):
        kword = self.request.query_params.get('kword','')
        if kword == '':
            return RutaTransporte.objects.none()

        return RutaTransporte.objects.filter(origen__icontains=kword)[:8]
    

class NominaList(ListView):
    model = Nomina
    template_name = 'procesosAdministrativos/nomina_list.html'
    paginate_by = 10