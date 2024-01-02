from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseNotFound, Http404, FileResponse
from django.urls import reverse_lazy
from django.db.models import Q
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from datetime import datetime
from openpyxl import Workbook
from django.http import HttpResponse
from django.core.paginator import Paginator

from .models import TicketsIt, Estados
from .forms import Tickets, IngresoSoporte, EditaSoporte
from .funciones import envioMail

# Create your views here.

def ordenesSoporte(request):
    queryset = request.GET.get("buscar")
    if queryset:
        orden = TicketsIt.objects.filter(
            Q(numticket__icontains = queryset)
        ).filter(usuario = request.user.id)
    else:
        orden = TicketsIt.objects.filter(usuario = request.user.id).order_by('-numticket')
    return render(request, 'soporteIt/listar_ordenes.html', {'form':orden, 'busqueda':queryset})

class IngresoTicket(CreateView):
    model = TicketsIt
    template_name = 'soporteIt/ingreso.html'
    form_class = Tickets
    success_url = reverse_lazy('soporte:listar_soporte')

    def get_context_data(self, **kwargs):
        context = super(IngresoTicket, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        fecha = datetime.now()
        fchtxt = datetime.strftime(fecha, '%Y%m%d')
        hrtxt = datetime.strftime(fecha, '%H:%M')
        
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.usuario = request.user
            solicitud.estado_solicitud = Estados.objects.get(id=1)
            solicitud.fch_solicita = fecha
            solicitud.fch_solicita_txt = fchtxt
            solicitud.hr_solicita = hrtxt
            solicitud.save()

            # queryset = TicketsIt.objects.all().last()
            # queryset2 = ''

            # envioMail('Solicitud de Soporte', 'desarrollo@ecofroz.com', 'email.html', queryset, queryset2)

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

def ordenesSoporteInt(request):
    queryset = request.GET.get("buscar")
    if queryset:
        orden = TicketsIt.objects.filter(
            Q(numticket__icontains = queryset)
        )
    else:
        ordenx = TicketsIt.objects.all().order_by('-numticket')

        paginator = Paginator(ordenx, 50)
        page = request.GET.get('page')
        orden = paginator.get_page(page)
    return render(request, 'soporteIt/listar_ordenes_int.html', {'form':orden, 'busqueda':queryset})

class IngresoTicketSop(CreateView):
    model = TicketsIt
    template_name = 'soporteIt/ingreso_int.html'
    form_class = IngresoSoporte
    success_url = reverse_lazy('soporte:listar_soporte_int')

    def get_context_data(self, **kwargs):
        context = super(IngresoTicketSop, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        fecha = datetime.now()
        fchtxt = datetime.strftime(fecha, '%Y%m%d')
        hrtxt = datetime.strftime(fecha, '%H:%M')
        
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.usuario = request.user.id
            solicitud.estado_solicitud = Estados.objects.get(id=1)
            solicitud.fch_solicita = fecha
            solicitud.fch_solicita_txt = fchtxt
            solicitud.hr_solicita = hrtxt
            solicitud.save()

            # queryset = TicketsIt.objects.all().last()
            # queryset2 = ''

            # envioMail('Solicitud de Soporte', 'desarrollo@ecofroz.com', 'email.html', queryset, queryset2)

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

class EditaTicketSop(UpdateView):
    model = TicketsIt
    template_name = 'soporteIt/edita_int.html'
    form_class = EditaSoporte
    success_url = reverse_lazy('soporte:listar_soporte_int')

    def get_context_data(self, **kwargs):
        context = super(EditaTicketSop, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        ticket = self.model.objects.get(numticket=pk)
        fecha = datetime.strftime(ticket.fch_solicita, '%Y/%m/%d')
        print(fecha)
        if ticket.fch_soluciona:
            fecha_sol = datetime.strftime(ticket.fch_soluciona, '%Y/%m/%d')
        else:
            fecha_sol = None
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'fecha' not in context:
            context['fecha'] = fecha
        if 'fechas' not in context:
            context['fechas'] = fecha_sol
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        ticket = kwargs['pk']
        soporte = self.model.objects.get(numticket=ticket)
        form = self.form_class(request.POST, request.FILES, instance=soporte)
        fecha = request.POST.get('fch_soluciona')
        fecha = datetime.strptime(fecha, '%Y-%m-%d')
        fchtxt = datetime.strftime(fecha, '%Y%m%d')
    
        
        if form.is_valid():
            solicitud = form.save(commit=False)
            # solicitud.estado_solicitud = Estados.objects.get(id=10)
            solicitud.fch_soluciona_txt = fchtxt
            solicitud.reasignado = True
            solicitud.save()

            # queryset = TicketsIt.objects.all().last()
            # queryset2 = ''

            # envioMail('Solicitud de Soporte', 'desarrollo@ecofroz.com', 'email.html', queryset, queryset2)

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))

def reporteSoporte(request):
    try:          
        queryset = TicketsIt.objects.all().order_by('-numticket')
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE SOPORTES'

        ws.merge_cells('A1:I1')
        ws['A3'] = 'N° TICKET'
        ws['B3'] = 'SOLICITANTE'
        ws['C3'] = 'UBICACIÓN'
        ws['D3'] = 'DEPARTAMENTO'
        ws['E3'] = 'FECHA SOLICITUD'
        ws['F3'] = 'HORA SOLICITUD'
        ws['G3'] = 'TIPO DE PROBLEMA'
        ws['H3'] = 'DESCRIPCIÓN DE PROBLEMA'
        ws['I3'] = 'FECHA DE SOLUCIÓN'
        ws['J3'] = 'HORA DE SOLUCIÓN'
        ws['K3'] = 'SOLUCIÓN'
        ws['L3'] = 'RESPONSABLE'
        ws['M3'] = 'AREA RESPONSABLE'
        ws['N3'] = 'TIEMPO DE SOLUCIÓN'
                
        count = 4
        rowcount = 1   

        for i in queryset:
            # fch_1 = i.fch_solicita
            # fch_2 = i.fch_soluciona
            # fch_txt_solicita = lambda fch_1: '' if (fch_1 == None) else (fch_1.strftime('%Y-%m-%d'))
            # fch_txt_soluciona = lambda fch_2: '' if (fch_2 == None) else (fch_2.strftime('%Y-%m-%d'))
            # texto_1 = str(fch_txt_solicita) + ' ' + str(i.hr_solicita)
            # texto_2 = str(fch_txt_soluciona) + ' ' + str(i.hr_soluciona_txt)
            # print(texto_1)
            # print(texto_2)
            # valor_1 = datetime.strptime(texto_1, '%Y-%m-%d %H:%M')
            # valor_2 = datetime.strptime(texto_2, '%Y-%m-%d %H:%M')
            ws.cell(row = count, column = 1).value =  str(i.numticket)
            ws.cell(row = count, column = 2).value =  str(i.solicita_nombre) + ' ' + str(i.solicita_apellido)
            ws.cell(row = count, column = 3).value =  str(i.ubicacion)
            ws.cell(row = count, column = 4).value =  str(i.departamento)
            ws.cell(row = count, column = 5).value =  str(i.fch_solicita.strftime('%Y-%m-%d'))
            ws.cell(row = count, column = 6).value =  str(i.hr_solicita)
            ws.cell(row = count, column = 7).value =  str(i.tipo_problema)
            ws.cell(row = count, column = 8).value =  str(i.descripcion_problema)
            if i.fch_soluciona == None:
                ws.cell(row = count, column = 9).value =  str(i.fch_soluciona)
            else:
                ws.cell(row = count, column = 9).value =  str(i.fch_soluciona.strftime('%Y-%m-%d'))
            ws.cell(row = count, column = 10).value =  str(i.hr_soluciona_txt)
            ws.cell(row = count, column = 11).value =  str(i.observa_solucion)
            ws.cell(row = count, column = 12).value =  str(i.usuario_responsabe)
            ws.cell(row = count, column = 13).value =  str(i.area_responsable)
            if i.hr_soluciona_txt == None:
                ws.cell(row = count, column = 14).value =  str('')
            else:
                fch_txt_solicita = i.fch_solicita.strftime('%Y-%m-%d')
                fch_txt_soluciona = i.fch_soluciona.strftime('%Y-%m-%d')
                texto_1 = str(fch_txt_solicita) + ' ' + str(i.hr_solicita)
                texto_2 = str(fch_txt_soluciona) + ' ' + str(i.hr_soluciona_txt)
                valor_1 = datetime.strptime(texto_1, '%Y-%m-%d %H:%M')
                valor_2 = datetime.strptime(texto_2, '%Y-%m-%d %H:%M')
                resultado = valor_2 - valor_1
                # print(valor_1)
                # print(valor_2)
                # print(resultado)
                ws.cell(row = count, column = 14).value =  str(resultado)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE SOPORTE" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        #ahora = datetime.now()
        #ayer = ahora - timedelta(days=1)
        #fecha = ahora.strftime('%Y-%m-%d')
        #fechatxt = ahora.strftime('%Y%m%d')
        #fechaacttxt = ayer.strftime('%Y%m%d')
        #queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        #paginator = Paginator(queryset.order_by('-ord'), 50)
        #page = request.GET.get('page')
        #ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Se ha encontrado un error'
        print(e)

        return redirect('soporte:listar_soporte_int')
