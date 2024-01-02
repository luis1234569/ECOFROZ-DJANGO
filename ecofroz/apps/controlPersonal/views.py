import random
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, permission_required
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Q
from django.urls import reverse_lazy, reverse
from reportlab.pdfgen import canvas
from datetime import datetime, date, timedelta
from django.db.models import F
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.template import loader
from django.contrib import messages
import logging



from .forms import PersonaRegistroForm, VehiculoRegistroForm, PersonaRegistroSForm, VehiculoRegistroSForm, \
    RegistroChoferForm, RegistroPlacaCabezal, cabIngresoReportes, detIngresoReportes, selectGuardia, DetMultiFormSet
from .models import PersonaRegistro, VehiculoRegistro, Vehiculos, Persona, Transportista, Cabezales, Chofer, \
                    ContenedorRegistro, AlmuerzoRegistro, ReporteNovedades, DetReporteNovedades, TrackBusqueda

from apps.usuarios.models import User
from django.http import JsonResponse


logger = logging.getLogger(__name__)

#EnvioCorreo
def envioMail(subject, email, template, consulta):
    html_message = loader.render_to_string(
        'controlPersonal/email/%s' %template,
            {
                'aprob':consulta,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except:
        logger.error("Unable to send mail.")


#Funcion AJAX para cargar fotografias

class uploadFotosView(View):
    def get(self, request):
        photos_list = DetReporteNovedades.objects.all()

        return render(self.request, 'personal/lista_fotos_cargada.html', {'photos': photos_list})
       

    def post(self, request):
        form = detIngresoReportes(self.request.POST, self.request.FILES)
        if form.is_valid():
            photo = form.save()
            data = {'is_valid': True, 'name': photo.file.name, 'url': photo.file.url}
        else:
            data = {'is_valid': False}
        return JsonResponse(data)




########### INGRESO DIARIO DE NOVEDADES GUARDIANIA #########
@login_required
def capacitacionSegu(request):
    return render(request, 'controlPersonal/capacitacion_seguridad.html')


@login_required
def editarReporte(request, pk):
    if request.method == 'GET':
        consulta = ReporteNovedades.objects.get(numreporte=pk)
        #consulta2 = DetReporteNovedades.objects.get(numreporte=pk)
        form = cabIngresoReportes(request.GET or None, instance=consulta)
        #form2 = detIngresoReportes(request.GET or None, instance=consulta2)
        form3 = selectGuardia(request.GET or None)
        
        queryg = User.objects.filter(guardia=True)
        #print(queryg)
        print(consulta, type(consulta.guardia2))

        if consulta.guardia2 == 'None':
            
            pguardia = 0
            for i in queryg:
                if i.last_name in consulta.guardia1:
                    pguardia = i.id
            sguardia=None
        
        else:
            print("Entró por else")
            pguardia = 0
            for i in queryg:
                if i.last_name in consulta.guardia1:
                    if i.first_name in consulta.guardia1:
                        pguardia = i.id
            
            for j in queryg:
                if j.last_name in consulta.guardia2:
                    if j.first_name in consulta.guardia2:
                        sguardia = j.id
            


        reporte = consulta.numreporte
        return render(request, 'controlPersonal/editar_reporte.html', {'form':form,'form3':form3,'persona_registra':pguardia,'sguardia':sguardia,'reporte':reporte, 'estado':consulta.estado})
    
    else:
        consulta = ReporteNovedades.objects.get(numreporte=pk)
        #consulta2 = DetReporteNovedades.objects.get(numreporte=pk)
        form = cabIngresoReportes(request.POST, instance=consulta)
        #form2 = detIngresoReportes(request.POST, request.FILES, instance=consulta2)
        form3 = selectGuardia(request.POST)

        if form.is_valid() and form3.is_valid():
            #det = form2.save(commit=False)
            cab = form.save(commit=False)
            nombre = form3.cleaned_data.get('guardia1')
            nombre2 = form3.cleaned_data.get('guardia2')
            # fecha = datetime.strptime(request.POST.get('fchregistro'),'%Y-%m-%d %H:%M')
            fechamod = datetime.today().strftime('%Y-%m-%d %H:%M:%S')
            
            cab.guardia1 = str(nombre)
            cab.guardia2 = str(nombre2)
            cab.fchregistro = fechamod
            cab.estado = 1
            cab.save()
            
            # det.numreporte = form.save()
            # det.save()
            
            #envioMail('Reporte Diario de Novedades Guardianía', 'dmencias@ecofroz.com,logistica@ecofroz.com', 'email_reporte_diario_novedades.html', consulta)

            return redirect('personal:listar_novedades')
        else:
            print(form.errors)
           
            return HttpResponse("No se pudo guardar. Comuníquese con IT")


@login_required
def verReporte(request, pk):
    if request.method == "GET":
        reporte = ReporteNovedades.objects.get(numreporte=pk)
        #Marca como leido el reporte
        #actualiza = ReporteNovedades.objects.filter(numreporte=pk).update(estado=3)
        return render(request, 'controlPersonal/ver_reportes_seguridad.html', {"query":reporte,'pk':pk})

    else:
        observaciones = request.POST.get('observaciones')
        detalle = request.POST.get('detalle')
        actualiza = ReporteNovedades.objects.filter(numreporte=pk).update(observaciones_seg=observaciones,observaciones_guardias=detalle)
        consulta = ReporteNovedades.objects.get(numreporte=pk)
        if consulta.observaciones_seg:
            if consulta.observaciones_seg != 'None':
                actualiza = ReporteNovedades.objects.filter(numreporte=pk).update(estado_observaciones_seg=True)

        return redirect('personal:listar_novedades')

@login_required
def cerrarEdicion(request, pk):
    if request.method == "GET":
        reporte = ReporteNovedades.objects.get(numreporte=pk)
        #Cierra la Edición del reporte para los guardias
        actualiza = ReporteNovedades.objects.filter(numreporte=pk).update(estado=3)
        return redirect('personal:listar_novedades')   
    

@login_required
def gallery_view(request, pk):
    reporte = ReporteNovedades.objects.get(numreporte=pk)
    return render(request, 'controlPersonal/seguridad_gallery_view.html', {"reporte":reporte})

@login_required
def ingresoNovedades(request):
    if request.method == "GET":
        usuario = str(request.user.username)
        print(usuario)

        usuarioq = User.objects.get(username=usuario)
        login = usuarioq.id

        form = cabIngresoReportes(request.GET or None)
        form2 = detIngresoReportes(request.GET or None)
        form3 = selectGuardia(request.GET or None)

        formdet = DetMultiFormSet(request.GET or None, queryset=DetReporteNovedades.objects.none())

        numrepo = 0
        fecha_registro = datetime.today().strftime('%Y-%m-%d %H:%M')
        print(fecha_registro)

        try:
            query = ReporteNovedades.objects.all().order_by('-numreporte')

        except:
            query = []

        if query:
            query = query[0]
            numrepo = query.numreporte + 1
        else:
            numrepo = 1


        return render(request, 'controlPersonal/ingreso_novedades.html',{'form':form,'form2':form2,'form3':form3,'formdet':formdet,'numrepo':numrepo,'fecha_registro':fecha_registro,'persona_registra':login})
    
    

    else:
        consulta = None
        
        form = cabIngresoReportes(request.POST)
        form2 = detIngresoReportes(request.POST, request.FILES)
        form3 = selectGuardia(request.POST)
        formdet = DetMultiFormSet(request.POST, request.FILES)

        #print(form2)
        
        for i in form2:
            print(i)

        # if form.is_valid() and form2.is_valid and form3.is_valid():
        #     det = form2.save(commit=False)
        #     cab = form.save(commit=False)
        #     nombre = form3.cleaned_data.get('guardia1')
        #     nombre2 = form3.cleaned_data.get('guardia2')
        #     fecha = datetime.strptime(request.POST.get('fchregistro'),'%Y-%m-%d %H:%M')
        #     fechamod = fecha.strftime('%Y-%m-%d %H:%M:%S')
            
        if form.is_valid() and formdet.is_valid and form3.is_valid():
            det = formdet.save(commit=False)
            cab = form.save()
            for i in formdet:
                detalle = i.save(commit=False)
                detalle.numreporte = cab
                detalle.save()

            nombre = form3.cleaned_data.get('guardia1')
            nombre2 = form3.cleaned_data.get('guardia2')
            fecha = datetime.strptime(request.POST.get('fchregistro'),'%Y-%m-%d %H:%M')
            fechamod = fecha.strftime('%Y-%m-%d %H:%M:%S')
            
            cab.guardia1 = str(nombre)
            cab.guardia2 = str(nombre2)
            cab.fchregistro = fechamod
            cab.estado = 1
            cab.save()
            
            # for i in det:
            #     det.numreporte = form.save()
            #     det.save()
            
            #envioMail('Reporte Diario de Novedades Guardianía', 'dmencias@ecofroz.comc,logistica@ecofroz.com', 'email_reporte_diario_novedades.html', consulta)

            

            return redirect('personal:listar_novedades')
        else:
            print(form.errors)
           
            return HttpResponse("No se pudo guardar. Comuníquese con IT")




       
@login_required
def listarNovedades(request):
    queryset = request.GET.get("buscar")
    if queryset:
        query = ReporteNovedades.objects.filter(
            Q(numreporte__iexact = queryset) |
            Q(asunto__icontains = queryset))
            
    else:
        query = ReporteNovedades.objects.all().order_by('-fchregistro')

    paginator = Paginator(query, 5)
    page = request.GET.get('page')
    solic = paginator.get_page(page)

    fotos=[]
    filtro = ReporteNovedades.objects.all()
    for i in filtro:
        fotos = i.detnovedades.all()
    
    
    return render(request, 'controlPersonal/listar_reportes.html', {'form':solic, 'busqueda':queryset,'fotos':fotos})





############ DIRECTORIO DE PERSONAL ############

@login_required
def personalDirectorio(request):
    busqueda = request.GET.get("buscar")
      
    if busqueda:
        queryset = Persona.objects.all().exclude(extension__isnull=True).order_by('area').filter(
            Q(first_name__icontains = busqueda) |
            Q(last_name__icontains = busqueda)).exclude(estado='0').order_by('area') 

        # queryset = Areas.objects.all().select_related('area_departamento').exclude(area_departamento__extension__isnull=True).filter(
        #     Q(area_departamento__first_name__icontains = busqueda) |
        #     Q(area_departamento__last_name__icontains = busqueda)).exclude(area_departamento__estado='0').order_by('area_departamento__area') 


        fechab = datetime.today()
        fuente = 'directorio'
        
        for i in range(1):
            r = TrackBusqueda(
                personabb = busqueda,
                personab = request.user.username,
                fecha = fechab,
                source = fuente,
                )
            r.save()           
    else:
        queryset = Persona.objects.all().exclude(extension__isnull=True).exclude(estado='0').order_by('area','first_name')
        paginator = Paginator(queryset, 25)

        # countpersonal = Persona.objects.all().exclude(estado_pl=False).count()

        page = request.GET.get('page')
        proveedores = paginator.get_page(page)

    contador = PersonaRegistro.objects.filter(persona__estado_pl=True, hr_salida__isnull=True, ubica='MCH').count()
    contadorv = VehiculoRegistro.objects.filter(vehiculo__estado_pl=True, hr_salida__isnull=True, ubica='MCH').count()

    return render(request,'controlPersonal/directorio_personal.html',{'personal':queryset, 'contadorp':contador, 'contadorv':contadorv})


############ INGRESO DE PERSONAL ############

@login_required
@permission_required('controlPersonal.registro_personal', raise_exception=True)
def ingresoPersonal(request):
    ahora = datetime.now()
    fecha = ahora.strftime('%Y-%m-%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso=fecha)
    paginator = Paginator(queryset, 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)
    # ingresos = paginator.object_list.reverse()
    # ingresos = list(reversed(ingresos.object_list))

    if request.method == 'POST':
        cont = len(request.POST.get('persona'))
        if cont > 7:
            form = PersonaRegistroForm(request.POST)
            if form.is_valid():
                form.save()
                return redirect('personal:ingreso_personal')
        else:
            veh = request.POST.get('persona')
            fch = request.POST.get('fch_ingreso')
            fch_txt = request.POST.get('fch_ingreso_txt')
            hr = request.POST.get('hr_ingreso')
            hr_txt = request.POST.get('hr_ingreso_txt')
            form2 = VehiculoRegistroForm
            return render (request, 'controlPersonal/ingreso_vehiculo.html', {'form':form2, 'fch':fch, 'fch_txt':fch_txt, 'hr':hr, 'hr_txt':hr_txt, 'veh':veh})
    else:
        form = PersonaRegistroForm()

    return render(request, 'controlPersonal/ingreso_personal.html', {'form':form, 'reg':ingresos})

@login_required
@permission_required('controlPersonal.registro_vehiculo', raise_exception=True)
def ingresoVehiculo(request):
    ahora = datetime.now()
    fecha = ahora.strftime('%Y-%m-%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso=fecha)
    paginator = Paginator(queryset, 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)
    ran = random.randint(1,7)
    print(ran)
    if request.method == 'POST':
        if request.POST.get('vehiculo'):
            placa = request.POST.get('vehiculo')
            rev = Vehiculos.objects.get(placa=placa)
            mensaje = {
                    'placa': placa,
                    'mensaje': 'PASE',
                }
            if rev.random == ran:
                mensaje = {
                    'placa': placa,
                    'mensaje': 'REVISIÓN DEL VEHÍCULO',
                }

            form = VehiculoRegistroForm(request.POST)
            form2 = PersonaRegistroForm()
            if form.is_valid():
                form.save()
                return  render(request, 'controlPersonal/ingreso_personal.html', {'ran':mensaje, 'form':form2, 'reg':ingresos})
            # return redirect('personal:ingreso_personal')
        else:
            return redirect('personal:ingreso_personal')
    else:
        form =VehiculoRegistroForm()

    return  render(request, 'controlPersonal/ingreso_vehiculo.html', {'form':form})

@login_required
@permission_required('controlPersonal.registro_personal', raise_exception=True)
def salidaPersona(request):
    ahora = datetime.now()
    fecha = ahora.strftime('%Y-%m-%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso=fecha)
    paginator = Paginator(queryset, 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    if request.method == 'POST':
        persona = request.POST.get('persona')
        form = PersonaRegistroSForm(request.POST)

        cont = len(request.POST.get('persona'))
        if cont > 7:
            actualiza = PersonaRegistro.objects.filter(persona=persona, fch_ingreso=fecha, hr_salida_txt__isnull=True)
            if form.is_valid():
                for i in actualiza:
                    i.hr_salida = request.POST.get('hr_salida')
                    i.hr_salida_txt = request.POST.get('hr_salida_txt')
                    i.save()
                return redirect('personal:salida_persona')
        else:
            veh = request.POST.get('persona')
            hr = request.POST.get('hr_salida')
            hr_txt = request.POST.get('hr_salida_txt')
            form2 = VehiculoRegistroSForm
            return render (request, 'controlPersonal/salida_vehiculo.html', {'form':form2, 'hr':hr, 'hr_txt':hr_txt, 'veh':veh})
    else:
        form = PersonaRegistroSForm()

    return  render(request, 'controlPersonal/salida_personal.html', {'form':form, 'reg':ingresos})

@login_required
@permission_required('controlPersonal.registro_vehiculo', raise_exception=True)
def salidaVehiculo(request):
    ahora = datetime.now()
    fecha = ahora.strftime('%Y-%m-%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso=fecha)
    paginator = Paginator(queryset, 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    if request.method == 'POST':
        form = VehiculoRegistroSForm(request.POST)
        form2 = PersonaRegistroSForm()
        vehiculo = request.POST.get('vehiculo')
        actualiza = VehiculoRegistro.objects.filter(vehiculo=vehiculo, fch_ingreso=fecha, hr_salida_txt__isnull=True)
        if actualiza:
            if form.is_valid():
                for i in actualiza:
                    i.hr_salida = request.POST.get('hr_salida')
                    i.hr_salida_txt = request.POST.get('hr_salida_txt')
                    i.save()
                    return render(request, 'controlPersonal/salida_personal.html', {'form':form2, 'reg':ingresos})
                    # return redirect('personal:salida_persona')
            else:
                return redirect('personal:salida_persona')
        else:
            mensaje = 'NO EXISTE EL REGISTRO'
            print(mensaje)
            return  render(request, 'controlPersonal/salida_personal.html', {'form':form2, 'reg':ingresos, 'msj':mensaje})
    else:
        form =VehiculoRegistroSForm()
    
    return  render(request, 'controlPersonal/salida_personal.html', {'form':form2}) 

@login_required
@permission_required('controlPersonal.registro_vehiculo', raise_exception=True)
def registroEmpleados(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='MCH').values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord', revision_vehiculo=F('carnet'))
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='MCH').values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord','revision_vehiculo')
    contador = PersonaRegistro.objects.filter(persona__estado_pl=True, hr_salida__isnull=True, ubica='MCH').count()
    contadorv = VehiculoRegistro.objects.filter(vehiculo__estado_pl=True, hr_salida__isnull=True, ubica='MCH').count()
    queryset = q1.union(q2)
    paginator = Paginator(queryset.order_by('-ord'), 5)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/ingreso.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')

        cont = len(persona_id)
        if cont > 7:
            try:
                persona = Persona.objects.get(cedula=persona_id, estado=1)  
            
                print(persona_id)
                try:
                    verifica = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                except:
                    verifica = None

                print (verifica)
                if verifica:
                    # queryset = Persona.objects.get(cedula=persona_id)
                    r = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    # print(queryset.estado_pl)
                    # queryset.estado_pl = False
                    # queryset.save()
                    queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=False)
                    r.save()
                else:
                    # queryset = Persona.objects.get(cedula=persona_id)
                    r = PersonaRegistro(
                        persona_id=persona_id,
                        fch_ingreso=fch,
                        fch_ingreso_txt=fch_txt,
                        hr_ingreso=hr,
                        hr_ingreso_txt=hr_txt,
                        usuario_registra_id=usuario_registra,
                        ubica='MCH',
                    )
                    # queryset.estado_pl=True
                    # queryset.save()
                    queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=True)
                    r.save()
        

                return redirect('personal:registros')
        
            except Exception as e:
                mensaje_error = str(e)
                # messages.error(request, mensaje_error)
                print('hola' + mensaje_error)
                # return redirect('personal:registros')
                return  render(request, template, {'ran':{
                    'mensaje': 'CARNET NO ESTA ACTIVO. DETENGA EL INGRESO Y CONSULTE CON RRHH','reg':None}})
        
        
        else:
            print(persona_id)
            try:
                get_estado = Vehiculos.objects.get(placa=persona_id, persona__estado=1, estado=1)  
                

                ran = random.randint(1,7)
                print(ran)
                rev = Vehiculos.objects.get(placa=persona_id)
                if rev.random == ran:
                    graba_estado_revision=True
                    mensaje = {
                        'mensaje': 'REVISIÓN DEL VEHÍCULO',
                    }
                    rev.maximo=0
                    rev.save()
                    
                else:
                    
                    rev.maximo+=1
                    rev.save()
                    
                    if rev.maximo == 9:
                        graba_estado_revision=True
                        mensaje = {
                        'mensaje': 'REVISIÓN DEL VEHÍCULO',
                        }
                    
                        rev.maximo=0
                        rev.save()
                    
                    else:
                        graba_estado_revision=False
                        mensaje = {
                        'mensaje': 'PASE',
                        }



                if rev.estado_pl == False:
                    r = VehiculoRegistro(
                        vehiculo_id=persona_id,
                        fch_ingreso=fch,
                        fch_ingreso_txt=fch_txt,
                        hr_ingreso=hr,
                        hr_ingreso_txt=hr_txt,
                        usuario_registra_id=usuario_registra,
                        ubica='MCH',
                        revision_vehiculo=graba_estado_revision,
                    ) 
                    rev.estado_pl=True
                    rev.save()
                    r.save()

                    return  render(request, template, {'ran':mensaje, 'reg':ingresos})
                
                elif rev.estado_pl == True:
                    r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    r.revision_vehiculo=graba_estado_revision
                    rev.estado_pl = False
                    rev.save()
                    r.save()
                    
                    return  render(request, template, {'ran':mensaje,'reg':ingresos})

            except Exception as e:
                mensaje_error = str(e)
                # messages.error(request, mensaje_error)
                print('hola2' + mensaje_error)
                # return redirect('personal:registros')
                return  render(request, template, {'ran':{
                    'mensaje': 'VEHICULO NO ESTA ACTIVO. DETENGA EL INGRESO Y CONSULTE CON SEGURIDAD','reg':None}})

    return render(request, template, {'reg':ingresos, 'contadorp':contador, 'contadorv':contadorv})

@login_required
@permission_required('controlPersonal.registro_ac', raise_exception=True)
def registroAguaClara(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='AC').values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord')
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='AC').values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord')
    contador = PersonaRegistro.objects.filter(persona__estado_pl=True, hr_salida__isnull=True, ubica='AC').count()
    contadorv = VehiculoRegistro.objects.filter(vehiculo__estado_pl=True, hr_salida__isnull=True, ubica='AC').count()
    queryset = q1.union(q2)
    paginator = Paginator(queryset.order_by('-ord'), 5)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/ingreso_ac.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        #persona_id = strip(persona_id)
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')

        cont = len(persona_id)
        if cont > 7:
            try:
                persona = Persona.objects.get(cedula=persona_id, estado=1)  
            
                print(persona_id)
                try:
                    verifica = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                except:
                    verifica = None

                print (verifica)
                if verifica:
                    # queryset = Persona.objects.get(cedula=persona_id)
                    r = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    # print(queryset.estado_pl)
                    # queryset.estado_pl = False
                    # queryset.save()
                    queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=False)
                    r.save()
                else:
                    # queryset = Persona.objects.get(cedula=persona_id)
                    r = PersonaRegistro(
                        persona_id=persona_id,
                        fch_ingreso=fch,
                        fch_ingreso_txt=fch_txt,
                        hr_ingreso=hr,
                        hr_ingreso_txt=hr_txt,
                        usuario_registra_id=usuario_registra,
                        ubica='AC',
                    )
                    # queryset.estado_pl=True
                    # queryset.save()
                    queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=True)
                    r.save()
        

                return redirect('personal:registros_ac')
        
            except Exception as e:
                mensaje_error = str(e)
                # messages.error(request, mensaje_error)
                print('hola' + mensaje_error)
                # return redirect('personal:registros')
                return  render(request, template, {'ran':{
                    'mensaje': 'CARNET NO ESTA ACTIVO. DETENGA EL INGRESO Y CONSULTE CON RRHH','reg':None}})
        
        
        else:
            print(persona_id)
            try:
                get_estado = Vehiculos.objects.get(placa=persona_id,persona__estado=1)  
                

                ran = random.randint(1,7)
                print(ran)
                rev = Vehiculos.objects.get(placa=persona_id)
                if rev.random == ran:
                    graba_estado_revision=True
                    mensaje = {
                        'mensaje': 'REVISIÓN DEL VEHÍCULO',
                    }
                    rev.maximo=0
                    rev.save()
                    
                else:
                    
                    rev.maximo+=1
                    rev.save()
                    
                    if rev.maximo == 9:
                        graba_estado_revision=True
                        mensaje = {
                        'mensaje': 'REVISIÓN DEL VEHÍCULO',
                        }
                    
                        rev.maximo=0
                        rev.save()
                    
                    else:
                        graba_estado_revision=False
                        mensaje = {
                        'mensaje': 'PASE',
                        }



                if rev.estado_pl == False:
                    r = VehiculoRegistro(
                        vehiculo_id=persona_id,
                        fch_ingreso=fch,
                        fch_ingreso_txt=fch_txt,
                        hr_ingreso=hr,
                        hr_ingreso_txt=hr_txt,
                        usuario_registra_id=usuario_registra,
                        ubica='AC',
                        revision_vehiculo=graba_estado_revision,
                    ) 
                    rev.estado_pl=True
                    rev.save()
                    r.save()

                    return  render(request, template, {'ran':mensaje, 'reg':ingresos})
                
                elif rev.estado_pl == True:
                    r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    r.revision_vehiculo=graba_estado_revision
                    rev.estado_pl = False
                    rev.save()
                    r.save()
                    
                    return  render(request, template, {'ran':mensaje,'reg':ingresos})

            except Exception as e:
                mensaje_error = str(e)
                # messages.error(request, mensaje_error)
                print('hola' + mensaje_error)
                # return redirect('personal:registros')
                return  render(request, template, {'ran':{
                    'mensaje': 'CARNET NO ESTA ACTIVO. DETENGA EL INGRESO Y CONSULTE CON RRHH','reg':None}})

    return render(request, template, {'reg':ingresos, 'contadorp':contador, 'contadorv':contadorv})


@login_required
@permission_required('controlPersonal.registro_lm', raise_exception=True)
def registroLaMerced(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LM').values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord')
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LM').values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord')
    contador = PersonaRegistro.objects.filter(persona__estado_pl=True, hr_salida__isnull=True, ubica='LM').count()
    contadorv = VehiculoRegistro.objects.filter(vehiculo__estado_pl=True, hr_salida__isnull=True, ubica='LM').count()
    queryset = q1.union(q2)
    paginator = Paginator(queryset.order_by('-ord'), 5)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/ingreso_lm.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')

        cont = len(persona_id)
        if cont > 7:
            print(persona_id)
            try:
                verifica = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
            except:
                verifica = None

            print (verifica)
            if verifica:
                r = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                
                queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=False)
                r.save()
            else:
               
                r = PersonaRegistro(
                    persona_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LM',
                )
               
                queryset = Persona.objects.filter(cedula=persona_id).update(estado_pl=True)
                r.save()
           

            return redirect('personal:registros_lm')
        else:
            ran = random.randint(1,7)
            rev = Vehiculos.objects.get(placa=persona_id)
            if rev.random == ran:
                mensaje = {
                    'mensaje': 'REVISIÓN DEL VEHÍCULO',
                }
            else:
                mensaje = {
                    'mensaje': 'PASE',
                }

            if rev.estado_pl == False:
                r = VehiculoRegistro(
                    vehiculo_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LM',
                ) 
                rev.estado_pl=True
                rev.save()
                r.save()

                return  render(request, template, {'ran':mensaje, 'reg':ingresos})
            elif rev.estado_pl == True:
                r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                rev.estado_pl = False
                rev.save()
                r.save()
                
                return  render(request, template, {'reg':ingresos})

    return render(request, template, {'reg':ingresos, 'contadorp':contador, 'contadorv':contadorv})



@login_required
@permission_required('controlPersonal.registro_ll', raise_exception=True)
def registroLaLaurita(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LL').values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord')
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LL').values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord')
    queryset = q1.union(q2)
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/ingreso_ll.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')

        cont = len(persona_id)
        if cont > 7:
            queryset = Persona.objects.get(cedula=persona_id)
            if queryset.estado_pl == False:
                r = PersonaRegistro(
                    persona_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LL',
                )
                queryset.estado_pl=True
                queryset.save()
                r.save()
            elif queryset.estado_pl == True:
                r = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                queryset.estado_pl = False
                queryset.save()
                r.save()

            return redirect('personal:registros_ll')
        else:
            ran = random.randint(1,7)
            rev = Vehiculos.objects.get(placa=persona_id)
            if rev.random == ran:
                mensaje = {
                    'mensaje': 'REVISIÓN DEL VEHÍCULO',
                }
            else:
                mensaje = {
                    'mensaje': 'PASE',
                }

            if rev.estado_pl == False:
                r = VehiculoRegistro(
                    vehiculo_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LL',
                ) 
                rev.estado_pl=True
                rev.save()
                r.save()

                return  render(request, template, {'ran':mensaje, 'reg':ingresos})
            elif rev.estado_pl == True:
                r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                rev.estado_pl = False
                rev.save()
                r.save()
                
                return  render(request, template, {'reg':ingresos})

    return render(request, template, {'reg':ingresos})

@login_required
@permission_required('controlPersonal.registro_la', raise_exception=True)
def registroLaAvelina(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LA').values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord')
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).filter(ubica='LA').values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord')
    queryset = q1.union(q2)
    contador = PersonaRegistro.objects.filter(persona__estado_pl=True, hr_salida__isnull=True, ubica='LA').count()
    paginator = Paginator(queryset.order_by('-ord'), 5)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)
    

    template = 'controlPersonal/ingreso_la.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')

        cont = len(persona_id)
        if cont > 7:
            queryset = Persona.objects.get(cedula=persona_id)
            if queryset.estado_pl == False:
                r = PersonaRegistro(
                    persona_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LA',
                )
                queryset.estado_pl=True
                queryset.save()
                r.save()
            elif queryset.estado_pl == True:
                r = PersonaRegistro.objects.get(persona=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                queryset.estado_pl = False
                queryset.save()
                r.save()

            return redirect('personal:registros_la')
        else:
            ran = random.randint(1,7)
            rev = Vehiculos.objects.get(placa=persona_id)
            if rev.random == ran:
                mensaje = {
                    'mensaje': 'REVISIÓN DEL VEHÍCULO',
                }
            else:
                mensaje = {
                    'mensaje': 'PASE',
                }

            if rev.estado_pl == False:
                r = VehiculoRegistro(
                    vehiculo_id=persona_id,
                    fch_ingreso=fch,
                    fch_ingreso_txt=fch_txt,
                    hr_ingreso=hr,
                    hr_ingreso_txt=hr_txt,
                    usuario_registra_id=usuario_registra,
                    ubica='LA',
                ) 
                rev.estado_pl=True
                rev.save()
                r.save()

                return  render(request, template, {'ran':mensaje, 'reg':ingresos,'contadorp':contador})
            elif rev.estado_pl == True:
                r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
                r.fch_salida = fch
                r.fch_salida_txt = fch_txt
                r.hr_salida = hr
                r.hr_salida_txt = hr_txt
                rev.estado_pl = False
                rev.save()
                r.save()
                
                return  render(request, template, {'reg':ingresos,'contadorp':contador})

    return render(request, template, {'reg':ingresos,'contadorp':contador})

@login_required
@permission_required('controlPersonal.registro_vehiculo', raise_exception=True)
def registroAlmuerzos(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    q1 = AlmuerzoRegistro.objects.filter(fecha_txt__gte=fechaacttxt, fecha_txt__lte=fechatxt)
    queryset = q1
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/almuerzo.html'

    if request.method == 'POST':
        persona_id = request.POST.get('persona')
        persona_id = persona_id.upper()
        fch = request.POST.get('fch')
        fch_txt = request.POST.get('fch_txt')
        hr = request.POST.get('hr')
        hr_txt = request.POST.get('hr_txt')
        usuario_registra = request.POST.get('user')
        
        # cont = len(persona_id)
        # if cont > 7:
        queryset = Persona.objects.get(cedula=persona_id)
        if queryset.estado_al == False:
            r = AlmuerzoRegistro(
                persona_id=persona_id,
                fecha=fch,
                fecha_txt=fch_txt,
                hr_salida_almuerzo=hr,
                hr_salida_almuerzo_txt=hr_txt,
                usuario_registra_id=usuario_registra,
            )
            queryset.estado_al=True
            queryset.save()
            r.save()
        elif queryset.estado_al == True:
            r = AlmuerzoRegistro.objects.get(persona=persona_id, hr_entrada_almuerzo_txt__isnull=True)
            r.hr_entrada_almuerzo = hr
            r.hr_entrada_almuerzo_txt = hr_txt
            queryset.estado_al = False
            queryset.save()
            r.save()

        return redirect('personal:almuerzo')
        # else:
        #     ran = random.randint(1,7)
        #     rev = Vehiculos.objects.get(placa=persona_id)
        #     if rev.random == ran:
        #         mensaje = {
        #             'mensaje': 'REVISIÓN DEL VEHÍCULO',
        #         }
        #     else:
        #         mensaje = {
        #             'mensaje': 'PASE',
        #         }

        #     if rev.estado_pl == False:
        #         r = VehiculoRegistro(
        #             vehiculo_id=persona_id,
        #             fch_ingreso=fch,
        #             fch_ingreso_txt=fch_txt,
        #             hr_ingreso=hr,
        #             hr_ingreso_txt=hr_txt,
        #             usuario_registra_id=usuario_registra,
        #         ) 
        #         rev.estado_pl=True
        #         rev.save()
        #         r.save()

        #         return  render(request, template, {'ran':mensaje, 'reg':ingresos})
        #     elif rev.estado_pl == True:
        #         r = VehiculoRegistro.objects.get(vehiculo=persona_id, hr_salida_txt__isnull=True)
        #         r.fch_salida = fch
        #         r.fch_salida_txt = fch_txt
        #         r.hr_salida = hr
        #         r.hr_salida_txt = hr_txt
        #         rev.estado_pl = False
        #         rev.save()
        #         r.save()
                
        #         return  render(request, template, {'reg':ingresos})

    return render(request, template, {'reg':ingresos})

@login_required
@permission_required('controlPersonal.reporte_ingresos', raise_exception=True)
def reporteTablet(request):
    fechadesde = request.GET.get("datetimepicker1")
    fechahasta = request.GET.get("datetimepicker2")
    print(fechahasta)
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    # q1 = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).values('persona', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'persona__first_name', 'persona__last_name', 'ord')
    q1 = PersonaRegistro.objects.all().select_related("persona").order_by("-ord")
    q2 = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt).values('vehiculo', 'fch_ingreso', 'hr_ingreso_txt', 'fch_salida', 'hr_salida_txt', 'vehiculo__persona__first_name', 'vehiculo__persona__last_name', 'ord')
    queryset = q1.union(q2)
    # paginator = Paginator(queryset.order_by('-ord'), 50)
    paginator = Paginator(q1, 6)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    template = 'controlPersonal/tablet.html'

    busqueda = request.GET.get("buscar")
    
    if busqueda:
        q1 = PersonaRegistro.objects.all().select_related("persona").order_by("-ord").filter(
                Q(persona__first_name__icontains = busqueda) | 
                Q(persona__last_name__icontains = busqueda) |
                Q(persona__cedula__icontains = busqueda))
        
        fechab = datetime.today()
        fuente = 'busqueda_rapida'
        
        for i in range(1):
            r = TrackBusqueda(
                personabb = busqueda,
                personab = request.user.username,
                fecha = fechab,
                source = fuente,
                )
            r.save()

        paginator = Paginator(q1, 6)
        page = request.GET.get('page')
        ingresos = paginator.get_page(page)

        

        return render(request, template, {'reg':ingresos,'busqueda':busqueda,'fechadesde':fechadesde,'fechahasta':fechahasta})

        
    return render(request, template, {'reg':ingresos})


@login_required
@permission_required('controlPersonal.reporte_ingresos', raise_exception=True)
def reporteIngresos(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)

    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    print(inicio)
    print(fin)
    # ingresos = ''
    mensaje = ''

    if inicio != 'None':
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        if inicio and fin:
            iniciotxt = iniciodate.strftime('%Y%m%d')
            fintxt = findate.strftime('%Y%m%d')
            filtro = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt)
            paginator = Paginator(filtro.order_by('-fch_ingreso','hr_ingreso_txt'), 50)
            page = request.GET.get('page')
            ingresos = paginator.get_page(page)
        else:
            mensaje = 'No selecciono las fechas de busqueda'

    return render(request, 'controlPersonal/reporte.html', {'reg':ingresos, 'msj':mensaje, 'inicio':inicio, 'fin':fin})
 
def exportExcel(request, inicio, fin):
    #try:
    inicio = str(inicio)
    fin = str(fin)
    iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
    findate = datetime.strptime(fin, '%Y-%m-%d')
    iniciotxt = iniciodate.strftime('%Y%m%d')
    fintxt = findate.strftime('%Y%m%d')
            
    filtro = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt).order_by('-fch_ingreso','hr_ingreso_txt')  

    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'REPORTE DE INGRESOS DE PERSONAL' + '  -  ' + inicio + '  -  ' + fin

    ws.merge_cells('A1:H1')
    ws['A3'] = 'N°'
    ws['B3'] = 'CEDULA'
    ws['C3'] = 'NOMBRES'
    ws['D3'] = 'CODIGO'
    ws['E3'] = 'AREA'
    ws['F3'] = 'FECHA INGRESO'
    ws['G3'] = 'HORA INGRESO'
    ws['H3'] = 'FECHA SALIDA'
    ws['I3'] = 'HORA SALIDA'
    ws['J3'] = 'UBICACIÓN'

    count = 4
    rowcount = 1  

    for i in filtro:
        ws.cell(row = count, column = 1).value =  rowcount
        ws.cell(row = count, column = 2).value =  str(i.persona_id)
        ws.cell(row = count, column = 3).value =  str(i.persona)
        ws.cell(row = count, column = 4).value =  str(i.persona.cod_empleado)
        ws.cell(row = count, column = 5).value =  str(i.persona.area.area_nombre)
        ws.cell(row = count, column = 6).value =  str(i.fch_ingreso)
        ws.cell(row = count, column = 7).value =  str(i.hr_ingreso_txt)
        ws.cell(row = count, column = 8).value =  str(i.fch_salida)
        ws.cell(row = count, column = 9).value =  str(i.hr_salida_txt)
        if i.ubica == 'MCH':
            ws.cell(row = count, column = 10).value =  'PLANTA MACHACHI'
        elif i.ubica == 'LA':
            ws.cell(row = count, column = 10).value =  'LA AVELINA'
        elif i.ubica == 'LL':
            ws.cell(row = count, column = 10).value =  'LA LAURITA'
        elif i.ubica == 'AC':
            ws.cell(row = count, column = 10).value =  'AGUA CLARA'
        else:
            ws.cell(row = count, column = 10).value =  'LA MERCED'

        count+=1
        rowcount+=1

    nombre_archivo = "REPORTE DE INGRESOS PERSONAL" + inicio + "_" + fin + ".xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response
    #except Exception as e:
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = PersonaRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    # mensaje = 'Error'
    # mensaje_e = 'Error en los valores, por favor verifique las fechas'
    #'msje':mensaje_e,

    return render(request, 'controlPersonal/reporte.html', {'reg':ingresos, 'msje2':mensaje,  'inicio':inicio, 'fin':fin})

@login_required
@permission_required('controlPersonal.reporte_vehiculos', raise_exception=True)
def reporteIngresosVehiculos(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)

    paginator = Paginator(queryset.order_by('-ord'), 50)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    revision = request.GET.get("revision")


    print(inicio)
    print(fin)
    print(revision)
    # ingresos = ''
    mensaje = ''

    if inicio != 'None':
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        if inicio and fin:
            if inicio and fin and revision:
                print("Si llega")
                iniciotxt = iniciodate.strftime('%Y%m%d')
                fintxt = findate.strftime('%Y%m%d')
                filtro = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt).filter(revision_vehiculo=True)
                paginator = Paginator(filtro.order_by('-fch_ingreso','hr_ingreso_txt'), 50)
                page = request.GET.get('page')
                ingresos = paginator.get_page(page)

            else:
                print("Solo inicio y fin")
                iniciotxt = iniciodate.strftime('%Y%m%d')
                fintxt = findate.strftime('%Y%m%d')
                filtro = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt)
                paginator = Paginator(filtro.order_by('-fch_ingreso','hr_ingreso_txt'), 50)
                page = request.GET.get('page')
                ingresos = paginator.get_page(page)
        else:
            mensaje = 'No selecciono las fechas de busqueda'

    return render(request, 'controlPersonal/reporte_vehiculos.html', {'reg':ingresos, 'msj':mensaje, 'inicio':inicio, 'fin':fin})
 
def exportExcelVehiculos(request, inicio, fin):
    try:
        inicio = str(inicio)
        fin = str(fin)
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        iniciotxt = iniciodate.strftime('%Y%m%d')
        fintxt = findate.strftime('%Y%m%d')
            
        filtro = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt).order_by('-fch_ingreso','hr_ingreso_txt')  

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE INGRESOS DE VEHÍCULOS' + '  -  ' + inicio + '  -  ' + fin

        ws.merge_cells('A1:H1')
        ws['A3'] = 'N°'
        ws['B3'] = 'PLACA'
        ws['C3'] = 'NOMBRES'
        ws['D3'] = 'CEDULA'
        ws['E3'] = 'CODIGO'
        ws['F3'] = 'AREAS'
        ws['G3'] = 'FECHA INGRESO'
        ws['H3'] = 'HORA INGRESO'
        ws['I3'] = 'FECHA SALIDA'
        ws['J3'] = 'HORA SALIDA'
        ws['K3'] = 'UBICACIÓN'

        count = 4
        rowcount = 1  

        for i in filtro:
            ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 2).value =  str(i.vehiculo)
            ws.cell(row = count, column = 3).value =  str(i.vehiculo.persona)
            ws.cell(row = count, column = 4).value =  str(i.vehiculo.persona.cedula)
            ws.cell(row = count, column = 5).value =  str(i.vehiculo.persona.cod_empleado)
            ws.cell(row = count, column = 6).value =  str(i.vehiculo.persona.area)
            ws.cell(row = count, column = 7).value =  str(i.fch_ingreso)
            ws.cell(row = count, column = 8).value =  str(i.hr_ingreso_txt)
            ws.cell(row = count, column = 9).value =  str(i.fch_salida)
            ws.cell(row = count, column = 10).value =  str(i.hr_salida_txt)
            if i.ubica == 'MCH':
                ws.cell(row = count, column = 11).value =  'PLANTA MACHACHI'
            elif i.ubica == 'LA':
                ws.cell(row = count, column = 11).value =  'LA AVELINA'
            elif i.ubica == 'LL':
                ws.cell(row = count, column = 11).value =  'LA LAURITA'
            elif i.ubica == 'AC':
                ws.cell(row = count, column = 11).value =  'AGUA CLARA'
            else:
                ws.cell(row = count, column = 11).value =  'LA MERCED'

            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE INGRESOS VEHICULOS " + inicio + "_" + fin + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        ahora = datetime.now()
        ayer = ahora - timedelta(days=1)
        fecha = ahora.strftime('%Y-%m-%d')
        fechatxt = ahora.strftime('%Y%m%d')
        fechaacttxt = ayer.strftime('%Y%m%d')
        queryset = VehiculoRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        paginator = Paginator(queryset.order_by('-ord'), 50)
        page = request.GET.get('page')
        ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'

    return render(request, 'controlPersonal/reporte_vehiculos.html', {'reg':ingresos, 'msje2':mensaje, 'msje':mensaje_e, 'inicio':inicio, 'fin':fin})

############ INGRESO DE CONTENEDORES ############

@login_required
@permission_required('controlPersonal.registro_contenedor', raise_exception=True)
def contenedorRegistro(request):
    try:
        # queryset = ContenedorRegistro.objects.all().order_by('-ord')
        # paginator = Paginator(queryset.order_by('-ord'), 50)
        # page = request.GET.get('page')
        # datos_consul = paginator.get_page(page)

        # datos_consul = ContenedorRegistro.objects.all().order_by('-ord')[:10]
        hasta = datetime.now()
        desde = hasta - timedelta(days=1)
        desdetxt = desde.strftime('%Y%m%d')
        hastatxt = hasta.strftime('%Y%m%d')
        datos_consul = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte = desdetxt, fch_ingreso_txt__lte=hastatxt).order_by('-ord')
        

        if request.GET.get('persona'):
            chofer = request.GET.get('persona')
            fch_txt = request.GET.get('fch_txts')
            queryset = ContenedorRegistro.objects.filter(chofer=chofer, fch_salida__isnull=True)
            
            if queryset:
                cedula = request.GET.get('persona')
                queryset = Chofer.objects.get(cedula=cedula)
                placas = Cabezales.objects.all()
                fch = request.GET.get('fch')
                hora = request.GET.get('hr_txt')
                consul = ContenedorRegistro.objects.get(chofer=chofer, fch_salida__isnull=True)

                return render(request, 'controlPersonal/ingreso_contenedor.html',{'datos':queryset, 'placa':placas, 'cedula':cedula, 'fch':fch, 'hora':hora, 'datos_consul':datos_consul, 'consul':consul, 'estado':'s'})
            else:
                cedula = request.GET.get('persona')
                queryset = Chofer.objects.get(cedula=cedula)
                placas = Cabezales.objects.all()
                fch = request.GET.get('fch')
                hora = request.GET.get('hr_txt')

                return render(request, 'controlPersonal/ingreso_contenedor.html',{'datos':queryset, 'placa':placas, 'cedula':cedula, 'fch':fch, 'hora':hora, 'datos_consul':datos_consul, 'estado':'i'})

        if request.method == 'POST':
            chofer = request.POST.get('chofer')
            fch = request.POST.get('fch2')
            fch_txt = request.POST.get('fch_txt')
            hr = request.POST.get('hr_txt3')
            hr_txt = request.POST.get('hr_txt2')
            contenedor = request.POST.get('contenedori')
            solocabezal = request.POST.get('solocabezal')
            if contenedor:
                contenedor = contenedor.upper()
            placa = request.POST.get('placa')
            placa = placa.upper()
            contenedor_salida = request.POST.get('contenedors')
            if contenedor_salida:
                contenedor_salida = contenedor_salida.upper()
            sello1 = request.POST.get('sello_1')
            sello2 = request.POST.get('sello_2')
            sello3 = request.POST.get('sello_3')
            sello_basc_1 = request.POST.get('sello_basc_1')
            sello_basc_2 = request.POST.get('sello_basc_2')
            sello_basc_3 = request.POST.get('sello_basc_3')
            sello_eco_1 = request.POST.get('sello_eco_1')
            sello_eco_2 = request.POST.get('sello_eco_2')
            sello_eco_3 = request.POST.get('sello_eco_3')
            sello_naviera = request.POST.get('sello_naviera')
            destino = request.POST.get('destino')

            queryset = ContenedorRegistro.objects.filter(chofer=chofer, fch_salida__isnull=True)

            if queryset:
                if solocabezal:
                    r = ContenedorRegistro.objects.get(chofer=chofer, fch_salida__isnull=True)
                    r.solo_cabezal = True
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    r.destino = None
                    r.save()

                else:
                    r = ContenedorRegistro.objects.get(chofer=chofer, fch_salida__isnull=True)
                    r.num_contenedor_salida = contenedor_salida
                    r.fch_salida = fch
                    r.fch_salida_txt = fch_txt
                    r.hr_salida = hr
                    r.hr_salida_txt = hr_txt
                    r.sello_basc_1 = sello_basc_1
                    r.sello_basc_2 = sello_basc_2
                    r.sello_basc_3 = sello_basc_3
                    r.sello_eco_1 = sello_eco_1
                    r.sello_eco_2 = sello_eco_2
                    r.sello_eco_3 = sello_eco_3
                    r.sello_naviera = sello_naviera
                    r.destino = destino
                    print(destino)
                    
                    if destino == 'GY':
                        fecha = datetime.strptime(hr,'%Y-%m-%d %H:%M')
                        r.tmaxgye = fecha + timedelta(hours=14,minutes=30)
                    elif destino == 'PS':
                        fecha = datetime.strptime(hr,'%Y-%m-%d %H:%M')
                        r.tmaxpsj = fecha + timedelta(hours=17)

                    r.save()
                    
            else:
                r = ContenedorRegistro(
                        chofer_id=chofer,
                        placa=placa,
                        num_contenedor_ingreso=contenedor,
                        fch_ingreso=fch,
                        fch_ingreso_txt=fch_txt,
                        hr_ingreso=hr,
                        hr_ingreso_txt=hr_txt,
                        sello_1=sello1,
                        sello_2=sello2,
                        sello_3=sello3,
                    )
                r.save()

        return render(request, 'controlPersonal/ingreso_contenedor.html', {'datos_consul':datos_consul})
    except Exception as e:
        print(e)
        
        datos_consul = ContenedorRegistro.objects.all().order_by('-ord')[:10]

        mensaje = 'Error'
        mensaje_e = 'Error, intente nuevamente, el número de cedula puede no existir'

        return render(request, 'controlPersonal/ingreso_contenedor.html', {'datos_consul':datos_consul, 'msje2':mensaje, 'msje':mensaje_e})

def editaRegConte(request, reg):
    queryset = ContenedorRegistro.objects.get(id=reg)
    template = 'controlPersonal/edita_contenedor.html'

    return render(request, template, {'form':queryset})

def registroGuardado(request):
    return render(request,'controlPersonal/registro_guardado_salida_chofer.html')

class editaChofer(UpdateView):
    model = Chofer
    template_name = 'controlPersonal/edita_chofer.html'
    form_class = RegistroChoferForm 
    success_url = reverse_lazy('personal:listar_chofer')

    def get_context_data(self, **kwargs):
        context = super(editaChofer, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        chofer = self.model.objects.get(cedula=pk)
        transpor = chofer.transportista.id
        print(transpor)
        if 'form' not in context:
            context['form'] = self.form_class()
            context['form2'] = 1
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        pk = kwargs['pk']
        chofer = self.model.objects.get(cedula=pk)
        form = self.form_class(request.POST, instance=chofer)
      
        if form.is_valid():
            form.save()
           
            return HttpResponseRedirect(self.get_success_url())
            #return HttpResponse("Exito!!")
        else:
            return HttpResponse("Error No se pudo guardar. Comuníquese con el Dep. de IT")

def registroChofer(request):
   
    if request.method == 'POST':
        form = RegistroChoferForm(request.POST)
        formdet = RegistroPlacaCabezal(request.POST)
       
        if form.is_valid():
            registro = form.save(commit=True)
            registro.save()
            
            return redirect('personal:registro_guardado_salida_chofer')
        
        else:
            return HttpResponse("No se pudo guardar. Comuníquese con del Departamento de IT")

    else:
        form = RegistroChoferForm(request.GET)
        form3 = RegistroPlacaCabezal()
        valor_estado = 1
        
        return render(request,'controlPersonal/ingreso_chofer.html', {'form':form,'form2':valor_estado,'form3':form3})


def ingresoNuevaPlaca(request):
    if request.method == 'GET':
        newplaca = request.GET.get("newplaca")
        cedula = request.GET.get('cedula')
        nombre = request.GET.get('nombres')
        celular = request.GET.get('celular')
        trans = request.GET.get('transportista')
        form = RegistroChoferForm()
        if newplaca:
            cabezales = Cabezales.objects.all().filter(placa=newplaca)
            if cabezales:
                print("Placa ya existe. Verifique")
            else:
                newregistro = Cabezales()
                newregistro.placa = newplaca
                newregistro.estado = 1
                newregistro.transportista_id = trans
                newregistro.save()
    
    print(cedula)
    print(nombre)
    print(celular)
    print(trans)
    # return redirect('personal:registro_chofer') 
    return render(request,'controlPersonal/ingreso_chofer.html', {'form':form, 'cedula':cedula, 'nombre':nombre, 'celular':celular, 'transportista':trans})

@login_required
def listarChoferes(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          choferes = Chofer.objects.all().order_by('transportista','nombres').filter(
              Q(cedula__icontains = queryset) |
              Q(nombres__icontains = queryset)).filter(estado=1)
              
          paginator = Paginator(choferes, 25)
          page = request.GET.get('page')
          choferx = paginator.get_page(page)
        
    else:
        choferes = Chofer.objects.filter(estado=1).order_by('transportista','nombres')
       
        paginator = Paginator(choferes, 25)

        page = request.GET.get('page')
        choferx = paginator.get_page(page)

    return render(request,'controlPersonal/busqueda_chofer.html',{'form':choferes})

@login_required
def listarVadministrativos(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
        contador = 1
        vehiculos = Vehiculos.objects.all().order_by('persona__last_name').filter(
            Q(placa__icontains = queryset) |
            Q(persona__cedula__icontains = queryset) |
            Q(persona__last_name__icontains = queryset) |
            Q(persona__first_name__icontains = queryset))
          
               
    else:
        contador = 0
        vehiculos = Vehiculos.objects.all().order_by('persona__last_name')
       
    paginator = Paginator(vehiculos, 25)
    page = request.GET.get('page')
    vehi = paginator.get_page(page)
    

    return render(request,'controlPersonal/gestiona_vadministrativos.html',{'form':vehi,'contador':contador})


@login_required
def nuevaPantallaReportesPersonal(request):
    return render(request,'controlPersonal/reportes_personal_new.html')

class choferDelete(DeleteView):
    model = Chofer
    template_name = 'controlPersonal/eliminar_chofer.html'
    success_url = reverse_lazy('personal:listar_chofer')

def activaChofer(request, cedula):
    cedula = cedula
    print(cedula)
    chofer = Chofer.objects.get(cedula=cedula)
    chofer.estado = 2
    chofer.save()

    return redirect('personal:listar_chofer')

@login_required
@permission_required('controlPersonal.reporte_contenedor', raise_exception=True)
def reporteContenedor(request):
    ahora = datetime.now()
    ayer = ahora - timedelta(days=1)
    fecha = ahora.strftime('%Y-%m-%d')
    fechatxt = ahora.strftime('%Y%m%d')
    fechaacttxt = ayer.strftime('%Y%m%d')
    queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
    paginator = Paginator(queryset.order_by('-ord'), 150)
    page = request.GET.get('page')
    ingresos = paginator.get_page(page)

    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    # ingresos = ''
    mensaje = ''
    if request.GET.get("inicio"):
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        if inicio and fin:
            iniciotxt = iniciodate.strftime('%Y%m%d')
            fintxt = findate.strftime('%Y%m%d')
            filtro = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt)
            paginator = Paginator(filtro.order_by('fch_ingreso','hr_ingreso_txt'), 150)
            page = request.GET.get('page')
            ingresos = paginator.get_page(page)
        else:
            mensaje = 'No selecciono las fechas de busqueda'

    return render(request, 'controlPersonal/reporte_contenedores.html', {'reg':ingresos, 'msj':mensaje, 'inicio':inicio, 'fin':fin, 'cod':'EC0F802'})

def exportExcelConte(request, inicio, fin):
    try:
        inicio = str(inicio)
        fin = str(fin)
        iniciodate = datetime.strptime(inicio, '%Y-%m-%d')
        findate = datetime.strptime(fin, '%Y-%m-%d')
        iniciotxt = iniciodate.strftime('%Y%m%d')
        fintxt = findate.strftime('%Y%m%d')
            
        filtro = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=iniciotxt, fch_ingreso_txt__lte=fintxt).order_by('-fch_ingreso','hr_ingreso_txt')  

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE INGRESOS DE CONTENEDORES' + '  -  ' + inicio + '  -  ' + fin

        ws.merge_cells('A1:U1')
        ws['A3'] = 'N°'
        ws['B3'] = 'CHOFER'
        ws['C3'] = 'PLACA'
        ws['D3'] = 'NOMBRES'
        ws['E3'] = 'TRANSPORTISTA'
        ws['F3'] = 'CONTENEDOR INGRESO'
        ws['G3'] = 'FECHA Y HORA INGRESO'
        #ws['H3'] = 'HORA INGRESO'
        ws['H3'] = 'SELLO 1'
        ws['I3'] = 'SELLO 2'
        ws['J3'] = 'SELLO 3'
        ws['K3'] = 'CONTENEDOR SALIDA'
        ws['L3'] = 'FECHA Y HORA SALIDA'
        #ws['N3'] = 'HORA SALIDA'
        ws['M3'] = 'SELLO BASC 1'
        ws['N3'] = 'SELLO BASC 2'
        ws['O3'] = 'SELLO BASC 3'
        ws['P3'] = 'SELLO ECOFROZ 1'
        ws['Q3'] = 'SELLO ECOFROZ 2'
        ws['R3'] = 'SELLO ECOFROZ 3'
        ws['S3'] = 'SELLO NAVIERA'
        ws['T3'] = 'DESTINO'
        ws['U3'] = 'TIEMPO LLEGADA GYE'
        ws['V3'] = 'TIEMPO LLEGADA PSJ'
      
        count = 4
        rowcount = 1  

        for i in filtro:
            ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 2).value =  str(i.chofer.cedula)
            ws.cell(row = count, column = 3).value =  str(i.placa)
            ws.cell(row = count, column = 4).value =  str(i.chofer.nombres)
            ws.cell(row = count, column = 5).value =  str(i.chofer.transportista.transportista)
            ws.cell(row = count, column = 6).value =  str(i.num_contenedor_ingreso)
            if i.fch_ingreso:
                ws.cell(row = count, column = 7).value =  str(i.fch_ingreso) + ' ' + str(i.hr_ingreso_txt)
            else:
                ws.cell(row = count, column = 7).value = "--"
            #ws.cell(row = count, column = 8).value =  str(i.hr_ingreso_txt)
            ws.cell(row = count, column = 8).value =  str(i.sello_1)
            ws.cell(row = count, column = 9).value =  str(i.sello_2)
            ws.cell(row = count, column = 10).value =  str(i.sello_3)
            ws.cell(row = count, column = 11).value =  str(i.num_contenedor_salida)
            if i.fch_salida:
                ws.cell(row = count, column = 12).value =  str(i.fch_salida) + ' ' + str(i.hr_salida_txt)
            else:
                ws.cell(row = count, column = 12).value = "--"
        
        #    ws.cell(row = count, column = 14).value =  str(i.hr_salida_txt)
            ws.cell(row = count, column = 13).value =  str(i.sello_basc_1)
            ws.cell(row = count, column = 14).value =  str(i.sello_basc_2)
            ws.cell(row = count, column = 15).value =  str(i.sello_basc_3)
            ws.cell(row = count, column = 16).value =  str(i.sello_eco_1)
            ws.cell(row = count, column = 17).value =  str(i.sello_eco_2)
            ws.cell(row = count, column = 18).value =  str(i.sello_eco_3)
            ws.cell(row = count, column = 19).value =  str(i.sello_naviera)
            if i.destino:
                ws.cell(row = count, column = 20).value =  str(i.get_destino_display())
            else:
                ws.cell(row = count, column = 20).value = "--"

            if i.tmaxgye:
                ws.cell(row = count, column = 21).value =  str(i.tmaxgye)
            else:
                ws.cell(row = count, column = 21).value = "--"
            if i.tmaxpsj:
                ws.cell(row = count, column = 22).value =  str(i.tmaxpsj)
            else:
                ws.cell(row = count, column = 22).value =  "--"
           
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE INGRESOS CONTENEDOR" + inicio + "_" + fin + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        ahora = datetime.now()
        ayer = ahora - timedelta(days=1)
        fecha = ahora.strftime('%Y-%m-%d')
        fechatxt = ahora.strftime('%Y%m%d')
        fechaacttxt = ayer.strftime('%Y%m%d')
        queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        paginator = Paginator(queryset.order_by('-ord'), 50)
        page = request.GET.get('page')
        ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'

        return render(request, 'controlPersonal/reporte_contenedores.html', {'reg':ingresos, 'msje2':mensaje, 'msje':mensaje_e, 'inicio':inicio, 'fin':fin})

@login_required
def ajax_call_placa(request):
    placa = request.GET["placa"]
    consulta = Vehiculos.objects.filter(placa=placa).select_related('persona') 
    
    context = []
    for i in consulta:
        d = {
            'propietario': i.persona.first_name + ' '+ i.persona.last_name,
            'placa': i.placa,
            'estado':i.estado, 
            'referencia':i.referencia,
        }
        context.append(d)

    return JsonResponse (context, safe=False)

@login_required
def actualizaEstadoVehiAdministrativo(request):
    if request.method == 'POST':
        valor_radio = request.POST.get("radio")
        placa = request.POST.get("placa")
        referencia = request.POST.get("referencia")
        fecha = datetime.today()

        if valor_radio == "SI":
            print("Ya!")
            actualiza = Vehiculos.objects.filter(placa=placa).update(estado=1,fecha_modifica=fecha,persona_edita=request.user.username,
            referencia=referencia)
        else:
            print("No!")
            actualiza = Vehiculos.objects.filter(placa=placa).update(estado=0,fecha_modifica=fecha,persona_edita=request.user.username,
            referencia=referencia)
        
    return redirect ('personal:listar_vadministrativos')
