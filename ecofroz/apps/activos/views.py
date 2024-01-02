import logging
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse, FileResponse
from django.urls import reverse_lazy, reverse
from .forms import RegistroForm,FiltrarForm,NoEncontradosForm,UbicacionesBodegaForm, ImageForm
from django.views.generic import ListView,CreateView,UpdateView,DeleteView,DetailView, TemplateView
from .models import desc_activo, detalle_desc_activo, activos_temp_excel, activo_nomenclatura, activo_estados_select, salida_activos, activo_depar, activo_grupo, activo_areas, historial_movimientos_internos, \
    activo_ubica,activo_areas,activo_depar,activo_subsector, no_encontrados,activo_tipo, Imagenes 
from apps.ordenPedido.models import SecuencialCodifica
from .forms import RegistroForm, EditaActivo, EditaActivoIngreso, RegistroFormDet, FiltrarForm, MoverActivosForm, \
 CambiaEstadoForm, RegistroSalida, RegistroSalidaAgri, FiltrarMarca, ImageForm, ActivoForm, FiltrarSerie
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.paginator import Paginator
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from django.views import View
from django.db import connection
from django.forms import DateTimeInput, modelformset_factory
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import A4 
from reportlab.platypus import Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from io import BytesIO
from django.conf import settings
import io
import os
from django.db.models import Q, Count
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.template import context, loader
from django.utils import formats, timezone
from apps.ordenTrabajo.models import OrdenesTrabajos,DetalleTrabajo
from apps.ordenPedido.models import OrdenesPedidos,DetallePedido
from apps.controlPersonal.models import Persona
from apps.proveedores.models import proveedor, proveedor_det
from apps.ordenPedido.models import SecuencialCodifica
from apps.usuarios.models import Departamento, Autorizador, User
from datetime import date
import unicodedata
from django.db.models import Subquery, OuterRef, CharField, Value
from .clases import *
import pytz


#from .filters import SnippetFilter

logger = logging.getLogger(__name__)



### FUNCIONES INTERNAS ###

def envioMail(subject, email, template, queryset, queryset2):
    html_message = loader.render_to_string(
        'activos/email/%s' %template,
            {
                'aprob':queryset,
                'aprob2':queryset2,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except:
        logger.error("Unable to send mail.")

def envioMailNoCustodio(subject, email, template, queryset, usuario):
    html_message = loader.render_to_string(
        'activos/email/%s' %template,
            {
                'aprob':queryset,
                'usuario':usuario,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except:
        logger.error("Unable to send mail.")

### VISTAS ###

#Función para mostrar activos encargados en custodio por persona

@login_required
def listar_mis_activos_en_custodio(request):

    usuario = request.user.username

    persona = Persona.objects.get(username_sia=usuario)

    activos = detalle_desc_activo.objects.filter(desc_activo_custodio__icontains=persona.last_name).select_related("desc_activo_codigo")

    return render(request, 'activos/listar_mis_activos.html',{'form':activos, 'usuario':request.user})



#Funcion para mostrar activos que fueron entregados y que requieren descargo en custodios por parte de seguridad

@login_required
def listarRecepcionesforSeg(request):
    queryset = request.GET.get("buscar")
    codigo = request.GET.get("activo_codigo")
    nombre_custodio = request.GET.get("nombre_custodio")

    if queryset:
        recepcion = DetallePedido.objects.filter(
            Q(numpedido__numpedido__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numpedido__genera_compra=1).filter(Q(numpedido__codigo_genera=None) | Q(numpedido__codigo_genera=False)).order_by('numpedido')
        
        entregado = DetallePedido.objects.filter(numpedido__genera_compra=1).filter(numpedido__entregado_bodega=True).order_by('numpedido')
    else:
        
        entregadox = DetallePedido.objects.filter(numpedido__genera_compra=1).filter(numpedido__entregado_bodega=True).order_by('-numpedido__fecha_entrega')
      
        paginator = Paginator(entregadox, 50)
        page = request.GET.get('page')
        entregado = paginator.get_page(page)

        parametros = FiltrarForm(request.GET or None) 

        if codigo:
            personal = Persona.objects.all().order_by('last_name')
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                Q(desc_activo_codigo__activo_codigo__iexact = codigo )|
                Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo)).exclude(desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO')

            return render(request, 'activos/descargo_activos.html', {'entregado':entregado,'form':activosx,'form2':parametros,'personal':personal})
        
        elif nombre_custodio:
            print("entró DM")
            nombre = nombre_custodio.replace("  "," ")
            print(nombre)
            personal = Persona.objects.all().order_by('last_name')
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                Q(desc_activo_custodio__icontains = nombre )).exclude(desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO')
           
            

            return render(request, 'activos/descargo_activos.html', {'entregado':entregado,'form':activosx,'form2':parametros,'personal':personal,'nombre':nombre})

        else:
            activosx=None
            personal = Persona.objects.all().order_by('last_name')

            return render(request, 'activos/descargo_activos.html', {'entregado':entregado,'form':activosx,'form2':parametros,'personal':personal})


#Función para realizar la confimación del custodio definitivo

@login_required
def descargo_custodio(request,numpedido):
    #pedido = request.GET.get('pedido')
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset_activo = desc_activo.objects.get(activo_codigo = queryset.cod_activo)
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    

    if request.method == 'POST':
        newcustodi = unicodedata.normalize('NFC', request.POST.get("nombre_custodio"))
        newcustodio = newcustodi.replace("  "," ") 
        nuevo_depar = activo_depar.objects.get(id=request.POST.get("activo_depar"))
        nueva_area = activo_areas.objects.get(area_codigo=request.POST.get("activo_area"))
        
        queryset.descargo_custodio_observa = request.POST.get("observaciones")
        queryset.fecha_descargo_custodio = datetime.now()
        queryset.persona_confirma_descargo = request.POST.get("usuario_registra")
        queryset.custodio_sugerido = request.POST.get("nombre_custodio")
        queryset.descargo_custodio = True
        queryset.save()

        actualiza = desc_activo.objects.filter(activo_codigo=queryset.cod_activo).update(activo_depar=nuevo_depar,activo_area=nueva_area)
        actualiza_custodio = detalle_desc_activo.objects.filter(desc_activo_codigo__activo_codigo=queryset.cod_activo).update(desc_activo_custodio=newcustodio)
        
        return redirect('activos:listar_entregas_activos')
    
    user = request.user
    personal = Persona.objects.all().order_by('last_name')

    return render(request, 'activos/descargo_custodio.html',{'queryset':queryset,'queryset_activo':queryset_activo,'depar':depar,'area':area,'user':user,'personal':personal})


#Función para realizar el reporte de activos que no están bajo custodio usuarios
@login_required
def reportar_no_custodio(request,id):
    #pedido = request.GET.get('pedido')
    #queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset_activo = desc_activo.objects.filter(id = id)
    #queryset_activo = detalle_desc_activo.objects.filter(desc_activo_codigo=id).select_related("desc_activo_codigo")
    
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    
    

    if request.method == 'POST':
        usuario = request.user
        queryset = desc_activo.objects.filter(id = id)
      
        observaciones_reporte_no_custodio = detalle_desc_activo.objects.filter(desc_activo_codigo__id=id).update(
            observaciones_reporte_no_custodio=request.POST.get("observaciones"),
            fecha_envia_reporte_no_custodio=date.today())

        envioMailNoCustodio('Reporte de Activo no encontrado o Custodio mal asignado', 'dmencias@ecofroz.com, fortiz@ecofroz.com', 'email_reporte_activo_custodio_no_encontrado.html', queryset, usuario)


        return redirect('activos:listar_mis_activos')

    
    user = request.user
    personal = Persona.objects.all().order_by('last_name')

    return render(request, 'activos/reportar_no_custodio.html',{'queryset_activo':queryset_activo,'depar':depar,'area':area,'user':user,'personal':personal})



#Función para realizar la confimación del custodio definitivo de activos usados

@login_required
def descargo_custodio_activold(request,id):
    #pedido = request.GET.get('pedido')
    #queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
    queryset_activo = desc_activo.objects.get(id = id)
    #queryset_activo = detalle_desc_activo.objects.filter(desc_activo_codigo=id).select_related("desc_activo_codigo")
    
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    

    if request.method == 'POST':
        newcustodi = unicodedata.normalize('NFC', request.POST.get("nombre_custodio"))
        newcustodio = newcustodi.replace("  "," ") 
        print(newcustodio)
        nuevo_depar = activo_depar.objects.get(id=request.POST.get("activo_depar"))
        nueva_area = activo_areas.objects.get(area_codigo=request.POST.get("activo_area"))
        

        actualiza = desc_activo.objects.filter(id=id).update(activo_depar=nuevo_depar,activo_area=nueva_area)
        borra_custodio_actual = detalle_desc_activo.objects.filter(desc_activo_codigo__id=id).update(
            desc_activo_custodio='')
        actualiza_custodio = detalle_desc_activo.objects.filter(desc_activo_codigo__id=id).update(
            desc_activo_custodio=newcustodio,
            desc_activo_usuario_registra_cambio_custodio=request.POST.get("usuario_registra"),
            desc_activo_observaciones_seg_cambio_custodio=request.POST.get("observaciones"),
            fecha_cambia_custodio=date.today())

        return redirect('activos:listar_entregas_activos')

    
    user = request.user
    personal = Persona.objects.all().order_by('last_name')

    return render(request, 'activos/descargo_custodiold.html',{'queryset_activo':queryset_activo,'depar':depar,'area':area,'user':user,'personal':personal})


@login_required
def descargo_custodio_activold_custom(request,id):
    queryset_activo = detalle_desc_activo.objects.filter(desc_activo_codigo=id).select_related("desc_activo_codigo")
    depar = activo_depar.objects.all()
    area = activo_areas.objects.all()
    
    if request.method == 'POST':
        newcustodi = unicodedata.normalize('NFC', request.POST.get("nombre_custodio"))
        newcustodio = newcustodi.replace("  "," ") 
        print("Hola mundo")
        nuevo_depar = activo_depar.objects.get(id=request.POST.get("activo_depar"))
        nueva_area = activo_areas.objects.get(area_codigo=request.POST.get("activo_area"))

        print(nuevo_depar)
        print(nueva_area)
        

        actualiza = desc_activo.objects.filter(id=id).update(activo_depar=nuevo_depar,activo_area=nueva_area)
        actualiza_custodio = detalle_desc_activo.objects.filter(desc_activo_codigo__id=id).update(
            desc_activo_custodio=newcustodio,desc_activo_usuario_registra_cambio_custodio=request.POST.get(
                "usuario_registra"),desc_activo_observaciones_seg_cambio_custodio=request.POST.get(
                    "observaciones"),fecha_cambia_custodio=date.today())
        
        return HttpResponseRedirect(reverse('activos:acta_de_entrega_custodio2',kwargs={'id':id}))
    

    



# Función para cambiar el estado del proceso de entrega de acta de recepción de activos

@login_required
def confirma_acta_recepcion_entregada(request,numpedido):
    #pedido = request.GET.get('pedido')
    queryset = OrdenesPedidos.objects.get(numpedido=numpedido)
        
    queryset.acta_entregada = True
    queryset.save()
        
    return redirect('activos:listar_entregas_activos')


# Vista para mostrar Fotografías de Activos

@login_required
def add_activo_view(request,pk):
    ImageFormSet = modelformset_factory(Imagenes, form=ImageForm, extra=3)
    codigo = desc_activo.objects.get(id=pk)
  
    if request.method == "GET":
        
        formset = ImageFormSet(queryset=Imagenes.objects.none())
        return render(request, 'activos/activos_gallery.html', {'formset':formset,'pk':pk})
    
    elif request.method == "POST":
        
       
        formset = ImageFormSet(request.POST, request.FILES)

        if formset.is_valid():
            
            for form in formset.cleaned_data:
                if form:
                    image = form['imagen_activo']
                    Imagenes.objects.create(imagen_activo=image, activo_codigo=codigo)
            return HttpResponse('Exito')
        else:
            print(formset.errors)
@login_required
def gallery_view(request, pk):
    activo = desc_activo.objects.get(id=pk)
    return render(request, 'activos/activos_gallery_view.html', {"activo":activo})

@login_required
def gallery_view_custom(request, pk):
    activo = desc_activo.objects.get(id=pk)
    return render(request, 'activos/activos_gallery_view_custom.html', {"activo":activo})

@login_required
def gallery_mis_activos(request, pk):
    activo = desc_activo.objects.get(id=pk)
    return render(request, 'activos/activos_gallery_mis_activos.html', {"activo":activo})

@login_required
def home(request):
    return render(request,"base.html")


@login_required
def registro_nuevo(request):
   
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        formdet = RegistroFormDet(request.POST)

        if form.is_valid() and formdet.is_valid():
            registro = formdet.save(commit=False)
            registro.desc_activo_codigo = form.save()
            registro.save()
            
            return redirect('activos:registro_guardado_salida')
        
        else:
            return HttpResponse("No se pudo guardar")

    else:
        form = RegistroForm()
        formdet = RegistroFormDet()
        
        return render(request,'activos/registro_activos.html', {'form':formdet,'form2':form})

@login_required
def registro_nuevov2(request):
   
    if request.method == 'POST':
        codigo = request.POST.get("codigo")
        numero = request.POST.get("numero")
        form = RegistroForm(request.POST)
        # formdet = EditaActivo(request.POST)
        formdet = EditaActivoIngreso(request.POST)

        if form.is_valid() and formdet.is_valid():
            registro = formdet.save(commit=False)
            registro.desc_activo_codigo = form.save()
            registro.save()
            for i in range(1):
                r = SecuencialCodifica(
                    codigo=codigo,
                    numeracion=numero,
                )
                r.save()
            
            return redirect('activos:registro_guardado')
        
        else:
            return HttpResponse("No se pudo guardar. Comuníquelo al Departamento de Tecnología")

    else:
        form = RegistroForm(initial={'grabado':'NO'})
        formdet = EditaActivoIngreso()
        
        return render(request,'activos/nuevo_ingreso_activos.html', {'form':formdet,'form2':form})

@login_required
def listarSalida(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = salida_activos.objects.all().order_by('-fecha_registro').filter(
            Q(id__icontains = queryset) |
              Q(detalle_activo__icontains = queryset) |
              Q(activo_num_serie__icontains = queryset) |
              Q(num_orden_trabajo__icontains = queryset) |
              Q(empresa_mantenimiento__icontains = queryset)).filter(solicitado_por=request.user.username)
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = salida_activos.objects.all().order_by('-fecha_registro').filter(solicitado_por=request.user.username)
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/listar_salidas.html',{'form':activos})

@login_required
def listarSalidaAgri(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = salida_activos.objects.all().order_by('-fecha_registro').filter(
              Q(detalle_activo__icontains = queryset) |
              Q(activo_num_serie__icontains = queryset) |
              Q(num_orden_trabajo__icontains = queryset) |
              Q(empresa_mantenimiento__icontains = queryset)).filter(persona_registra_en_proyecto=request.user.username)
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = salida_activos.objects.all().order_by('-fecha_registro').filter(persona_registra_en_proyecto=request.user.username)
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/listar_salidas_agri.html',{'form':activos})




@login_required
def listarSalidasAprobadas(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = salida_activos.objects.all().order_by('-fecha_registro').filter(
              Q(detalle_activo__icontains = queryset) |
              Q(activo_num_serie__icontains = queryset) |
              Q(num_orden_trabajo__icontains = queryset) |
              Q(empresa_mantenimiento__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = salida_activos.objects.all().order_by('-fecha_registro')
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/listar_salidas_aprobadas.html',{'form':activos})



@login_required
def listarMovimientosInternos(request,codigo):
    
    movimientos = historial_movimientos_internos.objects.all().filter(activo_codigo=codigo).order_by("-fecha_registro")

    
    return render(request,'activos/listar_historial_movimientos_internos.html',{'movimientos':movimientos,'codigo':codigo})

@login_required
def movimiento_a_bodega(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
        if request.POST.get('activo_subsec'):
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = activo_subsector.objects.filter(subsector_codigo=request.POST.get('activo_subsec')).distinct('subsector_nombre')[0]
            print(subsec)
            if form.is_valid():
                form.save()
                r = historial_movimientos_internos(
                activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                subsector = subsec,
                usuario_registra = request.POST.get('usuario_modifica'),
                justificacion_movimiento = request.POST.get('justificacion_modifica'),
                )
                r.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                registro.save()
                print ("Pasó por aquí cambio de custidio")

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")

        else:
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = ''
            print(subsec)
            if form.is_valid():
                form.save()
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.POST.get('usuario_modifica'),
                    justificacion_movimiento = request.POST.get('justificacion_modifica'),
                    )
                r.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                registro.desc_activo_motivo_modifica = 'cambio_ubicacion'
                registro.save()
                
            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")
 
        return redirect('activos:busqueda_activos_mantenimiento')
        
            
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.get(id=108)  #ADMINISTRACION
        area = activo_areas.objects.get(area_codigo=1010)  #BODEGA GENERAL
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':depar,
            'activo_area':area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        form = MoverActivosForm(initial=initial_data)

        return render(request,'activos/movimiento_a_bodega.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  


@login_required
def movimiento_a_bodega2(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
        if request.POST.get('activo_subsec'):
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = activo_subsector.objects.filter(subsector_codigo=request.POST.get('activo_subsec')).distinct('subsector_nombre')[0]
            print(subsec)
            if form.is_valid():
                form.save()
                r = historial_movimientos_internos(
                activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                subsector = subsec,
                usuario_registra = request.POST.get('usuario_modifica'),
                justificacion_movimiento = request.POST.get('justificacion_modifica'),
                )
                r.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                registro.save()
                print ("Pasó por aquí cambio de custidio")

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")

        else:
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = ''
            print(subsec)
            print(request.POST.get('activo_depar'))
            print(request.POST.get('activo_area'))

            if form.is_valid():
                form.save()
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.POST.get('usuario_modifica'),
                    justificacion_movimiento = request.POST.get('justificacion_modifica'),
                    )
                r.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                registro.desc_activo_motivo_modifica = 'cambio_ubicacion'
                registro.save()
                
            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")
 
        return redirect('activos:mueve_bodega')
        
            
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.get(id=108)  #ADMINISTRACION
        area = activo_areas.objects.get(area_codigo=1010)  #BODEGA GENERAL
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':depar,
            'activo_area':area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        form = MoverActivosForm(initial=initial_data)

        return render(request,'activos/movimiento_a_bodega.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  





@login_required
def movimiento_desde_bodega(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
        if request.POST.get('activo_subsec'):
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = activo_subsector.objects.filter(subsector_codigo=request.POST.get('activo_subsec')).distinct('subsector_nombre')[0]
            print(subsec)
            if form.is_valid():
                form.save()
                r = historial_movimientos_internos(
                activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                subsector = subsec,
                usuario_registra = request.POST.get('usuario_modifica'),
                usuario_retira = request.POST.get('usuario_retira'),
                justificacion_movimiento = request.POST.get('justificacion_modifica'),
                )
                r.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_retira')
                registro.save()
                print ("Pasó por aquí cambio de custidio")

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")

        else:
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = ''
            print(subsec)
            if form.is_valid():
                
                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_retira')
                registro.desc_activo_motivo_modifica = 'cambio_ubicacion'
                registro.save()

                #Cambia de ubicación al activo
                registroc = desc_activo.objects.get(activo_codigo=codigo)
                registroc.activo_ubica = activo_ubica.objects.get(id=4)
                registroc.activo_area = activo_areas.objects.get(area_codigo=request.POST.get('activo_area'))
                registroc.activo_depar = activo_depar.objects.get(id=request.POST.get('activo_depar'))
                registroc.activo_subsec = None
                registroc.save()

                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo_ubica.objects.get(id=4),
                    departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.POST.get('usuario_modifica'),
                    usuario_retira = request.POST.get('usuario_retira'),
                    justificacion_movimiento = request.POST.get('justificacion_modifica'),
                    )
                r.save()
         
            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")
 
        return redirect('activos:mov_success_bodega')
        
            
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.all()
        area = activo_areas.objects.all()
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':query.desc_activo_codigo.activo_depar,
            'activo_area':query.desc_activo_codigo.activo_area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        form = MoverActivosForm(initial=initial_data)

        return render(request,'activos/movimiento_desde_bodega.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  


@login_required
def cambio_estado_grabado_bodega(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
                
        #Cambia el estado grabado del activo
        registro = desc_activo.objects.get(activo_codigo=codigo)
        registro.grabado = request.POST.get('grabado')
        registro.save()

        #Falta incluir el registro del usuario que actualiza y el motivo de la actualización
        #Seguir Lunes desde aqui

        return redirect('activos:busqueda_movimientos_bodega')
        
            

    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.all()
        area = activo_areas.objects.all()
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':query.desc_activo_codigo.activo_depar,
            'activo_area':query.desc_activo_codigo.activo_area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        

        return render(request,'activos/cambio_estado_grabado_bodega.html',{'query':query, 'registra':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  

@login_required
def ver_detalle_bajas(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
        if request.POST.get('activo_subsec'):
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = activo_subsector.objects.filter(subsector_codigo=request.POST.get('activo_subsec')).distinct('subsector_nombre')[0]
            print(subsec)
            if form.is_valid():
                activo = desc_activo.objects.get(activo_codigo=codigo)
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo.activo_ubica,
                    departamento = activo.activo_depar,
                    sector = activo.activo_area, 
                    nuevo_ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    nuevo_departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    nuevo_sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.user,
                    )
                r.save()
                form.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                #registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                #registro.save()
                #print ("Pasó por aquí cambio de custidio")

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")

        else:
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = ''
            print(subsec)
        
            if form.is_valid():
                
                activo = desc_activo.objects.get(activo_codigo=codigo)
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo.activo_ubica,
                    departamento = activo.activo_depar,
                    sector = activo.activo_area, 
                    nuevo_ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    nuevo_departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    nuevo_sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.user,
                  
                    )
                r.save()
                form.save()

                #guarda el cambio de ubicación
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                #registro.desc_activo_custodio = request.POST.get('persona_recibe')
                registro.desc_activo_motivo_modifica = 'Cambio de ubicación y estado'
                registro.desc_activo_codigo.activo_ubica = activo_ubica.objects.get(id=request.POST.get('activo_ubica'))
                registro.desc_activo_codigo.activo_depar = activo_depar.objects.get(id=request.POST.get('activo_depar'))
                registro.desc_activo_codigo.activo_area = activo_areas.objects.get(area_codigo=request.POST.get('activo_area'))
                registro.usuario_modifica = request.user
                registro.save()
                

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar. Comuníquese con del departamento de IT")
 
        return redirect('activos:listar_bajas')
        
            
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.all()
        area = activo_areas.objects.all()
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':query.desc_activo_codigo.activo_depar,
            'activo_area':query.desc_activo_codigo.activo_area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        form = MoverActivosForm(initial=initial_data)

        return render(request,'activos/ver_detalle_bajas.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  



@login_required
def registro_movimiento_interno_func(request, codigo):
    print(codigo)
    activo = desc_activo.objects.get(activo_codigo=codigo)
    if request.method == 'POST':
        if request.POST.get('activo_subsec'):
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = activo_subsector.objects.filter(subsector_codigo=request.POST.get('activo_subsec')).distinct('subsector_nombre')[0]
            print(subsec)
            if form.is_valid():
                activo = desc_activo.objects.get(activo_codigo=codigo)
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo.activo_ubica,
                    departamento = activo.activo_depar,
                    sector = activo.activo_area, 
                    nuevo_ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    nuevo_departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    nuevo_sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.POST.get('usuario_modifica'),
                    justificacion_movimiento = request.POST.get('justificacion_modifica'),
                    )
                r.save()
                form.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('usuario_modifica')
                registro.save()
                print ("Pasó por aquí cambio de custidio")

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar")

        else:
            form = MoverActivosForm(request.POST,instance=activo)
            subsec = ''
            print(subsec)
        
            if form.is_valid():
                
                activo = desc_activo.objects.get(activo_codigo=codigo)
                r = historial_movimientos_internos(
                    activo_codigo = desc_activo.objects.get(activo_codigo=codigo),
                    ubicacion = activo.activo_ubica,
                    departamento = activo.activo_depar,
                    sector = activo.activo_area, 
                    nuevo_ubicacion = activo_ubica.objects.get(id=request.POST.get('activo_ubica')),
                    nuevo_departamento = activo_depar.objects.get(id=request.POST.get('activo_depar')),
                    nuevo_sector = activo_areas.objects.get(area_codigo=request.POST.get('activo_area')),
                    subsector = subsec,
                    usuario_registra = request.POST.get('usuario_modifica'),
                    usuario_retira = request.POST.get('persona_recibe'), #Se usa para registrar tambien al usaurio que recibe
                    justificacion_movimiento = request.POST.get('justificacion_modifica'),
                    )
                r.save()
                form.save()

                #Cambia de custodio al activo
                registro = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
                registro.desc_activo_custodio = request.POST.get('persona_recibe')
                registro.desc_activo_motivo_modifica = 'cambio_ubicacion'
                registro.save()
                

            else:
                print(form.errors)
                return HttpResponse("No se pudo guardar. Comuníquese con del departamento de IT")
 
        return HttpResponseRedirect(reverse('activos:mov_success',kwargs={'codigo':codigo}))
        
            
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.all()
        area = activo_areas.objects.all()
       

        initial_data = {
            'id':query.desc_activo_codigo.id,
            'activo_ubica':query.desc_activo_codigo.activo_ubica,
            'activo_depar':query.desc_activo_codigo.activo_depar,
            'activo_area':query.desc_activo_codigo.activo_area,
            'activo_subsec':query.desc_activo_codigo.activo_subsec,
        }

        user = request.user
        form = MoverActivosForm(initial=initial_data)

        return render(request,'activos/registro_movimiento_interno.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  

@login_required
def salida_activos_func(request, numtrabajo):
   
    if request.method == 'POST':
        form = RegistroSalida(request.POST)
        trabajos = OrdenesTrabajos.objects.get(numtrabajo=numtrabajo)
        if form.is_valid():
            print("Pasó")
            registro = form.save()
            registro.orden_mantenimiento = numtrabajo
            registro.estado_bodega = True
            trabajos.salida_genera = True
            registro.solicitado_por = request.user.username
            registro.save()
            trabajos.save()
            
            return redirect('activos:listar_salida')
        
        else:
            print(form.errors)
            return HttpResponse("No se pudo guardar")

    else:

        form = RegistroSalida()
        
        return render(request,'activos/registro_salida.html', {'form':form, 'mc':numtrabajo})




@login_required
def registro_no_encontrados(request,pk):
    query = desc_activo.objects.get(id=pk)
    if request.method == 'POST':
        
        r = no_encontrados(
            codigo = query,
            descripcion = request.POST.get("descripcion"),
            grabado = request.POST.get("grabado"),
            usuario_registra = request.POST.get("usuario_registra")
            )
        r.save()

        queryset = no_encontrados.objects.filter(codigo=pk).last()

        envioMail('Registro de Activo con Observaciones', 'dmencias@ecofroz.com,jefe.bodega@ecofroz.com', 'email_registro_activo_con_observaciones.html', queryset, '')

        return redirect('activos:filtrar_activos2')

    
    else:
        form = NoEncontradosForm()

        return render(request,'activos/registro_no_encontrados.html',{'form':form,'pk':query.activo_codigo})  





@login_required
def busqueda_mov_int(request):
    queryset = request.GET.get("buscar")
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                      Q(desc_activo_custodio__icontains = queryset) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                      Q(desc_activo_num_serie__icontains = queryset) |
                      Q(desc_activo_codigo_mba__icontains = queryset) |
                      Q(desc_activo_marca__icontains = queryset) |
                      Q(desc_activo_modelo__icontains = queryset) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset) 
                      )
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)

          return render(request,'activos/busqueda_mov_int.html',{'form':activos,'query':queryset})
      
    else:
        activosx = ''
        
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)
        queryset 

        return render(request,'activos/busqueda_mov_int.html',{'form':activos})


@login_required
def busqueda_salida_activos_agricola(request):
    
    codigo = request.GET.get("activo_codigo")
    
    if codigo:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = codigo) |
                      Q(desc_activo_custodio__icontains = codigo) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = codigo) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = codigo) |
                      Q(desc_activo_num_serie__icontains = codigo) |
                      Q(desc_activo_codigo_mba__icontains = codigo) |
                      Q(desc_activo_marca__icontains = codigo) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = codigo) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = codigo))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)

          parametros = FiltrarForm(request.GET or None)
          parametros2 = FiltrarMarca(request.GET or None) 

    else:
        
        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarMarca(request.GET or None) 
        activos = None

    return render(request,'activos/busqueda_activos_salida_agricola.html',{'form':activos, 'form2':parametros,'form4':parametros2})



@login_required
def busqueda_salida_activos(request):
    queryset = request.GET.get("buscar")
    codigo = request.GET.get("activo_codigo")
    tipo = request.GET.get("activo_tipo")
    area = request.GET.get("activo_area")
    marca = request.GET.get("desc_activo_marca")
    
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                      Q(desc_activo_custodio__icontains = queryset) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                      Q(desc_activo_num_serie__icontains = queryset) |
                      Q(desc_activo_codigo_mba__icontains = queryset) |
                      Q(desc_activo_marca__icontains = queryset) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)

    elif codigo:

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
            Q(desc_activo_codigo__activo_codigo__iexact = codigo )|
            Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo))

        
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarMarca(request.GET or None) 
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_activos_salida.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'tipo':tipo,'activosx':activosx})  

    elif tipo:
        parametros = FiltrarForm(request.GET or None) 
        print(parametros)
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_tipo__id = tipo).order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        parametros2 = FiltrarMarca(request.GET or None)
     
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)
        
        return render(request,'activos/busqueda_activos_salida.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'tipo':tipo})  

    elif area:
        parametros = FiltrarForm(request.GET or None)  
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_area__area_codigo = area).order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        parametros2 = FiltrarMarca(request.GET or None)
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_activos_salida.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'area':area})  
 
    elif marca:
        parametros2 = FiltrarMarca(request.GET or None)  
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_marca__icontains = marca).order_by('desc_activo_codigo__cod_activo_padre')
        
        parametros = FiltrarForm(request.GET or None)
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_activos_salida.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'area':area})  
 

    else:
        activosx = ''
        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo').exclude(
        #       desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
        #           desc_activo_codigo__activo_estado='DONADO').exclude(
        #               desc_activo_codigo__activo_estado='VENDIDO')

        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarMarca(request.GET or None) 
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/busqueda_activos_salida.html',{'form':activos, 'form2':parametros, 'query':queryset,'form4':parametros2})


@login_required
def busqueda_movimientos_bodega(request):
    queryset = request.GET.get("buscar")
    codigo = request.GET.get("activo_codigo")
    tipo = request.GET.get("activo_tipo")
    marca = request.GET.get("desc_activo_marca")
    ubica = request.GET.get("ubicacion")
    
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                      Q(desc_activo_custodio__icontains = queryset) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                      Q(desc_activo_num_serie__icontains = queryset) |
                      Q(desc_activo_codigo_mba__icontains = queryset) |
                      Q(desc_activo_marca__icontains = queryset) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)

    elif codigo:
        ubicacionbod = UbicacionesBodegaForm(request.GET or None)

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
            Q(desc_activo_codigo__activo_codigo__iexact = codigo )|
            Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo)).exclude(
               desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                   desc_activo_codigo__activo_estado='DONADO').exclude(
                       desc_activo_codigo__activo_estado='VENDIDO').filter(
                           Q(desc_activo_codigo__activo_area=1010) |
                           Q(desc_activo_codigo__activo_ubica=1 ) |
                           Q(desc_activo_codigo__activo_ubica=6 ) |
                           Q(desc_activo_codigo__activo_ubica=7 ) |
                           Q(desc_activo_codigo__activo_ubica=8 ) |
                           Q(desc_activo_codigo__activo_ubica=9 ))
    
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarMarca(request.GET or None) 
        activoscount = activosx.count()

        paginator = Paginator(activosx, 50)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_movimientos_bodega.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'tipo':tipo,'activosx':activosx,'ubicacionbod':ubicacionbod})  

    elif tipo:
        parametros = FiltrarForm(request.GET or None) 
        print(parametros)
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_tipo__id = tipo).order_by(
                    'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
               desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                   desc_activo_codigo__activo_estado='DONADO').exclude(
                       desc_activo_codigo__activo_estado='VENDIDO').filter(
                           Q(desc_activo_codigo__activo_area=1010) |
                           Q(desc_activo_codigo__activo_ubica=1 ) |
                           Q(desc_activo_codigo__activo_ubica=6 ) |
                           Q(desc_activo_codigo__activo_ubica=7 ) |
                           Q(desc_activo_codigo__activo_ubica=8 ) |
                           Q(desc_activo_codigo__activo_ubica=9 ))
        
        parametros2 = FiltrarMarca(request.GET or None)
     
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)
        
        return render(request,'activos/busqueda_movimientos_bodega.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount,'tipo':tipo})  


    elif ubica:
        if ubica == '9': #Reparacion Externa
            ubicacionbod = UbicacionesBodegaForm(request.GET or None)
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_ubica = ubica).order_by('desc_activo_codigo__cod_activo_padre')
            
            parametros = FiltrarForm(request.GET or None)
            activoscount = activosx.count()
            paginator = Paginator(activosx, 50)
            page = request.GET.get('page')
            activos = paginator.get_page(page)
           
        
        else:
            ubicacionbod = UbicacionesBodegaForm(request.GET or None)
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_ubica = ubica).filter(
                    desc_activo_codigo__activo_area=1010).order_by(
                                   'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
                                       desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                                           desc_activo_codigo__activo_estado='DONADO').exclude(
                                               desc_activo_codigo__activo_estado='VENDIDO')

            parametros = FiltrarForm(request.GET or None)
            activoscount = activosx.count()
            paginator = Paginator(activosx, 50)
            page = request.GET.get('page')
            activos = paginator.get_page(page)


        return render(request,'activos/busqueda_movimientos_bodega.html',{'form':activos,'form2':parametros,'var':activoscount,'ubicacionbod':ubicacionbod})  
 

    
    elif marca:
        parametros2 = FiltrarMarca(request.GET or None)  
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_marca__iexact = marca).order_by('desc_activo_codigo__cod_activo_padre').exclude(
               desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                   desc_activo_codigo__activo_estado='DONADO').exclude(
                       desc_activo_codigo__activo_estado='VENDIDO').filter(
                           Q(desc_activo_codigo__activo_area=1010) |
                           Q(desc_activo_codigo__activo_ubica=1 ) |
                           Q(desc_activo_codigo__activo_ubica=6 ) |
                           Q(desc_activo_codigo__activo_ubica=7 ) |
                           Q(desc_activo_codigo__activo_ubica=8 ) |
                           Q(desc_activo_codigo__activo_ubica=9 ))
        parametros = FiltrarForm(request.GET or None)
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_movimientos_bodega.html',{'form':activos,'form2':parametros,'form4':parametros2,'var':activoscount})  
 

    else:
        # activosx = ''
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo').exclude(
               desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                   desc_activo_codigo__activo_estado='DONADO').exclude(
                       desc_activo_codigo__activo_estado='VENDIDO').filter(Q(desc_activo_codigo__activo_area=1010) | Q(desc_activo_codigo__activo_ubica=9))

        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarMarca(request.GET or None) 
        ubicacionbod = UbicacionesBodegaForm(request.GET or None)
        paginator = Paginator(activosx, 50)
        activoscount = activosx.count()

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/busqueda_movimientos_bodega.html',{'form':activos, 'form2':parametros, 'query':queryset,'form4':parametros2,'ubicacionbod':ubicacionbod,'var':activoscount})



@login_required
def salida_activos_ind(request, codigo):
    if request.method == 'POST':
        form = RegistroSalida(request.POST)
        # grupo = activo_grupo.objects.get(id=request.POST.get('grupo'))
        # emailaprob = User.objects.get(id=request.user.id)
        # autorizador = Autorizador.objects.get(id=request.user.id)
        whoami = User.objects.get(id=request.user.id)
        emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
        email = emailaprob.e_mail
        print(email)
        # email = 'mantenimiento@ecofroz.com, operaciones@ecofroz.com, ' + email
        print(email)
        if isinstance(request.POST.get('num_orden_trabajo'), int): 
            ordenes = OrdenesTrabajos.objects.get(numtrabajo=request.POST.get('num_orden_trabajo'))
 
        if form.is_valid():
            print("Hola mundo")
            registro = form.save()
            registro.orden_mantenimiento = request.POST.get('num_orden_trabajo')
            registro.empresa_mantenimiento = request.POST.get('proveedor')
            registro.ubica_depar = request.POST.get('depar')
            registro.ubica_area = request.POST.get('areas')
            registro.solicitado_por = request.user.username
            registro.id_solicita = whoami
            registro.save()
            
            if isinstance(request.POST.get('num_orden_trabajo'), int): 
                ordenes.salida_genera = True
                ordenes.save()

            queryset = salida_activos.objects.filter(solicitado_por=request.user.username).last()
            queryset2 = ''

            envioMail('Solicitud de Salida de Activos', email, 'email_solicitud_salida_activo.html', queryset, queryset2)

            return redirect('activos:listar_salida')   
        else:     
            print(form.errors)
            return HttpResponse("No se pudo guardar")
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        
        if query.desc_activo_codigo.activo_ubica.id == 9:
            estado = 1
        else:
            estado = 0
        
        print(query.desc_activo_codigo.activo_ubica.id)
        print(estado)

        desc = ''
        
        desc = query.desc_activo_codigo.activo_descripcion

        initial_data = {
            'activo_tipo':query.desc_activo_codigo.activo_tipo,
            'activo_num_serie':query.desc_activo_num_serie,
            'marca':query.desc_activo_marca,
            'detalle_activo':query.desc_activo_codigo.activo_descripcion,
            'activo_codigo':query.desc_activo_codigo.id,
            'departamento':query.desc_activo_codigo.activo_depar.id,
            'grupo':query.desc_activo_codigo.activo_grupo.id,
            'retorno':'SI',
        }

        print(initial_data)

        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=45)

        #ordenes = OrdenesTrabajos.objects.filter(tipo_pedi='MC', aprobado__isnull=False,fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo').select_related('numtrabajo')
        
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(
            Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='CA') | Q(numtrabajo__tipo_pedi='MP') | Q(
                numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(
                    numtrabajo__aprobado__isnull=False,numtrabajo__fchsolicita__range=[enddate,startdate]).order_by(
                        '-numtrabajo')
        proveedores = proveedor_det.objects.filter(proveedor_mant_externo=True).order_by('codigo_id__nombre_empresa')
        departamento = activo_depar.objects.all()
        area = activo_areas.objects.all()
        user = request.user.id
        form = RegistroSalida(initial=initial_data)

        return render(request,'activos/registro_salida_ind.html',{'form':form,'query':query, 'solicita':user, 'desc':desc, 'codigo':codigo, 'orden':ordenes,'proveedor':proveedores, 'departamento':departamento, 'area':area, 'estado':estado})  


@login_required
def salida_activos_agri(request, codigo):
    if request.method == 'POST':
        form = RegistroSalidaAgri(request.POST)
        area = str(request.POST.get('ubica_area'))
        ubica = str(request.POST.get('ubica_ubica'))
        
        if isinstance(request.POST.get('ubica'), int): 
            print("Entra isinstance1")
            orden = OrdenesTrabajos.objects.get(numtrabajo=request.POST.get('num_orden_trabajo'))
            ubica_num = activo_ubica.objects.get(id=request.POST.get('ubica'))
        if isinstance(request.POST.get('areas'), int): 
            sector_num = activo_areas.objects.get(area_codigo=request.POST.get('areas'))
            print("Entra isinstance1")
        
        try:
            print(ubica_num)
            print(sector_num)
        except:
            pass

        if form.is_valid():
            # print("Hola mundo")
            registro = form.save()
            try:
                registro.ubica_area = sector_num
                registro.ubica_ubica = ubica_num
            except:
                registro.ubica_area = None
                registro.ubica_ubica = None
                registro.retorno = 'SI'

            registro.solicitado_por = request.POST.get('solicitado_por').upper()
            registro.persona_autoriza_dep = request.POST.get('persona_autoriza_dep').upper()
            registro.persona_registra_en_proyecto = request.user.username
            registro.estado_autoriza_dep = 1
            registro.estado_bodega = True
            registro.estado = 'SALIÓ'

            if request.POST.get('cbox1'):
                registro.empresa_mantenimiento = request.POST.get('otraempresa').upper()

            else:
                registro.empresa_mantenimiento = request.POST.get('proveedor')
            registro.save()
            
            queryset2 = ''
          

            # query_activo = desc_activo.objects.filter(activo_codigo=codigo).update(activo_ubica=ubi, activo_area=sec)

            #envioMail('Generación de Orden de Salida', email, 'email_solicitud_salida_activo.html', queryset, queryset2)

            return redirect('activos:listar_salida_agri')   
        else:     
            print(form.errors)
            return HttpResponse("No se pudo guardar")
    else:
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        
        if query.desc_activo_codigo.activo_ubica.id == 9:
            estado = 1
        else:
            estado = 0
        
        print(query.desc_activo_codigo.activo_ubica.id)
        print(estado)

        desc = ''
        
        desc = query.desc_activo_codigo.activo_descripcion
        ar = query.desc_activo_codigo.activo_area.area_nombre
        ubi = query.desc_activo_codigo.activo_ubica.ubica_nombre

        initial_data = {
            'activo_tipo':query.desc_activo_codigo.activo_tipo,
            'activo_num_serie':query.desc_activo_num_serie,
            'marca':query.desc_activo_marca,
            'detalle_activo':query.desc_activo_codigo.activo_descripcion,
            'activo_codigo':query.desc_activo_codigo.id,
            'departamento':query.desc_activo_codigo.activo_depar.id,
            'grupo':query.desc_activo_codigo.activo_grupo.id,
            # 'ubica_ubica':query.desc_activo_codigo.activo_ubica.id,
            # 'ubica_area':query.desc_activo_codigo.activo_area.area_codigo,
        }

        

        print(initial_data)

        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=45)

        #ordenes = OrdenesTrabajos.objects.filter(tipo_pedi='MC', aprobado__isnull=False,fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo').select_related('numtrabajo')
        
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') | Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(numtrabajo__aprobado__isnull=False,numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        proveedores = proveedor_det.objects.filter(proveedor_mant_externo=True).order_by('codigo_id__nombre_empresa')
        ubicacion = activo_ubica.objects.all().exclude(id__in=[5,9])
        area = activo_areas.objects.all()
        user = request.user.id
        form = RegistroSalidaAgri(initial=initial_data)

        return render(request,'activos/registro_salida_proyectos_agri.html',{'form':form,'query':query, 'registra':user, 'desc':desc, 'codigo':codigo, 'orden':ordenes,'proveedor':proveedores, 'ubicacion':ubicacion, 'area':area, 'ar':ar,'ubi':ubi,'estado':estado})  


@login_required
def AutorizaSalidasDep(request):
    queryset = request.GET.get("buscar")
    usu_depend = Autorizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(autorizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)
    print(id_usu)

    if queryset:
        solicitud = salida_activos.objects.filter(estado_autoriza_dep=None, id_solicita__in=id_usu).exclude(anula = True)
    else:
        solicitud = salida_activos.objects.filter(estado_autoriza_dep__isnull=True,id_solicita__in=id_usu).order_by("-fecha_registro").exclude(anula = True)
        aprobadas = salida_activos.objects.filter(estado_autoriza_dep=1,id_solicita__in=id_usu).order_by("-fecha_registro").exclude(anula = True)
        # solicitud = salida_activos.objects.filter(estado_autoriza_dep__isnull=True, grupo__in=grup).order_by("-fecha_registro")
        # aprobadas = salida_activos.objects.filter(estado_autoriza_dep=1, grupo__in=grup).order_by("-fecha_registro")
    return render(request, 'activos/listar_autorizador_salidas.html', {'form':solicitud, 'busqueda':queryset, 'aprobadas':aprobadas})

def depAutorizaSalida(request, numorden):
    if request.method == 'GET':
        detalle = salida_activos.objects.get(id=numorden)
        departamento = activo_depar.objects.get(id=detalle.departamento)
        proveedores = proveedor_det.objects.filter(proveedor_mant_externo=True).order_by('codigo_id__nombre_empresa')
        idpedido = numorden
        if detalle.ubica_depar:
            mov_depar = activo_depar.objects.get(id=detalle.ubica_depar)
        else:
            mov_depar = None

        if detalle.ubica_area:
            mov_area = activo_areas.objects.get(area_codigo=detalle.ubica_area)
        else:
            mov_area = None

        return render(request, 'activos/ver_salidas_autoriza.html', {'detalle': detalle, 'id':idpedido, 'dep':departamento.dep_nombre, 'proveedor':proveedores, 'depar':mov_depar, 'area':mov_area})

@login_required
def ver_autoriza_seguridad(request, numorden):
    if request.method == 'GET':
        detalle = salida_activos.objects.get(id=numorden)
        departamento = activo_depar.objects.get(id=detalle.departamento)
        if detalle.ubica_depar:
            mov_depar = activo_depar.objects.get(id=detalle.ubica_depar)
        else:
            mov_depar = None

        if detalle.ubica_area:
            mov_area = activo_areas.objects.get(area_codigo=detalle.ubica_area)
        else:
            mov_area = None
        idpedido = numorden

        return render(request, 'activos/ver_autoriza_seguridad.html', {'detalle': detalle, 'id':idpedido, 'dep':departamento.dep_nombre, 'depar':mov_depar, 'area':mov_area})

@login_required
def autorizaSalidaAutorizador(request, numorden):
    observa = request.GET.get('observa')
    proveedor = request.GET.get('proveedor')
    salida = salida_activos.objects.get(id=numorden)
    salida.estado_autoriza_dep = 1
    salida.observa_autoriza = observa
    salida.empresa_mantenimiento = proveedor
    salida.persona_autoriza_dep=request.user.username
    salida.save()

    queryset = salida_activos.objects.get(id=numorden)
    queryset2 = ''

    envioMail('Aprobación de Orden de Salida', 'bodega@ecofroz.com', 'email_aprobacion_salida_activo.html', queryset, queryset2)

    return redirect('activos:lista_autoriza_dep')

@login_required
def autoriza_salida_seguridad(request, numorden):
    observa = request.GET.get('observa')
    queryset = salida_activos.objects.get(id=numorden)
    queryset.estado = 'APROBADO'
    queryset.pers_autoriza_seguridad = request.user.username
    queryset.observa_autoriza = observa
    queryset.save()

    queryset2 = ''

    envioMail('Aprobación de Orden de Salida de Seguridad', 'bodega@ecofroz.com', 'email_aprobacion_salida_activo_seg.html', queryset, queryset2)
    
    return redirect('activos:control_retorno_activos')

@login_required
def recibeActivoBodega(request, codigo):
    if request.method == 'GET':
        query = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        desc = ''
        desc = query.desc_activo_codigo.activo_descripcion
        cod = query.desc_activo_codigo.activo_codigo
        depar = activo_depar.objects.all()
        area = activo_areas.objects.all()

        return render(request,'activos/recepcion_activos.html',{'query':query, 'desc':desc, 'codigo':codigo,'cod':cod,'depar':depar,'area':area})  
    else:
        custodio = User.objects.get(id=request.user.id)
        queryset = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        salida = salida_activos.objects.filter(codigo=codigo).last()
        queryset.desc_activo_custodio = custodio.first_name + ' ' + custodio.last_name
        queryset.desc_activo_codigo.activo_subsec = None
        # queryset.desc_activo_codigo.activo_area = activo_areas.objects.get(area_codigo=1010) 
        # queryset.desc_activo_codigo.activo_ubica = activo_ubica.objects.get(id=4)
        actualiza = desc_activo.objects.filter(activo_codigo=codigo).update(activo_ubica=4, activo_area=1010)

        salida.observa_recibe = request.POST.get('observacion')
        salida.estado = 'RETORNÓ'
        historial = historial_movimientos_internos()
        historial.activo_codigo = desc_activo.objects.get(activo_codigo=codigo)
        historial.ubicacion = queryset.desc_activo_codigo.activo_ubica
        historial.departamento = queryset.desc_activo_codigo.activo_depar
        historial.sector = queryset.desc_activo_codigo.activo_area
        historial.subsector = str(queryset.desc_activo_codigo.activo_subsec)
        historial.usuario_registra = request.user.username
        historial.justificacion_movimiento = salida.motivo
        historial.nuevo_ubicacion = activo_ubica.objects.get(id=4)
        historial.nuevo_sector = activo_areas.objects.get(area_codigo=1010)
        historial.nuevo_departamento = activo_depar.objects.get(id=111)
        salida.save()
        queryset.save()
        historial.save()

        # actualiza = desc_activo.objects.filter(activo_codigo=codigo).update(activo_ubica=4,activo_depar=108,activo_area=1010)
        print("Update ubica OK")

        #Actualiza tabla de movimientos de trazabilidad del activo

        q = desc_activo.objects.filter(activo_codigo=codigo)
        h = historial_movimientos_internos()

        #Aqui me quedé DDAAMMPP

        # desc_activo.objects.filter(activo_codigo=codigo).update(activo_depar=108)
        # desc_activo.objects.filter(activo_codigo=codigo).update(activo_area=1010)

        return redirect('activos:busqueda_movimientos_bodega')

@login_required
def apruebaBodegaSalida(request, id):
    queryset = salida_activos.objects.get(id=id)
    queryset2 = desc_activo.objects.get(activo_codigo=queryset.activo_codigo)
    historial = historial_movimientos_internos()
    historial.activo_codigo = queryset.activo_codigo
    historial.ubicacion = queryset2.activo_ubica
    historial.departamento = queryset2.activo_depar
    historial.sector = queryset2.activo_area
    historial.subsector = queryset2.activo_subsec
    historial.usuario_registra = request.user.username
    historial.justificacion_movimiento = queryset.motivo
    historial.nuevo_ubicacion = activo_ubica.objects.get(id=4)
    historial.nuevo_sector = activo_areas.objects.get(area_codigo=1010)
    historial.nuevo_departamento = activo_depar.objects.get(id=111)
    queryset2.activo_area = activo_areas.objects.get(area_codigo=1010) 
    queryset.pers_gestiona_bodega = request.user.id
    queryset.estado_bodega = True
    queryset.save()
    queryset2.save()
    historial.save()

    queryset = salida_activos.objects.get(id=id)
    queryset2 = ''

    envioMail('Solicitud de Orden de Salida de Activos', 'logistica@ecofroz.com, fortiz@ecofroz.com', 'email_verifica_salida_activo.html', queryset, queryset2)

    return redirect ('activos:solicitud_salida')

@login_required
def confirmaSalida(request, id):
    queryset = salida_activos.objects.get(id=id)
    queryset.estado = 'APROBADO'
    queryset.pers_autoriza_seguridad = request.user.id
    queryset.save()

    queryset = salida_activos.objects.get(id=id)
    queryset2 = ''

    email = User.objects.get(username=queryset.solicitado_por)
    email = 'bodega@ecofroz.com, ' + email
    print(email.email)

    envioMail('Confirmación de Orden de Salida', email, 'email_confirma_salida_activo.html', queryset, queryset2)

    return redirect('activos:control_retorno_activos')

@login_required
def saleActivo(request):
    if request.method == 'GET':
        try:
            codigo2 = str(request.GET.get('codigo'))
            codigo = codigo2.upper()
            if codigo != 'NONE':
                # queryset = salida_activos.objects.filter(codigo=codigo).last()
                cod = desc_activo.objects.get(activo_codigo=codigo)
                queryset = salida_activos.objects.filter(activo_codigo=cod.id).last()
                if queryset:
                    if queryset.pers_autoriza_seguridad:
                        autorizador = User.objects.get(username=queryset.pers_autoriza_seguridad)
                    else:
                        autorizador = None
                    salida_existe = 1
                else:
                    autorizador = None
                    salida_existe = 0

                return render(request, 'activos/salida_activo.html',{'form':queryset, 'codigo':codigo, 'autoriza':autorizador, 'salida_existe':salida_existe})
            else:
                return render(request, 'activos/salida_activo.html')
        except:
            return render(request, 'activos/salida_activo.html') 
    else:
        codigo = request.POST.get('codigoac')
        id = request.POST.get('cid')
        print(id)
        queryset = salida_activos.objects.get(id=id)
        seguridad = User.objects.get(username=queryset.pers_autoriza_seguridad)
        queryset2 = detalle_desc_activo.objects.get(desc_activo_codigo__activo_codigo=codigo)
        queryset3 = desc_activo.objects.get(activo_codigo=codigo)
        
        if queryset.ubica_area:
            ubicacion = activo_areas.objects.get(area_codigo=queryset.ubica_area)
        else:
            ubicacion = None
       
        historial = historial_movimientos_internos()
        queryset.estado = 'SALIÓ'
        queryset.guardia_control = request.user.id

        ubicacion_ori = queryset2.desc_activo_codigo.activo_ubica
        departamento_ori = queryset2.desc_activo_codigo.activo_depar
        sector_ori = queryset2.desc_activo_codigo.activo_area
        subsector_ori = str(queryset2.desc_activo_codigo.activo_subsec)
       
        if queryset.sale_por == 4:
            queryset2.desc_activo_custodio = str(seguridad.first_name) + ' ' + str(seguridad.last_name)
            queryset2.desc_activo_cod_responsable = None
            queryset3.activo_depar = activo_depar.objects.get(id=queryset.ubica_depar)
            queryset3.activo_area = activo_areas.objects.get(area_codigo=queryset.ubica_area)
            queryset3.activo_ubica = activo_ubica.objects.get(id=ubicacion.area_ubica)
            queryset2.save()
            queryset3.save()
        else:
            queryset2.desc_activo_custodio = str(seguridad.first_name) + ' ' + str(seguridad.last_name)
            queryset2.desc_activo_cod_responsable = None
            queryset3.activo_depar = None
            queryset3.activo_area = None
            queryset2.save()
            queryset3.save()
            
        historial.activo_codigo = desc_activo.objects.get(activo_codigo=queryset.codigo)
        historial.ubicacion = ubicacion_ori
        historial.departamento = departamento_ori
        historial.sector = sector_ori
        historial.subsector = subsector_ori
        historial.usuario_registra = request.user.username
        historial.justificacion_movimiento = queryset.motivo
        
        if queryset.sale_por == 4:
            historial.nuevo_ubicacion = activo_ubica.objects.get(id=ubicacion.area_ubica)
            historial.nuevo_sector = activo_areas.objects.get(area_codigo=queryset.ubica_area) #activo_areas.objects.get(area_codigo=1010)
            historial.nuevo_departamento = activo_depar.objects.get(id=queryset.ubica_depar) #activo_depar.objects.get(id=111)
        else:
            historial.nuevo_ubicacion = activo_ubica.objects.get(id=9)
            historial.nuevo_sector = None
            historial.nuevo_departamento = None

        historial.save()
        queryset.save()

        #Descomentar lineas para cambiar ubicacon por salida de activos
        if queryset.sale_por != 4:
            desc_activo.objects.filter(activo_codigo=codigo).update(activo_ubica=9)
        # desc_activo.objects.filter(activo_codigo=codigo).update(activo_area=None)
        # desc_activo.objects.filter(activo_codigo=codigo).update(activo_subsec=None)
        
        # queryset2.save()

        email = User.objects.get(username=queryset.solicitado_por)
        email = 'jefe.bodega@ecofroz.com, ' + email.email

        envioMail('Confirmación de Salida de Activo', email, 'email_salida_activo.html', queryset, '')

        return render(request, 'activos/salida_activo.html')

@login_required
def verSolicitudesSalida(request, numorden):
    if request.method == 'GET':
        detalle = salida_activos.objects.get(id=numorden)
        departamento = activo_depar.objects.get(id=detalle.departamento)
        if detalle.ubica_depar:
            mov_depar = activo_depar.objects.get(id=detalle.ubica_depar)
        else:
            mov_depar = None

        if detalle.ubica_area:
            mov_area = activo_areas.objects.get(area_codigo=detalle.ubica_area)
        else:
            mov_area = None
        
        idpedido = numorden

        return render(request, 'activos/ver_detalle_salidas.html', {'detalle': detalle, 'id':idpedido, 'dep':departamento.dep_nombre, 'depar':mov_depar, 'area':mov_area})



@login_required
# @permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def SolicitudesSalidaActivos(request):
    queryset = request.GET.get("buscar")
    #lista = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).order_by('-numtrabajo__ord')
    if queryset:
        solicitud = salida_activos.objects.filter(Q(id__icontains = queryset) |
                    Q(codigo__icontains = queryset)).order_by("-fecha_registro")

    else:
        solicitudi = salida_activos.objects.filter(estado_autoriza_dep=1).order_by("-fecha_registro")
        
        paginator = Paginator(solicitudi, 75)
        page = request.GET.get('page')
        solicitud = paginator.get_page(page)

    return render(request, 'activos/solicitudes_salida_activos.html', {'form':solicitud, 'busqueda':queryset})

@login_required
def registro_guardado(request):
    return render(request,'activos/registro_guardado.html')

@login_required
def registro_guardado_mov_int(request,codigo):
    return render(request,'activos/registro_guardado_mov_int.html', {'codigo':codigo})

@login_required
def registro_guardado_mov_int2(request,codigo):
    return render(request,'activos/registro_guardado_mov_int2.html', {'codigo':codigo})

@login_required
def mov_success_bodega(request):
    return render(request,'activos/mov_success_bodega.html')



@login_required
def registro_guardado_salida(request):
    return render(request,'activos/registro_guardado_salida.html')

@login_required
def eliminar_activos(request, codigo):
    acti = desc_activo.objects.get(activo_index=codigo)
    if request.method == 'POST':
        acti.delete()
        return redirect('activos:listar_activos')
    return render(request,'activos/eliminar_activos.html',{'activos':acti})

@login_required
def identifica_valores_mov_int(request):
    codigo = request.GET.get("activo_codigo")
    print(codigo)

    query = detalle_desc_activo.objects.filter(desc_activo_codigo=codigo).select_related("desc_activo_codigo")
    desc = ''
    for i in query:
        desc = i.desc_activo_codigo.activo_descripcion

    print(query)


    form = RegistroSalida(request.GET)

    return render(request,'activos/registro_movimiento_interno.html',{'form':form,'query':query, 'desc':desc})  

@login_required
def identifica_valores(request):
    codigo = request.GET.get("activo_codigo")
    # mc = request.GET.get('mc')
    print(codigo)

    query = detalle_desc_activo.objects.filter(desc_activo_codigo=codigo).select_related("desc_activo_codigo")
    desc = ''
    for i in query:
        desc = i.desc_activo_codigo.activo_descripcion

    print(query)

    # user = OrdenesTrabajos.objects.get(numtrabajo=mc)
    ordenes = OrdenesTrabajos.objects.filter()
    user = request.user
    print(user)

    form = RegistroSalida(request.GET)

    # return render(request,'activos/registro_salida.html',{'form':form,'query':query, 'mc':mc, 'solicita':user, 'desc':desc})  
    return render(request,'activos/registro_salida_ind.html',{'form':form,'query':query, 'solicita':user, 'desc':desc})  



@login_required
def genera_nomenclatura(request):
    nomenclatura = request.GET.get("activo_nomenclatura")
    get_nomenclatura_nombre = str(activo_nomenclatura.objects.get(id=nomenclatura)) 

    #Slicing para obtener la primera letra
    get_letra = get_nomenclatura_nombre[:1]
    
    if get_letra == 'P':
        print("entró")

        numcontinua = SecuencialCodifica.objects.filter(codigo='P').order_by('numeracion').last()
    
        codigo = 'P'
            
        num = ''
        if numcontinua:
            num = numcontinua.numeracion
        else:
            num = 999

        numero = int(num) + 1
        cod = len(str(numero))
        if cod == 1: 
            nomen = codigo + '000' + str(numero)
        elif cod == 2:
            nomen = codigo + '00' + str(numero)
        elif cod == 3:
            nomen = codigo + '0' + str(numero)
        else:
            nomen = codigo + str(numero)

    else:
        print("Por acá")
        if get_letra == 'S':

            numcontinua = SecuencialCodifica.objects.filter(codigo='S').order_by('numeracion').last()
        
            codigo = 'S'
                
            num = ''
            if numcontinua:
                num = numcontinua.numeracion
            else:
                num = 999

            numero = int(num) + 1
            cod = len(str(numero))
            if cod == 1: 
                nomen = codigo + '000' + str(numero)
            elif cod == 2:
                nomen = codigo + '00' + str(numero)
            elif cod == 3:
                nomen = codigo + '0' + str(numero)
            else:
                nomen = codigo + str(numero)

    parametros = FiltrarForm(request.GET or None) 

    form = RegistroForm(initial={'grabado':'NO'})    
    formdet = EditaActivoIngreso()
        
    return render(request,'activos/nuevo_ingreso_activos.html',{'form':formdet,'form2':form,'form3':nomen,'nomenclatura':nomenclatura,'codigo':codigo,'numero':numero})  


@login_required
def listar_activos(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
            desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                desc_activo_codigo__activo_estado='VENDIDO').exclude(
                    desc_activo_codigo__activo_estado='CODIGO ELIMINADO').filter(
                    Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                    Q(desc_activo_custodio__icontains = queryset) |
                    Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                    Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                    Q(desc_activo_num_serie__icontains = queryset) |
                    Q(desc_activo_codigo_mba__icontains = queryset) |
                    Q(desc_activo_marca__icontains = queryset) |
                    Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                    Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset)).annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1]))

            
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_activos.html',{'form':activos,'queryset':queryset})
            
    else:
        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1]))

        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/busqueda_activos.html',{'form':activos})



@login_required
def busqueda_activos_mantenimiento(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(Q(desc_activo_codigo__activo_area=41) | Q(desc_activo_codigo__activo_area=58)).order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                      Q(desc_activo_custodio__icontains = queryset) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                      Q(desc_activo_num_serie__icontains = queryset) |
                      Q(desc_activo_codigo_mba__icontains = queryset) |
                      Q(desc_activo_marca__icontains = queryset) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(Q(desc_activo_codigo__activo_area=41) | Q(desc_activo_codigo__activo_area=58)).order_by('desc_activo_codigo').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO')

                  
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/busqueda_activos_mantenimiento.html',{'form':activos})


@login_required
def mueve_bodega(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').filter(
                      Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                      Q(desc_activo_custodio__icontains = queryset) |
                      Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                      Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                      Q(desc_activo_num_serie__icontains = queryset) |
                      Q(desc_activo_codigo_mba__icontains = queryset) |
                      Q(desc_activo_marca__icontains = queryset) |
                      Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                      Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activos=None


    return render(request,'activos/mueve_bodega.html',{'form':activos})






def busqueda_activo_bodega(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by(
              'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                  desc_activo_codigo__activo_area=1010).exclude(desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                      desc_activo_codigo__activo_estado='DONADO').exclude(
                          desc_activo_codigo__activo_estado='VENDIDO').filter(
                              Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                              Q(desc_activo_custodio__icontains = queryset) |
                              Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                              Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                              Q(desc_activo_num_serie__icontains = queryset) |
                              Q(desc_activo_codigo_mba__icontains = queryset) |
                              Q(desc_activo_marca__icontains = queryset) |
                              Q(desc_activo_codigo__activo_area__area_nombre__icontains = queryset) |
                              Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = queryset))
                        
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                  desc_activo_codigo__activo_area=1010).order_by('desc_activo_codigo').exclude(
                      desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                          desc_activo_codigo__activo_estado='DONADO').exclude(
                              desc_activo_codigo__activo_estado='VENDIDO')

                  
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/busqueda_activo_bodega.html',{'form':activos})

@login_required
def listar_bajas(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='ACTIVO').filter( 
                  Q(desc_activo_codigo__activo_codigo__icontains = queryset) |
                  Q(desc_activo_custodio__icontains = queryset) |
                  Q(desc_activo_codigo__activo_descripcion__icontains = queryset) |
                  Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = queryset) |
                  Q(desc_activo_num_serie__icontains = queryset) |
                  Q(desc_activo_codigo_mba__icontains = queryset) |
                  Q(desc_activo_marca__icontains = queryset))
              
          paginator = Paginator(activosx, 25)
          page = request.GET.get('page')
          activos = paginator.get_page(page)
            
    else:
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('-fecha_baja_dona_vende').exclude(
              desc_activo_codigo__activo_estado='ACTIVO')
       
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

    return render(request,'activos/busqueda_bajas.html',{'form':activos})


@login_required
def cambiaEstadoRetorno(request,id,estado,codigo):
    if request.method == 'GET':
        if estado == 'SALIÓ':
            print("entra")
            captura = salida_activos.objects.get(id=id)
            captura.estado = 'RETORNÓ'
            captura.save()

            #Descomentar estas lineas para habilitar el cabio de ubicación de activos que retornan a la empresa

            #desc_activo.objects.filter(activo_codigo=codigo).update(activo_ubica=4)
            #desc_activo.objects.filter(activo_codigo=codigo).update(activo_area=1010)
            #desc_activo.objects.filter(activo_codigo=codigo).update(activo_subsec=None)

            print("pasó")
            return redirect('activos:solicitud_salida')

        else:
            captura = salida_activos.objects.get(id=id)
            captura.estado = 'SALIÓ'
            captura.save()

            activosx = salida_activos.objects.all().order_by('-fecha_registro')
       
            paginator = Paginator(activosx, 25)

            page = request.GET.get('page')
            activos = paginator.get_page(page)

            return redirect('activos:solicitud_salida')

@login_required        
def cambiaEstadoBajaActivos(request,id,estado):
    valor_destino = request.GET.get("destino")
    notas = request.GET.get("notas")
    print(valor_destino)
    print(notas)

    print(request.user.username)
    if request.method == 'GET':

        if estado == 'ACTIVO':
            print("entra")
            fec = date.today()
            #captura = desc_activo.objects.get(id=id)
            captura = detalle_desc_activo.objects.all().select_related('desc_activo_codigo').get(desc_activo_codigo__id=id)
            if valor_destino == 'baja':
                # captura.desc_activo_codigo.activo_estado = 'DADO DE BAJA'
                actualiza_estado = desc_activo.objects.filter(id=id).update(activo_estado='DADO DE BAJA')
                captura.fecha_baja_dona_vende = fec
                captura.usuario_realiza_baja_dona_vende = request.user.username
                captura.notas_baja_dona_vende = notas
                captura.save()
            elif valor_destino == 'donado':
                actualiza_estado = desc_activo.objects.filter(id=id).update(activo_estado='DONADO')
                captura.fecha_baja_dona_vende = fec
                captura.notas_baja_dona_vende = notas
                captura.usuario_realiza_baja_dona_vende = request.user.username
                captura.save()
            
            elif valor_destino == 'venta':
                actualiza_estado = desc_activo.objects.filter(id=id).update(activo_estado='VENDIDO')
                captura.fecha_baja_dona_vende = fec
                captura.notas_baja_dona_vende = notas
                captura.usuario_realiza_baja_dona_vende = request.user.username
                captura.save()
            
            elif valor_destino == 'eliminado':
                actualiza_estado = desc_activo.objects.filter(id=id).update(activo_estado='CODIGO ELIMINADO')
                captura.fecha_baja_dona_vende = fec
                captura.notas_baja_dona_vende = notas
                captura.usuario_realiza_baja_dona_vende = request.user.username
                captura.save()
            
            return redirect('activos:listar_activos')
    
            
        # elif estado == 'DADO DE BAJA':
        #     captura = desc_activo.objects.get(id=id)
        #     if valor_destino == 'baja':
        #         captura.activo_estado = 'DADO DE BAJA'
        #         captura.save()
        #     elif valor_destino == 'donado':
        #         captura.activo_estado = 'DONADO'
        #         captura.save()
            
        #     elif valor_destino == 'venta':
        #         print("exito")
        #         captura.activo_estado = 'VENDIDO'
        #         captura.save()
            
        #     return redirect('activos:listar_bajas')

        return redirect('activos:listar_bajas')     


@login_required
def control_retorno_activos(request):
    queryset = request.GET.get("buscar")
    print(queryset) 
    if queryset:
          activosx = salida_activos.objects.filter(estado_bodega=True).order_by('-fecha_registro').filter(
              Q(detalle_activo__icontains = queryset) |
              Q(activo_num_serie__icontains = queryset) |
              Q(id__icontains = queryset) |
              Q(empresa_mantenimiento__icontains = queryset) |
              Q(codigo__icontains = queryset))
              
        #   paginator = Paginator(activosx, 25)
        #   page = request.GET.get('page')
        #   activos = paginator.get_page(page)

          return render(request,'activos/control_salida_activos.html',{'form':activosx})
            
    else:
        activos = salida_activos.objects.filter(estado_bodega=True).filter(estado='PENDIENTE').order_by('-fecha_registro')
        aprobadas = salida_activos.objects.filter(Q(estado='APROBADO')|Q(estado='SALIÓ')|Q(estado='RETORNÓ')).order_by("-fecha_registro")[:200]
       
        try:
            countn = activos.count()
        except:
            countn = 0

        # paginator = Paginator(aprobadas, 25)

        # page = request.GET.get('page')
        # form2 = paginator.get_page(page)

        return render(request,'activos/control_salida_activos.html',{'form':activos,'form2':aprobadas,'countn':countn})

@login_required
def ultimos_movimientos(request):

    activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo")
       
    paginator = Paginator(activosx, 25)

    page = request.GET.get('page')
    activos = paginator.get_page(page)

    return render(request,'activos/ultimos_movimientos.html',{'form':activos})


@login_required
def filtrar_activos2(request):
    codigo = request.GET.get("activo_codigo")
    ubicacion = request.GET.get("activo_ubica")
    newcustodi = request.GET.get("custodio")
    serie = request.GET.get("desc_activo_num_serie")
    clave = request.GET.get("pclave")
  
    if codigo:
        parametros = FiltrarForm(request.GET or None)  
        form_serie = FiltrarSerie(request.GET or None)
        activosx = detalle_desc_activo.objects.all().select_related(
            "desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                Q(desc_activo_codigo__activo_codigo__iexact = codigo )|Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo))

        #Borra tabla temporal, abre tabla y guarda consulta filtrada de activos realizada por usuario        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.codigo = codigo
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')

        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'var':activoscount,'ubicacion':ubicacion,'activosx':activosx, 'personal':personal,'form3':form_serie})  

    elif clave:
        parametros = FiltrarForm(request.GET or None) 
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
            Q(desc_activo_codigo__activo_codigo__icontains = clave) |
            Q(desc_activo_custodio__icontains = clave) |
            Q(desc_activo_codigo__activo_descripcion__icontains = clave) |
            Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = clave) |
            Q(desc_activo_num_serie__icontains = clave) |
            Q(desc_activo_codigo_mba__icontains = clave) |
            Q(desc_activo_marca__icontains = clave) |
            Q(desc_activo_codigo__activo_area__area_nombre__icontains = clave) |
            Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = clave))
    
        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.pclave = clave
        activostemp.save()

        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')


        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx, 'personal':personal,'form3':clave}) 
    


    elif serie:
        parametros = FiltrarForm(request.GET or None) 
        form_serie = FiltrarSerie(request.GET or None) 
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                          Q(desc_activo_num_serie__iexact = serie ))

        #Borra tabla temporal, abre tabla y guarda consulta filtrada de activos realizada por usuario        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.codigo = codigo
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')

        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'var':activoscount,'ubicacion':ubicacion,'activosx':activosx, 'personal':personal,'form3':form_serie})  


    elif ubicacion:
        
        parametros = FiltrarForm(request.GET or None) 
        form_serie = FiltrarSerie(request.GET or None) 
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_ubica__id = ubicacion).order_by(
                'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        # activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        activostemp.save()

        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')

        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'var':activoscount,'ubicacion':ubicacion, 'personal':personal,'form3':form_serie})  
 
    elif newcustodi:
        print("Entró Custodio")
        custodio = unicodedata.normalize('NFC', newcustodi)
        print(custodio, len(custodio))
        newcustodio = custodio.replace("  "," ")
        parametros = FiltrarForm(request.GET or None)  
        form_serie = FiltrarSerie(request.GET or None) 
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(Q(
            desc_activo_custodio__icontains = custodio)).order_by(
                'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.custodio = custodio
        activostemp.save()
        
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')

        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'var':activoscount,'personal':personal,'form3':form_serie})  
 

    elif (ubicacion and codigo) is None:
        print("entro DM")
        #activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        parametros = FiltrarForm(request.GET or None) 
        form_serie = FiltrarSerie(request.GET or None) 
        activos = None
        # activosy = None


        #activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            #nuevas lineas optimizar listado de activos
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        
        activostemp.save()
        # activoscount = activosx.count() - activosy.count()
        # bajascount = activosy.count()
        # paginator = Paginator(activosx, 25)

        # page = request.GET.get('page')
        # activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')
   
        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros,'form3':form_serie,'personal':personal}) 

    else: 
        print("Ultima Opcion")
        #activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        parametros = FiltrarForm(request.GET or None) 
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by(
            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        #activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            #nuevas lineas optimizar listado de activos
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        
        
        activostemp.save()
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)
   
        return render(request,'activos/filtrar_activos_customized.html',{'form':activos,'form2':parametros}) 



@login_required
def filtrar_activos(request):
    codigo = request.GET.get("activo_codigo")
    grupo = request.GET.get("activo_grupo")
    tipo = request.GET.get("activo_tipo")
    ubicacion = request.GET.get("activo_ubica")
    area = request.GET.get("activo_area")
    departamento = request.GET.get("activo_depar")
    seguros = request.GET.get("pestanias_seguros")
    poliza = request.GET.get("poliza_seguros")           
    newcustodi = request.GET.get("custodio")

    cuenta_grab = desc_activo.objects.all().filter(grabado="GRABADO").count()
    cuenta_eti = desc_activo.objects.all().filter(grabado="ETIQUETA").count()
    cuenta_foto = Imagenes.objects.values('activo_codigo').annotate(dcount=Count('activo_codigo')).order_by().count()

    
      
    if codigo:
        parametros = FiltrarForm(request.GET or None)  
        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
        #       desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
        #           desc_activo_codigo__activo_estado='DONADO').exclude(
        #               desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
        #                   Q(desc_activo_codigo__activo_codigo__iexact = codigo )|Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo))

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
            desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                desc_activo_codigo__activo_estado='VENDIDO').exclude(
                    desc_activo_codigo__activo_estado='CODIGO ELIMINADO').filter(
                    Q(desc_activo_codigo__activo_codigo__iexact = codigo)).annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1]))



        #Borra tabla temporal, abre tabla y guarda consulta filtrada de activos realizada por usuario        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.codigo = codigo
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'tipo':tipo,'ubicacion':ubicacion,'activosx':activosx})  

    elif seguros:
        
        parametros = FiltrarForm(request.GET or None)
        q1 = Imagenes.objects.values('activo_codigo').distinct()
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__pestanias_seguros__id = seguros).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

    

        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.seguros_id = seguros
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'seguros':seguros}) 

    elif poliza:
        
        parametros = FiltrarForm(request.GET or None)


        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__poliza_seguros = poliza).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').annotate(
                            fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                                'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.poliza = poliza
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'poliza':poliza}) 



    elif grupo:
        if grupo and ubicacion:
            parametros = FiltrarForm(request.GET or None)

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_grupo__id = grupo).filter(desc_activo_codigo__activo_ubica__id = ubicacion)

            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            activostemp = activo_estados_select()
            activostemp.grupo_id = grupo
            activostemp.ubicacion_id= ubicacion
            activostemp.save()
        
            activoscount = activosx.count()

            paginator = Paginator(activosx, 25)
            page = request.GET.get('page')
            activos = paginator.get_page(page)

            return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'grupo':grupo,'ubicacion':ubicacion}) 

        
        elif grupo and departamento:
            parametros = FiltrarForm(request.GET or None)

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_grupo__id = grupo).filter(desc_activo_codigo__activo_depar__id = departamento)

            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            activostemp = activo_estados_select()
            activostemp.grupo = grupo
            activostemp.departamento = departamento
            activostemp.save()
        
            activoscount = activosx.count()

            paginator = Paginator(activosx, 25)
            page = request.GET.get('page')
            activos = paginator.get_page(page)

            return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'grupo':grupo,'departamento':departamento}) 

        elif grupo and area:
            parametros = FiltrarForm(request.GET or None)

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_grupo__id = grupo).filter(desc_activo_codigo__activo_area__area_codigo = area)

            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            activostemp = activo_estados_select()
            activostemp.grupo = grupo
            activostemp.area = area
            activostemp.save()
        
            activoscount = activosx.count()

            paginator = Paginator(activosx, 25)
            page = request.GET.get('page')
            activos = paginator.get_page(page)

            return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'grupo':grupo,'area':area}) 

        parametros = FiltrarForm(request.GET or None)

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_grupo__id = grupo).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').annotate(
                            fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                                'desc_activo_codigo__activo_codigo')

        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.grupo_id = grupo
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'activosx':activosx,'grupo':grupo}) 

    elif tipo:
        if tipo and ubicacion:
            parametros = FiltrarForm(request.GET or None) 
            
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_tipo__id = tipo).filter(desc_activo_codigo__activo_ubica__id = ubicacion).order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
            
            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            activostemp = activo_estados_select()
            activostemp.tipo_id = tipo
            activostemp.ubicacion_id = ubicacion
            activostemp.save()

            activoscount = activosx.count()
            paginator = Paginator(activosx, 25)
            page = request.GET.get('page')
            activos = paginator.get_page(page)

            return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'tipo':tipo,'ubicacion':ubicacion})  


        elif tipo and departamento:
             parametros = FiltrarForm(request.GET or None)
             activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                 desc_activo_codigo__activo_tipo__id = tipo).filter(desc_activo_codigo__activo_depar__id = departamento)
             
             activostemp = activos_temp_excel.objects.all().delete()
             activostemp = activos_temp_excel()
            
             for activos in activosx:
                activostemp.id = activos.desc_activo_codigo.id
                activostemp.activo_codigo = str(activos.desc_activo_codigo.activo_codigo)
                activostemp.cod_activo_padre = str(activos.desc_activo_codigo.cod_activo_padre)
                activostemp.activo_ubica = str(activos.desc_activo_codigo.activo_ubica)
                activostemp.activo_area = str(activos.desc_activo_codigo.activo_area) 
                activostemp.activo_descripcion = str(activos.desc_activo_codigo.activo_descripcion)
                activostemp.desc_activo_marca = str(activos.desc_activo_marca)
                activostemp.desc_activo_modelo = str(activos.desc_activo_modelo)
                activostemp.desc_activo_num_serie = str(activos.desc_activo_num_serie)
                activostemp.desc_activo_num_motor = str(activos.desc_activo_num_motor)
                activostemp.activo_valor = float(activos.desc_activo_codigo.activo_valor)
                activostemp.save()
             
             activoscount = activosx.count()
             paginator = Paginator(activosx, 25)
             page = request.GET.get('page')
             activos = paginator.get_page(page)

             return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'tipo':tipo,'departamento':departamento})  

        elif tipo and area:
             parametros = FiltrarForm(request.GET or None) 
             activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                 desc_activo_codigo__activo_tipo__id = tipo).filter(desc_activo_codigo__activo_area__area_codigo = area)
             
             activostemp = activos_temp_excel.objects.all().delete()
             activostemp = activos_temp_excel()
            
             for activos in activosx:
                activostemp.id = activos.desc_activo_codigo.id
                activostemp.activo_codigo = str(activos.desc_activo_codigo.activo_codigo)
                activostemp.cod_activo_padre = str(activos.desc_activo_codigo.cod_activo_padre)
                activostemp.activo_ubica = str(activos.desc_activo_codigo.activo_ubica)
                activostemp.activo_area = str(activos.desc_activo_codigo.activo_area) 
                activostemp.activo_descripcion = str(activos.desc_activo_codigo.activo_descripcion)
                activostemp.desc_activo_marca = str(activos.desc_activo_marca)
                activostemp.desc_activo_modelo = str(activos.desc_activo_modelo)
                activostemp.desc_activo_num_serie = str(activos.desc_activo_num_serie)
                activostemp.desc_activo_num_motor = str(activos.desc_activo_num_motor)
                activostemp.activo_valor = float(activos.desc_activo_codigo.activo_valor)
                activostemp.save()
             
             activoscount = activosx.count()
             paginator = Paginator(activosx, 25)
             page = request.GET.get('page')
             activos = paginator.get_page(page)

             return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'tipo':tipo,'area':area})  

        parametros = FiltrarForm(request.GET or None) 
        
        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_tipo__id = tipo).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.tipo_id = tipo
        activostemp.save()
        
        activoscount = activosx.count()

        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)
        
        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'tipo':tipo,'ubicacion':ubicacion})  

    elif ubicacion:
        if ubicacion and area:
            parametros = FiltrarForm(request.GET or None) 
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_ubica__id = ubicacion).filter(
                    desc_activo_codigo__activo_area__area_codigo = area).order_by('-desc_activo_codigo')
            
            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            activostemp = activo_estados_select()
            activostemp.area_id = area
            activostemp.ubicacion_id = ubicacion
            activostemp.save()
            
            activoscount = activosx.count()
            paginator = Paginator(activosx, 25)
            page = request.GET.get('page')
            activos = paginator.get_page(page)
        
            return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'departamento':departamento,'ubicacion':ubicacion})  

        parametros = FiltrarForm(request.GET or None) 

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_ubica__id = ubicacion).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        # activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        activostemp.save()

        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'ubicacion':ubicacion})  
 

    elif departamento:
        parametros = FiltrarForm(request.GET or None) 

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_depar__id = departamento).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').annotate(
                            fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                                'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.departamento_id = departamento
        activostemp.save()

        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'departamento':departamento})  
 

    elif area:
        parametros = FiltrarForm(request.GET or None)  

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_area__area_codigo = area).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.area_id = area
        activostemp.save()
        
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'area':area})  
 

    elif newcustodi:
        print("Entró Custodio")
        custodio = unicodedata.normalize('NFC', newcustodi)
        print(custodio, len(custodio))
        newcustodio = custodio.replace("  "," ")
        parametros = FiltrarForm(request.GET or None)  

        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(Q(
            desc_activo_custodio__icontains = custodio)).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')


        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
        activostemp = activo_estados_select()
        activostemp.custodio = custodio
        activostemp.save()
        
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)
        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')

        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'var':activoscount,'area':area,'personal':personal})  
 


    elif (area and ubicacion and codigo and tipo and departamento) is None:
        print("entro")
        #activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        parametros = FiltrarForm(request.GET or None) 
        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').exclude(
                        desc_activo_codigo__activo_estado='CODIGO ELIMINADO').annotate(
                            fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                                'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
        #       desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
        #           desc_activo_codigo__activo_estado='DONADO').exclude(
        #               desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        activosy = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='ACTIVO')


        #activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            #nuevas lineas optimizar listado de activos
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        activostemp.departamento_id = departamento
        activostemp.area_id = area
        activostemp.grupo_id = grupo
        activostemp.tipo_id = tipo
        
        activostemp.save()
        activoscount = activosx.count() - activosy.count()
        bajascount = activosy.count()
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)

        personal = Persona.objects.all().order_by('last_name')
   
        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'count':activoscount,'bajas':bajascount,'cuenta_grab':cuenta_grab,'cuenta_eti':cuenta_eti,'cuenta_foto':cuenta_foto,'personal':personal}) 

    else: 
        print("Ultima Opcion")
        #activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        parametros = FiltrarForm(request.GET or None) 
        q1 = Imagenes.objects.values('activo_codigo').distinct()

        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').annotate(
                        fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1])).order_by(
                            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
        #activostemp = activos_temp_excel.objects.all().delete()
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE activos.activo_estados_select")
            #nuevas lineas optimizar listado de activos
        activostemp = activo_estados_select()
        activostemp.ubicacion_id = ubicacion
        activostemp.departamento_id = departamento
        activostemp.area_id = area
        activostemp.grupo_id = grupo
        activostemp.tipo_id = tipo
        
        activostemp.save()
        activoscount = activosx.count()
        paginator = Paginator(activosx, 25)

        page = request.GET.get('page')
        activos = paginator.get_page(page)
   
        return render(request,'activos/filtrar_activos.html',{'form':activos,'form2':parametros,'count':activoscount,'cuenta_grab':cuenta_grab,'cuenta_eti':cuenta_eti}) 


#No uilizada
@login_required
def to_excel(request):

    codigo = request.GET.get("activo_codigo")
    tipo = request.GET.get("activo_tipo")
    ubicacion = request.GET.get("activo_ubica")
    area = request.GET.get("activo_area")
    departamento = request.GET.get("activo_depar")

    print(codigo)
    print(tipo)
    print(ubicacion)
    print(area)
    print(departamento)

    if (area and ubicacion and codigo and tipo and departamento) is None:
        
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo')
        
    
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-activos.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        worksheet = workbook.active
        worksheet.title = 'Activos'

        columns = [
            'No',
            'Codigo',
            'Parte de',
            'Grupo',
            'Tipo Activo',
            'Ubicación',
            'Departamento','Area',
            'Descripcion',
            ]
        row_num = 1
    
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

    
        for activo in activosx:
            row_num += 1
        
            row = [
                activo.desc_activo_codigo.id,
                activo.desc_activo_codigo.activo_codigo,
                activo.desc_activo_codigo.cod_activo_padre,
                activo.desc_activo_codigo.activo_grupo,activo.desc_activo_codigo.activo_tipo,
                activo.desc_activo_codigo.activo_ubica,
                activo.desc_activo_codigo.activo_depar,
                activo.desc_activo_codigo.activo_area,
                activo.desc_activo_codigo.activo_descripcion, 
                ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

                workbook.save(response)

                return response
    else:
        return HttpResponse("Se escogió otra cosa")

#Vistas basadas en clases

class ActivoList(LoginRequiredMixin,ListView):
    model = desc_activo
    template_name = 'activos/listar_activos.html'

class FiltrarActivos(LoginRequiredMixin,ListView):
    model = desc_activo
    template_name = 'activos/filtrar_activos.html'

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        context['filter'] = SnippetFilter(self.request.GET, queryset=self.get_queryset())
        return context

class FiltrarDetalleActivos(LoginRequiredMixin,DetailView):
    model = desc_activo
    template_name = 'activos/filtrar_activos.html'



class ActivoUpdate(LoginRequiredMixin,UpdateView):
    model = desc_activo
    second_model = detalle_desc_activo
    third_model = Imagenes
    template_name = 'activos/editar_activos.html'
    form_class = RegistroForm
    second_form_class = EditaActivo
    ImageFormSet = modelformset_factory(Imagenes, form=ImageForm, extra=3)
    success_url = reverse_lazy('activos:listar_activos')
    
    

    def dispatch(self, request, *args, **kwargs):
        
        return super(ActivoUpdate, self).dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super(ActivoUpdate, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        activo = self.model.objects.get(id=pk)
        detalle = self.second_model.objects.get(desc_activo_codigo=activo.id)

        
        if detalle.desc_activo_fecha_ingre:
            fecha = datetime.strftime(detalle.desc_activo_fecha_ingre, '%Y/%m/%d')
        else:
            fecha = None

        formset = self.ImageFormSet(queryset=Imagenes.objects.none())
    

        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'fecha' not in context:
            context['fecha'] = fecha
        if 'form3' not in context:
            context['form3'] = formset
        if 'pk' not in context:
            context['pk'] = pk
        if 'activo' not in context:
            context['activo'] = activo
        
        return context
    
    def post(self, request, *args, **kwargs):
        
        
        self.object = self.get_object
        id_activo = kwargs['pk']
        activor = self.model.objects.get(id=id_activo)
        codigo = desc_activo.objects.get(id=id_activo)
        
        detalle = self.second_model.objects.get(desc_activo_codigo=activor.id)
        imagenes = self.third_model.objects.filter(activo_codigo=activor.id)
        form = self.form_class(request.POST, instance=activor)
        form2 = self.second_form_class(request.POST, instance=detalle)
        
        formset = self.ImageFormSet(self.request.POST, self.request.FILES)      

        if form.is_valid() and form2.is_valid() and formset.is_valid():
            print("Pasó la validación")
            print (formset)
        
            form.save()
            form2.save()

            for formimage in formset.cleaned_data:
                if formimage:
                    print(formimage)
                    image = formimage['imagen_activo']
                    Imagenes.objects.create(imagen_activo=image, activo_codigo=codigo)
                else:
                    print("No hay form")

            return HttpResponseRedirect(self.get_success_url())


       
            # return HttpResponse("Exito!!")
        else:
            return HttpResponse("Error No se pudo guardar")
            print(formset.errors)
            #return HttpResponseRedirect(self.get_success_url())

class ActivoDelete(LoginRequiredMixin,DeleteView):
    model = desc_activo
    template_name = 'activos/eliminar_activos.html'
    success_url = reverse_lazy('activos:listar_activos')

class RegistroNuevo(LoginRequiredMixin,CreateView):
    model = detalle_desc_activo
    template_name ='activos/registro_activos.html'
    form_class = RegistroFormDet
    second_form_class = RegistroForm
    success_url = reverse_lazy('home')

    def get_context_data(self, **kwargs):
        context = super(RegistroNuevo, self).get_context_data(**kwargs)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        return  context
    
    def post(self,request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST)
        form2 = self.second_form_class(request.POST)
        if form.is_valid() and form2.is_valid():
            registro = form.save(commit=False)
            registro.desc_activo_codigo = form2.save()
            registro.save()
            return HttpResponseRedirect(self.get_success_url())
            # return HttpResponse("Exito!!")
        else:
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class ActivoMoverUbicacion(LoginRequiredMixin,UpdateView):
    model = desc_activo
    second_model = detalle_desc_activo
    template_name = 'activos/mueve_activos.html'
    form_class = MoverActivosForm
    #second_form_class = MoverActivosFormDet
    success_url = reverse_lazy('activos:listar_activos')

    def get_context_data(self, **kwargs):
        context = super(ActivoMoverUbicacion, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        activo = self.model.objects.get(id=pk)
        #detalle = self.second_model.objects.get(desc_activo_codigo=activo.id)
        if 'form' not in context:
            context['form'] = self.form_class()
        #if 'form2' not in context:
        #    context['form2'] = self.second_form_class(instance=detalle)
        return context
    

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_activo = kwargs['pk']
        activor = self.model.objects.get(id=id_activo)
        #detalle = self.second_model.objects.get(desc_activo_codigo=activor.id)
        form = self.form_class(request.POST, instance=activor)
        #form2 = self.second_form_class(request.POST, instance=detalle)
      
        #if form.is_valid() and form2.is_valid():
        if form.is_valid():
            form.save()
            
           # cursor = connection.cursor()
           # cursor.execute("UPDATE `activos_detalle_desc_activo` SET desc_activo_contador_cambio = 1 WHERE id=%s",[activor.id])
         #   form2.save()
      
            return HttpResponseRedirect(self.get_success_url())
            # return HttpResponse("Exito!!")
        else:
            return HttpResponse("Error No se pudo guardar")
            #return HttpResponseRedirect(self.get_success_url())

class ActivoCambiaEstado(LoginRequiredMixin,UpdateView):
    model = desc_activo
    second_model = detalle_desc_activo
    template_name = 'activos/dar_baja_activos.html'
    form_class = CambiaEstadoForm
    success_url = reverse_lazy('activos:listar_activos')   

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_activo = kwargs['pk']
        activor = self.model.objects.get(id=id_activo)
        
        cursor = connection.cursor()
        cursor.execute("UPDATE activos.desc_activo SET activo_estado = 'DADO DE BAJA' WHERE id=%s",[activor.id])
        
        return HttpResponseRedirect(self.get_success_url())

     

class custom_to_excelclass(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        # activosx = activos_temp_excel.objects.all()
        registro = activo_estados_select.objects.all().first()
        
        codigo = registro.codigo
        ubicacion = registro.ubicacion_id
        custodio = registro.custodio
        clave = registro.pclave

        if codigo:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                          Q(desc_activo_codigo__activo_codigo__iexact = codigo )|Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo))


            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'CUSTODIO'
            ws['M3'] = 'MARCA'
            ws['N3'] = 'MODELO'
            ws['O3'] = 'SERIE'
            ws['P3'] = 'CODIGO MBA'
            ws['Q3'] = 'FECHA DE COMPRA'
            ws['R3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
               
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 12).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 13).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 18).value = truncvalue
                else:
                    ws.cell(row = count, column = 18).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Custom.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response


        elif clave:
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
            Q(desc_activo_codigo__activo_codigo__icontains = clave) |
            Q(desc_activo_custodio__icontains = clave) |
            Q(desc_activo_codigo__activo_descripcion__icontains = clave) |
            Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__icontains = clave) |
            Q(desc_activo_num_serie__icontains = clave) |
            Q(desc_activo_codigo_mba__icontains = clave) |
            Q(desc_activo_marca__icontains = clave) |
            Q(desc_activo_codigo__activo_area__area_nombre__icontains = clave) |
            Q(desc_activo_codigo__activo_subsec__subsector_nombre__icontains = clave))

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Palabra Clave'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'CUSTODIO'
            ws['M3'] = 'MARCA'
            ws['N3'] = 'MODELO'
            ws['O3'] = 'SERIE'
            ws['P3'] = 'CODIGO MBA'
            ws['Q3'] = 'FECHA DE COMPRA'
            ws['R3'] = 'ANIO DE COMPRA'
            
            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 12).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 13).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 18).value = truncvalue
                else:
                    ws.cell(row = count, column = 18).value = "--"
                

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz_Custom.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        

        elif ubicacion:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_ubica__id = ubicacion).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Ubicación'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'CUSTODIO'
            ws['M3'] = 'MARCA'
            ws['N3'] = 'MODELO'
            ws['O3'] = 'SERIE'
            ws['P3'] = 'CODIGO MBA'
            ws['Q3'] = 'FECHA DE COMPRA'
            ws['R3'] = 'ANIO DE COMPRA'
            
            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 12).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 13).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 18).value = truncvalue
                else:
                    ws.cell(row = count, column = 18).value = "--"
                

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz_Custom.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        
        elif custodio:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_custodio = custodio).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Custodio'

            ws.merge_cells('B1:G1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'CUSTODIO'
            ws['M3'] = 'MARCA'
            ws['N3'] = 'MODELO'
            ws['O3'] = 'SERIE'
            ws['P3'] = 'CODIGO MBA'
            ws['Q3'] = 'FECHA DE COMPRA'
            ws['R3'] = 'ANIO DE COMPRA'
            ws['S3'] = 'CUSTODIO'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 12).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 13).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 16).value = "--"
               
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 18).value = truncvalue
                else:
                    ws.cell(row = count, column = 18).value = "--"
                
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 19).value = activo.desc_activo_custodio
                else:
                    ws.cell(row = count, column = 19).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz_Custom.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        else:
            print("######!!!! DM")

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('-desc_activo_codigo__activo_valor')
        


            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Total de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'CUSTODIO'
            ws['M3'] = 'MARCA'
            ws['N3'] = 'MODELO'
            ws['O3'] = 'SERIE'
            ws['P3'] = 'CODIGO MBA'
            ws['Q3'] = 'FECHA DE COMPRA'
            ws['R3'] = 'ANIO DE COMPRA'
           

            count = 4  
            rowcount = 1

            print("Va por aqui")

            
            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                
                if activo.desc_activo_codigo.activo_area:
                    ws.cell(row = count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
                else:
                    ws.cell(row = count, column = 9).value = "--"
                
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
               
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 11).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 12).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 13).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 17).value = truncvalue
                else:
                    ws.cell(row = count, column = 17).value = "--"
                
                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Total_Custom_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response



class to_excelclass(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        # activosx = activos_temp_excel.objects.all()
        registro = activo_estados_select.objects.all().first()
        
        codigo = registro.codigo
        ubicacion = registro.ubicacion_id
        ubica_nombre = registro.ubicacion
        departamento = registro.departamento_id
        depar_nombre = registro.departamento
        area = registro.area_id
        area_nombre = registro.area
        tipo = registro.tipo_id
        tipo_nombre = registro.tipo
        grupo = registro.grupo_id
        grupo_nombre = registro.grupo
        seguros = registro.seguros_id
        seguros_nombre = registro.seguros
        poliza = registro.poliza
        custodio = registro.custodio

        if codigo:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').filter(
                          Q(desc_activo_codigo__activo_codigo__iexact = codigo )|Q(desc_activo_codigo__cod_activo_padre__cod_activo_padre__iexact = codigo))


            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        
        if ubicacion:
            if ubicacion and departamento:
                activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                    desc_activo_codigo__activo_ubica__id = ubicacion).filter(
                        desc_activo_codigo__activo_depar__id = departamento).order_by('-desc_activo_codigo')
            
                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'CODIGO'
                ws['D3'] = 'PARTE DE'
                ws['E3'] = 'GRUPO'
                ws['F3'] = 'TIPO ACTIVO'
                ws['G3'] = 'UBICACION'
                ws['H3'] = 'DEPARTAMENTO'
                ws['I3'] = 'SECTOR'
                ws['J3'] = 'SUB SECTOR'
                ws['K3'] = 'DESCRIPCION'
                ws['L3'] = 'VALOR REPOSICION'
                ws['M3'] = 'VALOR DE COMPRA'
                ws['N3'] = 'CUSTODIO'
                ws['O3'] = 'MARCA'
                ws['P3'] = 'MODELO'
                ws['Q3'] = 'SERIE'
                ws['R3'] = 'CODIGO MBA'
                ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
                ws['T3'] = 'NUMERO FACTURA'
                ws['U3'] = 'FECHA DE COMPRA'
                ws['V3'] = 'ANIO DE COMPRA'
                
                count = 4
                rowcount = 1  

                for activo in activosx:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                    if activo.desc_activo_codigo.cod_activo_padre:
                        ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                    else:
                        ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                    if activo.desc_activo_codigo.activo_subsec:
                        ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                    else:
                        ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                    if activo.desc_activo_codigo.activo_valor:
                        ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    else:
                        ws.cell(row = count, column = 12).value = "" 
                    
                    if activo.desc_activo_codigo.activo_valor_compra:
                        ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    else:
                        ws.cell(row = count, column = 13).value = ""
                    if activo.desc_activo_custodio:
                        ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                    else:
                        ws.cell(row = count, column = 14).value = "--"
                    if activo.desc_activo_marca:
                        ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                    else:
                        ws.cell(row = count, column = 15).value = "--"
                    if activo.desc_activo_modelo:
                        ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                    else:
                        ws.cell(row = count, column = 16).value = "--"
                    if activo.desc_activo_num_serie:
                        ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                    else:
                        ws.cell(row = count, column = 17).value = "--"
                    if activo.desc_activo_codigo_mba:
                        ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                    else:
                        ws.cell(row = count, column = 18).value = "--"
                    if activo.desc_activo_codigo.pestanias_seguros:
                        ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                    else:
                        ws.cell(row = count, column = 19).value = "--"
                    if activo.desc_activo_codigo.numero_factura:
                        ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                    else:
                        ws.cell(row = count, column = 20).value = "--"
                    
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                    else:
                        ws.cell(row = count, column = 21).value = "--"
                    truncfechacompra = str(activo.desc_activo_fecha_ingre)
                    truncvalue = truncfechacompra[0:4]
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 22).value = truncvalue
                    else:
                        ws.cell(row = count, column = 22).value = "--"
                    

                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response
            
            elif ubicacion and area:
                activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                    desc_activo_codigo__activo_ubica__id = ubicacion).filter(
                        desc_activo_codigo__activo_area__area_codigo = area).order_by('-desc_activo_codigo')
            
                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'CODIGO'
                ws['D3'] = 'PARTE DE'
                ws['E3'] = 'GRUPO'
                ws['F3'] = 'TIPO ACTIVO'
                ws['G3'] = 'UBICACION'
                ws['H3'] = 'DEPARTAMENTO'
                ws['I3'] = 'SECTOR'
                ws['J3'] = 'SUB SECTOR'
                ws['K3'] = 'DESCRIPCION'
                ws['L3'] = 'VALOR REPOSICION'
                ws['M3'] = 'VALOR DE COMPRA'
                ws['N3'] = 'CUSTODIO'
                ws['O3'] = 'MARCA'
                ws['P3'] = 'MODELO'
                ws['Q3'] = 'SERIE'
                ws['R3'] = 'CODIGO MBA'
                ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
                ws['T3'] = 'NUMERO FACTURA'
                ws['U3'] = 'FECHA DE COMPRA'
                ws['V3'] = 'ANIO DE COMPRA'

                count = 4
                rowcount = 1  

                for activo in activosx:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                    if activo.desc_activo_codigo.cod_activo_padre:
                        ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                    else:
                        ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                    if activo.desc_activo_codigo.activo_subsec:
                        ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                    else:
                        ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                    if activo.desc_activo_codigo.activo_valor:
                        ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    else:
                        ws.cell(row = count, column = 12).value = "" 
                    
                    if activo.desc_activo_codigo.activo_valor_compra:
                        ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    else:
                        ws.cell(row = count, column = 13).value = ""
                    if activo.desc_activo_custodio:
                        ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                    else:
                        ws.cell(row = count, column = 14).value = "--"
                    if activo.desc_activo_marca:
                        ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                    else:
                        ws.cell(row = count, column = 15).value = "--"
                    if activo.desc_activo_modelo:
                        ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                    else:
                        ws.cell(row = count, column = 16).value = "--"
                    if activo.desc_activo_num_serie:
                        ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                    else:
                        ws.cell(row = count, column = 17).value = "--"
                    if activo.desc_activo_codigo_mba:
                        ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                    else:
                        ws.cell(row = count, column = 18).value = "--"
                    if activo.desc_activo_codigo.pestanias_seguros:
                        ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                    else:
                        ws.cell(row = count, column = 19).value = "--"
                    if activo.desc_activo_codigo.numero_factura:
                        ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                    else:
                        ws.cell(row = count, column = 20).value = "--"
                    
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                    else:
                        ws.cell(row = count, column = 21).value = "--"
                    truncfechacompra = str(activo.desc_activo_fecha_ingre)
                    truncvalue = truncfechacompra[0:4]
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 22).value = truncvalue
                    else:
                        ws.cell(row = count, column = 22).value = "--"
                    

                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response

                
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_ubica__id = ubicacion).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Ubicación'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
                    


                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz_Ubicacion.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        
        elif area:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_area__area_codigo = area).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Area'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response


        elif custodio:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_custodio = custodio).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz por Custodio'

            ws.merge_cells('B1:G1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'
            ws['W3'] = 'CUSTODIO'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
                
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 23).value = activo.desc_activo_custodio
                else:
                    ws.cell(row = count, column = 23).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        

        elif seguros:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__pestanias_seguros__id = seguros).exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                  desc_activo_codigo__activo_estado='DONADO').exclude(
                      desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

            
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz para manejo de Seguros'

            ws.merge_cells('B1:H1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'SECTOR'
            ws['F3'] = 'SUB SECTOR'
            ws['G3'] = 'TIPO ACTIVO'
            ws['H3'] = 'DESCRIPCION'
            ws['I3'] = 'VALOR REPOSICION'
            ws['J3'] = 'VALOR DE COMPRA'
            ws['K3'] = 'MARCA'
            ws['L3'] = 'MODELO'
            ws['M3'] = 'SERIE'
            ws['N3'] = 'NUMERO FACTURA'
            ws['O3'] = 'FECHA DE COMPRA'
            ws['P3'] = 'POLIZA DE SEGUROS'
            ws['Q3'] = 'INCENDIOS'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 6).value = str(activo.desc_activo_codigo.activo_subsec )
                else:
                    ws.cell(row = count, column = 6).value = "--"
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 8).value = str(activo.desc_activo_codigo.activo_descripcion)
                
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 9).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 9).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 10).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 10).value = ""
                
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 11).value = "--" 
                
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 12).value = "--" 
                
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 13).value = "--" 
                
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 14).value = "--" 
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_codigo.poliza_seguros:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_codigo.get_poliza_seguros_display())
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_codigo.incendios:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_codigo.incendios)
                else:
                    ws.cell(row = count, column = 17).value = "NO"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_para_Seguros.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        

        elif poliza:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__poliza_seguros = poliza).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
 
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz para manejo de Pólizas de Seguros'

            ws.merge_cells('B1:H1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'SECTOR'
            ws['F3'] = 'SUB SECTOR'
            ws['G3'] = 'TIPO ACTIVO'
            ws['H3'] = 'DESCRIPCION'
            ws['I3'] = 'VALOR REPOSICION'
            ws['J3'] = 'VALOR DE COMPRA'
            ws['K3'] = 'MARCA'
            ws['L3'] = 'MODELO'
            ws['M3'] = 'SERIE'
            ws['N3'] = 'NUMERO FACTURA'
            ws['O3'] = 'FECHA DE COMPRA'
            ws['P3'] = 'POLIZA DE SEGUROS'
            ws['Q3'] = 'INCENDIOS'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 6).value = str(activo.desc_activo_codigo.activo_subsec )
                else:
                    ws.cell(row = count, column = 6).value = "--"
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 8).value = str(activo.desc_activo_codigo.activo_descripcion)
                
                if activo.desc_activo_codigo.activo_valor:
                    # ws.cell(row = count, column = 9).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    ws.cell(row = count, column = 9).value = activo.desc_activo_codigo.activo_valor
                
                else:
                    ws.cell(row = count, column = 9).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    # ws.cell(row = count, column = 10).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    ws.cell(row = count, column = 10).value = activo.desc_activo_codigo.activo_valor_compra
                else:
                    ws.cell(row = count, column = 10).value = ""
                
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 11).value = "--" 
                
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 12).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 12).value = "--" 
                
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 13).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 13).value = "--" 
                
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 14).value = "--" 
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                
                if activo.desc_activo_codigo.poliza_seguros:
                    ws.cell(row= count, column = 16).value = str(activo.desc_activo_codigo.get_poliza_seguros_display())
                else:
                    ws.cell(row= count, column = 16).value = "--"
                if activo.desc_activo_codigo.incendios:
                    ws.cell(row= count, column = 17).value = str(activo.desc_activo_codigo.incendios)
                else:
                    ws.cell(row= count, column = 17).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Polizas_de_Seguros.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        
        elif tipo:
            
            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
                desc_activo_codigo__activo_tipo__id = tipo).exclude(
                    desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                        desc_activo_codigo__activo_estado='DONADO').exclude(
                            desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            if tipo == 83: #VEHICULOS

                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'CODIGO'
                ws['D3'] = 'PARTE DE'
                ws['E3'] = 'GRUPO'
                ws['F3'] = 'TIPO ACTIVO'
                ws['G3'] = 'UBICACION'
                ws['H3'] = 'DEPARTAMENTO'
                ws['I3'] = 'SECTOR'
                ws['J3'] = 'SUB SECTOR'
                ws['K3'] = 'DESCRIPCION'
                ws['L3'] = 'VALOR REPOSICION'
                ws['M3'] = 'VALOR DE COMPRA'
                ws['N3'] = 'CUSTODIO'
                ws['O3'] = 'MARCA'
                ws['P3'] = 'MODELO'
                ws['Q3'] = 'SERIE'
                ws['R3'] = 'CODIGO MBA'
                ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
                ws['T3'] = 'NUMERO FACTURA'
                ws['U3'] = 'FECHA DE COMPRA'
                ws['V3'] = 'ANIO DE COMPRA'
                ws['W3'] = 'PLACA'

                count = 4
                rowcount = 1  

                for activo in activosx:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                    if activo.desc_activo_codigo.cod_activo_padre:
                        ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                    else:
                        ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                    if activo.desc_activo_codigo.activo_subsec:
                        ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                    else:
                        ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                    if activo.desc_activo_codigo.activo_valor:
                        ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    else:
                        ws.cell(row = count, column = 12).value = "" 
                    
                    if activo.desc_activo_codigo.activo_valor_compra:
                        ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    else:
                        ws.cell(row = count, column = 13).value = ""
                    if activo.desc_activo_custodio:
                        ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                    else:
                        ws.cell(row = count, column = 14).value = "--"
                    if activo.desc_activo_marca:
                        ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                    else:
                        ws.cell(row = count, column = 15).value = "--"
                    if activo.desc_activo_modelo:
                        ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                    else:
                        ws.cell(row = count, column = 16).value = "--"
                    if activo.desc_activo_num_serie:
                        ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                    else:
                        ws.cell(row = count, column = 17).value = "--"
                    if activo.desc_activo_codigo_mba:
                        ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                    else:
                        ws.cell(row = count, column = 18).value = "--"
                    if activo.desc_activo_codigo.pestanias_seguros:
                        ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                    else:
                        ws.cell(row = count, column = 19).value = "--"
                    if activo.desc_activo_codigo.numero_factura:
                        ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                    else:
                        ws.cell(row = count, column = 20).value = "--"
                    
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                    else:
                        ws.cell(row = count, column = 21).value = "--"
                    truncfechacompra = str(activo.desc_activo_fecha_ingre)
                    truncvalue = truncfechacompra[0:4]
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 22).value = truncvalue
                    else:
                        ws.cell(row = count, column = 22).value = "--"
                
                    if activo.desc_activo_num_placa_vehiculo:
                        ws.cell(row = count, column = 23).value = activo.desc_activo_num_placa_vehiculo
                    else:
                        ws.cell(row = count, column = 23).value = "--"


                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response
            
            else:
                
                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'CODIGO'
                ws['D3'] = 'PARTE DE'
                ws['E3'] = 'GRUPO'
                ws['F3'] = 'TIPO ACTIVO'
                ws['G3'] = 'UBICACION'
                ws['H3'] = 'DEPARTAMENTO'
                ws['I3'] = 'SECTOR'
                ws['J3'] = 'SUB SECTOR'
                ws['K3'] = 'DESCRIPCION'
                ws['L3'] = 'VALOR REPOSICION'
                ws['M3'] = 'VALOR DE COMPRA'
                ws['N3'] = 'CUSTODIO'
                ws['O3'] = 'MARCA'
                ws['P3'] = 'MODELO'
                ws['Q3'] = 'SERIE'
                ws['R3'] = 'CODIGO MBA'
                ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
                ws['T3'] = 'NUMERO FACTURA'
                ws['U3'] = 'FECHA DE COMPRA'
                ws['V3'] = 'ANIO DE COMPRA'

                count = 4
                rowcount = 1  

                for activo in activosx:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                    if activo.desc_activo_codigo.cod_activo_padre:
                        ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                    else:
                        ws.cell(row = count, column = 4).value = "--"
                    ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                    ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                    ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                    ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                    ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                    if activo.desc_activo_codigo.activo_subsec:
                        ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                    else:
                        ws.cell(row = count, column = 10).value = "--"
                    ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                    if activo.desc_activo_codigo.activo_valor:
                        ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    else:
                        ws.cell(row = count, column = 12).value = "" 
                    
                    if activo.desc_activo_codigo.activo_valor_compra:
                        ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    else:
                        ws.cell(row = count, column = 13).value = ""
                    if activo.desc_activo_custodio:
                        ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                    else:
                        ws.cell(row = count, column = 14).value = "--"
                    if activo.desc_activo_marca:
                        ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                    else:
                        ws.cell(row = count, column = 15).value = "--"
                    if activo.desc_activo_modelo:
                        ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                    else:
                        ws.cell(row = count, column = 16).value = "--"
                    if activo.desc_activo_num_serie:
                        ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                    else:
                        ws.cell(row = count, column = 17).value = "--"
                    if activo.desc_activo_codigo_mba:
                        ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                    else:
                        ws.cell(row = count, column = 18).value = "--"
                    if activo.desc_activo_codigo.pestanias_seguros:
                        ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                    else:
                        ws.cell(row = count, column = 19).value = "--"
                    if activo.desc_activo_codigo.numero_factura:
                        ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                    else:
                        ws.cell(row = count, column = 20).value = "--"
                    
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                    else:
                        ws.cell(row = count, column = 21).value = "--"
                    truncfechacompra = str(activo.desc_activo_fecha_ingre)
                    truncvalue = truncfechacompra[0:4]
                    if activo.desc_activo_fecha_ingre:
                        ws.cell(row = count, column = 22).value = truncvalue
                    else:
                        ws.cell(row = count, column = 22).value = "--"
                

                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response



        elif grupo:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_grupo__id = grupo).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo')

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
        

        elif departamento:

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(
            desc_activo_codigo__activo_depar__id = departamento).exclude(
                desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(
                    desc_activo_codigo__activo_estado='DONADO').exclude(
                        desc_activo_codigo__activo_estado='VENDIDO').order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')
        
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Filtrado de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'

            count = 4
            rowcount = 1  

            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                ws.cell(row = count, column = 9).value =  str(activo.desc_activo_codigo.activo_area)
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
               

                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response
 
        #SEcción para pantalla principal de reportes
        else:
            print("######!!!!")

            activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('-desc_activo_codigo__activo_valor')
        


            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte Total de Activos Ecofroz'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'CODIGO'
            ws['D3'] = 'PARTE DE'
            ws['E3'] = 'GRUPO'
            ws['F3'] = 'TIPO ACTIVO'
            ws['G3'] = 'UBICACION'
            ws['H3'] = 'DEPARTAMENTO'
            ws['I3'] = 'SECTOR'
            ws['J3'] = 'SUB SECTOR'
            ws['K3'] = 'DESCRIPCION'
            ws['L3'] = 'VALOR REPOSICION'
            ws['M3'] = 'VALOR DE COMPRA'
            ws['N3'] = 'CUSTODIO'
            ws['O3'] = 'MARCA'
            ws['P3'] = 'MODELO'
            ws['Q3'] = 'SERIE'
            ws['R3'] = 'CODIGO MBA'
            ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
            ws['T3'] = 'NUMERO FACTURA'
            ws['U3'] = 'FECHA DE COMPRA'
            ws['V3'] = 'ANIO DE COMPRA'
            ws['W3'] = 'POLIZA DE SEGUROS'
            ws['X3'] = 'INCENDIOS'

            count = 4  
            rowcount = 1

            print("Va por aqui")

            
            for activo in activosx:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
                if activo.desc_activo_codigo.cod_activo_padre:
                    ws.cell(row = count, column = 4).value = str(activo.desc_activo_codigo.cod_activo_padre )
                else:
                    ws.cell(row = count, column = 4).value = "--"
                ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
                ws.cell(row = count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
                ws.cell(row = count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
                ws.cell(row = count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)
                
                if activo.desc_activo_codigo.activo_area:
                    ws.cell(row = count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
                else:
                    ws.cell(row = count, column = 9).value = "--"
                
                if activo.desc_activo_codigo.activo_subsec:
                    ws.cell(row = count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
                else:
                    ws.cell(row = count, column = 10).value = "--"
                ws.cell(row = count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
                if activo.desc_activo_codigo.activo_valor:
                    # ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',','.')
                    ws.cell(row = count, column = 12).value = activo.desc_activo_codigo.activo_valor
                else:
                    ws.cell(row = count, column = 12).value = "" 
                
                if activo.desc_activo_codigo.activo_valor_compra:
                    # ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',','.')
                    ws.cell(row = count, column = 13).value = activo.desc_activo_codigo.activo_valor_compra
                else:
                    ws.cell(row = count, column = 13).value = ""
                if activo.desc_activo_custodio:
                    ws.cell(row = count, column = 14).value = str(activo.desc_activo_custodio)
                else:
                    ws.cell(row = count, column = 14).value = "--"
                if activo.desc_activo_marca:
                    ws.cell(row = count, column = 15).value = str(activo.desc_activo_marca)
                else:
                    ws.cell(row = count, column = 15).value = "--"
                if activo.desc_activo_modelo:
                    ws.cell(row = count, column = 16).value = str(activo.desc_activo_modelo)
                else:
                    ws.cell(row = count, column = 16).value = "--"
                if activo.desc_activo_num_serie:
                    ws.cell(row = count, column = 17).value = str(activo.desc_activo_num_serie)
                else:
                    ws.cell(row = count, column = 17).value = "--"
                if activo.desc_activo_codigo_mba:
                    ws.cell(row = count, column = 18).value = str(activo.desc_activo_codigo_mba)
                else:
                    ws.cell(row = count, column = 18).value = "--"
                if activo.desc_activo_codigo.pestanias_seguros:
                    ws.cell(row = count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
                else:
                    ws.cell(row = count, column = 19).value = "--"
                if activo.desc_activo_codigo.numero_factura:
                    ws.cell(row = count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
                else:
                    ws.cell(row = count, column = 20).value = "--"
                
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 21).value = str(activo.desc_activo_fecha_ingre)
                else:
                    ws.cell(row = count, column = 21).value = "--"
                truncfechacompra = str(activo.desc_activo_fecha_ingre)
                truncvalue = truncfechacompra[0:4]
                if activo.desc_activo_fecha_ingre:
                    ws.cell(row = count, column = 22).value = truncvalue
                else:
                    ws.cell(row = count, column = 22).value = "--"
                if activo.desc_activo_codigo.poliza_seguros:
                    ws.cell(row = count, column = 23).value = str(activo.desc_activo_codigo.get_poliza_seguros_display())
                else:
                    ws.cell(row = count, column = 23).value = "--"
                if activo.desc_activo_codigo.incendios:
                    ws.cell(row = count, column = 24).value = str(activo.desc_activo_codigo.incendios)
                else:
                    ws.cell(row = count, column = 24).value = "--"
                
                    
                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Activos_Filtrados_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

class to_mba(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        # activosx = activos_temp_excel.objects.all()
        
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo')

        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte Total de Activos Ecofroz'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'PARTE DE'
        ws['D3'] = 'CODIGO MBA'
        ws['E3'] = 'DESCRIPCION'
        ws['F3'] = 'MARCA'
        ws['G3'] = 'MODELO'
        ws['H3'] = 'SERIE'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'FECHA COMPRA'
        ws['K3'] = 'No FACTURA'
        ws['L3'] = 'VALOR REPOSICION'
        ws['M3'] = 'TIPO'
        ws['N3'] = 'GRUPO'
        ws['O3'] = 'DEPARTAMENTO'
        ws['P3'] = 'UBICACION'
        ws['Q3'] = 'AREA'
        ws['R3'] = 'RESPONSABLE'


        count = 4
        rowcount = 1  

        for activo in activosx:
            #ws.cell(row = count, column = 2).value =  rowcount
            ws.cell(row = count, column = 2).value =  str(activo.desc_activo_codigo.activo_codigo)
            ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.cod_activo_padre)
            ws.cell(row = count, column = 4).value =  str(activo.desc_activo_codigo_mba)
            ws.cell(row = count, column = 5).value =  str(activo.desc_activo_codigo.activo_descripcion)
            ws.cell(row = count, column = 6).value =  str(activo.desc_activo_marca)
            ws.cell(row = count, column = 7).value =  str(activo.desc_activo_modelo)
            ws.cell(row = count, column = 8).value =  str(activo.desc_activo_num_serie)
            ws.cell(row = count, column = 9).value =  str(activo.desc_activo_proveedor)
            ws.cell(row = count, column = 10).value =  str(activo.desc_activo_fecha_ingre)
            ws.cell(row = count, column = 11).value =  str(activo.desc_activo_codigo.numero_factura)
            ws.cell(row = count, column = 12).value =  str(activo.desc_activo_codigo.activo_valor)
            ws.cell(row = count, column = 13).value =  str(activo.desc_activo_codigo.activo_tipo)
            ws.cell(row = count, column = 14).value =  str(activo.desc_activo_codigo.activo_grupo)
            ws.cell(row = count, column = 15).value =  str(activo.desc_activo_codigo.activo_depar)
            ws.cell(row = count, column = 16).value =  str(activo.desc_activo_codigo.activo_ubica)
            ws.cell(row = count, column = 17).value =  str(activo.desc_activo_codigo.activo_area)
            ws.cell(row = count, column = 18).value = str(activo.desc_activo_custodio)

            count+=1
            rowcount+=1

        nombre_archivo = "Reporte_de_Activos_para_MBA_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class to_ExcelAll(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        q1 = Imagenes.objects.values('activo_codigo').distinct()

        #ANNOTATE CREA UNA COLUMNA ADICINAL EN LA QUE SE COLOCAN LOS IDS DE LOS ACTIVOS QUE TIENEN FOTO
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').annotate(
                    fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1]))

        

        # q1 = Imagenes.objects.values('activo_codigo').distinct().annotate(
        #     imagen_activo=Subquery(activosx.filter(desc_activo_codigo=OuterRef('activo_codigo')).values('desc_activo_codigo')[:1]))
                

        # print(q1, type(q1))

                
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte Total de Activos Ecofroz'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'No'
        ws['C3'] = 'CODIGO'
        ws['D3'] = 'PARTE DE'
        ws['E3'] = 'GRUPO'
        ws['F3'] = 'TIPO ACTIVO'
        ws['G3'] = 'UBICACION'
        ws['H3'] = 'DEPARTAMENTO'
        ws['I3'] = 'SECTOR'
        ws['J3'] = 'SUB SECTOR'
        ws['K3'] = 'DESCRIPCION'
        ws['L3'] = 'VALOR REPOSICION'
        ws['M3'] = 'VALOR DE COMPRA'
        ws['N3'] = 'CUSTODIO'
        ws['O3'] = 'MARCA'
        ws['P3'] = 'MODELO'
        ws['Q3'] = 'SERIE'
        ws['R3'] = 'CODIGO MBA'
        ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
        ws['T3'] = 'NUMERO FACTURA'
        ws['U3'] = 'FECHA DE COMPRA'
        ws['V3'] = 'ANIO DE COMPRA'
        ws['W3'] = 'POLIZA DE SEGUROS'
        ws['X3'] = 'INCENDIOS'
        ws['Y3'] = 'IDENTIFICADO'
        ws['Z3'] = 'MOTIVO NO SE PUEDE IDENTIFICAR'
        ws['AA3'] = 'FOTOS'

        count = 4  
        rowcount = 1

        # print()

            
        for activo in activosx:
        
            ws.cell(row = count, column = 2).value =  rowcount
            
            ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
            
            
            if activo.desc_activo_codigo.cod_activo_padre:
                ws.cell(row=count, column=4).value = str(
                    activo.desc_activo_codigo.cod_activo_padre)
            else:
                ws.cell(row= count, column = 4).value = "--"
            ws.cell(row= count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
            ws.cell(row= count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
            ws.cell(row= count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
            ws.cell(row= count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)

            if activo.desc_activo_codigo.activo_area:
                ws.cell(row= count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
            else:
                ws.cell(row= count, column = 9).value = "--"

            # if activo.desc_activo_codigo.activo_subsec:
            #     ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            # else:
            #     ws.cell(row= count, column = 10).value = "--"
            if activo.desc_activo_codigo.activo_subsec:
                ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            else:
                ws.cell(row=count, column = 10).value = '--'

            ws.cell(row= count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
            
            if activo.desc_activo_codigo.activo_valor:
                # ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',', '.')
                ws.cell(row = count, column = 12).value = activo.desc_activo_codigo.activo_valor
            else:
                ws.cell(row = count, column = 12).value = ""

            if activo.desc_activo_codigo.activo_valor_compra:
                # ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',', '.')
                ws.cell(row = count, column = 13).value = activo.desc_activo_codigo.activo_valor_compra
            else:
                ws.cell(row= count, column = 13).value = ""
            
            if activo.desc_activo_custodio:
                ws.cell(row= count, column = 14).value = str(activo.desc_activo_custodio)
            else:
                ws.cell(row= count, column = 14).value = "--"
            
            if activo.desc_activo_marca:
                ws.cell(row= count, column = 15).value = str(activo.desc_activo_marca)
            else:
                ws.cell(row= count, column = 15).value = "--"
            
            if activo.desc_activo_modelo:
                ws.cell(row= count, column = 16).value = str(activo.desc_activo_modelo)
            else:
                ws.cell(row= count, column = 16).value = "--"
            
            if activo.desc_activo_num_serie:
                ws.cell(row= count, column = 17).value = str(activo.desc_activo_num_serie)
            else:
                ws.cell(row= count, column = 17).value = "--"
            
            if activo.desc_activo_codigo_mba:
                ws.cell(row= count, column = 18).value = str(activo.desc_activo_codigo_mba)
            else:
                ws.cell(row= count, column = 18).value = "--"
            
            if activo.desc_activo_codigo.pestanias_seguros:
                ws.cell(row= count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
            else:
                ws.cell(row= count, column = 19).value = "--"
            
            if activo.desc_activo_codigo.numero_factura:
                ws.cell(row= count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
            else:
                ws.cell(row= count, column = 20).value = "--"

            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 21).value = str(activo.desc_activo_fecha_ingre)
            else:
                ws.cell(row= count, column = 21).value = "--"
           
            truncfechacompra = str(activo.desc_activo_fecha_ingre)
            truncvalue = truncfechacompra[0:4]
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 22).value = truncvalue
            else:
                ws.cell(row= count, column = 22).value = "--"
            
            if activo.desc_activo_codigo.poliza_seguros:
                ws.cell(row= count, column = 23).value = str(activo.desc_activo_codigo.get_poliza_seguros_display())
            else:
                ws.cell(row= count, column = 23).value = "--"
            if activo.desc_activo_codigo.incendios:
                ws.cell(row= count, column = 24).value = str(activo.desc_activo_codigo.incendios)
            else:
                ws.cell(row= count, column = 24).value = "NO"

            if activo.desc_activo_codigo.grabado:
                ws.cell(row= count, column = 25).value =  str(activo.desc_activo_codigo.grabado)
            else:
                ws.cell(row= count, column = 25).value = "--"
            
            if activo.motivo_no_se_puede_identificar:
                ws.cell(row= count, column = 26).value =  str(activo.motivo_no_se_puede_identificar)
            else:
                ws.cell(row= count, column = 26).value = "--"
            
            # ws.cell(row= count, column = 27).value = "SI"

            if activo.fotos:
                ws.cell(row= count, column = 27).value = "SI"
            else:
                ws.cell(row= count, column = 27).value = "--"

            count += 1
            rowcount += 1

        nombre_archivo = "Reporte_de_Activos_Total_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

class CustomExcelAll(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):

        q1 = Imagenes.objects.values('activo_codigo').distinct()
        
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='DADO DE BAJA').exclude(desc_activo_codigo__activo_estado='DONADO').exclude(
                  desc_activo_codigo__activo_estado='VENDIDO').annotate(
                    fotos=Subquery(q1.filter(activo_codigo=OuterRef('desc_activo_codigo__id')).values('activo_codigo')[:1]))

        
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte Total de Activos Ecofroz'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'No'
        ws['C3'] = 'CODIGO'
        ws['D3'] = 'PARTE DE'
        ws['E3'] = 'GRUPO'
        ws['F3'] = 'TIPO ACTIVO'
        ws['G3'] = 'UBICACION'
        ws['H3'] = 'DEPARTAMENTO'
        ws['I3'] = 'SECTOR'
        ws['J3'] = 'SUB SECTOR'
        ws['K3'] = 'DESCRIPCION'
        ws['L3'] = 'CUSTODIO'
        ws['M3'] = 'MARCA'
        ws['N3'] = 'MODELO'
        ws['O3'] = 'SERIE'
        ws['P3'] = 'CODIGO MBA'
        ws['Q3'] = 'FECHA DE COMPRA'
        ws['R3'] = 'ANIO DE COMPRA'
        ws['S3'] = 'IDENTIFICADO'
        ws['T3'] = 'MOTIVO NO SE PUEDE IDENTIFICAR'
        ws['U3'] = 'FOTO'

        count = 4  
        rowcount = 1

        # print()

            
        for activo in activosx:
            ws.cell(row = count, column = 2).value =  rowcount
            ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
            if activo.desc_activo_codigo.cod_activo_padre:
                ws.cell(row=count, column=4).value = str(
                    activo.desc_activo_codigo.cod_activo_padre)
            else:
                ws.cell(row= count, column = 4).value = "--"
            ws.cell(row= count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
            ws.cell(row= count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
            ws.cell(row= count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
            ws.cell(row= count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)

            if activo.desc_activo_codigo.activo_area:
                ws.cell(row= count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
            else:
                ws.cell(row= count, column = 9).value = "--"

            # if activo.desc_activo_codigo.activo_subsec:
            #     ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            # else:
            #     ws.cell(row= count, column = 10).value = "--"
            if activo.desc_activo_codigo.activo_subsec:
                ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            else:
                ws.cell(row=count, column = 10).value = '--'

            ws.cell(row= count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
            
            if activo.desc_activo_custodio:
                ws.cell(row= count, column = 12).value = str(activo.desc_activo_custodio)
            else:
                ws.cell(row= count, column = 12).value = "--"
            
            if activo.desc_activo_marca:
                ws.cell(row= count, column = 13).value = str(activo.desc_activo_marca)
            else:
                ws.cell(row= count, column = 13).value = "--"
            
            if activo.desc_activo_modelo:
                ws.cell(row= count, column = 14).value = str(activo.desc_activo_modelo)
            else:
                ws.cell(row= count, column = 14).value = "--"
            
            if activo.desc_activo_num_serie:
                ws.cell(row= count, column = 15).value = str(activo.desc_activo_num_serie)
            else:
                ws.cell(row= count, column = 15).value = "--"
            
            if activo.desc_activo_codigo_mba:
                ws.cell(row= count, column = 16).value = str(activo.desc_activo_codigo_mba)
            else:
                ws.cell(row= count, column = 16).value = "--"
            
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 17).value = str(activo.desc_activo_fecha_ingre)
            else:
                ws.cell(row= count, column = 17).value = "--"
           
            truncfechacompra = str(activo.desc_activo_fecha_ingre)
            truncvalue = truncfechacompra[0:4]
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 18).value = truncvalue
            else:
                ws.cell(row= count, column = 18).value = "--"
            
            if activo.desc_activo_codigo.grabado:
                ws.cell(row= count, column = 19).value =  str(activo.desc_activo_codigo.grabado)
            else:
                ws.cell(row= count, column = 19).value = "--"
            
            if activo.motivo_no_se_puede_identificar:
                ws.cell(row= count, column = 20).value =  str(activo.motivo_no_se_puede_identificar)
            else:
                ws.cell(row= count, column = 20).value = "--"
            
            if activo.fotos:
                ws.cell(row= count, column = 21).value = "SI"
            else:
                ws.cell(row= count, column = 21).value = "--"
            


            count += 1
            rowcount += 1

        nombre_archivo = "Reporte_de_Activos_Total_Ecofroz_Custom-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


class CustomExcelBajas(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='ACTIVO')

        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Bajas, Donaciones y Ventas Activos Ecofroz'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'No'
        ws['C3'] = 'CODIGO'
        ws['D3'] = 'PARTE DE'
        ws['E3'] = 'GRUPO'
        ws['F3'] = 'TIPO ACTIVO'
        ws['G3'] = 'UBICACION'
        ws['H3'] = 'DEPARTAMENTO'
        ws['I3'] = 'SECTOR'
        ws['J3'] = 'SUB SECTOR'
        ws['K3'] = 'DESCRIPCION'
        ws['L3'] = 'CUSTODIO'
        ws['M3'] = 'MARCA'
        ws['N3'] = 'MODELO'
        ws['O3'] = 'SERIE'
        ws['P3'] = 'CODIGO MBA'
        ws['Q3'] = 'FECHA DE COMPRA'
        ws['R3'] = 'ANIO DE COMPRA'
        ws['S3'] = 'DESCRIPCION MOVIMIENTO'
        ws['T3'] = 'FECHA BAJA VENTA O DONACION'
        ws['U3'] = 'USUARIO REALIZA BAJA VENTA O DONACION'
        ws['V3'] = 'NOTAS BAJA VENTA O DONACION'

        count = 4  
        rowcount = 1

        # print()

            
        for activo in activosx:
            ws.cell(row = count, column = 2).value =  rowcount
            ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
            if activo.desc_activo_codigo.cod_activo_padre:
                ws.cell(row=count, column=4).value = str(
                    activo.desc_activo_codigo.cod_activo_padre)
            else:
                ws.cell(row= count, column = 4).value = "--"
            ws.cell(row= count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
            ws.cell(row= count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
            ws.cell(row= count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
            ws.cell(row= count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)

            if activo.desc_activo_codigo.activo_area:
                ws.cell(row= count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
            else:
                ws.cell(row= count, column = 9).value = "--"

            if activo.desc_activo_codigo.activo_subsec:
                ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            else:
                ws.cell(row= count, column = 10).value = "--"
            ws.cell(row= count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
            
            if activo.desc_activo_custodio:
                ws.cell(row= count, column = 12).value = str(activo.desc_activo_custodio)
            else:
                ws.cell(row= count, column = 12).value = "--"
            if activo.desc_activo_marca:
                ws.cell(row= count, column = 13).value = str(activo.desc_activo_marca)
            else:
                ws.cell(row= count, column = 13).value = "--"
            if activo.desc_activo_modelo:
                ws.cell(row= count, column = 14).value = str(activo.desc_activo_modelo)
            else:
                ws.cell(row= count, column = 14).value = "--"
            if activo.desc_activo_num_serie:
                ws.cell(row= count, column = 15).value = str(activo.desc_activo_num_serie)
            else:
                ws.cell(row= count, column = 15).value = "--"
            if activo.desc_activo_codigo_mba:
                ws.cell(row= count, column = 16).value = str(activo.desc_activo_codigo_mba)
            else:
                ws.cell(row= count, column = 16).value = "--"
            
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 17).value = str(activo.desc_activo_fecha_ingre)
            else:
                ws.cell(row= count, column = 17).value = "--"
            truncfechacompra = str(activo.desc_activo_fecha_ingre)
            truncvalue = truncfechacompra[0:4]
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 18).value = truncvalue
            else:
                ws.cell(row= count, column = 18).value = "--"
            
            
            ws.cell(row= count, column = 19).value =  str(activo.desc_activo_codigo.activo_estado)
            ws.cell(row= count, column = 20).value =  str(activo.fecha_baja_dona_vende)
            ws.cell(row= count, column = 21).value =  str(activo.usuario_realiza_baja_dona_vende)
            if activo.notas_baja_dona_vende:
                ws.cell(row= count, column = 22).value =  str(activo.notas_baja_dona_vende)
            else:
                ws.cell(row= count, column = 22).value = "--"


            count += 1
            rowcount += 1

        nombre_archivo = "Reporte_de_Bajas_Donaciones_Ventas_Ecofroz_Custom-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response



#Reporte de Excel de Bajas

class to_ExcelBajas(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").order_by('desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre').exclude(
              desc_activo_codigo__activo_estado='ACTIVO')

        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Bajas, Donaciones y Ventas Activos Ecofroz'

        ws.merge_cells('B1:E1')
        ws['B3'] = 'No'
        ws['C3'] = 'CODIGO'
        ws['D3'] = 'PARTE DE'
        ws['E3'] = 'GRUPO'
        ws['F3'] = 'TIPO ACTIVO'
        ws['G3'] = 'UBICACION'
        ws['H3'] = 'DEPARTAMENTO'
        ws['I3'] = 'SECTOR'
        ws['J3'] = 'SUB SECTOR'
        ws['K3'] = 'DESCRIPCION'
        ws['L3'] = 'VALOR REPOSICION'
        ws['M3'] = 'VALOR DE COMPRA'
        ws['N3'] = 'CUSTODIO'
        ws['O3'] = 'MARCA'
        ws['P3'] = 'MODELO'
        ws['Q3'] = 'SERIE'
        ws['R3'] = 'CODIGO MBA'
        ws['S3'] = 'PESTAÑA ARCHIVO SEGUROS'
        ws['T3'] = 'NUMERO FACTURA'
        ws['U3'] = 'FECHA DE COMPRA'
        ws['V3'] = 'ANIO DE COMPRA'
        ws['W3'] = 'POLIZA DE SEGUROS'
        ws['X3'] = 'INCENDIOS'
        ws['Y3'] = 'DESCRIPCION MOVIMIENTO'
        ws['Z3'] = 'FECHA BAJA VENTA O DONACION'
        ws['AA3'] = 'USUARIO REALIZA BAJA VENTA O DONACION'
        ws['AB3'] = 'NOTAS BAJA VENTA O DONACION'

        count = 4  
        rowcount = 1

        # print()

            
        for activo in activosx:
            ws.cell(row = count, column = 2).value =  rowcount
            ws.cell(row = count, column = 3).value =  str(activo.desc_activo_codigo.activo_codigo)
            if activo.desc_activo_codigo.cod_activo_padre:
                ws.cell(row=count, column=4).value = str(
                    activo.desc_activo_codigo.cod_activo_padre)
            else:
                ws.cell(row= count, column = 4).value = "--"
            ws.cell(row= count, column = 5).value =  str(activo.desc_activo_codigo.activo_grupo)
            ws.cell(row= count, column = 6).value =  str(activo.desc_activo_codigo.activo_tipo)
            ws.cell(row= count, column = 7).value =  str(activo.desc_activo_codigo.activo_ubica)
            ws.cell(row= count, column = 8).value =  str(activo.desc_activo_codigo.activo_depar)

            if activo.desc_activo_codigo.activo_area:
                ws.cell(row= count, column = 9).value = str(activo.desc_activo_codigo.activo_area)
            else:
                ws.cell(row= count, column = 9).value = "--"

            if activo.desc_activo_codigo.activo_subsec:
                ws.cell(row= count, column = 10).value = str(activo.desc_activo_codigo.activo_subsec)
            else:
                ws.cell(row= count, column = 10).value = "--"
            ws.cell(row= count, column = 11).value = str(activo.desc_activo_codigo.activo_descripcion)
            if activo.desc_activo_codigo.activo_valor:
                ws.cell(row = count, column = 12).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor).replace(',', '.')
            else:
                ws.cell(row = count, column = 12).value = ""

            if activo.desc_activo_codigo.activo_valor_compra:
                ws.cell(row = count, column = 13).value = "{:,.2f}".format(activo.desc_activo_codigo.activo_valor_compra).replace(',', '.')
            else:
                ws.cell(row= count, column = 13).value = ""
            if activo.desc_activo_custodio:
                ws.cell(row= count, column = 14).value = str(activo.desc_activo_custodio)
            else:
                ws.cell(row= count, column = 14).value = "--"
            if activo.desc_activo_marca:
                ws.cell(row= count, column = 15).value = str(activo.desc_activo_marca)
            else:
                ws.cell(row= count, column = 15).value = "--"
            if activo.desc_activo_modelo:
                ws.cell(row= count, column = 16).value = str(activo.desc_activo_modelo)
            else:
                ws.cell(row= count, column = 16).value = "--"
            if activo.desc_activo_num_serie:
                ws.cell(row= count, column = 17).value = str(activo.desc_activo_num_serie)
            else:
                ws.cell(row= count, column = 17).value = "--"
            if activo.desc_activo_codigo_mba:
                ws.cell(row= count, column = 18).value = str(activo.desc_activo_codigo_mba)
            else:
                ws.cell(row= count, column = 18).value = "--"
            if activo.desc_activo_codigo.pestanias_seguros:
                ws.cell(row= count, column = 19).value = str(activo.desc_activo_codigo.pestanias_seguros)
            else:
                ws.cell(row= count, column = 19).value = "--"
            if activo.desc_activo_codigo.numero_factura:
                ws.cell(row= count, column = 20).value = str(activo.desc_activo_codigo.numero_factura)
            else:
                ws.cell(row= count, column = 20).value = "--"

            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 21).value = str(activo.desc_activo_fecha_ingre)
            else:
                ws.cell(row= count, column = 21).value = "--"
            truncfechacompra = str(activo.desc_activo_fecha_ingre)
            truncvalue = truncfechacompra[0:4]
            if activo.desc_activo_fecha_ingre:
                ws.cell(row= count, column = 22).value = truncvalue
            else:
                ws.cell(row= count, column = 22).value = "--"
            if activo.desc_activo_codigo.poliza_seguros:
                ws.cell(row= count, column = 23).value = str(activo.desc_activo_codigo.get_poliza_seguros_display())
            else:
                ws.cell(row= count, column = 23).value = "--"
            if activo.desc_activo_codigo.incendios:
                ws.cell(row= count, column = 24).value = str(activo.desc_activo_codigo.incendios)
            else:
                ws.cell(row= count, column = 24).value = "--"
            ws.cell(row= count, column = 25).value =  str(activo.desc_activo_codigo.activo_estado)
            ws.cell(row= count, column = 26).value =  str(activo.fecha_baja_dona_vende)
            ws.cell(row= count, column = 27).value =  str(activo.usuario_realiza_baja_dona_vende)
            if activo.notas_baja_dona_vende:
                ws.cell(row= count, column = 28).value =  str(activo.notas_baja_dona_vende)
            else:
                ws.cell(row= count, column = 28).value = "--"


            count += 1
            rowcount += 1

        nombre_archivo = "Reporte_de_Bajas_Donaciones_Ventas_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


#Vistas de generacion de reportes PDF

@login_required
def imprimeSalidaActivos(request,id):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        reg = salida_activos.objects.get(id=id)
        autorizador = User.objects.get(username=reg.persona_autoriza_dep)
        solicitador = User.objects.get(username=reg.solicitado_por)
        seguridad_fisica = User.objects.get(username=reg.pers_autoriza_seguridad)
        
        num = str(reg.id)
        orden = str(reg.orden_mantenimiento)
        fec = str(reg.fecha_registro)
        fec_estimada = str(reg.fecha_estimada_retorno)
        tipo = str(reg.activo_tipo)
        serie = str(reg.activo_num_serie)
        marca = str(reg.marca)
        activo_codigo_draw =  '*' + str(reg.activo_codigo) + '*' 
        activo_codigo = str(reg.activo_codigo)  
        
        detalle = reg.detalle_activo
        empresa_mantenimiento = str(reg.empresa_mantenimiento)
        motivo = str(reg.motivo)
        autoriza = str(autorizador.first_name + ' ' + autorizador.last_name)
        responsable = str(solicitador.first_name + ' ' + solicitador.last_name)
        seguridad = str(seguridad_fisica.first_name + ' ' + seguridad_fisica.last_name)
        if reg.ubica_depar:
            depar = str(reg.ubica_depar)
        else:
            depar = ''
        if reg.ubica_area:
            area = str(reg.ubica_area)
        else:
            area=''
        #response = HttpResponse(content_type='application/pdf')
        #response['Content-Disposition'] = 'attachment; filename="autorizacion_salida_activos.pdf"'
        w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, activo_codigo_draw)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Guía de Autorización de Salida de Activos (ECSEG-R   1.2-5)")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Guía #:")
        p.drawString(50,710, "Requisición #:")
        p.drawString(50,690, "Fecha Solicitud Salida:")
        p.drawString(50,670, "Fecha Aproximada de Retorno:")
        p.drawString(50,650, "Tipo de Activo:")
        p.drawString(50,630, "Serie:")
        p.drawString(50,610, "Marca:")
        p.drawString(50,580, "Código de Activo:")
        p.drawString(50,560, "Empresa de Mantenimiento:")
        p.drawString(50,540, "Movimiento Departamento:")
        p.drawString(50,520, "Movimiento Área:")

        p.drawString(50,500, "Justificación de Salida:")
        p.drawString(50,440, "Descripción:")
        
        p.drawString(30,194, "Persona Solicita:")
        p.drawString(220,194, "Persona Autoriza Area:")
        p.drawString(395,194, "Persona Autoriza Seguridad:")
        p.drawString(190,100, "---------------------------------------------------")
        p.drawString(215,90, "Nombre y Firma Transportista ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in detalle:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        moti = ''
        c = 1
        for d in motivo:
            if c == 46:
                moti = moti + d
                moti = moti + '\n'
                c = 0
            else:
                moti = moti + d
                c = c + 1
        moti = moti.split('\n')

        # styles = getSampleStyleSheet()
        # estilos = styles['Normal']
        # estilos.alignment = TA_JUSTIFY
        # estilos = ParagraphStyle(name='Justify', alignment=TA_JUSTIFY, fontName='Arial')

        # parrafo = []

        # texto = Paragraph(texto, style=estilos)

        # parrafo.append(texto)

        p.drawString(250,730, num)
        p.drawString(250,710, orden)
        p.drawString(250,690, fec)
        p.drawString(250,670, fec_estimada)
        p.drawString(250,650, tipo)
        p.drawString(250,630, serie)
        p.drawString(250,610, marca)
        p.drawString(250,580, activo_codigo)
        p.drawString(250,560, empresa_mantenimiento)
        p.drawString(250,540, depar)
        p.drawString(250,520, area)
        # p.drawString(250,580, parrafo)
        # text = p.beginText(250, 580)
        # text.textLine(str(Paragraph(detalle, estilos)))
        # text.textLine(str(detalle))
        # p.drawText(text)
        # textobject = p.beginText()  # Iniciamos el textobject
        # textobject.setTextOrigin(250, 580)  # Ubicamos el cursor donde dibujar
        # textobject.setFont("Helvetica-Oblique", 13)
        # for i in texto:
        #     i = i.replace("\n","")  # Remplazamos el salto de linea
        #     textobject.setCharSpace(0.5)  # Espacio entre caracteres 
        #     textobject.setWordSpace(1)  # Espacio entre palabras
        #     textobject.textLine(i)  # Dibujamos la linea
        #     textobject.moveCursor(5, 10)  # Movemos el cursor para luego escribir
        # textobject.setFillGray(0.4)  # Rellenamos con Escala de grises
        # textobject.setLeading(20)  # Espaciado entre lineas
        # textobject.textLines(detalle)  # Insertamos el texto y con color gris
        # p.drawText(textobject)  # Dibujamos el texto pasando un objeto de texto
        # p.drawString(250,580, ''.join(texto))
        a = 0
        b = 440
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
        # p.drawString(250, 580, str(texto))
        # p.drawString(250,440, motivo)
        f = 0
        g = 500
        for i in moti:
            p.drawString(250, g, moti[f])
            f = f + 1
            g = g - 15
        p.drawString(30,180, responsable)
        p.drawString(220,180, autoriza)
        p.drawString(395,180, seguridad)
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="salida.pdf")


@login_required
def imprimeSalidaActivosAgri(request,id):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        reg = salida_activos.objects.get(id=id)
        autorizador = User.objects.get(username=reg.persona_autoriza_dep)
        solicitador = User.objects.get(username=reg.solicitado_por)
        seguridad_fisica = User.objects.get(username=reg.pers_autoriza_seguridad)
        
        num = str(reg.id)
        orden = str(reg.orden_mantenimiento)
        fec = str(reg.fecha_registro)
        fec_estimada = str(reg.fecha_estimada_retorno)
        tipo = str(reg.activo_tipo)
        serie = str(reg.activo_num_serie)
        marca = str(reg.marca)
        activo_codigo_draw =  '*' + str(reg.activo_codigo) + '*' 
        activo_codigo = str(reg.activo_codigo)  
        
        detalle = reg.detalle_activo
        empresa_mantenimiento = str(reg.empresa_mantenimiento)
        motivo = str(reg.motivo)
        autoriza = str(autorizador.first_name + ' ' + autorizador.last_name)
        responsable = str(solicitador.first_name + ' ' + solicitador.last_name)
        seguridad = str(seguridad_fisica.first_name + ' ' + seguridad_fisica.last_name)
        if reg.ubica_depar:
            depar = str(reg.ubica_depar)
        else:
            depar = ''
        if reg.ubica_area:
            area = str(reg.ubica_area)
        else:
            area=''
        
        w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, activo_codigo_draw)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Guía de Autorización de Salida de Activos (ECSEG-R   1.2-5)")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Guía #:")
        # p.drawString(50,710, "Requisición #:")
        p.drawString(50,690, "Fecha Solicitud Salida:")
        p.drawString(50,670, "Fecha Aproximada de Retorno:")
        p.drawString(50,650, "Tipo de Activo:")
        p.drawString(50,630, "Serie:")
        p.drawString(50,610, "Marca:")
        p.drawString(50,580, "Código de Activo:")
        p.drawString(50,560, "Empresa de Mantenimiento:")
        p.drawString(50,540, "Movimiento Departamento:")
        p.drawString(50,520, "Movimiento Área:")

        p.drawString(50,500, "Justificación de Salida/Movimiento:")
        p.drawString(50,440, "Descripción del Activo:")
        
        p.drawString(90,194, "Persona Solicita:")
        p.drawString(280,194, "Persona Registra Bodega:")
        # p.drawString(395,194, "Persona Autoriza Seguridad:")
        # p.drawString(190,100, "---------------------------------------------------")
        # p.drawString(215,90, "Nombre y Firma Transportista ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in detalle:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        moti = ''
        c = 1
        for d in motivo:
            if c == 46:
                moti = moti + d
                moti = moti + '\n'
                c = 0
            else:
                moti = moti + d
                c = c + 1
        moti = moti.split('\n')


        p.drawString(250,730, num)
        # p.drawString(250,710, orden)
        p.drawString(250,690, fec)
        p.drawString(250,670, fec_estimada)
        p.drawString(250,650, tipo)
        p.drawString(250,630, serie)
        p.drawString(250,610, marca)
        p.drawString(250,580, activo_codigo)
        p.drawString(250,560, empresa_mantenimiento)
        p.drawString(250,540, depar)
        p.drawString(250,520, area)
       
        a = 0
        b = 440
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
        # p.drawString(250, 580, str(texto))
        # p.drawString(250,440, motivo)
        f = 0
        g = 500
        for i in moti:
            p.drawString(250, g, moti[f])
            f = f + 1
            g = g - 15
        p.drawString(90,180, responsable)
        p.drawString(280,180, autoriza)
        # p.drawString(395,180, seguridad)
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="salida.pdf")




# '''''''''''''''ACTA DE MOVIMIENTOS INTERNOS DE ACTIVOS''''''''''''''''
@login_required
def imprimeActaMovimientosInternos(request,codigo):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        #reg = DetallePedido.objects.filter(numpedido=numpedido).select_related("numpedido")
        
        activo = desc_activo.objects.get(activo_codigo=codigo)

        cod = activo.id
        
        entrega = request.user.username

        reg = historial_movimientos_internos.objects.filter(activo_codigo = cod).order_by("fecha_registro").last()

        print (reg)

        
        
        #for i in reg:
        num = str(reg.id)
        fec = formats.date_format(timezone.localtime(reg.fecha_registro), "SHORT_DATETIME_FORMAT")
        codigo_genera = str(reg.activo_codigo.activo_codigo)
        desc_string = str(reg.activo_codigo.activo_descripcion)
        entrega_obs = str(reg.justificacion_movimiento)
        tipo = str(activo.activo_tipo.tipo_nombre)
        dep = str(activo.activo_depar.dep_nombre)
        sector = str(activo.activo_area.area_nombre)
        nuevo_depar = str(reg.nuevo_departamento.dep_nombre)
        nuevo_sector = str(reg.nuevo_sector.area_nombre)  
        

        #w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, codigo)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Acta de Entrega-Recepción de Activos")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Guía #:")
        p.drawString(50,710, "Fecha de Entrega:")
        p.drawString(50,690, "Código del Activo:")
        p.drawString(50,670, "Descripción del Activo:")

        p.drawString(50,510, "Tipo Activo:")
        #p.drawString(50,580, "Código de Activo:")
        #p.drawString(50,560, "Empresa de Mantenimiento:")
        p.drawString(50,490, "Departamento Original:")
        p.drawString(50,470, "Sector Original:")
        p.drawString(50,440, "Nuevo Departamento:")
        p.drawString(50,420, "Nuevo Sector:")
       
       
        p.drawString(50,400, "Observaciones de entrega:")
        #p.drawString(30,194, "Solicitante:")
        p.drawString(150,194, "Entrega:")
        p.drawString(350,194, "Recibe:")
        p.drawString(190,100, "---------------------------------------------------")
        p.drawString(215,90, "Nombre y Firma del Custodio ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in desc_string:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        texto2 = ''
        e = 1
        for i in entrega_obs:
            if e == 46:
                texto2 = texto2 + i
                texto2 = texto2 + '\n'
                e = 0
            else:
                texto2 = texto2 + i
                e = e + 1
        texto2 = texto2.split('\n')

        p.drawString(250,730, num)
        p.drawString(250,710, fec)
        p.drawString(250,690, codigo_genera)
        p.drawString(250,510, tipo)
        p.drawString(250,490, dep)
        p.drawString(250,470, sector)

        p.drawString(250,440, nuevo_depar)
        p.drawString(250,420, nuevo_sector)



        
        a = 0
        b = 670
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
       
        a = 0
        b = 400
        for i in texto2:
            p.drawString(250, b, texto2[a])
            a = a + 1
            b = b - 15
       
       
       
        #p.drawString(30,180, "solicita")
        p.drawString(150,180, "entrega")
        p.drawString(350,180, "recibe")
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="entrega.pdf")



@login_required
def autMovimientosActivosExcel(request):
    try:          
        cursor = connection.cursor()
     
        cursor.execute("""select id,solicitado_por,fecha_registro,fecha_estimada_retorno,codigo,activo_tipo,activo_num_serie,
            marca,detalle_activo,sale_por,motivo,empresa_mantenimiento,pers_autoriza_seguridad,estado
            from activos.salida_activos
            order by fecha_registro desc""")

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE SALIDA-RETORNO DE ACTIVOS'

        ws.merge_cells('A1:L1')
        ws['A3'] = 'N°'
        ws['B3'] = 'ORDEN SALIDA N°'
        ws['C3'] = 'SOLICITADO POR'
        ws['D3'] = 'FECHA REGISTRO SOLICITUD'
        ws['E3'] = 'FECHA APROXIMADA DE RETORNO'
        ws['F3'] = 'CODIGO DE ACTIVO'
        ws['G3'] = 'TIPO'
        ws['H3'] = 'SERIE'
        ws['I3'] = 'MARCA'
        ws['J3'] = 'DETALLE'
        ws['K3'] = 'SALE POR'
        ws['L3'] = 'MOTIVO'
        ws['M3'] = 'EMPRESA DE MANTENIMIENTO'
        ws['N3'] = 'AUTORIZA'
        ws['O3'] = 'ESTADO'
       

        count = 4
        rowcount = 1  
        
        for i in cursor:
            print(i)
            ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 2).value =  str(i[0])
            ws.cell(row = count, column = 3).value =  str(i[1])
            ws.cell(row = count, column = 4).value =  str(i[2])[0:10]
            ws.cell(row = count, column = 5).value =  str(i[3])
            ws.cell(row = count, column = 6).value =  str(i[4])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[7])
            ws.cell(row = count, column = 10).value =  str(i[8])
            if i[9] == 4:
                ws.cell(row = count, column = 11).value =  "Movimiento entre ubicaciones propias"
            else:
                ws.cell(row = count, column = 11).value = "Reparación Externa"
            ws.cell(row = count, column = 12).value =  str(i[10])
            ws.cell(row = count, column = 13).value =  str(i[11])
            ws.cell(row = count, column = 14).value =  str(i[12])
            ws.cell(row = count, column = 15).value =  str(i[13])
    
            count+=1
            rowcount+=1
        

        nombre_archivo = "REPORTE DE SALIDA-RETORNO DE ACTIVOS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    
    except Exception as e:
        
        mensaje = 'Error'
       
        return redirect('activos:aut_movimientos_activos_excel')

@login_required
#Entrega de Acta Recepción de Activos para Custodio

@login_required
def imprimeActaEntregaActivosCustodio(request,numpedido):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        reg = DetallePedido.objects.filter(numpedido=numpedido).select_related("numpedido")
        
        for i in reg:
            num = str(i.numpedido.numpedido)
            fec = str(i.numpedido.fecha_entrega)
            desc = i.descripcion
            codigo_genera = str(i.numpedido.cod_activo)
            tipo = str(i.numpedido.tipoactivo.tipo_nombre)
            depar = str(i.numpedido.departamento.dep_nombre)
            sec = str(i.numpedido.area.area_nombre)
            solicita = str(i.numpedido.usuario_solicita.first_name+' '+i.numpedido.usuario_solicita.last_name)
            retira = str(i.numpedido.persona_retira_de_bodega)
            seguridad = str(i.numpedido.persona_confirma_descargo)
            custodio = str(i.numpedido.custodio_sugerido)
            entrega_obs = str(i.numpedido.entrega_observa)
            entrega = str(i.numpedido.persona_entrega_en_bodega)
            seguridad_obs = str(i.numpedido.descargo_custodio_observa)

        desc_string = ''
        for i in desc:
            desc_string = desc_string + i
        
        desc_string = (desc_string[:150] + '..') if len(desc_string) > 200 else desc_string
            
        #w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, codigo_genera)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Acta de entrega de equipos y bienes ECSEG-R-2.1-15")
        p.setFont('Arial', 12)
        p.drawString(50,730, "Requisición #:")
        p.drawString(50,710, "Fecha de Entrega:")
        p.drawString(50,690, "Código del Activo:")
        p.drawString(50,670, "Descripción del Activo:")

        p.drawString(50,510, "Tipo Activo:")
        #p.drawString(50,580, "Código de Activo:")
        #p.drawString(50,560, "Empresa de Mantenimiento:")
        p.drawString(50,490, "Departamento:")
        p.drawString(50,470, "Sector:")
        p.drawString(50,450, "Persona Retira de Bodega:")
        p.drawString(50,430, "Observaciones de Seguridad:")

        p.drawString(50,550, "Observaciones de Bodega:")
        #p.drawString(30,194, "Solicitante:")
        p.drawString(120,194, "Responsable Seguridad:")
        p.drawString(340,194, "Custodio:")
        p.drawString(190,100, "---------------------------------------------------")
        p.drawString(235,90, "Firma Custodio ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in desc_string:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        texto2 = ''
        e = 1
        for i in seguridad_obs:
            if e == 46:
                texto2 = texto2 + i
                texto2 = texto2 + '\n'
                e = 0
            else:
                texto2 = texto2 + i
                e = e + 1
        texto2 = texto2.split('\n')

        p.drawString(250,730, num)
        p.drawString(250,710, fec)
        p.drawString(250,690, codigo_genera)
        p.drawString(250,510, tipo)
        p.drawString(250,490, depar)
        p.drawString(250,470, sec)
        p.drawString(250,450, retira)
        p.drawString(250,430, seguridad_obs)
        
        a = 0
        b = 670
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
       
        a = 0
        b = 550
        for i in texto2:
            p.drawString(250, b, texto2[a])
            a = a + 1
            b = b - 15
       
       
       
        #p.drawString(30,180, solicita)
        p.drawString(120,180, seguridad)
        p.drawString(340,180, custodio)
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="entrega_custodio.pdf")

@login_required
def imprimeActaEntregaActivosCustodio2(request,id):

    if request.method == 'GET':
        pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()
    
        #reg = DetallePedido.objects.filter(numpedido=numpedido).select_related("numpedido")
        reg = detalle_desc_activo.objects.filter(desc_activo_codigo=id).select_related("desc_activo_codigo")

        for i in reg:
            
            fec = str(i.fecha_cambia_custodio)
            desc = i.desc_activo_codigo.activo_descripcion
            codigo_genera = str(i.desc_activo_codigo.activo_codigo)
            tipo = str(i.desc_activo_codigo.activo_tipo)
            depar = str(i.desc_activo_codigo.activo_depar)
            sec = str(i.desc_activo_codigo.activo_area)
            seguridad = str(i.desc_activo_usuario_registra_cambio_custodio)
            custodio = str(i.desc_activo_custodio)
            seguridad_obs = str(i.desc_activo_observaciones_seg_cambio_custodio)

        desc_string = ''
        for i in desc:
            desc_string = desc_string + i
        
        desc_string = (desc_string[:150] + '..') if len(desc_string) > 200 else desc_string
            
        #w, h = A4
        p = canvas.Canvas(buffer)
        p.setLineWidth(.3)
        p.setFont('3 of 9 Barcode', 32)
        p.drawString(410, 790, codigo_genera)
        p.setFont('Arial', 16)
        p.drawString(50, 760, "Acta de entrega de equipos y bienes ECSEG-R-2.1-15")
        p.setFont('Arial', 12)
        #p.drawString(50,730, "Requisición #:")
        p.drawString(50,730, "Fecha de Entrega:")
        p.drawString(50,710, "Código del Activo:")
        p.drawString(50,690, "Tipo Activo:")
        p.drawString(50,670, "Departamento:")
        p.drawString(50,650, "Sector:")
        p.drawString(50,630, "Descripción del Activo:")

        p.drawString(50,550, "Condición:")
        p.drawString(50,530, "Observaciones de Seguridad:")

        p.drawString(120,194, "Responsable Seguridad:")
        p.drawString(340,194, "Custodio:")
        p.drawString(190,100, "---------------------------------------------------")
        p.drawString(235,90, "Firma Custodio ")
        
        p.setFont('Arial', 11)

        texto = ''
        e = 1
        for i in desc_string:
            if e == 46:
                texto = texto + i
                texto = texto + '\n'
                e = 0
            else:
                texto = texto + i
                e = e + 1
        texto = texto.split('\n')

        texto2 = ''
        e = 1
        for i in seguridad_obs:
            if e == 46:
                texto2 = texto2 + i
                texto2 = texto2 + '\n'
                e = 0
            else:
                texto2 = texto2 + i
                e = e + 1
        texto2 = texto2.split('\n')

        #p.drawString(250,730, num)
        p.drawString(250,730, fec)
        p.drawString(250,710, codigo_genera)
        p.drawString(250,690, tipo)
        p.drawString(250,670, depar)
        p.drawString(250,650, sec)
        p.drawString(250,550, "Reuso")
        #p.drawString(250,450, seguridad_obs)
        
        a = 0
        b = 630
        for i in texto:
            p.drawString(250, b, texto[a])
            a = a + 1
            b = b - 15
       
        a = 0
        b = 530
        for i in texto2:
            p.drawString(250, b, texto2[a])
            a = a + 1
            b = b - 15
       
       
        #p.drawString(30,180, solicita)
        p.drawString(120,180, seguridad)
        p.drawString(340,180, custodio)
        

        p.showPage()
        p.save()

        # pdf = buffer.getvalue()
        
        buffer.seek(0)
        # buffer.close()
        # response.write(pdf)
        
        return FileResponse(buffer,as_attachment=False,filename="entrega_custodio.pdf")


@login_required
def imprimeActaEntregaActivosCustodioGlobal(request,nombre):
    now = datetime.now()
    print(nombre)
    # fec = formats.date_format(timezone.localtime(now, pytz.timezone('America/Guayaquil')),"SHORT_DATETIME_FORMAT")
    if request.method == 'GET':
        reg = detalle_desc_activo.objects.filter(desc_activo_custodio__icontains=nombre).order_by(
            'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

        identifica = Persona.objects.get(last_name__icontains=nombre)

        for i in reg:
            desc = i.desc_activo_codigo.activo_descripcion
            codigo_genera = str(i.desc_activo_codigo.activo_codigo)
            depar = str(i.desc_activo_codigo.activo_depar)
            sec = str(i.desc_activo_codigo.activo_area)
            custodio = str(i.desc_activo_custodio)
            seguridad_obs = str(i.desc_activo_observaciones_seg_cambio_custodio)
            ubica = str(i.desc_activo_codigo.activo_ubica)


        data = {
            'reg': reg,
            'identifica': identifica,
            'fch_entrega': now,
            'descripcion': desc,
            'custodio': custodio,
            'sector': sec,
            'departamento': depar,
        }

        template_url = 'activos/reports/entrega_activos.html'

        pdf = render_to_pdf(template_url, data)

        return HttpResponse(pdf, content_type='application/pdf')

        # width, height = A4
        # styles = getSampleStyleSheet()
        # styleN = styles["BodyText"]
        # styleBH = styles["Normal"]
        
        
        # pdfmetrics.registerFont(TTFont('3 of 9 Barcode', os.path.join(settings.BASE_DIR, 'static') + '/fonts/3of9_new.TTF'))
        # pdfmetrics.registerFont(TTFont('Arial', os.path.join(settings.BASE_DIR, 'static') + '/fonts/arial.TTF'))
        # # Create a file-like buffer to receive PDF data.
        # buffer = io.BytesIO()
    
        # #reg = DetallePedido.objects.filter(numpedido=numpedido).select_related("numpedido")
        # reg = detalle_desc_activo.objects.filter(desc_activo_custodio__icontains=nombre).select_related("desc_activo_codigo").order_by(
        #     'desc_activo_codigo__activo_codigo','desc_activo_codigo__cod_activo_padre')

        # identifica = Persona.objects.get(last_name__icontains=nombre)
        # persona = str(identifica.cedula)

        # def li(x):
        #     return str(x, 'utf-8')

        # fec = str(datetime.now().strftime('%Y-%m-%d'))
        # seguridad = str(request.user)

        # for i in reg:
            
        #     #fec = str(i.fecha_cambia_custodio)
        #     desc = i.desc_activo_codigo.activo_descripcion
        #     codigo_genera = str(i.desc_activo_codigo.activo_codigo)
        #     #tipo = str(i.desc_activo_codigo.activo_tipo)
        #     depar = str(i.desc_activo_codigo.activo_depar)
        #     sec = str(i.desc_activo_codigo.activo_area)
        #     #seguridad = str(i.desc_activo_usuario_registra_cambio_custodio)
        #     custodio = str(i.desc_activo_custodio)
        #     seguridad_obs = str(i.desc_activo_observaciones_seg_cambio_custodio)

        # desc_string = ''
        # for i in desc:
        #     desc_string = desc_string + i
        
        # desc_string = (desc_string[:150] + '..') if len(desc_string) > 200 else desc_string
            
        # #w, h = A4
        # p = canvas.Canvas(buffer)
        # p.setLineWidth(.3)
        # p.setFont('3 of 9 Barcode', 32)
        # p.drawString(350, 790, persona)
        # p.setFont('Arial', 16)
        # p.drawString(50, 760, "Acta de entrega de equipos y bienes ECSEG-R-2.1-15")
        # p.setFont('Arial', 12)
        # #p.drawString(50,730, "Requisición #:")
        # p.drawString(50,730, "Fecha de Entrega:")
        # #p.drawString(50,710, "Código del Activo:")
        # p.drawString(50,710, "Custodio:")
        # p.drawString(50,690, "Departamento:")
        # p.drawString(50,670, "Sector:")
        # p.drawString(50,650, "Descripción de el(los) Activo(s):")
        
        # #p.drawString(50,250, "Observaciones de Seguridad:")
        # p.drawString(120,194, "Responsable Seguridad:")
        # p.drawString(340,194, "Custodio:")
        # p.drawString(190,100, "---------------------------------------------------")
        # p.drawString(235,90, "Firma Custodio ")
        
        # p.setFont('Arial', 11)


        # styleSheet = getSampleStyleSheet()
        # style = styleSheet['Normal']

        # header = Paragraph("<bold><font size='18'>TPS Report</font></bold>", style)
        # data = [['N°','Activo Código', 'Tipo', 'Descripción','Grabado/Etiquetado'],
        # ]

        # datas = []
        # rowcount = 1

        # for i in reg:
        #     rt = []
        #     rt.append(str(rowcount))
        #     rt.append(str(i.desc_activo_codigo.activo_codigo))
        #     rt.append(str(i.desc_activo_codigo.activo_tipo))
        #     rt.append(str(i.desc_activo_codigo.activo_descripcion))
        #     rt.append(str(i.desc_activo_codigo.grabado))
        #     datas.append(rt)
        #     rowcount+=1
        
        # for j in datas:
        #     print(j)
        #     data.append(j)
        #     data2 = [[Paragraph(li(cell.encode('utf-8')), styleSheet['Normal']) for cell in row] for row in data]



        # u = Table(data2, colWidths=(30,50, 98, 248, 82))

        # u.setStyle(TableStyle([('FONT',(0, 0), (-1, -1), 'Arial'),
        # ('FONTSIZE',(0, 0), (-1, -1), 9),
        # ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
        # ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black), ] ) )

        # u.wrapOn(p, width, height)
        
        # #La Tabla se dibuja desde el fondo hacia arriba.
        # if reg.count() == 1:
        #     u.drawOn(p,50,510)

        # elif reg.count() == 2:
        #     u.drawOn(p,50,490)
        
        # elif reg.count() == 3:
        #     u.drawOn(p,50,460)

        # elif reg.count() > 3:
        #     u.drawOn(p,50,300)


        # texto2 = ''
        # e = 1
        # for i in seguridad_obs:
        #     if e == 46:
        #         texto2 = texto2 + i
        #         texto2 = texto2 + '\n'
        #         e = 0
        #     else:
        #         texto2 = texto2 + i
        #         e = e + 1
        # texto2 = texto2.split('\n')

        # #p.drawString(250,730, num)
        # p.drawString(250,730, fec)
        # #p.drawString(250,710, codigo_genera)
        # p.drawString(250,710, custodio)
        # p.drawString(250,690, depar)
        # p.drawString(250,670, sec)
       
       
        
        # p.drawString(120,180, seguridad)
        # p.drawString(340,180, custodio)
        

        # p.showPage()
        # p.save()

        # # pdf = buffer.getvalue()
        
        # buffer.seek(0)
        # # buffer.close()
        # # response.write(pdf)
        
        # return FileResponse(buffer,as_attachment=False,filename="entrega_custodio_global.pdf")

