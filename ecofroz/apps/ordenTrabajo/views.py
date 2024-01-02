import logging
import io
import os
from io import BytesIO
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, HttpResponseNotFound, Http404, FileResponse, JsonResponse
from django.urls import reverse_lazy
from django.db.models import Q, F
from django.views.generic import TemplateView, ListView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.auth.decorators import login_required, permission_required
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.template import context, loader
from django.contrib.sites.shortcuts import get_current_site
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import Workbook
from django.db import connection
from django.core.paginator import Paginator
from collections import defaultdict
from apps.ordenPedido.forms import selectAsignadoa
from apps.ordenPedido.models import CotizaPedido, DetallePedido, OrdenesPedidos
from .models import OrdenesTrabajos, DetalleTrabajo, CotizaTrabajo, PedidosMba, TrabajosPago, TipoPedido, ConsultaExpertoTrab, ConsultadosTrab, PreCotiza, PreCotizaDet, SolicitudesAdqui
from apps.activos.models import salida_activos, activo_ubica, activo_grupo, activo_areas
from apps.proveedores.models import proveedor
from .forms import TrabajoForm, TrabajoFormLab, DetTrForm, EditaDetTrForm, DetCotizaFormSet, TrabajoAutoriForm, IngresoTrabajos, PedidosMbaForm, IngresoPagosForm, DetCotizaFormSet_LabCa, selectAsignadoa, Predocumentos, cargaDoc, cargaComprobante
from apps.usuarios.models import User, Cotizador, Autorizador, Generador, Email
from apps.proveedores.models import proveedor
from datetime import date, timedelta
from django.db.models import Subquery, OuterRef
from apps.parametrosGlobales.models import proyectos_contabilidad, tipo_notificacion, notificaciones_globales, modelo_App, consolida_ordenes_proyectos
from django.core.files.storage import FileSystemStorage

# Create your views here.

logger = logging.getLogger(__name__)

### FUNCIONES INTERNAS ###

def envioMail(subject, email, template, queryset, queryset2):
    html_message = loader.render_to_string(
        'ordenTrabajo/email/%s' %template,
            {
                'aprob':queryset,
                'aprob2':queryset2,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except:
        logger.error("Unable to send mail.")

#PARA ENVIAR MENSAJES AL GRUPO ADQUISICIONES DE TELEGRAM

idBot = '1948489247:AAGNGKEKBssZn2DDbapUtPyhuVATuVWoilM'
idGrupo = '-488335650'

def enviarMensajeTelegramAdquisiciones(mensaje):
    requests.post('https://api.telegram.org/bot' + idBot + '/sendMessage',
              data={'chat_id': idGrupo, 'text': mensaje,'parse_mode': 'HTML'})




### VISTAS ###


@login_required
def AnularOrdenes(request):
    busqueda = request.GET.get("buscar")
    anuladas = DetalleTrabajo.objects.filter(numtrabajo__genera_compra=3)[:25]
    if busqueda:
        consulta = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = busqueda) | Q(descripcion__icontains = busqueda))

        return render(request,'ordenTrabajo/anula_ordenes.html', {'consulta':consulta, 'anuladas':anuladas})

    return render(request,'ordenTrabajo/anula_ordenes.html', {'anuladas':anuladas})

@login_required
# @permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
@permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def AnulaOrdenesConfirma(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        idpedido = numpedido
        aprobado = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numpedido)
        aprob = aprobado.numtrabajo.genera_compra
        print(aprob)
        return render(request, 'ordenTrabajo/anula_ordenes_confirma.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})

@login_required        
def confirma_anulacion(request, numpedido):
    
    observa = request.GET.get('text')
    orden = OrdenesTrabajos.objects.get(numtrabajo = numpedido)
    orden.genera_compra = 3
    orden.observa_compra = observa
    orden.aprobado = 2
    orden.fchgenerac = datetime.now()
    orden.estado_cotiza = None
    qrealiza = request.user.username
    orden.usuario_anula = qrealiza
    orden.save()
    
    identificau = User.objects.get(username=qrealiza)
    print(identificau.rol)
    if str(identificau.rol) == 'Analista de Compras':
        
        return redirect('trabajos:cotiza_trabajos')
    else:
        return redirect('trabajos:anular_ordenes')

@login_required
@permission_required('ordenPedido.view_ordenespedidos', raise_exception=True)
def ordenesTrabajo(request):
    queryset = request.GET.get("buscar")
    if queryset:
        orden = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            # Q(numtrabajo__numproyecto__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numtrabajo__usuario_solicita_id = request.user.id)
    else:
        orden2 = DetalleTrabajo.objects.filter(numtrabajo__usuario_solicita_id = request.user.id).order_by('-numtrabajo__numtrabajo')

        paginator = Paginator(orden2, 100)
        page = request.GET.get('page')
        orden = paginator.get_page(page)
    
    return render(request, 'ordenTrabajo/listar_ordenes.html', {'form':orden, 'busqueda':queryset})

@login_required
@permission_required('ordenPedido.view_ordenespedidos', raise_exception=True)
def listar_pedidos_laboratorio(request):
    queryset = request.GET.get("buscar")
    if queryset:
        q1 = DetalleTrabajo.objects.all().filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(numtrabajo__numproyecto__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numtrabajo__usuario_solicita_id = request.user.id).filter(
            Q(numtrabajo__tipo_pedi = 'AM') |
            Q(numtrabajo__tipo_pedi = 'AP') |
            Q(numtrabajo__tipo_pedi = 'CA') |
            Q(numtrabajo__tipo_pedi = 'CP'))
    else:
        q1 = DetalleTrabajo.objects.all().filter(numtrabajo__usuario_solicita_id = request.user.id).filter(
            Q(numtrabajo__tipo_pedi = 'AM') |
            Q(numtrabajo__tipo_pedi = 'AP') |
            Q(numtrabajo__tipo_pedi = 'CA') |
            Q(numtrabajo__tipo_pedi = 'CP')
            ).order_by('-numtrabajo__numtrabajo')

    if q1:
        orden = []
        for data in q1:
            q2 = CotizaTrabajo.objects.get(numtrabajo__numtrabajo = data.numtrabajo.numtrabajo)
            d={
                'pedido':data.numtrabajo,
                'usuario':data.numtrabajo.usuario_solicita.first_name + data.numtrabajo.usuario_solicita.last_name,
                'departamento':data.numtrabajo.departamento,
                'proyecto':data.numtrabajo.numproyecto,
                'fecha':data.numtrabajo.fchsolicita,
                'tipo':data.numtrabajo.tipo_pedi,
                'orden':data.numtrabajo.orden_referencial,
                'descripcion':data.descripcion,
                'valor':q2.valor,
                'estado':data.numtrabajo.genera_compra,
                'proveedor':q2.empresa_cotiza.nombre_empresa,
            }
            orden.append(d)

            paginator = Paginator(orden, 100)
            page = request.GET.get('page')
            orden2 = paginator.get_page(page)
    else:
        orden2 = None

    return render(request, 'ordenTrabajo/listar_pedidos_laboratorio.html', {'form':orden2, 'busqueda':queryset})

def ubicaAreaAjax(request):
    data = []
    action = request.GET['action']
    area = request.GET['id']
    print('hola' + ' ' + area)
    try:
        for i in activo_areas.objects.filter(area_ubica=area):
            data.append({
                'area_codigo': i.area_codigo,
                'area_nombre': i.area_nombre,
            })

        return JsonResponse(data, safe=False)
    except Exception as e:
        print(e)


class validaCotizacion(View):
    model = PreCotiza
    template_name = 'ordenTrabajo/pre_cotiza.html'
    success_url = reverse_lazy('trabajos:cotiza_trabajos')
    
    
    def get(self,request,**kwargs):
        pk = self.kwargs.get('pk',0)
        ruta = request.GET.get('path')
        
        archivos_list = PreCotiza.objects.filter(numtrabajo = pk)
        for i in archivos_list:
            print(i.archivos.url)
        q = DetalleTrabajo.objects.filter(numtrabajo = pk).select_related('numtrabajo')
        
        return render(self.request, 'ordenTrabajo/pre_cotiza.html', {'pk':pk,'archivos_list':archivos_list, 'query':q, 'ruta':ruta})


    def post(self, request, **kwargs):

        doc_id = kwargs['pk']  
        form = Predocumentos(self.request.POST, self.request.FILES)
        
        print(form)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.numtrabajo_id = doc_id
            
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()
        
        else:
            data = {'is_valid': False}
        return JsonResponse(data)

def save_estado_precotiza(request):    #PARA GURDAR EL ESTADO DE LA SELECCION INDIVIDUAL
    id = request.GET.get("id")
    valor = request.GET.get("valor")
    print(valor)
    if valor == 'true':
        valor = True
    else:
        valor = False

    actualiza = PreCotiza.objects.filter(id = id).update(seleccion=valor)

    data = 'OK'

    return JsonResponse(data,safe=False)

def eliminar_documentos(request, pk, trabajo):
    docu = PreCotiza.objects.get(id=pk)
    if request.method == 'POST':
        docu.delete()
        return redirect('trabajos:valida_cotizacion',trabajo)
    return render(request,'ordenTrabajo/eliminar_documentos.html',{'documento':docu})


class revisaCotizacion(View):
    model = PreCotiza
    template_name = 'ordenTrabajo/rev_pre_cotiza.html'
    
    def get(self,request,**kwargs):
        pk = self.kwargs.get('pk',0)
        archivos_list = PreCotiza.objects.filter(numtrabajo = pk)
        queryset = DetalleTrabajo.objects.get(numtrabajo=pk)
        
        return render(self.request, 'ordenTrabajo/rev_pre_cotiza.html', {'queryset':queryset,'archivos_list':archivos_list})


    def post(self, request, **kwargs):
        success_url = reverse_lazy('trabajos:listar_trabajos')
        pk = self.kwargs.get('pk',0)
        observa_responde = request.POST.get('observaciones')

        trabajo = OrdenesTrabajos.objects.filter(numtrabajo=pk)
        trab = OrdenesTrabajos.objects.get(numtrabajo=pk)
        detrabajo = DetalleTrabajo.objects.filter(numtrabajo=pk)

        try:
            guarda_obs_responde = detrabajo.update(observa_responde_precotiza=observa_responde)
            guarda_precotiza_cab = trabajo.update(fchpreresponde = date.today(),estado_responde_precotiza = True)
            
        except:
            
            guarda_precotiza_cab = trabajo.update(fchpreresponde = date.today(),estado_responde_precotiza = True)
            

        #Obtiene dirección de correo de compradora

        
        queryset = DetalleTrabajo.objects.get(numtrabajo=pk)
        queryset2 = PreCotiza.objects.filter(numtrabajo=pk)
        
        id_usuario = queryset.numtrabajo.usuario_solicita.id
        id_solicita = User.objects.get(id=id_usuario)

        id_precotiza =  User.objects.get(username=trab.usuario_precotiza)
     
            
        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=22) # 1. TRABAJOS RESPONDE PRE COTIZACION
        app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = queryset.numtrabajo.numtrabajo,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_precotiza,
            )

        r.save()

        ### FIN DE SECCION DE NOTIFICACIONES ##

        emailaprob = id_precotiza
        html_message = loader.render_to_string(
            'ordenTrabajo/email/email_responde_precotiza.html',
            {
                'aprob':queryset,
                'aprob2':queryset2
            
            }
        )
        email_subject = 'Respuesta de Pre Cotización Trabajos N°'+' '+str(pk)
        to_list = emailaprob.email
        mail = EmailMultiAlternatives(
                email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        mail.attach_alternative(html_message, "text/html")
        try:
            mail.send()
        except:
            logger.error("Unable to send mail.")

        return HttpResponseRedirect(success_url)

def save_record_pre_cotiza(request, pk):
    trabajo = OrdenesTrabajos.objects.filter(numtrabajo=pk)
    detrabajo = DetalleTrabajo.objects.filter(numtrabajo=pk)
    
    observaciones_adqui = request.POST.get("observaciones")
    success_url = reverse_lazy('trabajos:cotiza_trabajos')
    compradora = request.user.username
   
    if request.method == 'POST':
        ruta_prev = request.POST.get('preruta')

        try:
            guarda_obs_adqui = detrabajo.update(observa_envia_precotiza=observaciones_adqui)
            guarda_fecha_precotiza = PreCotiza.objects.filter(numtrabajo=pk).update(escrito=True)
            guarda_precotiza_cab = trabajo.update(fchprecotiza = date.today(),usuario_precotiza=compradora,estado_precotiza = True)
            
        except:
            guarda_fecha_precotiza = PreCotiza.objects.filter(numtrabajo=pk).update(escrito=True)
            guarda_precotiza_cab = trabajo.update(fchprecotiza = date.today(),usuario_precotiza=compradora,estado_precotiza = True)
            

        #Obtiene dirección de correo de solicitante

        
        queryset = DetalleTrabajo.objects.get(numtrabajo=pk)
        
        id_usuario = queryset.numtrabajo.usuario_solicita.id
        id_solicita = User.objects.get(id=id_usuario)
        id_precotiza =  User.objects.get(username=request.user.username)
     
            
        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=21) # 1. TRABAJOS PRE COTIZACION
        app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = queryset.numtrabajo.numtrabajo,
            tipo = noti, 
            usuario_activa = id_precotiza,
            autorizador_id = id_solicita,
            )

        r.save()

        ### FIN DE SECCION DE NOTIFICACIONES ##

        #LINEAS DE CORREO COMENTADAS POR DM EL 16-MAY-2023

        # emailaprob = id_solicita
        # html_message = loader.render_to_string(
        #     'ordenTrabajo/email/email_envia_precotiza.html',
        #     {
        #         'aprob':queryset
            
        #     }
        # )
        # email_subject = 'Envío de Pre Cotización Trabajos N°'+' '+str(pk)
        # to_list = emailaprob.email
        # mail = EmailMultiAlternatives(
        #         email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        # mail.attach_alternative(html_message, "text/html")
        # try:
        #     mail.send()
        # except:
        #     logger.error("Unable to send mail.")


        # return HttpResponseRedirect(liga)
        if ruta_prev != 'None':

            return redirect(ruta_prev)
        
        else:
            return HttpResponseRedirect(success_url)

       

def save_record_responde_pre_cotiza(request, pk):
    trabajo = OrdenesTrabajos.objects.filter(numtrabajo=pk)
    observaciones_iniciador = request.POST.get("observaciones")
    success_url = reverse_lazy('trabajos:listar_trabajos')
   
    if request.method == 'POST':

        try:
            guarda_obs_fecha = PreCotiza.objects.filter(numtrabajo=pk).update(observa_solicita=observaciones_iniciador,fecha_pre_responde=date.today())
            guarda_estado_responde_pre_cotiza = trabajo.update(estado_responde_precotiza = True)

        except:
            guarda_fecha = PreCotiza.objects.filter(numtrabajo=pk).update(fecha_pre_responde=date.today())
            guarda_estado_preresponde = trabajo.update(estado_responde_precotiza = True)

        #Envía correo a adquisiciones

        queryset = OrdenesTrabajos.objects.get(numtrabajo=pk)

        username_compradora = queryset.usuario_precotiza
        usuario_solicita = queryset.usuario_solicita_id

        id_compradora = User.objects.get(username=username_compradora)
        id_solicita = User.objects.get(id=usuario_solicita)
            
        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=22) # 1. TRABAJOS RESPONDE PRE COTIZACION
        app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = queryset.numtrabajo,
            tipo = noti, 
            usuario_activa = request.user.id,
            autorizador_id = id_compradora.id,
            )

        r.save()

        ### FIN DE SECCION DE NOTIFICACIONES ##

        # emailaprob = id_solicita

        html_message = loader.render_to_string(
            'ordenTrabajo/email/email_responde_precotiza.html',
            {
                'aprob':queryset
            }
        )
        email_subject = 'Respuesta de Pre Cotización Trabajos N°'+' '+str(pk)
        to_list = adquisiciones@ecofroz.com
        mail = EmailMultiAlternatives(
                email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
        mail.attach_alternative(html_message, "text/html")
        try:
            mail.send()
        except:
            logger.error("Unable to send mail.")

        return HttpResponseRedirect(success_url)
            


class IngresoTrabajo(CreateView):
    # permission_required = 'ordenPedido.add_ordenespedidos'
    model = OrdenesTrabajos
    template_name = 'ordenTrabajo/ingreso_trabajo.html'
    form_class = DetTrForm
    second_form_class = TrabajoForm
    success_url = reverse_lazy('trabajos:listar_trabajos')

    def get_context_data(self, **kwargs):
        context = super(IngresoTrabajo, self).get_context_data(**kwargs)
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=300)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') |
         Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')
         | Q(numtrabajo__tipo_pedi='OT')).filter(numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        #ordenes = OrdenesTrabajos.objects.all()
        ubica = activo_ubica.objects.filter(ubica_estado=1)
        area = activo_areas.objects.none()
        grupo = activo_grupo.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'proy' not in context:
            context['proy'] = proy
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        form2 = self.second_form_class(request.POST)
        if request.POST.get('orden_referencial'):
            ref = int(request.POST.get('orden_referencial'))
        else:
            ref = None
        
        if request.POST.get('coninsumos'):
            insu = True
        else:
            insu = False

        now = datetime.now()
        fchtxt = now.strftime('%Y%m%d')
        print(fchtxt)

        proyecto = request.POST.get('numproyecto') 
        
        try:
            consulta = proyectos_contabilidad.objects.get(id=int(proyecto))
        except:
            consulta = None
        
        
        if form.is_valid() and form2.is_valid():
            solicitud = form.save(commit=False)
            cab = form2.save()
            solicitud.numtrabajo = cab
            # solicitud.usuario_solicita = request.user
            cab.orden_referencial = ref
            cab.fchsolicitatxt = fchtxt
            cab.insumos = insu
            if consulta:
                cab.numproyecto = consulta
            else:
                cab.numproyecto = None
                
            solicitud.save()
            cab.save()

            queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__usuario_solicita=request.user.id).last()
            
            ### SECCION PARA GRABAR NOTICACIONES
            
            id_usuario = queryset.numtrabajo.usuario_solicita_id
                

            id_solicita = User.objects.get(id=id_usuario)
            aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
            id_aprobador = User.objects.get(id=aprobador.user_id.id)

            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=1) # 1. PEDIDOS INSUMOS AUTORIZACION GERENTE AREA
            app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = queryset.numtrabajo.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_aprobador,
                )

            r.save()

            ### FIN DE SECCION DE NOTIFICACIONES ##

            emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
            html_message = loader.render_to_string(
                'ordenTrabajo/email/email.html',
                {
                    'mail':queryset
                }
            )
            email_subject = 'Solicitud de orden de pedido'
            to_list = emailaprob.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class generar_ingreso_laboratorio(CreateView):
    # permission_required = 'ordenPedido.add_ordenespedidos'
    model = OrdenesTrabajos
    template_name = 'ordenTrabajo/ingreso_pedidos_laboratorio.html'
    form_class = DetTrForm
    # second_form_class = TrabajoForm
    second_form_class = TrabajoFormLab
    
    success_url = reverse_lazy('trabajos:listar_pedidos_laboratorio')

    def get_context_data(self, **kwargs):
        context = super(generar_ingreso_laboratorio, self).get_context_data(**kwargs)
        ordenes = OrdenesTrabajos.objects.all()
        ubica = activo_ubica.objects.all()
        area = activo_areas.objects.all()
        grupo = activo_grupo.objects.all()
        provedores = proveedor.objects.filter(Q(categoria__id=3) | Q(categoria__id=8) | Q(categoria__id=9))
        tipos = TipoPedido.objects.filter(tipo='L')
    
        lavrestak = proveedor.objects.filter(id=398)

        prove = provedores | lavrestak

        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'form2' not in context:
            context['form2'] = self.second_form_class(self.request.GET)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'proveedor' not in context:
            context['proveedor'] = prove
        if 'tipo' not in context:
            context['tipo'] = tipos
        
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        
        form = self.form_class(request.POST, request.FILES)
        form2 = self.second_form_class(request.POST)
        valor_factura = request.POST.get('valor')
        proveedor_nombre = request.POST.get('proveedor')
        tipo_pedi = request.POST.get('tipos')

        print(valor_factura)
        print(proveedor_nombre)
        
        now = datetime.now()
        fchtxt = now.strftime('%Y%m%d')
        print(fchtxt)
        if form.is_valid() and form2.is_valid():
            solicitud = form.save(commit=False)
            solicitud.numtrabajo = form2.save()
            solicitud.usuario_solicita = request.user
            solicitud.fchsolicitatxt = fchtxt
            solicitud.save()
        

            queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__usuario_solicita=request.user.id).last()
            
            query = OrdenesTrabajos.objects.all().filter(usuario_solicita=request.user.id).last()
            

            #Guarda los datos recogidos en el formulario de ingreso y actualiza la tabla de Cotización

            q2 = OrdenesTrabajos.objects.get(numtrabajo=query.numtrabajo)
            q3 = proveedor.objects.get(nombre_empresa__icontains=proveedor_nombre)

            upd = CotizaTrabajo(
                    valor = valor_factura,
                    pdf_cotiza = queryset.cotiza_Ref,
                    cotiza_seleccion = True,
                    empresa_cotiza_id = q3.id,
                    numtrabajo_id = q2.numtrabajo,
            )
            upd.save()

            ### SECCION PARA GRABAR NOTICACIONES 2-JUN-2022 - DM.
            
            id_usuario = queryset.numtrabajo.usuario_solicita_id

            id_solicita = User.objects.get(id=id_usuario)
            aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
            id_aprobador = User.objects.get(id=aprobador.user_id.id)

            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=1) # 1. PEDIDOS INSUMOS AUTORIZACION GERENTE AREA
            app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = queryset.numtrabajo.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_aprobador,
                )

            r.save()

            ### FIN DE SECCION DE NOTIFICACIONES ##

            emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
            html_message = loader.render_to_string(
                'ordenTrabajo/email/email.html',
                {
                    'mail':queryset
                }
            )
            email_subject = 'Solicitud de aprobación facturas Laboratorio de Calidad'
            to_list = emailaprob.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            print(form.errors)
            print(form2.errors)
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class DuplicaTrabajo(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.add_ordenespedidos'
    model = OrdenesTrabajos
    second_model = DetalleTrabajo
    template_name = 'ordenTrabajo/duplica_trabajo.html'
    form_class = TrabajoForm
    second_form_class = DetTrForm
    success_url = reverse_lazy('trabajos:listar_trabajos')

    def get_context_data(self, **kwargs):
        context = super(DuplicaTrabajo, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numtrabajo=pk)
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=60)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') | Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        #ordenes = OrdenesTrabajos.objects.all()
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        area = orden.area.area_codigo
        grup = orden.grupo
        ubica = activo_ubica.objects.filter(ubica_estado=1)
        areas = activo_areas.objects.all()
        grupo = activo_grupo.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'refer' not in context:
            context['refer'] = orden.orden_referencial
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'areas' not in context:
            context['areas'] = areas
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'grup' not in context:
            context['grup'] = grup
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)
        form2 = self.second_form_class(request.POST, request.FILES)
        if request.POST.get('orden_referencial'):
            ref = int(request.POST.get('orden_referencial'))
        else:
            ref = None

        try:
            proy = proyectos_contabilidad.objects.get(id=request.POST.get('numproyecto'))
            # proyn = proy.nombre_proyecto
        except Exception as e:
            proy = None

        if form.is_valid() and form2.is_valid():
            solicitud = form2.save(commit=False)
            cab = form.save()
            solicitud.numtrabajo = cab
            # solicitud.usuario_solicita = request.user
            cab.orden_referencial = ref
            cab.numproyecto = proy
            solicitud.save()
            cab.save()

            queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__usuario_solicita=request.user.id).last()
            
            ### SECCION PARA GRABAR NOTICACIONES 1-JUN-2022 - DM.
            
            id_usuario = queryset.numtrabajo.usuario_solicita_id
                

            id_solicita = User.objects.get(id=id_usuario)
            aprobador =  User.objects.get(id=id_usuario).autorizador.get(status=True)
            id_aprobador = User.objects.get(id=aprobador.user_id.id)

            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=1) # 1. PEDIDOS INSUMOS AUTORIZACION GERENTE AREA
            app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = queryset.numtrabajo.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_aprobador,
                )

            r.save()

            ### FIN DE SECCION DE NOTIFICACIONES ##


            emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
            html_message = loader.render_to_string(
                'ordenTrabajo/email/email.html',
                {
                    'mail':queryset
                }
            )
            email_subject = 'Solicitud de orden de trabajo'
            to_list = emailaprob.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form, form2=form2))

class EditaTrabajo(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.change_ordenespedidos'
    model = OrdenesTrabajos
    second_model = DetalleTrabajo
    template_name = 'ordenTrabajo/edita_trabajo.html'
    form_class = TrabajoForm
    second_form_class = EditaDetTrForm
    success_url = reverse_lazy('trabajos:listar_trabajos')

    def get_context_data(self, **kwargs):
        context = super(EditaTrabajo, self).get_context_data(**kwargs)
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=60)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numtrabajo=pk)
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') | Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        #ordenes = OrdenesTrabajos.objects.all()
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        area = orden.area.area_codigo
        grup = orden.grupo
        ubica = activo_ubica.objects.filter(ubica_estado=1)
        areas = activo_areas.objects.all()
        grupo = activo_grupo.objects.all()
        tipo = TipoPedido.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'refer' not in context:
            context['refer'] = orden.orden_referencial
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'area' not in context:
            context['area'] = area
        if 'areas' not in context:
            context['areas'] = areas
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'grup' not in context:
            context['grup'] = grup
        if 'tipo' not in context:
            context['grup'] = tipo
        if 'tip' not in context:
            context['tip'] = orden.tipo_pedi
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numtrabajo=id_orden)
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        form = self.form_class(request.POST, instance=orden)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        estado = OrdenesTrabajos.objects.get(numtrabajo=id_orden)

        try:
            proy = proyectos_contabilidad.objects.get(id=request.POST.get('numproyecto'))
            # proyn = proy.nombre_proyecto
        except Exception as e:
            proy = None

        orden.numproyecto = proy

        if form.is_valid() and form2.is_valid():
            form.save()
            form2.save()
            orden.save()
    
            if orden.aprobado == 0:
                queryset = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=id_orden)
                emailaprob = User.objects.get(id=request.user.id).autorizador.get(status=True)
                html_message = loader.render_to_string(
                    'ordenTrabajo/email/email.html',
                    {
                        'mail':queryset
                    }
                )
                email_subject = 'Solicitud de orden de trabajo'
                to_list = emailaprob.e_mail
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())

class EliminaTrabajo(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    permission_required = 'ordenPedido.delete_ordenespedidos'
    model = OrdenesTrabajos
    success_url = reverse_lazy('trabajos:listar_trabajos')

def recibeServicio(request, numtrabajo):
    queryset = OrdenesTrabajos.objects.get(numtrabajo=numtrabajo)
    if request.method == 'GET':
        queryset.entrega_solicita = True
        queryset.fecha_recepcion = datetime.now()
        queryset.usuario_recibe = request.user.id
        queryset.save()

    return redirect('trabajos:listar_trabajos')

# -----------SUPERVISIÓN DE PEDIDOS 1---------------  

def supervisaPedidos1(request):
    usuarios = User.objects.filter(supervisor1=request.user)
    id_usu = []
    for i in usuarios:
        d = i.id
        id_usu.append(d)
    queryset = DetalleTrabajo.objects.filter(numtrabajo__usuario_solicita__in=id_usu).order_by('-numtrabajo__fchsolicita', '-numtrabajo__fchaprueba')

    return render(request, 'ordenTrabajo/para_supervisar1.html', {'form':queryset})

# -----------SUPERVISIÓN DE PEDIDOS MANTENIMIENTO---------------  

def supervisaPedidos(request):
    usuarios = User.objects.filter(mantenimiento=request.user)
    id_usu = []
    for i in usuarios:
        d = i.id
        id_usu.append(d)
    queryset = DetalleTrabajo.objects.filter(numtrabajo__usuario_solicita__in=id_usu).order_by('-numtrabajo__fchsolicita', '-numtrabajo__fchaprueba')

    return render(request, 'ordenTrabajo/para_supervisar.html', {'form':queryset})

# -----------SUPERVISIÓN DE PEDIDOS 2---------------  

def supervisaPedidos2(request):
    usuarios = User.objects.filter(supervisor2=request.user)
    id_usu = []
    for i in usuarios:
        d = i.id
        id_usu.append(d)
    queryset = DetalleTrabajo.objects.filter(numtrabajo__usuario_solicita__in=id_usu).order_by('-numtrabajo__fchsolicita', '-numtrabajo__fchaprueba')

    return render(request, 'ordenTrabajo/para_supervisar2.html', {'form':queryset})

# -----------APROBACIÓN DE TRABAJOS GERENTE AREA---------------

@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def trabajosAprueba(request):
    queryset = request.GET.get("buscar")
    usu_depend = Autorizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(autorizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)
    print(id_usu)
    ordenx = DetalleTrabajo.objects.filter(Q(numtrabajo__usuario_solicita__in=id_usu) & ~Q(numtrabajo__aprobado=0)).order_by('-numtrabajo__numtrabajo')
    
    if queryset:
        ordenx = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) 
        ).filter(Q(numtrabajo__usuario_solicita__in=id_usu) & ~Q(numtrabajo__aprobado=0))
    
    paginator = Paginator(ordenx, 100)
    page = request.GET.get('page')
    orden = paginator.get_page(page)

    return render(request, 'ordenTrabajo/para_aprobar.html', {'form':orden, 'busqueda':queryset})

class EditaTrabajoAutoriza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.change_ordenespedidos'
    model = OrdenesTrabajos
    second_model = DetalleTrabajo
    template_name = 'ordenTrabajo/edita_trabajo_autori.html'
    form_class = TrabajoAutoriForm
    second_form_class = EditaDetTrForm
    success_url = reverse_lazy('trabajos:aprueba_trabajos')

    def get_context_data(self, **kwargs):
        context = super(EditaTrabajoAutoriza, self).get_context_data(**kwargs)
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=60)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numtrabajo=pk)
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') | Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        #ordenes = OrdenesTrabajos.objects.all()
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        grup = orden.grupo
        tipo = TipoPedido.objects.filter(estado=True)
        ubica = activo_ubica.objects.filter(ubica_estado=1)
        grupo = activo_grupo.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'refer' not in context:
            context['refer'] = orden.orden_referencial
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'grup' not in context:
            context['grup'] = grup
        if 'tipo' not in context:
            context['tipo'] = tipo
        if 'tip' not in context:
            context['tip'] = orden.tipo_pedi
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        fch = datetime.now()
        orden = self.model.objects.get(numtrabajo=id_orden)
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        form = self.form_class(request.POST, instance=orden)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        orden.aprobado = 1
        print(orden.tipo_pedi.cod)
        if orden.tipo_pedi.cod == 'AM' or orden.tipo_pedi.cod == 'AP' or orden.tipo_pedi.cod == 'CA' or orden.tipo_pedi.cod == 'CP':
            orden.estado_cotiza = True
            orden.select_cotiza = True
            orden.genera_compra = 1
        orden.usuario_aprueba = request.user.id
        orden.fchaprueba = fch

        try:
            proy = proyectos_contabilidad.objects.get(id=request.POST.get('numproyecto'))
            # proyn = proy.nombre_proyecto
        except Exception as e:
            proy = None

        orden.numproyecto = proy

        if form.is_valid() and form2.is_valid():
            form.save()
            form2.save()
            orden.save()
            query = OrdenesTrabajos.objects.get(numtrabajo=id_orden)
            query.save()

            email_usuario = User.objects.get(id=orden.usuario_solicita.id)

            queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden)

            for i in queryset:
                tipo = i.numtrabajo.tipo_pedi.cod
                id_usuario = i.numtrabajo.usuario_solicita_id 
                

            if tipo == 'AM' or tipo == 'AP' or tipo == 'CA' or tipo == 'CP':
                envioMail('Aprobación de Factura de Laboratorio de Calidad', 'contabilidad_sis@ecofroz.com,dmencias@ecofroz.com', 'email_aprobado_factura_calidad.html', queryset, '')
            else:
                         
                id_solicita = User.objects.get(id=id_usuario)
                # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
                
                id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
                id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
                id_compra3 = User.objects.get(id=22) # EVELY ROBLES
                id_compra4 = User.objects.get(id=118) # SOL RAMOS


                ##Guarda cambio de estado en tabla de mensajeria
                noti = tipo_notificacion.objects.get(id=2) # 2. PEDIDOS INSUMOS COTIZACION

                app_number = modelo_App.objects.get(id=1) # 1. INSUMOS / TRABAJOS

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = query.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = id_compra1,
                    )
                r.save()

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = query.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = id_compra2,
                    )
                r.save()

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = query.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = id_compra3,
                    )
                r.save()

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = query.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = id_compra4,
                    )
                r.save()

                html_message = loader.render_to_string(
                    'ordenTrabajo/email/email_aprobado.html',
                    {
                        'aprob':queryset
                    }
                )

                email_subject = 'Aprobación de orden de trabajo'
                to_list = 'adquisiciones@ecofroz.com'
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")

            envioMail('Aprobación de Orden de Pedido Gerente de Area', email_usuario.email, 'email_aprobado_gerente.html', queryset, '')

            if request.user.id == 67 or request.user.id == 70:
                print('hola:' + str(request.user.id))
                envioMail('Aprobación de Orden de Pedido', 'asvegetal1@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')
                # envioMail('Aprobación de Orden de Pedido', 'dmencias@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')

            return HttpResponseRedirect(self.get_success_url())
        else:
            print(form.errors)
            return HttpResponseRedirect(self.get_success_url())



class RevisaTrabajoAutoriza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.change_ordenespedidos'
    model = OrdenesTrabajos
    second_model = DetalleTrabajo
    template_name = 'ordenTrabajo/revisa_trabajo_autori.html'
    form_class = TrabajoAutoriForm
    second_form_class = EditaDetTrForm
    success_url = reverse_lazy('trabajos:aprueba_trabajos')

    def get_context_data(self, **kwargs):
        context = super(RevisaTrabajoAutoriza, self).get_context_data(**kwargs)
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=60)
        pk = self.kwargs.get('pk',0)
        orden = self.model.objects.get(numtrabajo=pk)
        detalle = self.second_model.objects.get(numtrabajo=orden.numtrabajo)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(Q(numtrabajo__tipo_pedi='MC') | Q(numtrabajo__tipo_pedi='MP') | Q(numtrabajo__tipo_pedi='PR') | Q(numtrabajo__tipo_pedi='OS')).filter(numtrabajo__fchsolicita__range=[enddate,startdate]).order_by('-numtrabajo')
        #ordenes = OrdenesTrabajos.objects.all()
        try:
            proy_select = proyectos_contabilidad.objects.get(nombre_proyecto=orden.numproyecto)
            proy_select = proy_select.id
        except Exception as e:
            proy_select = None

        ubicacion = orden.ubica
        grup = orden.grupo
        tipo = TipoPedido.objects.filter(estado=True)
        ubica = activo_ubica.objects.filter(ubica_estado=1)
        grupo = activo_grupo.objects.all()
        proy = proyectos_contabilidad.objects.filter(activo=True)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'orden' not in context:
            context['orden'] = ordenes
        if 'refer' not in context:
            context['refer'] = orden.orden_referencial
        if 'ubica' not in context:
            context['ubica'] = ubica
        if 'grupo' not in context:
            context['grupo'] = grupo
        if 'ubicacion' not in context:
            context['ubicacion'] = ubicacion
        if 'grup' not in context:
            context['grup'] = grup
        if 'tipo' not in context:
            context['tipo'] = tipo
        if 'tip' not in context:
            context['tip'] = orden.tipo_pedi
        if 'proy' not in context:
            context['proy'] = proy
        if 'proy_select' not in context:
            context['proy_select'] = proy_select
        return context






@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def apruebaDirect(request, numpedido):
    aprueba = OrdenesTrabajos.objects.get(numtrabajo=numpedido)
    fch = datetime.now()
    if request.method == 'GET':
        aprueba.aprobado = 1
        if aprueba.tipo_pedi.cod == 'AM' or aprueba.tipo_pedi.cod == 'AP' or aprueba.tipo_pedi.cod == 'CA' or aprueba.tipo_pedi.cod == 'CP':
            aprueba.estado_cotiza = True
            aprueba.select_cotiza = True
            aprueba.genera_compra = 1
        aprueba.usuario_aprueba = request.user.id
        aprueba.fchaprueba = fch
        aprueba.save()
        
        email_usuario = User.objects.get(id=aprueba.usuario_solicita.id)

        queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        for i in queryset:
            tipo = i.numtrabajo.tipo_pedi.cod
            id_usuario = i.numtrabajo.usuario_solicita_id 

        if tipo == 'AM' or tipo == 'AP' or tipo == 'CA' or tipo == 'CP':
            envioMail('Aprobación de Factura de Laboratorio de Calidad', 'contabilidad_sis@ecofroz.com,dmencias@ecofroz.com', 'email_aprobado_factura_calidad.html', queryset, '')
            
        else:
            ##Inicio de seccion de notificaciones Globales
            id_solicita = User.objects.get(id=id_usuario)
            # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
            
            id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
            id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
            id_compra3 = User.objects.get(id=22) # EVELY ROBLES
            id_compra4 = User.objects.get(id=118) # SOL RAMOS


            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=2) # 2. PEDIDOS INSUMOS COTIZACION

            app_number = modelo_App.objects.get(id=1) # 1. INSUMOS / TRABAJOS

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra1,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra2,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra3,
                )
            r.save()

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = aprueba.numtrabajo,
                tipo = noti, 
                usuario_activa = id_solicita,
                autorizador_id = id_compra4,
                )
            r.save()

            ##Fin Seccion Notificaciones Globales

            html_message = loader.render_to_string(
                'ordenTrabajo/email/email_aprobado.html',
                {
                    'aprob':queryset
                }
            )

            email_subject = 'Aprobación de orden de trabajo'
            to_list = 'adquisiciones@ecofroz.com'
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

        envioMail('Aprobación de Orden de Pedido Gerente de Area', email_usuario.email, 'email_aprobado_gerente.html', queryset, '')
        
        if request.user.id == 67 or request.user.id == 70:
            print('hola:' + str(request.user.id))
            envioMail('Aprobaciòn de Orden de Pedido', 'asvegetal1@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')
            envioMail('Aprobaciòn de Orden de Pedido', 'dmencias@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')

        # if request.user.id == 10:
        #     print('hola:' + str(request.user.id))
        #     # envioMail('Aprobaciòn de Orden de Pedido', 'asvegetal1@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')
        #     envioMail('Aprobaciòn de Orden de Pedido', 'dmencias@ecofroz.com', 'email_aprobado_gerente.html', queryset, '')
                
    return redirect('trabajos:aprueba_trabajos')

@login_required
@permission_required('ordenPedido.aprueba_ordenespedidos', raise_exception=True)
def rechazaDirect(request, numpedido):
    pedido = request.GET.get("pedido")
    rechaza = OrdenesTrabajos.objects.get(numtrabajo=pedido)
    if request.method == 'GET':
        rechaza.aprobado = 0
        rechaza.motRechaza = request.GET.get("text")
        rechaza.save()

        email_usuario = User.objects.get(id=rechaza.usuario_solicita.id)
        queryset = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=pedido)

        envioMail('Rechazo de Orden de PeTrabajodido', email_usuario.email, 'email_rechaza_pedido.html', queryset, '')

    return redirect('trabajos:aprueba_trabajos')

###########------------COTIZACIONES DE PEDIDOS-----------###########

@login_required
@permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def cotizaTrabajos(request):
    queryset = request.GET.get("buscar")
    print(queryset)
    if queryset:
        aprobadas = OrdenesTrabajos.objects.filter(aprobado=1)
        cotizadas = OrdenesTrabajos.objects.filter(estado_cotiza=True).filter(aprobado=1)

        q1 = DetalleTrabajo.objects.filter(numtrabajo__aprobado=1).annotate(
            usuario_cotiza=Subquery(aprobadas.filter(numtrabajo=OuterRef('numtrabajo')).values('usuario_cotiza')[:1]
            ))
        
        val2 = q1.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(numtrabajo__departamento__dep_nombre__icontains = queryset) |
            Q(numtrabajo__usuario_solicita__username__icontains = queryset)).filter(
                numtrabajo__estado_cotiza__isnull=True).filter(numtrabajo__aprobado=1).exclude(numtrabajo__tipo_pedi__cod__in = ['AM','AP','CA','CP']).order_by('-numtrabajo__fchaprueba')
    
        paginator = Paginator(val2, 100)
        page = request.GET.get('page')
        orden2 = paginator.get_page(page)

        q2 = DetalleTrabajo.objects.filter(numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).annotate(
            usuario_cotiza=Subquery(aprobadas.filter(numtrabajo=OuterRef('numtrabajo')).values('usuario_cotiza')[:1]
            ))

        val2 = q2.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(numtrabajo__usuario_cotiza__icontains = queryset) |
            Q(numtrabajo__departamento__dep_nombre__icontains = queryset) |
            Q(numtrabajo__usuario_solicita__username__icontains = queryset)).filter(
                numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__aprobado=1).exclude(
                        numtrabajo__tipo_pedi__cod__in = ['AM','AP','CA','CP']).order_by('-numtrabajo__fchcotiza')
    
        paginator = Paginator(val2, 100)
        page = request.GET.get('page')
        aprobados2 = paginator.get_page(page)
 
    else:
        orden2 = DetalleTrabajo.objects.select_related('numtrabajo').filter(numtrabajo__estado_cotiza__isnull=True).filter(numtrabajo__aprobado=1).order_by('-numtrabajo__fchaprueba')[:100]

        aprobados2 = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).exclude(
                        numtrabajo__tipo_pedi__cod__in = ['AM','AP','CA','CP']).order_by('-numtrabajo__fchcotiza')[:100]
        
   
        
    return render(request, 'ordenTrabajo/para_cotizar.html', {'form':orden2, 'form2': aprobados2, 'busqueda':queryset})

class IngresoCotizaciones(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    permission_required = 'ordenPedido.add_cotizapedido'
    model = OrdenesTrabajos
    template_name = 'ordenTrabajo/cotiza_trabajos.html'
    form_class = TrabajoForm
    success_url = reverse_lazy('trabajos:cotiza_trabajos')

    def get(self, request, *args, **kwargs):
        self.object = None
        pk = self.kwargs.get('pk',0)
        cotiza = DetalleTrabajo.objects.filter(numtrabajo__numtrabajo=pk)
        # cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=pk)
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=pk)
        # estado = cotiza
        # print(cotiza.numtrabajo.numtrabajo)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        # detalle = DetCotizaFormSet()
        return self.render_to_response(self.get_context_data(detalle = detalle_cotiza, cot = cotiza))

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numtrabajo=id_orden)
        aprob = None
        prove = None
        seleccion = None
        pago_mba = None
        fch = datetime.now()
        if request.POST.get('pago_mba'):
            pago_mba = True
        print(pago_mba)
        if request.POST.get('compra_directa'):
            cd = True
            aprob = 1
            seleccion = True
            prove = request.POST.get('cotizatrabajo_set-0-empresa_cotiza')
            print('Aqui el proveedor' + prove)
            orden.proveedor = proveedor.objects.get(id=prove)
            fchgenera = datetime.now()
        else:
            cd = False
            fchgenera = None
        devuelto = orden.genera_compra
        orden.estado_cotiza = True
        orden.cotiza_observa = request.POST.get('observa_cotiza')
        orden.compra_directa = cd
        orden.genera_compra = aprob
        orden.select_cotiza = seleccion
        orden.usuario_cotiza = request.user.username
        orden.pago_mba = pago_mba
        orden.fchcotiza = fch
        orden.fchgenerac = fchgenera
        form2 = DetCotizaFormSet(request.POST, request.FILES, instance=orden)
        
        # print("Hola error",form2.errors)
       
        if form2.is_valid():
            if devuelto != 2:
                for form in form2.forms:
                    if form.cleaned_data.get('valor'):
                        form.save()
            if cd == True:
                cotiza = CotizaTrabajo.objects.get(numtrabajo_id=id_orden)
                cotiza.cotiza_seleccion = True
                cotiza.save()
            print(request.FILES)
            try:
                otros_docs_adqui = request.FILES['otros_docs_adqui']
                storage = FileSystemStorage()
                file_path = storage.save('otros_doc/'+otros_docs_adqui.name, otros_docs_adqui)
                modelo = DetalleTrabajo.objects.get(numtrabajo=id_orden)
                modelo.otros_doc_adqui = file_path
                modelo.save()
            except Exception as e:
                print(e)

            orden.save()

            if cd != True and devuelto != 2:
                queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden)
                for i in queryset:
                    usuario_id = i.numtrabajo.usuario_solicita_id
                    analista = i.numtrabajo.usuario_cotiza
                
                queryset2 = OrdenesTrabajos.objects.get(numtrabajo=id_orden)
                # email_selectcotiza = User.objects.get(id=usuario_id)
                email_selectcotiza = User.objects.filter(id=usuario_id)
                for i in email_selectcotiza:
                    for a in i.autorizador.filter(status=True):
                        email_autoriza = a.e_mail

                email_selectcotiza = User.objects.get(id=usuario_id)
                
                ### SECCION PARA GRABAR NOTICACIONES 1-JUN-2022 - DM.
                id_solicita = User.objects.get(id=usuario_id)
                analista_compras = User.objects.get(username=analista)
                id_analista_compras = User.objects.get(id=analista_compras.id)
                ger_selecciona =  Cotizador.objects.get(id=id_solicita.cotizador_id)
                id_select_cotiza = User.objects.get(id=ger_selecciona.user_id.id)

                ##Guarda cambio de estado en tabla de mensajeria
                noti = tipo_notificacion.objects.get(id=3) # 3. PEDIDOS INSUMOS SELECCION COTIZACION
                app_number = modelo_App.objects.get(id=1) # 1. Insumos/Trabajos

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = queryset2.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_analista_compras,
                    autorizador_id = id_select_cotiza,
                    )

                r.save()

                ### FIN DE SECCION DE NOTIFICACIONES ##



                html_message = loader.render_to_string(
                    'ordenTrabajo/email/email_cotizado.html',
                    {
                        'aprob':queryset
                    }
                )

                email_subject = 'Cotización de orden de trabajo'
                to_list = email_selectcotiza.cotizador.e_mail
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    pass
                    #mail.send()
                except:
                    logger.error("Unable to send mail.")

                return HttpResponseRedirect(self.get_success_url())
            elif cd != True and devuelto == 2:
                queryset = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden).filter(cotiza_seleccion=True)
                queryset2 = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden)
                for i in queryset2:
                    id_usuario = i.numtrabajo.usuario_solicita_id

                id_genera = User.objects.get(id=id_usuario)
                emailgenera =  Generador.objects.get(id=id_genera.generador_id)

                html_message = loader.render_to_string(
                    'ordenTrabajo/email/email_selec_cotiza.html',
                    {
                        'aprob':queryset,
                        'aprob2':queryset2,
                    }
                )

                email_subject = 'Rectificación de Cotizaciones Rechazadas'
                to_list = emailgenera.e_mail
                mail = EmailMultiAlternatives(
                        email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
                mail.attach_alternative(html_message, "text/html")
                try:
                    mail.send()
                except:
                    logger.error("Unable to send mail.")
                
                return HttpResponseRedirect(self.get_success_url())
            else:
                return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())

@login_required
def observaReferencial(request, numtrabajo):
    if request.method == 'GET':
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numtrabajo)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numtrabajo)
        idpedido = numtrabajo
        aprobado = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numtrabajo)
        aprob = aprobado.numtrabajo.genera_compra
        print(aprob)
        return render(request, 'ordenTrabajo/referencial_cotiza.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})


@login_required
def lookCotiza(request, numtrabajo):
    if request.method == 'GET':
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numtrabajo)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numtrabajo)
        idpedido = numtrabajo
        aprobado = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numtrabajo)
        aprob = aprobado.numtrabajo.genera_compra
        print(aprob)
        return render(request, 'ordenTrabajo/look_cotiza.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})






@login_required
def regresaCDCotizacion(request,numtrabajo):
    if request.method == 'GET':

        #Actualiza campos para regresar a sección de cotización
        trabajo = OrdenesTrabajos.objects.get(numtrabajo=numtrabajo)

        trabajo.estado_cotiza = None
        trabajo.cotiza_observa = None
        trabajo.select_cotiza = None
        trabajo.genera_compra = None
        trabajo.compra_directa = None
        trabajo.fchgenerac = None
        trabajo.fchcotiza = None
        trabajo.usuario_cotiza = None
        trabajo.save()

        #Borra la/las cotizaciones
        cotizacion = CotizaTrabajo.objects.filter(numtrabajo=numtrabajo).delete()

        return redirect('trabajos:ver_compras')


@login_required
@permission_required('ordenPedido.view_cotizapedido', raise_exception=True)
def compraTrabajos(request):
    queryset = request.GET.get("buscar")
    
    if queryset:
        aprobados = OrdenesTrabajos.objects.filter(estado_cotiza=True).filter(aprobado=1)

        q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__select_cotiza=True).filter(numtrabajo__aprobado=1).annotate(
            usuario_cotiza=Subquery(aprobados.filter(numtrabajo=OuterRef('numtrabajo')).values('usuario_cotiza')[:1]
            ))
        
        val2 = q1.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(numtrabajo__proveedor__nombre_empresa__icontains = queryset) |
            Q(numtrabajo__usuario_cotiza__icontains = queryset) |
            Q(numtrabajo__departamento__dep_nombre__icontains = queryset) |
            Q(numtrabajo__usuario_solicita__username__icontains = queryset)).filter(
                numtrabajo__estado_cotiza=True).filter(numtrabajo__select_cotiza=True).filter(
                    numtrabajo__aprobado=1).exclude(numtrabajo__tipo_pedi__cod__in = ['AM','AP','CA','CP']).order_by('-numtrabajo__fchcotiza')
    
        paginator = Paginator(val2, 50)
        page = request.GET.get('page')
        orden = paginator.get_page(page)
    
    else:
        orden = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__estado_cotiza=True).filter(
            numtrabajo__select_cotiza=True).filter(
                numtrabajo__aprobado=1).exclude(numtrabajo__tipo_pedi__cod__in = ['AM','AP','CA','CP']).order_by('-numtrabajo__fchcotiza')[:25]
    
    
    return render(request, 'ordenTrabajo/para_comprar.html', {'form2': orden, 'busqueda':queryset})


@login_required
@permission_required('ordenPedido.selecciona_cotizapedido', raise_exception=True)
def trabajosCotizados(request):
    queryset = request.GET.get("buscar")
    usu_depend = Cotizador.objects.get(user_id=request.user.id)
    usu_depend2 = User.objects.filter(cotizador=usu_depend.id)
    id_usu = []
    for i in usu_depend2:
        d = i.id
        id_usu.append(d)
    cotizados = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo__select_cotiza=True).filter(numtrabajo__usuario_solicita__in=id_usu).filter(numtrabajo__compra_directa=False).order_by('-numtrabajo__ord')
    if queryset:
        orden = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion = queryset) 
        ).filter(numtrabajo__select_cotiza__isnull=True).filter(numtrabajo__usuario_solicita__in=id_usu).order_by('-numtrabajo__ord')
    else:
        orden = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__select_cotiza__isnull=True).filter(numtrabajo__compra_directa=False).filter(numtrabajo__usuario_solicita__in=id_usu).order_by('-numtrabajo__ord')

    return render(request, 'ordenTrabajo/selec_cotizacion.html', {'form':orden, 'form2':cotizados, 'busqueda':queryset})

class ApruebaCotiza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.selecciona_cotizapedido'
    model = OrdenesTrabajos
    template_name = 'ordenTrabajo/selec_cotiza.html'
    form_class = TrabajoForm
    success_url = reverse_lazy('trabajos:selec_cotiza')

    def get_context_data(self, **kwargs):
        context = super(ApruebaCotiza, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        detalle = CotizaTrabajo.objects.filter(numtrabajo=pk)
        detalle_data = []
        for det in detalle:
            d={
               'id':det.id,
               'valor':det.valor,
               'empresa_cotiza':det.empresa_cotiza,
               'pdf_cotiza':det.pdf_cotiza,
               'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=pk)
        if 'detalle' not in context:
            context['detalle'] = detalle_cotiza
        if 'cot' not in context:
            context['cot'] = cotiza
        if 'pk' not in context:
            context['pk'] = pk
         
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        id_orden = kwargs['pk']
        fch = datetime.now()
        orden = self.model.objects.get(numtrabajo=id_orden)
        cotiza = CotizaTrabajo.objects.filter(numtrabajo=orden.numtrabajo)
        form2 = DetCotizaFormSet(request.POST, queryset=cotiza)
        orden.observa_selec_cot = request.POST.get('observa_aut_cot')
        orden.proveedor = proveedor.objects.get(id=request.POST.get('prove'))
        orden.select_cotiza = True
        orden.fchselectcotiza = fch
        if form2.is_valid():
            for form in form2.forms:
                clave = form.cleaned_data.get('id')
                if clave:
                    if form.cleaned_data.get('cotiza_seleccion'):
                        queryset = CotizaTrabajo.objects.get(id=clave.id)
                        queryset.cotiza_seleccion = True
                        queryset.save()
                    else:
                        queryset = CotizaTrabajo.objects.get(id=clave.id)
                        queryset.cotiza_seleccion = False
                        queryset.save()
                
            orden.save()

            queryset = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden).filter(cotiza_seleccion=True)
            queryset2 = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden)
            for i in queryset2:
                id_usuario = i.numtrabajo.usuario_solicita_id
                

            id_genera = User.objects.get(id=id_usuario)
            emailgenera =  Generador.objects.get(id=id_genera.generador_id)
            id_aprueba = User.objects.get(id=emailgenera.user_id.id)

            ##Guarda cambio de estado en tabla de mensajeria
            noti = tipo_notificacion.objects.get(id=4) # 4. PEDIDOS INSUMOS APROBACION GER. ADMIN

            app_number = modelo_App.objects.get(id=1) # 1. INSUMOS / TRABAJOS

            r = notificaciones_globales(
                app_origen = app_number,
                estado = True,
                identificador = orden.numtrabajo,
                tipo = noti, 
                usuario_activa = id_genera,
                autorizador_id = id_aprueba,
                )

            r.save()


            html_message = loader.render_to_string(
                'ordenTrabajo/email/email_selec_cotiza.html',
                {
                    'aprob':queryset,
                    'aprob2':queryset2,
                }
            )

            email_subject = 'Selección de cotización de orden de trabajo'
            to_list = emailgenera.e_mail
            mail = EmailMultiAlternatives(
                    email_subject, '', '', [to_list], bcc=['desarrolloecofroz@gmail.com'])
            mail.attach_alternative(html_message, "text/html")
            try:
                mail.send()
            except:
                logger.error("Unable to send mail.")

            #envioMail('Selecciòn de Cotizaciòn Gerente de Area', id_genera.email, 'email_selec_cotiza_gerente.html', queryset, queryset2)

            return HttpResponseRedirect(self.get_success_url())
        else:
            return HttpResponseRedirect(self.get_success_url())


class DevuelveCotiza(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    permission_required = 'ordenPedido.selecciona_cotizapedido'
    model = OrdenesTrabajos
    template_name = 'ordenTrabajo/selec_cotiza.html'
    form_class = TrabajoForm
    success_url = reverse_lazy('trabajos:selec_cotiza')

    def get_context_data(self, **kwargs):
        print("Hola mundo",pk)

        context = super(DevuelveCotiza, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        detalle = CotizaTrabajo.objects.filter(numtrabajo=pk)
        detalle_data = []
        for det in detalle:
            d={
               'id':det.id,
               'valor':det.valor,
               'empresa_cotiza':det.empresa_cotiza,
               'pdf_cotiza':det.pdf_cotiza,
               'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=pk)
        if 'detalle' not in context:
            context['detalle'] = detalle_cotiza
        if 'cot' not in context:
            context['cot'] = cotiza
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        id_orden = kwargs['pk']
        orden = self.model.objects.get(numtrabajo=id_orden)
        cotiza = CotizaTrabajo.objects.filter(numtrabajo=orden.numtrabajo).delete()
        #form2 = DetCotizaFormSet(request.POST, queryset=cotiza)
        orden.observa_selec_cot = request.POST.get('observa_aut_cot')
        #orden.proveedor = proveedor.objects.get(id=request.POST.get('prove'))
        orden.select_cotiza = None
        orden.estado_cotiza = None
        orden.compra_directa = None
        orden.usuario_cotiza = None
                
        orden.save()

        #id_genera = User.objects.get(id=id_usuario)
        queryset = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=id_orden)
       
        envioMail('Devolución de Cotización Gerente de Area', 'adquisiciones@ecofroz.com,dmencias@ecofroz.com', 'email_cotizacion_devuelta_gerente.html', queryset, '')

            
        return HttpResponseRedirect(self.get_success_url())
        

################ORDENES DE COMPRA##########################

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def ordenesCompra(request):
    queryset = request.GET.get("buscar")
    form2 = selectAsignadoa()
    if queryset:
        query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(~Q(numtrabajo__compra_directa=True))
        
        q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).filter(~Q(numtrabajo__compra_directa=True)).annotate(
            valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]
            ))
        
        q2 = q1.annotate(cotiza_seleccion=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]
        ))
        
        q3 = q2.annotate(empresa_cotiza=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]
        ))
        
        val2 = q3.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(empresa_cotiza__icontains = queryset) |
            Q(numtrabajo__usuario_cotiza__icontains = queryset)).filter(cotiza_seleccion=True).order_by('-numtrabajo__genera_compra', '-numtrabajo__fchsolicita')
    
        preguntas_experto = ConsultaExpertoTrab.objects.all()

    else:
        query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(~Q(numtrabajo__compra_directa=True))
        
        q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).filter(~Q(numtrabajo__compra_directa=True)).annotate(
            valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]

            ))
        
        q2 = q1.annotate(cotiza_seleccion=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]
        
        ))
        
        q3 = q2.annotate(empresa_cotiza=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]
        
        ))
        
        val2 = q3.filter(cotiza_seleccion=True).order_by('-numtrabajo__genera_compra', '-numtrabajo__fchsolicita')
        preguntas_experto = ConsultaExpertoTrab.objects.all()

    paginator = Paginator(val2, 80)
    page = request.GET.get('page')
    orden2 = paginator.get_page(page)

    return render(request, 'ordenTrabajo/listar_compras.html', {'form':orden2, 'busqueda':queryset, 'form2':form2, 'preguntas':preguntas_experto})

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def ordenesCompraDirecta(request):
    queryset = request.GET.get("buscar")
    if queryset:
        # val2 = DetalleTrabajo.objects.filter(
        #     Q(numtrabajo__numtrabajo__icontains = queryset)
        # ).filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__compra_directa=True).order_by('-numtrabajo__fchsolicita')
    
        query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__compra_directa=True)
        
        q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__compra_directa=True).annotate(
            valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]
            ))
        
        q2 = q1.annotate(cotiza_seleccion=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]
        ))
        
        q3 = q2.annotate(empresa_cotiza=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]
        ))
        
        val2 = q3.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset) |
            Q(empresa_cotiza__icontains = queryset) |
            Q(numtrabajo__usuario_cotiza__icontains = queryset)).filter(cotiza_seleccion=True).order_by('-numtrabajo__fchsolicita')

    
    else:
        query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(numtrabajo__compra_directa=True)
      
       
        q1 = DetalleTrabajo.objects.filter(numtrabajo__compra_directa=True,numtrabajo__estado_cotiza=True,numtrabajo__aprobado=1).annotate(valor=Subquery(
            query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]

        ))


        q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]
        
        ))

        val2 = q2.filter(cotiza_seleccion=True).order_by('-numtrabajo__fchsolicita')

        #val = val.order_by('-numtrabajo__fchsolicita')
        # orden = []

        # for data in val:
    
        #     d={
        #         'pedido':data.numtrabajo.numtrabajo,
        #         'usuario':data.numtrabajo.usuario_solicita.first_name + data.numtrabajo.usuario_solicita.last_name,
        #         'departamento':data.numtrabajo.departamento,
        #         'proyecto':data.numtrabajo.numproyecto,
        #         'fecha':data.numtrabajo.fchsolicita,
        #         'tipo':data.numtrabajo.tipo_pedi,
        #         'orden':data.numtrabajo.orden_referencial,
        #         'descripcion':data.descripcion,
        #         'valor':data.valor,
        #         'compradora':data.numtrabajo.usuario_cotiza,
        #         'estado':data.numtrabajo.genera_compra,
        #         'proveedor':data.numtrabajo.proveedor,
        #     }
        #     orden.append(d)
    
    paginator = Paginator(val2, 25)
    page = request.GET.get('page')
    orden2 = paginator.get_page(page)

    return render(request, 'ordenTrabajo/listar_compras_directas.html', {'form':orden2, 'busqueda':queryset})

def autComprasExcel(request):
    try:          
        cursor = connection.cursor()
        # cursor.execute("""select o.numtrabajo, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
        #     (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto, o.fchsolicita,
        #     d.descripcion, c.valor, o.usuario_cotiza,
        #     	case
		#             when o.genera_compra = 1 then 'APROBADO'
		#             else 'PENDIENTE'
	    #         end as estado,
        #         case
        #             when o.tipo_pedi. = 'PR' then 'PROYECTO'
        #             when o.tipo_pedi = 'MC' then 'MANTENIMIENTO CORRECTIVO'
        #             when o.tipo_pedi = 'MP' then 'MANTENIMIENTO PREVENTIVO'
        #             when o.tipo_pedi = 'OT' then 'OTROS INSUMOS'
        #             when o.tipo_pedi = 'OS' then 'ORDEN DE SERVICIO'
        #             else 'DESCONOCIDO'
        #         end as tipo, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
        #         o.fchaprueba, o.fchgenerac, o.observa_compra
        #     from pedidos.ordenestrabajos o inner join pedidos.detalletrabajo d on (o.numtrabajo = d.numtrabajo_id)
        #     inner join pedidos.cotizatrabajo c on (o.numtrabajo = c.numtrabajo_id)
        #     where c.cotiza_seleccion = true order by o.fchsolicita desc""")
        
        cursor.execute("""select o.numtrabajo, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
            (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto_id, o.fchsolicita,
            d.descripcion, c.valor, o.usuario_cotiza,
            	case
		            when o.genera_compra = 1 then 'APROBADO'
		            else 'PENDIENTE'
	            end as estado,
                t.nombre, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
                o.fchaprueba, o.fchgenerac, o.observa_compra
            from pedidos.ordenestrabajos o inner join pedidos.detalletrabajo d on (o.numtrabajo = d.numtrabajo_id)
            inner join pedidos.cotizatrabajo c on (o.numtrabajo = c.numtrabajo_id)
            inner join pedidos.tipopedido t on (o.tipo_pedi_id = t.cod)
            where o.compra_directa != true and c.cotiza_seleccion = true order by o.fchsolicita desc""")

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE COMPRAS DE INSUMOS Y ORDENES DE SERVICIO'

        ws.merge_cells('A1:L1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'USUARIO SOLICITUD'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'TIPO'
        ws['E3'] = 'FECHA SOLICITUD'
        ws['F3'] = 'FECHA APROBACIÓN'
        ws['G3'] = 'DESCRIPCIÒN'
        ws['H3'] = 'VALOR'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'COMPRADORA'
        ws['K3'] = 'ESTADO GENERACIÒN COMPRA'
        ws['L3'] = 'FECHA APROBACIÓN ADMINISTRATIVA'
        ws['M3'] = 'OBSERVACIÓN ADMINISTRATIVA'
              
        count = 4
        rowcount = 1   

        for i in cursor:
            print(i)
            #for (k, v) in i.items():
            #    print()
            #ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 1).value =  str(i[0])
            ws.cell(row = count, column = 2).value =  str(i[1])
            ws.cell(row = count, column = 3).value =  str(i[2])
            ws.cell(row = count, column = 4).value =  str(i[9])
            ws.cell(row = count, column = 5).value =  str(i[4])
            ws.cell(row = count, column = 6).value =  str(i[11])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[10])
            ws.cell(row = count, column = 10).value =  str(i[7])
            ws.cell(row = count, column = 11).value =  str(i[8])
            if str(i[12]) == 'None':
                ws.cell(row = count, column = 12).value =  '--'
            else:    
                ws.cell(row = count, column = 12).value =  str(i[12])
            if str(i[13]) == 'None':
                ws.cell(row = count, column = 13).value =  '--'
            else:
                ws.cell(row = count, column = 13).value =  str(i[13])
            # ws.cell(row = count, column = 11).value =  str(i.compradora)
            # ws.cell(row = count, column = 12).value =  str(i.genera_compra)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE COMPRAS DE INSUMOS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        #ahora = datetime.now()
        #ayer = ahora - timedelta(days=1)
        #fecha = ahora.strftime('%Y-%m-%d')
        #fechatxt = ahora.strftime('%Y%m%d')
        #fechaacttxt = ayer.strftime('%Y%m%d')
        #queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        #paginator = Paginator(queryset.order_by('-ord'), 50)
        #page = request.GET.get('page')
        #ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'
        print(e)

        return redirect('trabajos:ordenescompra')

def autCompraDirectaExcel(request):
    try:          
        cursor = connection.cursor()
        cursor.execute("""select o.numtrabajo, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
            (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto_id, o.fchsolicita,
            d.descripcion, c.valor, o.usuario_cotiza,
            	case
		            when o.genera_compra = 1 then 'APROBADO'
		            else 'PENDIENTE'
	            end as estado,
                case
                    when o.tipo_pedi_id = 'PR' then 'PROYECTO'
                    when o.tipo_pedi_id = 'MC' then 'MANTENIMIENTO CORRECTIVO'
                    when o.tipo_pedi_id = 'MP' then 'MANTENIMIENTO PREVENTIVO'
                    when o.tipo_pedi_id = 'OT' then 'OTROS INSUMOS'
                    when o.tipo_pedi_id = 'OS' then 'ORDEN DE SERVICIO'
                    else 'DESCONOCIDO'
                end as tipo, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
            o.fchaprueba, o.fchgenerac
            from pedidos.ordenestrabajos o inner join pedidos.detalletrabajo d on (o.numtrabajo = d.numtrabajo_id)
            inner join pedidos.cotizatrabajo c on (o.numtrabajo = c.numtrabajo_id)
            where c.cotiza_seleccion = true and o.compra_directa = true order by o.fchsolicita desc""")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE COMPRAS DIRECTAS DE INSUMOS Y ORDENES DE SERVICIO'

        ws.merge_cells('A1:I1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'USUARIO SOLICITA'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'TIPO'
        ws['E3'] = 'FECHA SOLICITA'
        ws['F3'] = 'FECHA APRUEBA'
        ws['G3'] = 'DESCRIPCIÒN'
        ws['H3'] = 'VALOR'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'COMPRADORA'
        ws['K3'] = 'ESTADO GENERACIÒN COMPRA'
        ws['L3'] = 'FECHA GENERACIÓN COMPRA'
              
        count = 4
        rowcount = 1   

        for i in cursor:
            print(i)
            #for (k, v) in i.items():
            #    print()
            #ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 1).value =  str(i[0])
            ws.cell(row = count, column = 2).value =  str(i[1])
            ws.cell(row = count, column = 3).value =  str(i[2])
            ws.cell(row = count, column = 4).value =  str(i[9])
            ws.cell(row = count, column = 5).value =  str(i[4])
            ws.cell(row = count, column = 6).value =  str(i[11])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[10])
            ws.cell(row = count, column = 10).value =  str(i[7])
            ws.cell(row = count, column = 11).value =  str(i[8])
            if str(i[12]) == 'None':
                ws.cell(row = count, column = 12).value =  '--'
            else:    
                ws.cell(row = count, column = 12).value =  str(i[12])
            # ws.cell(row = count, column = 11).value =  str(i.compradora)
            # ws.cell(row = counta, column = 12).value =  str(i.genera_compra)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE COMPRAS DIRECTAS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        #ahora = datetime.now()
        #ayer = ahora - timedelta(days=1)
        #fecha = ahora.strftime('%Y-%m-%d')
        #fechatxt = ahora.strftime('%Y%m%d')
        #fechaacttxt = ayer.strftime('%Y%m%d')
        #queryset = ContenedorRegistro.objects.filter(fch_ingreso_txt__gte=fechaacttxt, fch_ingreso_txt__lte=fechatxt)
        #paginator = Paginator(queryset.order_by('-ord'), 50)
        #page = request.GET.get('page')
        #ingresos = paginator.get_page(page)

        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'
        print(e)

        return redirect('trabajos:comprasdirectas')

def gaComprasExcel(request):
    try:          
        cursor = connection.cursor()
        
        cursor.execute("""select o.numtrabajo, (select u.first_name || ' ' || u.last_name from public.usuarios_user u where id = o.usuario_solicita_id) as usuario,
            (select a.dep_nombre from activos.activo_depar a where id = departamento_id) as depar, o.numproyecto, o.fchsolicita,
            d.descripcion, c.valor, o.usuario_cotiza,
            	case
		            when o.genera_compra = 1 then 'APROBADO'
		            else 'PENDIENTE'
	            end as estado,
                t.nombre, (select p.nombre_empresa from proveedores.proveedor p where id = c.empresa_cotiza_id) as prove,
                o.fchaprueba, o.fchgenerac, o.observa_compra
            from pedidos.ordenestrabajos o inner join pedidos.detalletrabajo d on (o.numtrabajo = d.numtrabajo_id)
            inner join pedidos.cotizatrabajo c on (o.numtrabajo = c.numtrabajo_id)
            inner join pedidos.tipopedido t on (o.tipo_pedi_id = t.cod)
            where o.usuario_aprueba = """ + str(request.user.id)  + """ and c.cotiza_seleccion = true order by o.fchsolicita desc""")

        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'REPORTE DE COMPRAS DE INSUMOS Y ORDENES DE SERVICIO'

        ws.merge_cells('A1:L1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'USUARIO SOLICITUD'
        ws['C3'] = 'DEPARTAMENTO'
        ws['D3'] = 'TIPO'
        ws['E3'] = 'FECHA SOLICITUD'
        ws['F3'] = 'FECHA APROBACIÓN'
        ws['G3'] = 'DESCRIPCIÒN'
        ws['H3'] = 'VALOR'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'COMPRADORA'
        ws['K3'] = 'ESTADO GENERACIÒN COMPRA'
        ws['L3'] = 'FECHA APROBACIÓN ADMINISTRATIVA'
        ws['M3'] = 'OBSERVACIÓN ADMINISTRATIVA'
              
        count = 4
        rowcount = 1   

        for i in cursor:
            ws.cell(row = count, column = 1).value =  str(i[0])
            ws.cell(row = count, column = 2).value =  str(i[1])
            ws.cell(row = count, column = 3).value =  str(i[2])
            ws.cell(row = count, column = 4).value =  str(i[9])
            ws.cell(row = count, column = 5).value =  str(i[4])
            ws.cell(row = count, column = 6).value =  str(i[11])
            ws.cell(row = count, column = 7).value =  str(i[5])
            ws.cell(row = count, column = 8).value =  str(i[6])
            ws.cell(row = count, column = 9).value =  str(i[10])
            ws.cell(row = count, column = 10).value =  str(i[7])
            ws.cell(row = count, column = 11).value =  str(i[8])
            if str(i[12]) == 'None':
                ws.cell(row = count, column = 12).value =  '--'
            else:    
                ws.cell(row = count, column = 12).value =  str(i[12])
            if str(i[13]) == 'None':
                ws.cell(row = count, column = 13).value =  '--'
            else:
                ws.cell(row = count, column = 13).value =  str(i[13])
            # ws.cell(row = count, column = 11).value =  str(i.compradora)
            # ws.cell(row = count, column = 12).value =  str(i.genera_compra)
        
            count+=1
            rowcount+=1

        nombre_archivo = "REPORTE DE COMPRAS DE INSUMOS" + ".xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
    except Exception as e:
        mensaje = 'Error'
        mensaje_e = 'Error en los valores, por favor verifique las fechas'
        print(e)

        return redirect('trabajos:ordenescompra')

@login_required
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def generaCompra(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        detalle_data = []
        for det in detalle:
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        idpedido = numpedido
        aprobado = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numpedido)
        aprob = aprobado.numtrabajo.genera_compra
        print(aprob)
        return render(request, 'ordenTrabajo/genera_compras.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob})

@login_required 
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def apruebaCompra(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        fch = datetime.now()    
        compra = OrdenesTrabajos.objects.get(numtrabajo=numpedido)
        compra.genera_compra = 1
        compra.observa_compra = observa
        compra.usuario_genera = request.user.id
        compra.fchgenerac = fch
        compra.save()

        queryset = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido).filter(cotiza_seleccion=True)
        queryset2 = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        
        for i in queryset2:
            id_usuario = i.numtrabajo.usuario_solicita_id      

        id_solicita = User.objects.get(id=id_usuario)
        # id_genera =  Generador.objects.get(id=id_genera.generador_id)  #Rol de Generador
        
        id_compra1 = User.objects.get(id=38) # MONICA CALVOPIÑA
        id_compra2 = User.objects.get(id=43) # PAOLA CALVACHE
        id_compra3 = User.objects.get(id=22) # EVELY ROBLES
        id_compra4 = User.objects.get(id=118) # SOL RAMOS


        ##Guarda cambio de estado en tabla de mensajeria
        noti = tipo_notificacion.objects.get(id=5) # 5. INSUMOS/TRAB. AUTORIZA COMPRA GER. ADMIN

        app_number = modelo_App.objects.get(id=1) # 1. INSUMOS / TRABAJOS

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numtrabajo,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra1,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numtrabajo,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra2,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numtrabajo,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra3,
            )
        r.save()

        r = notificaciones_globales(
            app_origen = app_number,
            estado = True,
            identificador = compra.numtrabajo,
            tipo = noti, 
            usuario_activa = id_solicita,
            autorizador_id = id_compra4,
            )
        r.save()


        envioMail('Generación de Orden de Compra', 'adquisiciones@ecofroz.com, bodega@ecofroz.com', 'email_orden_compra.html', queryset, queryset2)
        
        
        return redirect('trabajos:ordenescompra')
    else:
        return redirect('trabajos:ordenescompra')

@login_required
def recuperaPreguntas(request):
    numpe = request.GET['numpe']
    data = []
    
    print(numpe)
    try:
        query = ConsultaExpertoTrab.objects.filter(numtrabajo=numpe)
        data = []
    
        for queryset in query:
            d={
                'id':queryset.numtrabajo.numtrabajo,
                'pregunta':queryset.pregunta,
                'respuesta':queryset.respuesta,
                'estado':queryset.numtrabajo.estado_consulta_experto,

            }   
            data.append(d)
         
    except:
        data=None
        
    return JsonResponse(data, safe=False)

@login_required
def consulta_expertos(request):
    
    pregunta = request.GET.get("pregunta")
    numtrabajo = request.GET.get("numpe")
    
    if request.method == 'GET':
        
        fec = datetime.today()
        form = selectAsignadoa(request.GET or None)

        if form.is_valid():
            data = form.cleaned_data['asignacion']
            
            ids = []
            for i in data:
                id = i.id
                ids.append(id)
            
            print(ids)
                        
             
            fecha=datetime.today()
            pregunta_compuesta = str(datetime.strftime(fecha, '%Y-%m-%d %H:%M')) + " " + pregunta 

            numero = OrdenesTrabajos.objects.get(numtrabajo=numtrabajo)
            
            usuarios = User.objects.filter(id__in=ids)

            r = ConsultaExpertoTrab(
                numtrabajo = numero,
                fecha_pregunta = fecha,
                pregunta = pregunta_compuesta,
                persona_pregunta = User.objects.get(genera_consultas_a_expertos=True),
                )
            r.save()

            # Actualiza cabecera de pedidos con el estado de la consulta
            actualiza_estado = OrdenesTrabajos.objects.filter(numtrabajo=numtrabajo).update(estado_consulta_experto=1)
            
            queryset = ConsultaExpertoTrab.objects.filter(numtrabajo=numtrabajo).last()
            
            for i in usuarios:
                responsable = ConsultadosTrab(responsable=i)
                responsable.save()
                responsable.consultadoa.add(queryset)

            
            queryset2 = ConsultadosTrab.objects.filter(consultadoa=queryset.id)
            
            d= []
            emailaprob=""
            for i in usuarios:
                emailaprob = emailaprob + i.email + ','
                d.append(emailaprob)
            
            print("Lista de correo, emailaprob")

            id_solicita = User.objects.get(genera_consultas_a_expertos=True) # MONICA SEVILLA
            
            ##Guarda cambio de estado en tabla de mensajeria
            
            noti = tipo_notificacion.objects.get(id=20) # 19. PEDIDOS INSUMOS CONSULTA A ROL EXPERTO

            app_number = modelo_App.objects.get(id=5) # 5. ordenTrabajo - ConsultaExperto

            for i in usuarios:

                r = notificaciones_globales(
                    app_origen = app_number,
                    estado = True,
                    identificador = queryset.numtrabajo.numtrabajo,
                    tipo = noti, 
                    usuario_activa = id_solicita,
                    autorizador_id = User.objects.get(id=i.id),
                    )
                r.save()

            envioMail('Consulta para Experto Técnico en orden de pedido de insumos', emailaprob, 'email_consulta_experto.html', queryset, queryset2)
            
        else:
            return HttpResponse("Errores")
        
        return redirect('trabajos:ordenescompra')

@login_required
def listarPreguntasExperto(request):

    usuario = request.user.id
    us = User.objects.get(id=usuario)
    print(usuario, us)

    form = ConsultaExpertoTrab.objects.filter(consul_trab__responsable=us).filter(numtrabajo__estado_consulta_experto=1).order_by('-fecha_pregunta')
    respondidas = ConsultaExpertoTrab.objects.filter(consul_trab__responsable=us).filter(numtrabajo__estado_consulta_experto=2).order_by('-fecha_respuesta')
    print(respondidas)

    return render(request, 'ordenTrabajo/listar_preguntas_experto.html', {'form':form,'form2':respondidas})

@login_required
def verPreguntasExperto(request,id):
    resp = request.POST.get('respuesta')
    try:
        query = ConsultaExpertoTrab.objects.get(id=id)
        numpe = query.numtrabajo.numtrabajo
    except: 
        query = ConsultaExpertoTrab.objects.get(numtrabajo=id)
        numpe = query.numtrabajo.numtrabajo
        id = query.id


    if request.method == 'GET':
        pregunta = query.pregunta
        query = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numpe)
        return render(request, 'ordenTrabajo/ver_preguntas_rol_experto.html', {'solic':query, 'pregunta':pregunta}) 

    else:
        fecha = datetime.now()
        respuesta_compuesta = str(datetime.strftime(fecha, '%Y-%m-%d %H:%M')) + " " + resp 

        actualiza_respuesta = ConsultaExpertoTrab.objects.filter(id=id).update(respuesta=respuesta_compuesta,
        fecha_respuesta=fecha)

        actualiza_estado = OrdenesTrabajos.objects.filter(numtrabajo=numpe).update(estado_consulta_experto=2)
        mail = User.objects.get(genera_consultas_a_expertos=True) # MONICA SEVILLA
        #mail = User.objects.get(id=10) # DM
        recibemail = mail.email

        queryset = ConsultaExpertoTrab.objects.get(id=id)
        
        queryset2 = ConsultadosTrab.objects.filter(consultadoa=queryset.id)
        

        envioMail('Ha recibido una respuesta de Consulta a Rol Experto en orden de pedido de insumos', recibemail, 'email_responde_experto.html', queryset, queryset2)
        #envioMail('Ha recibido una respuesta de Consulta a Rol Experto en orden de pedido de insumos', recibemail, 'email_responde_experto.html', queryset, "")

        return redirect('trabajos:listar_preguntas_experto')


@login_required
def reporteProyectos(request):
    # ahora = datetime.now()
    # ayer = ahora - timedelta(days=1)
    # fecha = ahora.strftime('%Y-%m-%d')
    # fechatxt = ahora.strftime('%Y%m%d')
    # fechaacttxt = ayer.strftime('%Y%m%d')
    proy = proyectos_contabilidad.objects.all().order_by('nombre_proyecto')
    proyecto_select = request.GET.get('numproyecto')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')

    print(proyecto_select)
    
    if proyecto_select:
        if proyecto_select == '35': # TODOS LOS PROYECTOS
            pselect = proyectos_contabilidad.objects.get(id=35)
            if desde and hasta:
                query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(
                    numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(
                        numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto__isnull=False).filter(
                            numtrabajo__fchgenerac__gte=desde,numtrabajo__fchgenerac__lte=hasta)
            
                q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(
                numtrabajo__aprobado=1).filter(numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto__isnull=False).filter(
                    numtrabajo__fchgenerac__gte=desde,numtrabajo__fchgenerac__lte=hasta).annotate(
                        valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]))
                
                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion.filter(
                numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]))
                
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion.filter(
                numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]))
            
                val2 = q3.filter(cotiza_seleccion=True).filter(numtrabajo__genera_compra=True)

                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE parametros.consolida_ordenes_proyectos")

                for i in val2:
                    r = consolida_ordenes_proyectos(
                        orden = i.numtrabajo.numtrabajo,
                        app_origen = '2',
                        solicita = i.numtrabajo.usuario_solicita.username,
                        departamento = i.numtrabajo.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numtrabajo.fchgenerac,
                        tipo = i.numtrabajo.tipo_pedi,
                        proyecto = i.numtrabajo.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numtrabajo.usuario_cotiza,
                        
                        )
                    r.save()

                query_cotiza_seleccion_acti = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(
                    numpedido__estado_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto__isnull=False).filter(
                        numpedido__fchgenerac__gte=desde,numpedido__fchgenerac__lte=hasta)

                q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(
                    numpedido__select_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto__isnull=False).filter(
                        numpedido__fchgenerac__gte=desde,numpedido__fchgenerac__lte=hasta).annotate(
                            valor=Subquery(query_cotiza_seleccion_acti.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]))

                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]))
                        
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]))


                val3 = q3.filter(cotiza_seleccion=True).filter(numpedido__genera_compra=True)

                for i in val3:

                    r = consolida_ordenes_proyectos(
                        orden = i.numpedido.numpedido,
                        app_origen = '1',
                        solicita = i.numpedido.usuario_solicita.username,
                        departamento = i.numpedido.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numpedido.fchgenerac,
                        tipo = i.numpedido.tipoactivo.tipo_nombre,
                        proyecto = i.numpedido.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numpedido.usuario_cotiza,    
                        )
                    r.save()      

            else:
            
                query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(
                numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto__isnull=False)
                
                q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__aprobado=1).filter(numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto__isnull=False).annotate(
                    valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]))
                    
                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]))
                    
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]))
                
                val2 = q3.filter(cotiza_seleccion=True).filter(numtrabajo__genera_compra=True)

                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE parametros.consolida_ordenes_proyectos")

                for i in val2:
                    r = consolida_ordenes_proyectos(
                        orden = i.numtrabajo.numtrabajo,
                        app_origen = '2',
                        solicita = i.numtrabajo.usuario_solicita.username,
                        departamento = i.numtrabajo.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numtrabajo.fchgenerac,
                        tipo = i.numtrabajo.tipo_pedi,
                        proyecto = i.numtrabajo.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numtrabajo.usuario_cotiza,
                        
                        )
                    r.save()
                
                query_cotiza_seleccion_acti = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(
                numpedido__estado_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto__isnull=False)

                q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(
                    numpedido__select_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto__isnull=False).annotate(
                        valor=Subquery(query_cotiza_seleccion_acti.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]))

                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]))
                        
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]))


                val3 = q3.filter(cotiza_seleccion=True).filter(numpedido__genera_compra=True)

                for i in val3:

                    r = consolida_ordenes_proyectos(
                        orden = i.numpedido.numpedido,
                        app_origen = '1',
                        solicita = i.numpedido.usuario_solicita.username,
                        departamento = i.numpedido.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numpedido.fchgenerac,
                        tipo = i.numpedido.tipoactivo.tipo_nombre,
                        proyecto = i.numpedido.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numpedido.usuario_cotiza,    
                        )
                    r.save()  

        else:

            pselect = proyectos_contabilidad.objects.get(id=proyecto_select)
            
            if desde and hasta:
                query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(
                numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto=proyecto_select).filter(
                        numtrabajo__fchgenerac__gte=desde,numtrabajo__fchgenerac__lte=hasta)
                     
                q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__aprobado=1).filter(numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto=proyecto_select).filter(
                        numtrabajo__fchgenerac__gte=desde,numtrabajo__fchgenerac__lte=hasta).annotate(
                            valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]))
                    
                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]))
                    
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]))
                
                val2 = q3.filter(cotiza_seleccion=True).filter(numtrabajo__genera_compra=True)

                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE parametros.consolida_ordenes_proyectos")

                for i in val2:
                    r = consolida_ordenes_proyectos(
                        orden = i.numtrabajo.numtrabajo,
                        app_origen = '2',
                        solicita = i.numtrabajo.usuario_solicita.username,
                        departamento = i.numtrabajo.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numtrabajo.fchgenerac,
                        tipo = i.numtrabajo.tipo_pedi,
                        proyecto = i.numtrabajo.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numtrabajo.usuario_cotiza,
                        
                        )
                    r.save()

                query_cotiza_seleccion_acti = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(
                    numpedido__estado_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto=proyecto_select).filter(
                        numpedido__fchgenerac__gte=desde,numpedido__fchgenerac__lte=hasta)
                    
                q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(
                    numpedido__select_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto=proyecto_select).filter(
                        numpedido__fchgenerac__gte=desde,numpedido__fchgenerac__lte=hasta).annotate(
                            valor=Subquery(query_cotiza_seleccion_acti.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]))

                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]))
                            
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]))


                val3 = q3.filter(cotiza_seleccion=True).filter(numpedido__genera_compra=True)

                for i in val3:

                    r = consolida_ordenes_proyectos(
                        orden = i.numpedido.numpedido,
                        app_origen = '1',
                        solicita = i.numpedido.usuario_solicita.username,
                        departamento = i.numpedido.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numpedido.fchgenerac,
                        tipo = i.numpedido.tipoactivo.tipo_nombre,
                        proyecto = i.numpedido.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numpedido.usuario_cotiza,    
                        )
                    r.save()
            
            else:

                query_cotiza_seleccion = CotizaTrabajo.objects.filter(cotiza_seleccion=True).filter(
                    numtrabajo__aprobado=1).filter(numtrabajo__estado_cotiza=True).filter(
                        numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto=proyecto_select)
                    
                q1 = DetalleTrabajo.objects.filter(numtrabajo__estado_cotiza=True).filter(
                    numtrabajo__aprobado=1).filter(numtrabajo__genera_compra=1).filter(numtrabajo__numproyecto=proyecto_select).annotate(
                    valor=Subquery(query_cotiza_seleccion.filter(numtrabajo=OuterRef('numtrabajo')).values('valor')[:1]))
                    
                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('cotiza_seleccion')[:1]))
                    
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion.filter(
                    numtrabajo=OuterRef('numtrabajo')).values('empresa_cotiza__nombre_empresa')[:1]))
                
                val2 = q3.filter(cotiza_seleccion=True).filter(numtrabajo__genera_compra=True)

                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE parametros.consolida_ordenes_proyectos")

                for i in val2:
                    r = consolida_ordenes_proyectos(
                        orden = i.numtrabajo.numtrabajo,
                        app_origen = '2',
                        solicita = i.numtrabajo.usuario_solicita.username,
                        departamento = i.numtrabajo.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numtrabajo.fchgenerac,
                        tipo = i.numtrabajo.tipo_pedi,
                        proyecto = i.numtrabajo.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numtrabajo.usuario_cotiza,
                        
                        )
                    r.save()

                query_cotiza_seleccion_acti = CotizaPedido.objects.filter(cotiza_seleccion=True).filter(numpedido__aprobado=1).filter(
                    numpedido__estado_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto=proyecto_select)

                q1 = DetallePedido.objects.filter(numpedido__aprobado=1).filter(numpedido__estado_cotiza=True).filter(
                    numpedido__select_cotiza=True).filter(numpedido__genera_compra=1).filter(numpedido__numproyecto=proyecto_select).annotate(
                        valor=Subquery(query_cotiza_seleccion_acti.filter(numpedido=OuterRef('numpedido')).values('valor')[:1]))

                q2 = q1.annotate(cotiza_seleccion=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('cotiza_seleccion')[:1]))
                            
                q3 = q2.annotate(empresa_cotiza=Subquery(query_cotiza_seleccion_acti.filter(
                    numpedido=OuterRef('numpedido')).values('empresa_cotiza__nombre_empresa')[:1]))


                val3 = q3.filter(cotiza_seleccion=True).filter(numpedido__genera_compra=True)

                for i in val3:

                    r = consolida_ordenes_proyectos(
                        orden = i.numpedido.numpedido,
                        app_origen = '1',
                        solicita = i.numpedido.usuario_solicita.username,
                        departamento = i.numpedido.departamento.dep_nombre,
                        fecha_aprueba_compra = i.numpedido.fchgenerac,
                        tipo = i.numpedido.tipoactivo.tipo_nombre,
                        proyecto = i.numpedido.numproyecto.nombre_proyecto,
                        descripcion = i.descripcion,
                        proveedor = i.empresa_cotiza,
                        valor = i.valor,
                        analista = i.numpedido.usuario_cotiza,    
                        )
                    r.save()
                
        query_final = consolida_ordenes_proyectos.objects.all().order_by('-fecha_aprueba_compra')

        paginator = Paginator(query_final, 50)
        page = request.GET.get('page')
        orden2 = paginator.get_page(page)
        
    else:
        orden2 = None
        pselect = None
        desde = None
        hasta = None
        
    return render(request, 'ordenTrabajo/reporte_proyectos.html', {'activos':orden2,'proy':proy,'pseleccionado':pselect,'inicio':desde,'fin':hasta})


def comprasPorProyectoExcel(request):
    
        query = consolida_ordenes_proyectos.objects.all().order_by('-fecha_aprueba_compra')
        
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Requisiciones de Activos e Insumos/Trabajos por Proyecto'

        ws.merge_cells('B1:K1')
        ws['A3'] = 'N° PEDIDO'
        ws['B3'] = 'FEC. AUTORIZA COMPRA GER. ADMIN'
        ws['C3'] = 'ORIGEN'
        ws['D3'] = 'SOLICITA'
        ws['E3'] = 'DEPARTAMENTO'
        ws['F3'] = 'TIPO'
        ws['G3'] = 'PROYECTO'
        ws['H3'] = 'DESCRIPCION'
        ws['I3'] = 'PROVEEDOR'
        ws['J3'] = 'VALOR'
        ws['K3'] = 'COMPRADORA'
        

            
        count = 4
        # column_num = 3
        # column_preg = 2
        # contador = 0
        # rowcount = 1  

        for registro in query:

            ws.cell(row = count, column = 1).value =  registro.orden
            ws.cell(row = count, column = 2).value =  datetime.strftime(registro.fecha_aprueba_compra, '%d-%m-%Y')
            if registro.app_origen == '1':
                ws.cell(row = count, column = 3).value =  'ACTIVOS'
            elif registro.app_origen == '2':
                ws.cell(row = count, column = 3).value =  'INSUMOS/TRAB'
            ws.cell(row = count, column = 4).value =  registro.solicita
            ws.cell(row = count, column = 5).value =  registro.departamento
            ws.cell(row = count, column = 6).value =  registro.tipo
            ws.cell(row = count, column = 7).value =  registro.proyecto
            ws.cell(row = count, column = 8).value =  registro.descripcion
            ws.cell(row = count, column = 9).value =  registro.proveedor
            ws.cell(row = count, column = 10).value =  registro.valor
            ws.cell(row = count, column = 11).value =  registro.analista
          
            count+=1
            # contador+=1
            # rowcount+=1

        nombre_archivo = "Reporte_de_Compras_por_Proyecto.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response


@login_required 
@permission_required('ordenPedido.genera_ordenescompra', raise_exception=True)
def actualizaObserva(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        compra = OrdenesTrabajos.objects.get(numtrabajo=numpedido)
        compra.observa_compra = observa
        # print('Actualiza Con observación: ' + observa)
        compra.save()

        return redirect('trabajos:ordenescompra')
    else:
        return redirect('trabajos:ordenescompra')

def cancelaCompra(request, numpedido):
    if request.method == 'GET':
        observa = request.GET.get('text')
        compra = OrdenesTrabajos.objects.get(numtrabajo=numpedido)
        compra.genera_compra = 2
        compra.observa_compra = observa
        # compra.aprobado = None
        compra.estado_cotiza = None
        # compra.select_cotiza = None
        compra.save()

        # cotiza = CotizaTrabajo.objects.filter(numtrabajo=numpedido).delete()

        # emailaprob = User.objects.get(id=compra.usuario_solicita.id).autorizador.get(status=True)
        # emailaprob_auto = emailaprob.e_mail
        emailaprob_auto = 'adquisiciones@ecofroz.com'
        print(emailaprob_auto)
        queryset = observa
        queryset2 = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)

        envioMail('Rechazo de Cotización', emailaprob_auto, 'email_rechaza_cotizacion.html', queryset, queryset2)

        return redirect('trabajos:ordenescompra')
    else:
        return redirect('trabajos:ordenescompra')

###### ACTIVAR ACTIVOS ######

@login_required
@permission_required('ordenPedido.facturacion', raise_exception=True)
def listarComprasConta(request):
    queryset = request.GET.get("buscar")
    # aprobados = DetallePedido.objects.all().select_related('numpedido').filter(numpedido__estado_cotiza=True).filter(numpedido__aprobado=1).order_by('-numpedido__ord')
    # aprobados = OrdenesPedidos.objects.filter(estado_cotiza = True).filter(aprobado=1)
    if queryset:
        orden = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset)
        ).filter(numtrabajo__genera_compra=1).order_by('-numtrabajo__ord')
    else:
        # orden = OrdenesPedidos.objects.filter(estado_cotiza__isnull=True).filter(aprobado=1)
        ordenx = DetalleTrabajo.objects.filter(numtrabajo__genera_compra=1).order_by('-numtrabajo__ord')

        paginator = Paginator(ordenx, 25)
        page = request.GET.get('page')
        orden = paginator.get_page(page)

    return render(request, 'ordenTrabajo/listar_compras_conta.html', {'form':orden, 'busqueda':queryset})

#######CONTAILIDAD###############

def datosCompraConta(request, numpedido):
    if request.method == 'GET':
        detalle = CotizaTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        valor = 0
        detalle_data = []
        for det in detalle:
            if det.cotiza_seleccion == True:
                valor = det.valor
            d={
                'id':det.id,
                'valor':det.valor,
                'empresa_cotiza':det.empresa_cotiza,
                'pdf_cotiza':det.pdf_cotiza,
                'cotiza_seleccion':det.cotiza_seleccion, 
            }
            detalle_data.append(d)
        detalle_cotiza = DetCotizaFormSet(initial=detalle_data)
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numpedido)
        idpedido = numpedido
        aprobado = DetalleTrabajo.objects.get(numtrabajo__numtrabajo=numpedido)
        aprob = aprobado.numtrabajo.genera_compra
        fact = aprobado.numtrabajo.aceptacion_factura
        print(aprob)
        return render(request, 'ordenTrabajo/compras_conta.html', {'detalle': detalle_cotiza, 'cot': cotiza, 'id':idpedido, 'aprobado':aprob, 'valor':valor, 'fact':fact})

## BODEGA ##
@login_required
def ordenesBodega(request):
    queryset = request.GET.get("buscar")
    activate = 0
    if queryset:
        # recepcion = DetalleTrabajo.objects.filter(
        #     Q(numtrabajo__numtrabajo__icontains = queryset) |
        #     Q(descripcion__icontains = queryset)
        # ).filter(numtrabajo__genera_compra=1).filter(Q(numtrabajo__tipo_liquida__isnull=True)|Q(numtrabajo__tipo_liquida=1)).filter(~Q(numtrabajo__tipo_pedi='OS') & ~Q(numtrabajo__tipo_pedi='MP') & ~Q(numtrabajo__tipo_pedi='MC') & ~Q(numtrabajo__tipo_pedi='PR')).order_by('-numtrabajo__ord')
       
       #Nueva consulta para listar productos que han sido identificados desde su soliitud como que sí requieren
       #recepcion de insumos. Se usa el campo "insumos" en la cabecera de Trabajos
        recepcion = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numtrabajo__genera_compra=1).filter(Q(numtrabajo__tipo_liquida__isnull=True)|Q(numtrabajo__tipo_liquida=1)).filter(numtrabajo__insumos=True) 

        entregado = DetalleTrabajo.objects.filter(
            Q(numtrabajo__numtrabajo__icontains = queryset) |
            Q(descripcion__icontains = queryset)
        ).filter(numtrabajo__genera_compra=1).filter(numtrabajo__entregado_bodega=True, numtrabajo__pago_mba__isnull=True).filter(~Q(numtrabajo__tipo_pedi='OS') & ~Q(numtrabajo__tipo_pedi='MP') & ~Q(numtrabajo__tipo_pedi='MC') & ~Q(numtrabajo__tipo_pedi='PR')).order_by('-numtrabajo__ord')

        q1 = recepcion.count()
        # q2 = generado.count()
        q3 = entregado.count()

        if q1 > q3:
            activate = 'pendientes'
        elif q3 > q1:
            activate = 'entregadas'
        else:
            activate = 0
    else:
        queryset = None
        recepcion = DetalleTrabajo.objects.filter(numtrabajo__genera_compra=1).filter(Q(numtrabajo__tipo_liquida__isnull=True)|Q(numtrabajo__tipo_liquida=1)).filter(numtrabajo__insumos=True).order_by('-numtrabajo__ord')
       
        entregado = DetalleTrabajo.objects.filter(numtrabajo__genera_compra=1).filter(numtrabajo__entregado_bodega=True, numtrabajo__pago_mba__isnull=True).filter(~Q(numtrabajo__tipo_pedi='OS') & ~Q(numtrabajo__tipo_pedi='MP') & ~Q(numtrabajo__tipo_pedi='MC') & ~Q(numtrabajo__tipo_pedi='PR')).order_by('-numtrabajo__ord')

    paginator_recep = Paginator(recepcion, 25) # Show 25 contacts per page.
    page_number_recep = request.GET.get('page_recep')
    page_obj_recep = paginator_recep.get_page(page_number_recep)

    # paginator_gen = Paginator(generado, 25) # Show 25 contacts per page.
    # page_number_gen = request.GET.get('page_gen')
    # page_obj_gen = paginator_gen.get_page(page_number_gen)

    paginator_ent = Paginator(entregado, 25) # Show 25 contacts per page.
    page_number_ent = request.GET.get('page_ent')
    page_obj_ent = paginator_ent.get_page(page_number_ent)

    if page_number_recep:
        activate = 'pendientes'
    elif page_number_ent:
        activate = 'entregadas'

    return render(request, 'ordenTrabajo/listar_recepciones.html', {'form':page_obj_recep, 'form3':page_obj_ent, 'activate':activate, 'queryset':queryset})

@login_required
def recibePedido(request, numtrabajo):
    queryset = OrdenesTrabajos.objects.get(numtrabajo=numtrabajo)
    queryset2 = DetalleTrabajo.objects.filter(numtrabajo__numtrabajo=numtrabajo)
    if request.method == 'GET':
        queryset.recibido_bodega = True
        queryset.fecha_recepcion = datetime.now()
        queryset.usuario_recibe = request.user.id
        queryset.save()

        email_usuario = User.objects.get(id=queryset.usuario_solicita.id)
        print(email_usuario.email)

        envioMail('Recepción de Pedido', email_usuario.email, 'email_recepcion.html', queryset, queryset2)

    return redirect('trabajos:recepcion')

def recepParcial(request, numtrabajo):
    pedido = request.GET.get('pedidoparcial')
    queryset = OrdenesTrabajos.objects.get(numtrabajo=pedido)
    if queryset.entrega_observa:
        observacion = str(queryset.entrega_observa) + ', ' + request.GET.get('text')
    else:
        observacion = request.GET.get('text')

    queryset.tipo_liquida = 1
    queryset.entrega_observa = observacion
    queryset.save()

    return redirect('trabajos:recepcion')

def recepTotal(request, numtrabajo):
    pedido = request.GET.get('pedido')
    queryset = OrdenesTrabajos.objects.get(numtrabajo=pedido)
    queryset2 = DetalleTrabajo.objects.filter(numtrabajo__numtrabajo=pedido)
    if request.method == 'GET':
        if queryset.total_observa:
            observacion = str(queryset.total_observa) + ', ' + request.GET.get('texttotal')
        else:
            observacion = request.GET.get('texttotal')

        queryset.recibido_bodega = True
        queryset.fecha_recepcion = datetime.now()
        queryset.tipo_liquida = 2
        queryset.total_observa = observacion
        queryset.entregado_bodega = True
        queryset.fecha_entrega = datetime.now()
        queryset.save()

    return redirect('trabajos:recepcion')

def detalleEntrega(request, numtrabajo):
    if request.method == 'GET':
        cotiza = DetalleTrabajo.objects.all().select_related('numtrabajo').filter(numtrabajo=numtrabajo)
        return render(request, 'ordenTrabajo/detalle_entrega.html', {'cot': cotiza})

def apruebaFactura(request):
    pedido = request.GET.get('pedido')
    valor = request.GET.get('valor')
    factura = request.GET.get('factura')
    fecha = datetime.now()
    ordenes = OrdenesTrabajos.objects.get(numtrabajo=pedido)
    ordenes.aceptacion_factura = True
    ordenes.valor_factura = valor
    ordenes.numero_factura = factura
    ordenes.usuario_factura = request.user.username
    ordenes.fch_factura = fecha
    ordenes.fch_factura_txt = fecha.strftime('%Y%m%d')
    ordenes.save()
    ordenes.save()

    # if ordenes.cod_activo is not None:
    #     activos = desc_activo.objects.get(activo_codigo=ordenes.cod_activo)
    #     activos.activo_valor_compra = valor
    #     activos.numero_factura = factura
    #     activos.save()
    
    return redirect('trabajos:compras_conta')

def creaTrabajo(request):
    ubica = request.GET.get('ubica_nombre')
    form2 = IngresoTrabajos(request.GET or None)
    form = DetTrForm(request.GET or None)
    success_url = reverse_lazy('trabajos:listar_trabajos')
    # print(ubica)
    if ubica:
        # print(ubica)
        x = activo_areas.objects.filter(area_ubica=ubica)
        form2 = IngresoTrabajos(request.GET or None, ubica_nombre = ubica, area_nombre=x)
        form = DetTrForm(request.GET or None)

        return render(request, 'ordenTrabajo/ingreso_trabajo2.html', {'form': form, 'form2':form2})

    return render(request, 'ordenTrabajo/ingreso_trabajo2.html', {'form': form, 'form2': form2})

# def creaTrabajo(request):
#     form = IngresoTrabajos
#     # if request.method == 'POST':
#         #guardo
    
#     return render(request, 'ordenTrabajo/ingreso_trabajo_op2.html', {'form': form})

def handler404(request, *args, **kwargs):
    return render(request, '404.html', status=404)

class CargaMbaPedidos(View):
    # model = PedidosMba
    def get(self, request, *args, **kwargs):
        user = kwargs['pk']
        docs = PedidosMba.objects.filter(usuario_solicita=user)      
             
        return render(self.request, 'ordenTrabajo/pedidos_mba.html', {'documentosmba': docs})

    def post(self, request, *args, **kwargs):
        user = kwargs['pk']  
        form = PedidosMbaForm(self.request.POST, self.request.FILES)
        fecha = datetime.now()
        fecha_txt =  fecha.strftime('%Y%m%d')
        if form.is_valid():
            doc = form.save(commit=False)
            doc.usuario_solicita = User.objects.get(id=user)
            doc.fch_carga = fecha
            doc.fch_carga_txt = fecha_txt
            doc.save() 
            data = {'is_valid': True, 'name': doc.documentos.name, 'url': doc.documentos.url}
        else:
            data = {'is_valid': False}
        return JsonResponse(data)

def listarMba(request):
    buscar = request.GET.get("buscar")
    if buscar:
        queryset = PedidosMba.objects.filter(
            Q(documentos__icontains = buscar)
        ).filter(estado_gestion__isnull=True)
        queryset2 = PedidosMba.objects.filter(estado_gestion=True)
    else:
        queryset = PedidosMba.objects.filter(estado_gestion__isnull=True).order_by('-ordenar')
        queryset2 = PedidosMba.objects.filter(estado_gestion=True).order_by('-ordenar')

    return render(request, 'ordenTrabajo/listar_pedidos_mba.html', {'pedidos':queryset, 'gestionados':queryset2})

def gestionaMba(request, id):
    queryset = PedidosMba.objects.get(id=id)
    queryset.estado_gestion = True
    queryset.save()

    return redirect('trabajos:pedidos_mba')

def actualizaMbaPedidos(request):

    if request.method == 'POST':
        id = request.POST.get('id')
        doc = PedidosMba.objects.get(id=id)
        form = PedidosMbaForm(request.POST, request.FILES, instance=doc)

        if form.is_valid():
            doc = form.save()
            doc.save() 
            return redirect('trabajos:pedidos_mba')

    return redirect('trabajos:pedidos_mba')

# ''''''''''ORDENES DE PAGO'''''''''''''''''''#

# '''''''''''ADQUISICIONES''''''''''''''''''''#

def muestraOrdenes(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = TrabajosPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = TrabajosPago.objects.all().order_by('fch_genera')

    return render(request, 'ordenTrabajo/listar_pagos_sol.html', {'ordenes':ordenes,})

class GeneraPagos(CreateView):
    model = TrabajosPago
    template_name = 'ordenTrabajo/ingreso_pagos.html'
    form_class = IngresoPagosForm
    success_url = reverse_lazy('trabajos:ingreso_pagos')

    def get_context_data(self, **kwargs):
        context = super(GeneraPagos, self).get_context_data(**kwargs)
        numero = OrdenesTrabajos.objects.all()
        if 'form' not in context:
            context['form'] = self.form_class(self.request.GET)
        if 'numero' not in context:
            context['numero'] = numero
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        form = self.form_class(request.POST, request.FILES)

        if form.is_valid():
            form.save()

            queryset = TrabajosPago.objects.all().filter(usuario_solicita=request.user.id).last()
            queryset2 = ""

            envioMail('Generación de Orden de Pago', 'gerencia.administrativa@ecofroz.com', 'email_orden_pago_sol.html', queryset, queryset2)

        return HttpResponseRedirect(self.get_success_url())

# '''''''''''GERENCIA ADMINISTRATIVA''''''''''''''''''''#

def muestraOrdenesGa(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = TrabajosPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = TrabajosPago.objects.all().order_by('fch_genera')

    return render(request, 'ordenTrabajo/listar_pagos_ga.html', {'ordenes':ordenes,})

def detallePagos(request, orden):
    orden = TrabajosPago.objects.get(id=orden)
    
    return render(request, 'ordenTrabajo/aprueba_pagos.html', {'orden':orden,})

def apruebaPagos(request, orden):
    orden =  TrabajosPago.objects.get(id=orden)
    orden.observaciones_ga = request.GET.get("text")
    orden.estado = True
    orden.save()

    queryset = orden
    queryset2 = ""

    envioMail('Generación de Orden de Pago', 'adquisiciones@ecofroz.com', 'email_orden_pago_ga_ap.html', queryset, queryset2)
    envioMail('Generación de Orden de Pago', 'contabilidad_sis@ecofroz.com', 'email_orden_pago_ga_ap.html', queryset, queryset2)

    return redirect('trabajos:aprueba_pagos')

def rechazaPagos(request, orden):
    orden =  TrabajosPago.objects.get(id=orden)
    orden.observaciones_ga = request.GET.get("text")
    orden.estado = False
    orden.save()

    queryset = orden
    queryset2 = ""

    envioMail('Generación de Orden de Pago', 'adquisiciones@ecofroz.com', 'email_orden_pago_ga_re.html', queryset, queryset2)
    envioMail('Generación de Orden de Pago', 'contabilidad_sis@ecofroz.com', 'email_orden_pago_ga_ap.html', queryset, queryset2)

    return redirect('trabajos:aprueba_pagos')

# '''''''''''CONTABILIDAD''''''''''''''''''''#

def muestraOrdenesConta(request):
    queryset = request.GET.get("buscar")
    if queryset:
        ordenes = TrabajosPago.objects.filter(numero__icontains=queryset).order_by('fch_genera')
    else:
        ordenes = TrabajosPago.objects.all().filter(~Q(estado=False)).order_by('fch_genera')

    return render(request, 'ordenTrabajo/listar_pagos_conta.html', {'ordenes':ordenes,})


@login_required
def listarSolicitudesAnticipos(request):
    queryset = request.GET.get("buscar")
    if queryset:
        
        query = SolicitudesAdqui.objects.filter(numsolic = queryset).filter(
            psolicita = request.user.username).exclude(anula=True)
    else:
        
        query = SolicitudesAdqui.objects.filter(psolicita = request.user.username).exclude(
            anula=True).order_by('-fch_solicita')
    
    qprove = proveedor.objects.all()
    
    paginator = Paginator(query, 50)
    page = request.GET.get('page')
    solic = paginator.get_page(page)

    return render(request, 'ordenTrabajo/listar_solicitudes_anticipos.html', {'form':solic,'busqueda':queryset,'prove':qprove})

@login_required
def anulaSolicitudesAnticipos(request, numsolic):
    if request.method == 'GET':
        anular = SolicitudesAdqui.objects.filter(numsolic=numsolic).update(anula=True)
    
    return redirect('trabajos:listar_solicitudes_anticipos')

@login_required
def ajax_call_proveedor(request):
    if 'term' in request.GET:
        qs =  proveedor.objects.filter(Q(nombre_empresa__icontains=request.GET.get('term')) | Q(ruc__icontains=request.GET.get('term')))
        # print(qs)
        prove = list()
        
        for empresa in qs:
            d={
            
                'label':empresa.nombre_empresa,
                'value':empresa.id,

            }
            prove.append(d)
            
        return JsonResponse(data=prove,safe=False)
    
    return render('ordenTrabajo/listar_solicitudes_anticipos.html')


@login_required
def ajax_call_trabajos(request):
    if 'term' in request.GET:
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=360)
        ordenes = DetalleTrabajo.objects.all().select_related('numtrabajo').exclude(
            Q(numtrabajo__tipo_pedi='AM') | Q(numtrabajo__tipo_pedi='AP')
              | Q(numtrabajo__tipo_pedi='CA') | Q(numtrabajo__tipo_pedi='CP')).filter(
                  numtrabajo__fchsolicita__range=[enddate,startdate]).filter(
                      numtrabajo__numtrabajo__icontains=request.GET.get('term')).order_by('-numtrabajo')
        trab_list = list()
        for trab in ordenes:
            d={
            
                'label':str(trab.numtrabajo.numtrabajo) +'  '+ trab.descripcion[:80]+'...',
                'value':trab.numtrabajo.numtrabajo,

            }
            trab_list.append(d)
        
            
        return JsonResponse(data=trab_list,safe=False)
    
    return render('ordenTrabajo/listar_solicitudes_anticipos.html')



@login_required
def ajax_consulta_solic(request):
    solic = request.GET.get('solic',None)

    orden = SolicitudesAdqui.objects.filter(numsolic=solic)                  
    trab_list = list()
    for trab in orden:
        if trab.numtrabajo:
            trabajo = trab.numtrabajo.numtrabajo
            if trab.docs:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':trabajo,
                    'valor':trab.valor,
                    'origen':'SIA Trabajos',
                    'notas':trab.observa_solicita,
                    'url':trab.docs.url,
                    'notasg':trab.observa_autoriza,
                }
            else:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':trabajo,
                    'valor':trab.valor,
                    'origen':'SIA Trabajos',
                    'notas':trab.observa_solicita,
                    'notasg':trab.observa_autoriza,
                }
            trab_list.append(d)
        elif trab.numpedido:
            pedido = trab.numpedido.numpedido
            if trab.docs:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':pedido,
                    'valor':trab.valor,
                    'origen':'SIA Activos',
                    'notas':trab.observa_solicita,
                    'url':trab.docs.url,
                    'notasg':trab.observa_autoriza,
                }
            else:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':pedido,
                    'valor':trab.valor,
                    'origen':'SIA Activos',
                    'notas':trab.observa_solicita,
                    'notasg':trab.observa_autoriza,
                    
                }

            trab_list.append(d)
        elif trab.num_mba:
            mba = trab.num_mba
            if trab.docs:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':mba,
                    'valor':trab.valor,
                    'origen':'MBA',
                    'notas':trab.observa_solicita,
                    'url':trab.docs.url,
                    'notasg':trab.observa_autoriza,
                    
                }
            else:
                d={
                    'numsolic':trab.numsolic,
                    'psolicita':trab.psolicita,
                    'proveedor':trab.prove.nombre_empresa,
                    'req':mba,
                    'valor':trab.valor,
                    'origen':'MBA',
                    'notas':trab.observa_solicita,
                    'notasg':trab.observa_autoriza,
                }
            
            trab_list.append(d)
        
    
    return JsonResponse(data=trab_list,safe=False)
    
    # return render('ordenTrabajo/listar_solicitudes_anticipos_1n.html')


@login_required
def ajax_aprueba_1n(request):
    
    solic = request.GET.get('solic',None)
    notas = request.GET.get('notas',None)

    try:
        actualiza = SolicitudesAdqui.objects.filter(numsolic=solic).update(
            aprobado_ger=True,observa_autoriza=notas,fch_aprueba=datetime.now(),usuario_aprueba=request.user.id)                
        
        data = {
                'message': 'Autorización enviada..'
            }

    except:
        data = {
                'message': 'Error. No se pudo realizar. Comuníquese con IT'
            }
    
    return JsonResponse(data,safe=False)

@login_required
def ajax_aprueba_2n(request):
    # form = cargaDoc(request.POST, request.FILES)
    
    solic = request.GET.get('solic',None)
    notas = request.GET.get('notas',None)

    # if form.is_valid():
    #     cab = form.save(commit=False)


    try:
        actualiza = SolicitudesAdqui.objects.filter(numsolic=solic).update(
            aceptado_conta=True,observa_cont=notas,fch_gestiona_cont=datetime.now(),pgestiona_cont=request.user.id)                
        
        data = {
                'message': 'Autorización aceptada..'
            }

    except:
        data = {
                'message': 'Error. No se pudo realizar. Comuníquese con IT'
            }
        
    return JsonResponse(data,safe=False)


@login_required
def ajax_rechaza_2n(request):
  
    
    solic = request.GET.get('solic',None)
    notas = request.GET.get('notas',None)

    try:
        actualiza = SolicitudesAdqui.objects.filter(numsolic=solic).update(
            aceptado_conta=False,observa_cont=notas,fch_gestiona_cont=datetime.now(),pgestiona_cont=request.user.id)                
        
        data = {
                'message': 'Autorización rechazada..'
            }

    except:
        data = {
                'message': 'Error. No se pudo realizar. Comuníquese con IT'
            }
        
    return JsonResponse(data,safe=False)


@login_required
def ajax_rechaza_1n(request):
    solic = request.GET.get('solic',None)
    notas = request.GET.get('notas',None)

    try:
        actualiza = SolicitudesAdqui.objects.filter(numsolic=solic).update(
            aprobado_ger=False,observa_autoriza=notas,fch_aprueba=datetime.now(),usuario_aprueba=request.user.id)                
        
        data = {
                'message': 'Rechazo generado..'
            }

    except:
        data = {
                'message': 'Error. No se pudo realizar. Comuníquese con IT'
            }
    
    return JsonResponse(data,safe=False)



@login_required
def ajax_call_activos(request):
    if 'term' in request.GET:
        startdate = date.today() + timedelta(days=1) 
        enddate = startdate - timedelta(days=360)
        ordenes = DetallePedido.objects.all().select_related('numpedido').filter(
            numpedido__fchsolicita__range=[enddate,startdate]).filter(
                      numpedido__numpedido__icontains=request.GET.get('term')).order_by('-numpedido')
        ped_list = list()
        for pedi in ordenes:
            d={
            
                'label':str(pedi.numpedido.numpedido) +'  '+ pedi.descripcion[:80]+'...',
                'value':pedi.numpedido.numpedido,

            }
            ped_list.append(d)
        
            
        return JsonResponse(data=ped_list,safe=False)
    
    return render('ordenTrabajo/listar_solicitudes_anticipos.html')


@login_required
def add_solic(request):
    
    if request.method == 'POST':
        
        form = cargaDoc(request.POST, request.FILES)
    
        origen = request.POST.get("origen", None)
        id_prove = request.POST.get("ruta", None)
        num_req = request.POST.get("num-req", None)
        valor_anticipo = request.POST.get("tarifa",None)
        notas = request.POST.get("desc_evento",None)
        num_mba = request.POST.get("mba",None)
        
        if form.is_valid():
            cab = form.save(commit=False)
            if origen == 'mba':
                cab.origen = 1
                cab.num_mba = num_mba
            elif origen == 'trabajos':
                cab.origen = 3
                numtra = OrdenesTrabajos.objects.get(numtrabajo=num_req)
                cab.numtrabajo = numtra
            elif origen == 'activos':
                cab.origen = 2
                numpe = OrdenesPedidos.objects.get(numpedido=num_req)
                cab.numpedido = numpe
            empresa = proveedor.objects.get(id=id_prove)
            cab.prove = empresa
            cab.valor = valor_anticipo
            cab.observa_solicita = notas
            cab.psolicita = request.user.username
            cab.fch_solicita = datetime.now()
            cab.aprobado = 0
            cab.enviada = True
            cab.save()

                        
        data = {
            'message': 'Salio por json response'
        }

        return JsonResponse(data)

@login_required
def listarApruebaAnticipos(request):
    
    usu_depend = Autorizador.objects.get(user_id=request.user.id)
    print(usu_depend)
    usu_depend2 = User.objects.filter(autorizador=usu_depend.id)
    print(usu_depend2)
    username_usu = []
    for i in usu_depend2:
        d = i.username
        username_usu.append(d)

    solic = SolicitudesAdqui.objects.filter(psolicita__in=username_usu).filter(enviada=True,aprobado_ger=None).exclude(
        anula=True).order_by('-fch_solicita')
    
    
    query = SolicitudesAdqui.objects.filter(psolicita__in=username_usu).filter(aprobado_ger__in=[True,False]).exclude(
            anula=True).order_by('-fch_aprueba')

    paginator = Paginator(query, 50)
    page = request.GET.get('page')
    query2 = paginator.get_page(page)

    return render(request, 'ordenTrabajo/listar_solicitudes_anticipos_1n.html', {'form':solic,'procesadas':query2})


@login_required
def listarGestionaAnticipos(request):
    queryset = request.GET.get("buscar")
    if queryset:
        pass
    else:
        solic = SolicitudesAdqui.objects.filter(aprobado_ger=True,aceptado_conta=None).exclude(
            anula=True).order_by('-fch_solicita')
    
        query = SolicitudesAdqui.objects.filter(aceptado_conta__in=[True,False]).exclude(
            anula=True).order_by('-fch_aprueba')

    paginator = Paginator(query, 50)
    page = request.GET.get('page')
    query2 = paginator.get_page(page)
   
    return render(request, 'ordenTrabajo/gestionar_anticipos.html', {'form':solic,'procesadas':query2})



@login_required
def revisaConta(request,numsolic):
    if request.method == 'GET':
        solic = SolicitudesAdqui.objects.get(numsolic=numsolic)

        return render(request, 'ordenTrabajo/revisa_conta.html', {'solic':solic})
    else:
        confirma = request.POST.get("confirmav")
        fecha = datetime.today()
        numero = request.POST.get("numvale")
        valor = request.POST.get("valorvale")

        if confirma == "registro":
            actualiza_campo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                estado_vale=1,fecha_gestiona_vale=fecha,
                 usu_gestiona_vale=request.user.username,observaciones_vale=request.POST.get("notas_registro"),
                 num_vale=numero,valor_vale=valor)
            query = SolicitudTransporte.objects.get(numpedido=numpedido)
            usuario_solic = query.usuario_solicita.email
            envioMail('Vale de Combustible Generado', usuario_solic, 'vale_generado.html', query, '')

        else:
        
            if request.POST.get('notas_rechazo'):
                actualiza_rechazo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                    estado_vale=2,fecha_gestiona_vale=fecha,usu_gestiona_vale=request.user.id, observaciones_vale=str(request.POST.get('notas_rechazo')))
            else:
                actualiza_campo = SolicitudTransporte.objects.filter(numpedido=numpedido).update(
                    estado_vale=2,fecha_gestiona_vale=fecha, usu_gestiona_vale=request.user.id)

            # queryset = SolicitudTransporte.objects.get(numpedido=numpedido)
            # emailaprobb = User.objects.get(rol='GTR')
            # emailaprob = emailaprobb.email
            
            # copia = 'mantenimiento@ecofroz.com'

            ## Sección para Notificaciones Globales
            
                
            # id_genera = User.objects.get(id=id_usuario)
            # emailgenera =  Generador.objects.get(id=id_genera.generador_id)
            # id_aprueba = User.objects.get(rol='GTR')  #ROL = GESTOR DE TRANSPORTE

            ##Guarda cambio de estado en tabla de mensajeria
            # noti = tipo_notificacion.objects.get(id=25) # 25. TRANSPORTE AUTORIZACION GERENTE AREA

            # app_number = modelo_App.objects.get(id=6) # 6. PROCESOS ADMINISTRATIVOS 

            # r = notificaciones_globales(
            #     app_origen = app_number,
            #     estado = True,
            #     identificador = query.numpedido,
            #     tipo = noti, 
            #     usuario_activa = id_genera,
            #     autorizador_id = id_aprueba,
            #     )

            # r.save()


            
        #     #envioMail('Solicitud Autorizada de Orden de Trabajo Interna', copia, 'email_generado.html', queryset, '')

        return redirect('trabajos:listar_gestiona_anticipos_2n')

#Guardar Archivo sin Form 
def saveComprobante(request,pk):
    registro = SolicitudesAdqui.objects.get(numsolic=pk)
    if request.method == 'POST' and request.FILES.get('comprobante'):
        archivo = request.FILES['comprobante']

        registro.comprobante = archivo
        registro.save()
    
        
    return redirect('trabajos:listar_gestiona_anticipos_2n')
