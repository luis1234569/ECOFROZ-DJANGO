from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView,CreateView,UpdateView,DeleteView,DetailView, TemplateView, View
from django.db.models import Q
from django.db import connection
from .forms import RegistroFormP, RegistroFormDetP, FiltrarForm, FiltrarFormDet, Documentos, Fichas, ProveedorN, \
    ResultadoConsultaForm,  FiltrarFormSegmento, ParaExportarExcelForm, NombreEmpresaForm, NombreEmpresaConRespuestasForm, \
        SegmentoForm, CargaFichasForm, HojasMSDS,EtiquetasProductos, ExpiraDocumentos
from .models import *
from .models import proveedor
from django.core.paginator import Paginator
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from psycopg2.extras import NamedTupleCursor
from django.urls import reverse
from django.template import context, loader

from django.views.generic import View
from django.views import View
#from .filters import SnippetFilter
from django.db import connection, connections
from tablib import Dataset 
from .resources import RespuestasResource
import random
import logging
from django.contrib.auth.decorators import login_required, permission_required
from django.core.mail import EmailMessage, EmailMultiAlternatives, send_mail
from django.forms.models import inlineformset_factory
from django.contrib import messages
import csv
from os import system
# import pandas as pd



from braces.views import (
    AjaxResponseMixin,
    JSONResponseMixin,
    LoginRequiredMixin,
    SuperuserRequiredMixin,
)

def home(request):
     return render(request,'proveedores/base_proveedores.html')

logger = logging.getLogger(__name__)

### FUNCIONES INTERNAS ###
def envioMail(subject, email, template, prove_nombre, categoria, args):
    html_message = loader.render_to_string(
        'proveedores/email/%s' %template,
            {
                'prove_nombre':prove_nombre,
                'categoria':categoria,
                'args':args,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except Exception as e:
        # print(e)
        logging.basicConfig(level=logging.DEBUG, filename='/home/django/ecofroz/ecofroz/registro_error_mail.log')
        logger.error(e)

def envioMailParaRevisionCalidad(subject, email, template, prove_nombre, categoria, msj_reenvio, args):
    html_message = loader.render_to_string(
        'proveedores/email/%s' %template,
            {
                'prove_nombre':prove_nombre,
                'categoria':categoria,
                'msj_reenvio':msj_reenvio,
                'args':args,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except Exception as e:
        # print(e)
        logging.basicConfig(level=logging.DEBUG, filename='/home/django/ecofroz/ecofroz/registro_error_mail.log')
        logger.error(e)

def envioMailRespuestaCalidad(subject, email, template, prove_nombre, categoria, observaciones, valor_aprueba, args):
    html_message = loader.render_to_string(
        'proveedores/email/%s' %template,
            {
                'prove_nombre':prove_nombre,
                'categoria':categoria,
                'observaciones':observaciones,
                'valor_aprueba':valor_aprueba,
                'args':args,
            }
        )
    email_subject = subject
    to_list = email.split(',')
    mail = EmailMultiAlternatives(
            email_subject, '', '', (to_list), bcc=['desarrolloecofroz@gmail.com'])
    mail.attach_alternative(html_message, "text/html")
    try:
        mail.send()
    except Exception as e:
        # print(e)
        logging.basicConfig(level=logging.DEBUG, filename='/home/django/ecofroz/ecofroz/registro_error_mail.log')
        logger.error(e)

#Eliminar Documentos subidos


def eliminar_documentos(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    if request.method == 'POST':
        docu.delete()
        return redirect('proveedores:ver_documentos',prove)
    return render(request,'proveedores/eliminar_documentos.html',{'documento':docu})

#Guardar respuestas de revisiòn de Fichas Calidad

def save_record_ficha_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    valor_revisa = request.POST.getlist("revisado")
    valor_marca = request.POST.getlist("marca")
    sub = request.POST.getlist("subcate")

    prove = proveedor.objects.get(id=prove)
    subcate = sub[0]
    
    

    print(valor_revisa)
    print(valor_marca)
    print(subcate,type(subcate))
    

    # SUBCATEGORIAS AGRICOLAS (NO BORRAR):
    # 1 FICHAS TECNICAS Pesticidas
    # 2 FICHAS TECNICAS Foliares
    # 3. FICHAS TECNICAS Fertilizantes
    # 4. FICHAS TECNICAS MATERIA ORGANICA

    # 13. FICHAS TECNICAS SEMILLAS

    # 5. REGISTROS AGROCALIDAD PESTICIDAS
    # 6. REGISTROS AGROCALIDAD Foliares
    # 7. REGISTROS AGROCALIDAD Fertilizantes
    # 8. REGISTROS AGROCALIDAD MAteria Orgánica

    # 14. REGISTROS AGROCALIDAD SEMILLAS

    # 9. MSDS PESTICIDAS
    # 10. MSDS Foliares
    # 11. MSDS Fertilizantes
    # 12. MSDS Materia Orgánica

    # 15. MSDS Semillas
    # 16.Etiquetas de productos Foliares
    # 17. Etiquetas de productos Pesticidas
    # 18. Etiquetas de productos Fertilizantes
    # 19. Etiquetas de productos MAteria Orgánica
    # 20. Etiquetas de productis Semillas

    # 21. Análisis Pesticidas
    # 22. Análsis Foliares
    # 23. Análisis Fertilizantes
    # 24. Analsiis Materia Orgánica
    # 25. Análisis Semillas

    # 26. Certificados Pesticidas
    # 27. Certificados Foliares
    # 28. Certificados Fertilizantes
    # 29. Certificados MAO
    # 30. Certificados Semillas



    if request.method == 'POST':

        if subcate == '1':

            query = documentos_prove.objects.filter(proveedor=prove,es_ficha_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '2':

            query = documentos_prove.objects.filter(proveedor=prove,es_ficha_fol=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '3':

            query = documentos_prove.objects.filter(proveedor=prove,es_ficha_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '4':

            query = documentos_prove.objects.filter(proveedor=prove,es_ficha_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '5':

            query = documentos_prove.objects.filter(proveedor=prove,es_agro_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()


        elif subcate == '6':

            query = documentos_prove.objects.filter(proveedor=prove,es_agro_fol=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '7':

            query = documentos_prove.objects.filter(proveedor=prove,es_agro_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '8':

            query = documentos_prove.objects.filter(proveedor=prove,es_agro_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '9':

            query = documentos_prove.objects.filter(proveedor=prove,es_msds_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '10':

            query = documentos_prove.objects.filter(proveedor=prove,es_msds_fol=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '11':

            query = documentos_prove.objects.filter(proveedor=prove,es_msds_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '12':

            query = documentos_prove.objects.filter(proveedor=prove,es_msds_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '13':

            query = documentos_prove.objects.filter(proveedor=prove,es_ficha_sem=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '14':

            query = documentos_prove.objects.filter(proveedor=prove,es_agro_sem=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '15':

            query = documentos_prove.objects.filter(proveedor=prove,es_msds_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

    # SUBCATEGORIAS AGRICOLAS (NO BORRAR):
    # 1 FICHAS TECNICAS Pesticidas
    # 2 FICHAS TECNICAS Foliares
    # 3. FICHAS TECNICAS Fertilizantes
    # 4. FICHAS TECNICAS MATERIA ORGANICA

    # 13. FICHAS TECNICAS SEMILLAS

    # 5. REGISTROS AGROCALIDAD PESTICIDAS
    # 6. REGISTROS AGROCALIDAD Foliares
    # 7. REGISTROS AGROCALIDAD Fertilizantes
    # 8. REGISTROS AGROCALIDAD MAteria Orgánica

    # 14. REGISTROS AGROCALIDAD SEMILLAS

    # 9. MSDS PESTICIDAS
    # 10. MSDS Foliares
    # 11. MSDS Fertilizantes
    # 12. MSDS Materia Orgánica

    # 15. MSDS Semillas
    # 16.Etiquetas de productos Foliares
    # 17. Etiquetas de productos Pesticidas
    # 18. Etiquetas de productos Fertilizantes
    # 19. Etiquetas de productos MAteria Orgánica
    # 20. Etiquetas de productis Semillas

    # 21. Análisis Pesticidas
    # 22. Análsis Foliares
    # 23. Análisis Fertilizantes
    # 24. Analsiis Materia Orgánica
    # 25. Análisis Semillas

    # 26. Certificados Pesticidas
    # 27. Certificados Foliares
    # 28. Certificados Fertilizantes
    # 29. Certificados MAO
    # 30. Certificados Semillas

        elif subcate == '16':

            query = documentos_prove.objects.filter(proveedor=prove,es_etiqueta_foliar=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '17':

            query = documentos_prove.objects.filter(proveedor=prove,es_eti_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '18':

            query = documentos_prove.objects.filter(proveedor=prove,es_eti_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '19':

            query = documentos_prove.objects.filter(proveedor=prove,es_eti_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '20':

            query = documentos_prove.objects.filter(proveedor=prove,es_eti_sem=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '21':

            query = documentos_prove.objects.filter(proveedor=prove,es_analisis_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '22':

            query = documentos_prove.objects.filter(proveedor=prove,es_analisis_fol=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '23':

            query = documentos_prove.objects.filter(proveedor=prove,es_analisis_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '24':

            query = documentos_prove.objects.filter(proveedor=prove,es_analisis_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '25':

            query = documentos_prove.objects.filter(proveedor=prove,es_analisis_sem=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '26':

            query = documentos_prove.objects.filter(proveedor=prove,es_certificado_pes=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '27':

            query = documentos_prove.objects.filter(proveedor=prove,es_certificado_fol=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()
        
        elif subcate == '28':

            query = documentos_prove.objects.filter(proveedor=prove,es_certificado_fer=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '29':

            query = documentos_prove.objects.filter(proveedor=prove,es_certificado_mao=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

        elif subcate == '30':

            query = documentos_prove.objects.filter(proveedor=prove,es_certificado_sem=True).order_by('nombre_corto')
            for actualiza,v in zip(query,valor_revisa):
                actualiza.revisado_cal = v
                actualiza.save()
            
            for actualiza,v in zip(query,valor_marca):
                actualiza.marca_cal = v
                actualiza.save()

    return HttpResponseRedirect(reverse('proveedores:ver_docu_subcates_calidad',kwargs={'pk':prove.id,'id':subcate}))


#Guarda la fecha de los documentos ingresada por Administración en la pantalla de Cargar Doocumentos
def save_record_fecha_documento(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    valor_revisa_cabecera = request.POST.getlist("fecha-documento-cabecera")
    valor_revisa_cuerpo = request.POST.getlist("fecha-documento-cuerpo")
    
    print("Cabecera",valor_revisa_cabecera)
    print("Cuerpo",valor_revisa_cuerpo,len(valor_revisa_cuerpo))
    

    if request.method == 'POST':
        
        if valor_revisa_cabecera:

            cabecera = documentos_prove.objects.filter(proveedor=prove).exclude(Q(es_ficha=True) | Q(es_hoja_msds=True)| Q(es_etiqueta_producto=True)).filter(
                Q(archivos__icontains='RUC') | Q(archivos__icontains='IESS') | Q(
                    archivos__icontains='NOMBRAMIENTO_REP')).order_by('-archivos')
    

            for actualiza,v in zip(cabecera,valor_revisa_cabecera):
                if v == '':
                    actualiza.fecha_documento = None
                else:
                    actualiza.fecha_documento = v
                actualiza.save()
        
        elif valor_revisa_cuerpo:
            print("Entra DM")

            cuerpo = documentos_prove.objects.filter(proveedor=prove).exclude(Q(es_ficha=True)
         | Q(es_hoja_msds=True)| Q(es_etiqueta_producto=True)| Q(es_ficha_pes=True) | Q(es_ficha_fol=True)
         | Q(es_ficha_fer=True) | Q(es_ficha_mao=True) | Q(es_ficha_sem=True) | Q(es_agro_pes=True) 
         | Q(es_agro_fol=True) | Q(es_agro_fer=True) | Q(es_agro_mao=True) | Q(es_agro_sem=True) 
         | Q(es_msds_pes=True) | Q(es_msds_fol=True) | Q(es_msds_fer=True) | Q(es_msds_mao=True)
         | Q(es_msds_sem=True) | Q(es_etiqueta_foliar=True) | Q(es_etiqueta_fer=True) | Q(es_etiqueta_mao=True)
         | Q(es_etiqueta_pes=True) | Q(es_etiqueta_sem=True)| Q(es_analisis_fer=True) | Q(es_analisis_fol=True)
         | Q(es_analisis_mao=True) | Q(es_analisis_pes=True) | Q(es_analisis_sem=True) | Q(es_certificado_fer=True)
         | Q(es_certificado_fol=True) | Q(es_certificado_mao=True) | Q(es_certificado_pes=True) | Q(es_certificado_sem=True)
         | Q(archivos__icontains='RUC')
         | Q(archivos__icontains="NOMBRAMIENTO_REP_LEGA")
         | Q(archivos__icontains="CUMPLIMIENTO_IES")).order_by('archivos')

            print(cuerpo)
        
            for actualiza,v in zip(cuerpo,valor_revisa_cuerpo):
                if v == '':
                    actualiza.fecha_documento = None
                else:
                    actualiza.fecha_documento = v
                actualiza.save()

            

    return HttpResponseRedirect(reverse('proveedores:ver_documentos',kwargs={'pk':prove}))


def save_record_msds_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    valor_revisa = request.POST.getlist("revisado")
    valor_marca = request.POST.getlist("marca")

    print(valor_revisa)
    print(valor_marca)

    if request.method == 'POST':

        query = documentos_prove.objects.filter(proveedor=prove,es_hoja_msds=True).order_by('nombre_corto')
        for actualiza,v in zip(query,valor_revisa):
            actualiza.revisado_cal = v
            actualiza.save()
        
        for actualiza,v in zip(query,valor_marca):
            actualiza.marca_cal = v
            actualiza.save()
        

    return HttpResponseRedirect(reverse('proveedores:ver_msds_calidad',kwargs={'pk':prove}))

def save_record_eti_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    valor_revisa = request.POST.getlist("revisado")
    valor_marca = request.POST.getlist("marca")

    print(valor_revisa)
    print(valor_marca)

    if request.method == 'POST':

        query = documentos_prove.objects.filter(proveedor=prove,es_etiqueta_producto=True).order_by('nombre_corto')
        for actualiza,v in zip(query,valor_revisa):
            actualiza.revisado_cal = v
            actualiza.save()
        
        for actualiza,v in zip(query,valor_marca):
            actualiza.marca_cal = v
            actualiza.save()
        

    return HttpResponseRedirect(reverse('proveedores:ver_etiquetas_calidad',kwargs={'pk':prove}))



#Guarda comentarios individuales de fichas 
def save_observa_ficha_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    observaciones = request.POST.get("comentarios")
    cate = proveedor.objects.get(id=prove)
    prove_nombre=cate.nombre_empresa
    categoria = cate.categoria

    prove = proveedor.objects.get(id=prove)
    
    print(subcate)
    
    if request.method == 'POST':
        print(observaciones)
        print(prove)
        print(pk)
        print(subcate,type(subcate))
        
        actualiza = documentos_prove.objects.filter(proveedor=prove,id=pk).update(observaciones_cal=observaciones)

        args=documentos_prove.objects.filter(proveedor=prove,id=pk)
        

        #envioMail('Nuevo comentario de observaciones de Calidad en Fichas de Proveedores', 'asistente.administrativa@ecofroz.com,dmencias@ecofroz.com', 'email_observa_ficha_proveedor_calidad.html', prove_nombre, categoria,args)

        
        return HttpResponseRedirect(reverse('proveedores:ver_docu_subcates_calidad',kwargs={'pk':prove.id}))
    
    return render(request,'proveedores/registra_observaciones_ficha_individual.html',{'documento':docu})


def save_observa_docu_agro(request, pk, prove,subcate):
    docu = documentos_prove.objects.get(id=pk)
    observaciones = request.POST.get("comentarios")
    cate = proveedor.objects.get(id=prove)
    prove_nombre=cate.nombre_empresa
    categoria = cate.categoria

    prove = proveedor.objects.get(id=prove)
    
    print(subcate)
    
    if request.method == 'POST':
        print(observaciones)
        print(prove)
        print(pk)
        print(subcate,type(subcate))
        
        actualiza = documentos_prove.objects.filter(proveedor=prove,id=pk).update(observaciones_cal=observaciones)

        args=documentos_prove.objects.filter(proveedor=prove,id=pk)
        

        #envioMail('Nuevo comentario de observaciones de Calidad en Fichas de Proveedores', 'asistente.administrativa@ecofroz.com,dmencias@ecofroz.com', 'email_observa_ficha_proveedor_calidad.html', prove_nombre, categoria,args)

        
        return HttpResponseRedirect(reverse('proveedores:ver_docu_subcates_calidad',kwargs={'pk':prove.id,'id':subcate}))
    
    return render(request,'proveedores/registra_observaciones_ficha_individual.html',{'documento':docu,'subcate':subcate})




def save_observa_msds_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    observaciones = request.POST.get("comentarios")
    cate = proveedor.objects.get(id=prove)
    prove_nombre=cate.nombre_empresa
    categoria = cate.categoria
    
    if request.method == 'POST':
        print(observaciones)
        print(prove)
        print(pk)
        
        actualiza = documentos_prove.objects.filter(proveedor=prove,id=pk).update(observaciones_msds_cal=observaciones)

        args=documentos_prove.objects.filter(proveedor=prove,id=pk)
        

        envioMail('Nuevo comentario de observaciones de Calidad en Hojas MSDS de Proveedores', 'asistente.administrativa@ecofroz.com,dmencias@ecofroz.com', 'email_observa_ficha_proveedor_calidad.html', prove_nombre, categoria,args)

        
        return HttpResponseRedirect(reverse('proveedores:ver_msds_calidad',kwargs={'pk':prove}))
    return render(request,'proveedores/registra_observaciones_msds_individual.html',{'documento':docu})


def save_observa_eti_cal(request, pk, prove):
    docu = documentos_prove.objects.get(id=pk)
    observaciones = request.POST.get("comentarios")
    cate = proveedor.objects.get(id=prove)
    prove_nombre=cate.nombre_empresa
    categoria = cate.categoria
    
    if request.method == 'POST':
        print(observaciones)
        print(prove)
        print(pk)
        
        actualiza = documentos_prove.objects.filter(proveedor=prove,id=pk).update(observaciones_eti_cal=observaciones)

        args=documentos_prove.objects.filter(proveedor=prove,id=pk)
        

        envioMail('Nuevo comentario de observaciones de Etiquetas de Productos Proveedores', 'asistente.administrativa@ecofroz.com,dmencias@ecofroz.com', 'email_observa_ficha_proveedor_calidad.html', prove_nombre, categoria,args)

        
        return HttpResponseRedirect(reverse('proveedores:ver_eti_calidad',kwargs={'pk':prove}))
    return render(request,'proveedores/registra_observaciones_eti_individual.html',{'documento':docu})



#Auditoria
def auditoriayCambios(request):

    nuevo = proveedor_det.objects.all().select_related('codigo_id')
    antig = proveedor.history.all()
    # viejo = proveedor.objects.all().history.most_recent()

    paginator = Paginator(antig, 150)
    page = request.GET.get('page')
    antiguo = paginator.get_page(page)
   
    return render(request,'proveedores/auditoria.html',{'form':antiguo})

def buscaCambios(request,pk):
    query = proveedor.history.all().filter(id=pk)[:2]

    query2 = proveedor_det.history.all().filter(codigo_id=pk)[:2]

    query3 = proveedor_respuestas.history.all().order_by('history_date').filter(proveedor_id=pk)[:2000]
    return render(request,'proveedores/auditoria_detalle.html',{'query':query,'query2':query2,'query3':query3})


#Auditoria a CSV

def auditaCabecera(request):
    output = []
    response = HttpResponse (content_type='text/csv')
    writer = csv.writer(response)
    query_set = proveedor.history.filter(history_date__range=["2022-06-13", "2022-06-30"])
    #Header
    writer.writerow(['Proveedor', 'Fecha Acción', 'Usuario Realiza', 'Tipo Acción'])
    for i in query_set:
        output.append([i.nombre_empresa, i.history_date, i.history_user, i.history_type])
    #CSV Data
    writer.writerows(output)
    return response

#AUDITA CAMBIOS PREGUNTAS/RESPUESTAS
def auditaRespuestas(request):
    output = []
    response = HttpResponse (content_type='text/csv')
    writer = csv.writer(response)
    query_set = proveedor_respuestas.history.filter(history_date__range=["2022-06-27", "2022-06-29"]).select_related('proveedor_id')
    #Header
    writer.writerow(['Proveedor', 'Fecha Acción', 'Usuario Realiza', 'Pregunta', 'Respuesta', 'Calificacion'])
    for i in query_set:
        output.append([i.proveedor_id.nombre_empresa, i.history_date, i.history_user, i.pregunta, i.respuesta, i.calificacion])
    #CSV Data
    writer.writerows(output)
    return response



#Edición de respuestas de proveedores

def editarRespuestaEncuesta(request,pregunta,respuesta,prove):
    if request.method == 'POST':
        segmento_escogido = request.POST.get("segmento_escogido")
        respuesta_id = respuesta
        nueva_respuesta = request.POST.get("nueva-respuesta")

        if segmento_escogido:
            print("Ediciòn de preguntas con seleccion de segmento")
            query = proveedor_respuestas.objects.all().select_related('pregunta').select_related('proveedor_id').filter(
                pregunta__id = pregunta).filter(proveedor_id__id = prove)
        
            quest = proveedor_encuesta.objects.get(id=pregunta)
            id_prove = proveedor.objects.get(id=prove)

            actualiza = proveedor_respuestas.objects.filter(id = respuesta_id,proveedor_id=id_prove,pregunta=quest).update(respuesta=nueva_respuesta)
            print("Pasò por el update con segmento")

            # r = proveedor_respuestas(
            #     id = respuesta_id,
            #     respuesta = nueva_respuesta.upper(),
            #     proveedor_id = id_prove,
            #     pregunta = quest,
            #     )
            # r.save()
  
        
            prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=prove).order_by('pregunta__pregunta_campo_modelo_id').filter(
                    pregunta__segmento__id=segmento_escogido).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

            calculaCalificacion(prove1)

            q = proveedor.objects.filter(id=prove)

            for cat in q:
                cate = cat.categoria.id

            empresa_tip = proveedor.objects.all().filter(id=prove)
            for i in empresa_tip:
                tipo = i.tipo_empresa
            
            print("Hola Tipo",tipo)
            
            if tipo == 'Persona Jurídica':
                print("Entra papà")

                prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=prove).order_by('pregunta__pregunta_campo_modelo_id').filter(
                    pregunta__segmento__id=segmento_escogido).filter(
                        proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(
                            pregunta__pregunta__icontains='*')
            
            else:
                prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(proveedor_id__id=prove).order_by('pregunta__pregunta_campo_modelo_id').filter(
                        pregunta__segmento__id=segmento_escogido).filter(
                            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').filter(
                                Q(pregunta__pregunta__icontains="*"))
            
            calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=prove).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)


            nombre = proveedor.objects.get(id=prove)
            parametros = NombreEmpresaConRespuestasForm(request.GET or None)
            parametros2 = FiltrarFormDet(request.GET or None)
            parametros3 = SegmentoForm(request.GET or None)

            suma_prove1 = 0
            for nota in prove1:
                suma_prove1+=nota.calificacion
            
            if suma_prove1 > 100:
                suma_prove1 = 100

            suma_total = 0
            for nota in calif_total:
                suma_total+=nota.calificacion

            if suma_total > 100:
                suma_total = 100
            
            v = proveedor.objects.filter(id=prove)

            for nota in v:
                califica = nota.grade 
            
            consulta = proveedor.objects.get(id=prove)
            consulta.calificacion = suma_total
            consulta.save()
            

            return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form2':prove1,'form3':parametros2,'form4':parametros3,'nombre':nombre,'id_prove':prove,'suma_prove1':suma_prove1,'segmento_escogido':segmento_escogido,'suma_total':suma_total,'califica':califica})

   



        #ELSE (NO SE ESCOGIO SEGMENTO)
        

        query = proveedor_respuestas.objects.all().select_related('pregunta').select_related('proveedor_id').filter(
            pregunta__id = pregunta).filter(proveedor_id__id = prove)
        
        quest = proveedor_encuesta.objects.get(id=pregunta)
        id_prove = proveedor.objects.get(id=prove)


        actualiza = proveedor_respuestas.objects.filter(id = respuesta_id,proveedor_id=id_prove,pregunta=quest).update(respuesta=nueva_respuesta)
        print("Pasò por el update")

        # r = proveedor_respuestas(
        #     id = respuesta_id,
        #     respuesta = nueva_respuesta,
        #     proveedor_id = id_prove,
        #     pregunta = quest,
        #     )
        # r.save()
        

        prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=prove).order_by('pregunta__pregunta_campo_modelo_id').exclude(respuesta='xxx')

        calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=prove)

        calculaCalificacion(prove1) 
        calculaCalificacion(calif_total) 

        suma_prove1 = 0
        for nota in prove1:
            suma_prove1+=nota.calificacion
        
        if suma_prove1 > 100:
            suma_prove1 = 100

        suma_total = 0
        for nota in calif_total:
            suma_total+=nota.calificacion
        
        if suma_total > 100:
            suma_total = 100
        
        consulta = proveedor.objects.get(id=prove)
        consulta.calificacion = suma_total
        consulta.save()
            
            
        nombre = proveedor.objects.all().get(id=prove)
        parametros = NombreEmpresaConRespuestasForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None)
        parametros3 = SegmentoForm(request.GET or None)

        return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form2':prove1,'form3':parametros2,'form4':parametros3,'nombre':nombre,'id_prove':prove,'suma_prove1':suma_prove1,'suma_total':suma_total})
    
    if request.method == 'GET':
        print("pregunta",pregunta)
        print("respuesta",respuesta)
        print("proveedor",prove)
        

        
        query = proveedor_respuestas.objects.all().select_related('pregunta').select_related('proveedor_id').filter(
            pregunta__id = pregunta).filter(proveedor_id__id = prove)

        nombre_empresa = proveedor.objects.get(id=prove)
        texto_pregunta = proveedor_encuesta.objects.get(id=pregunta)
        segmento_choose = proveedor_encuesta.objects.filter(id=pregunta).values('segmento').distinct()

        for valor in segmento_choose:
            segmento_escogido = valor['segmento']

        print(query)
        print(nombre_empresa)
        print(segmento_escogido)
        

        return render(request,'proveedores/edita_respuestas.html',{'form':query,'pregunta':pregunta,'respuesta':respuesta,'prove':prove,'nombre_empresa':nombre_empresa,'texto_pregunta':texto_pregunta,'segmento_escogido':segmento_escogido})

#Carga de respuestas via import de archivos Excel

# def cargaRespuestasPru(request): 
#     if request.method == 'POST':
        
#         # respuestas_resource = RespuestasResource()  
#         # dataset = Dataset()  
#         # print(dataset)    
#         nuevas_respuestas = request.FILES['xlsfile'] 
#         # imported_data = dataset.load(nuevas_respuestas.read())
#         data = pd.read_excel(nuevas_respuestas)
        
#         data_transpo = data.transpose()


#         # headers = data.columns.tolist()
#         # body = data.values.tolist()

#         # for i,j in zip(body,headers):
#         #     for k in i:
#         #         print(k,j,'\n')

#         # headers = data.columns.tolist()
#         # headers = data.tolist()
        
#         print(data_transpo)
        

#         # header = system.dataset.getColumnHeaders(imported_data)

#         # for i in imported_data:
#         #     print(i,'\n')

#         # print(imported_data)  
    
#     return render(request,'proveedores/importarpru.html')




def cargaRespuestas(request): 
    if request.method == 'POST':
        empresa = request.POST.get('nombre_empresa')
        
        #Obtiene la categoria del proveedor

        proveedor_id = proveedor.objects.get(id=request.POST.get('nombre_empresa'))

        print(proveedor_id.id)
        print(proveedor_id.categoria)

        
        respuestas_resource = RespuestasResource()  
        dataset = Dataset()  
        # print(dataset)    
        nuevas_respuestas = request.FILES['xlsfile']  
        #print(nuevas_respuestas)  
        imported_data = dataset.load(nuevas_respuestas.read())
        
        #Nuevos controles para determinar la categoría y aplicar los headers al archivo de respuestas
        

        #PRODUCTOS Y SERVICIOS VARIOS
        if proveedor_id.categoria.id == 19:
        
            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087','p088','p089']  

        #FUNDAS / ROLLOS
        elif proveedor_id.categoria.id == 4:

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087'] 
        
        #GESTORES AMBIENTALES DE DESECHO
        elif proveedor_id.categoria.id == 5:

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086'] 


        #Control de Plagas
        elif proveedor_id.categoria.id == 3:
            

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086'] 

        #Transporte Otros
        elif proveedor_id.categoria.id == 17:
            

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087'] 



        #Productos de Uso Agricola
        elif proveedor_id.categoria.id == 20:
            

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087','p088','p089','p090','p091'] 

        #Insumos de Laboratorio
        elif proveedor_id.categoria.id == 11:
            

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086'] 

        #Laboratorios De Análisis
        elif proveedor_id.categoria.id == 9:

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086'] 

        #Quimicos para el Caldero
        elif proveedor_id.categoria.id == 14:

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086'] 

        #CONSTRUCTORES
        elif proveedor_id.categoria.id == 21:
        
            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087','p088','p089']  

        #QUIMICOS DE LIMPIEZA (DESINFECTANTES)
        elif proveedor_id.categoria.id == 12:
        
            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086']  

        #CAJAS
        elif proveedor_id.categoria.id == 18:

            imported_data.headers = ['marca_temporal', 'p001','p002','p003','p004','p005','p006','p007','p008','p009','p010',
            'p011','p012','p013','p014','p015','p016','p017','p018','p019','p020','p021','p022','p023','p024','p025','p026','p027','p028','p029',
            'p030','p031','p032','p033','p034','p035','p036','p037','p038','p039','p040','p041','p042','p043','p044','p045','p046','p047','p048',
            'p049','p050','p051','p052','p053','p054','p055','p056','p057','p058','p059','p060','p061','p062','p063','p064','p065','p066','p067',
            'p068','p069','p070','p071','p072','p073','p074','p075','p076','p077','p078','p079','p080','p081','p082','p083','p084','p085','p086',
            'p087'] 

        else:
            imported_data.headers = []
            return HttpResponse("Categoria de proveedor no configurada")

        print(imported_data)
        result = respuestas_resource.import_data(imported_data, dry_run=True, raise_errors=True) # Test the data import 
        print(result) 
        print(result.row_errors())  
        if not result.has_errors():
            cursor = connection.cursor()
            cursor.execute("TRUNCATE TABLE proveedores.carga_respuestas_prod_serv_varios")  #carga_respuestas_prod_serv_varios se utiliza para cargar las respuestas de todas las categorias  
            respuestas_resource.import_data(imported_data, dry_run=False,raise_errors=True) # Actually import now  

        #Lineas para tomar la data cargada y asociarla al Proveedor
        #La tabla carga_respuestas_prod_serv_varios sirve para todas las categorias
        
        respuestas_cargadas = carga_respuestas_prod_serv_varios.objects.all()

        #ESTAS SON LAS CATEGORIAS CONFIGURADAS Y REVISADAS CON CONSTATACION DE DOCUMENTOS Y PREGUNTAS 

        #PROD Y SERV AGRICOLAS
        if proveedor_id.categoria.id == 20:
            preguntas_encuesta_q = proveedor_encuesta_prod_agricolas.objects.all().order_by('ordenamiento')
        
        #QUIMICOS PARA EL CALDERO
        elif proveedor_id.categoria.id == 14:
            preguntas_encuesta_q = proveedor_encuesta_quimicos_caldero.objects.all().order_by('ordenamiento')
        
        #LABORATORIOS DE ANALISIS
        elif proveedor_id.categoria.id == 9:
            preguntas_encuesta_q = proveedor_encuesta_laboratorio_analisis.objects.all().order_by('ordenamiento')

       #PRODUCTOS Y SERVICIOS VARIOS
        elif proveedor_id.categoria.id == 19:
            preguntas_encuesta_q = proveedor_encuesta_prod_serv_varios.objects.all().order_by('ordenamiento')
        
        #PLAGAS
        elif proveedor_id.categoria.id == 3:
            preguntas_encuesta_q = proveedor_encuesta_control_plagas.objects.all().order_by('ordenamiento')

        #FUNDAS / ROLLOS
        elif proveedor_id.categoria.id == 4:
            preguntas_encuesta_q = proveedor_encuesta_fundas_rollos.objects.all().order_by('ordenamiento')
        
        #GESTORES AMBIENTALES DE DESECHO
        elif proveedor_id.categoria.id == 5:
            preguntas_encuesta_q = proveedor_encuesta_gestores_desechos.objects.all().order_by('ordenamiento')
        
        #CONSTRUCTORES
        elif proveedor_id.categoria.id == 21:
            preguntas_encuesta_q = proveedor_encuesta_constructores.objects.all().order_by('ordenamiento')
        
        #QUIMICOS DESINFECTANTES
        elif proveedor_id.categoria.id == 12:
            preguntas_encuesta_q = proveedor_encuesta_quimicos_desinfectantes.objects.all().order_by('ordenamiento')

        #CAJAS
        elif proveedor_id.categoria.id == 18:
            preguntas_encuesta_q = proveedor_encuesta_cajas.objects.all().order_by('ordenamiento')


 #ESTAS CATEGORIAS ESTAN SUBIENDO BIEN RESPUESTAS PERO ES NECESARIO VALIDAR COHERENCIA DE BUQUEDA DE DOCUMENTOS Y PREGUNTAS PARA CALIDAD

        elif proveedor_id.categoria.id == 17:
            preguntas_encuesta_q = proveedor_encuesta_transporte_otros.objects.all().order_by('ordenamiento')

        
        elif proveedor_id.categoria.id == 11:
            preguntas_encuesta_q = proveedor_encuesta_laboratorio_insumos.objects.all().order_by('ordenamiento')

        else:
            
            return HttpResponse("Categoria de proveedor no configurada")


        preguntas_encuesta_ids = []
        for i in preguntas_encuesta_q:
            valor = i.pregunta_id
            preguntas_encuesta_ids.append(valor)
        

        datos = []
        for respuestas in respuestas_cargadas:
            p001 = respuestas.p001
            p002 = respuestas.p002
            p003 = respuestas.p003
            p004 = respuestas.p004
            p005 = respuestas.p005
            p006 = respuestas.p006
            p007 = respuestas.p007
            p008 = respuestas.p008
            p009 = respuestas.p009
            p010 = respuestas.p010
            p011 = respuestas.p011
            p012 = respuestas.p012
            p013 = respuestas.p013
            p014 = respuestas.p014
            p015 = respuestas.p015
            p016 = respuestas.p016
            p017 = respuestas.p017
            p018 = respuestas.p018
            p019 = respuestas.p019
            p020 = respuestas.p020
            p021 = respuestas.p021
            p022 = respuestas.p022
            p023 = respuestas.p023
            p024 = respuestas.p024
            p025 = respuestas.p025
            p026 = respuestas.p026
            p027 = respuestas.p027
            p028 = respuestas.p028
            p029 = respuestas.p029
            p030 = respuestas.p030
            p031 = respuestas.p031
            p032 = respuestas.p032
            p033 = respuestas.p033
            p034 = respuestas.p034
            p035 = respuestas.p035
            p036 = respuestas.p036
            p037 = respuestas.p037
            p038 = respuestas.p038
            p039 = respuestas.p039
            p040 = respuestas.p040
            p041 = respuestas.p041
            p042 = respuestas.p042
            p043 = respuestas.p043
            p044 = respuestas.p044
            p045 = respuestas.p045
            p046 = respuestas.p046
            p047 = respuestas.p047
            p048 = respuestas.p048

            p049 = respuestas.p049
            p050 = respuestas.p050
            p051 = respuestas.p051
            p052 = respuestas.p052
            p053 = respuestas.p053
            p054 = respuestas.p054
            p055 = respuestas.p055
            p056 = respuestas.p056
            p057 = respuestas.p057
            p058 = respuestas.p058
            p059 = respuestas.p059
            p060 = respuestas.p060
            p061 = respuestas.p061
            p062 = respuestas.p062
            p063 = respuestas.p063
            p064 = respuestas.p064
            p065 = respuestas.p065
            p066 = respuestas.p066
            p067 = respuestas.p067
            p068 = respuestas.p068
            p069 = respuestas.p069
            p070 = respuestas.p070
            p071 = respuestas.p071
            p072 = respuestas.p072
            p073 = respuestas.p073
            p074 = respuestas.p074
            p075 = respuestas.p075
            p076 = respuestas.p076
            p077 = respuestas.p077
            p078 = respuestas.p078
            p079 = respuestas.p079
            p080 = respuestas.p080
            p081 = respuestas.p081
            p082 = respuestas.p082
            p083 = respuestas.p083
            p084 = respuestas.p084
            p085 = respuestas.p085
            p086 = respuestas.p086
            p087 = respuestas.p087
            p088 = respuestas.p088
            p089 = respuestas.p089
            
            #SE Excluye el valor de la primera columna de nombre marca
            datos.append(p001)
            datos.append(p002)
            datos.append(p003)
            datos.append(p004)
            datos.append(p005)
            datos.append(p006)
            datos.append(p007)
            datos.append(p008)
            datos.append(p009)
            datos.append(p010)
            datos.append(p011)
            datos.append(p012)
            datos.append(p013)
            datos.append(p014)
            datos.append(p015)
            datos.append(p016)
            datos.append(p017)
            datos.append(p018)
            datos.append(p019)
            datos.append(p020)
            datos.append(p021)
            datos.append(p022)
            datos.append(p023)
            datos.append(p024)
            datos.append(p025)
            datos.append(p026)
            datos.append(p027)
            datos.append(p028)
            datos.append(p029)
            datos.append(p030)
            datos.append(p031)
            datos.append(p032)
            datos.append(p033)
            datos.append(p034)
            datos.append(p035)
            datos.append(p036)
            datos.append(p037)
            datos.append(p038)
            datos.append(p039)
            datos.append(p040)
            datos.append(p041)
            datos.append(p042)
            datos.append(p043)
            datos.append(p044)
            datos.append(p045)
            datos.append(p046)
            datos.append(p047)
            datos.append(p048)
            datos.append(p049)
            datos.append(p050)
            datos.append(p051)
            datos.append(p052)
            datos.append(p053)
            datos.append(p054)
            datos.append(p055)
            datos.append(p056)
            datos.append(p057)
            datos.append(p058)
            datos.append(p059)
            datos.append(p060)
            datos.append(p061)
            datos.append(p062)
            datos.append(p063)
            datos.append(p064)
            datos.append(p065)
            datos.append(p066)
            datos.append(p067)
            datos.append(p068)
            datos.append(p069)
            datos.append(p070)
            datos.append(p071)
            datos.append(p072)
            datos.append(p073)
            datos.append(p074)
            datos.append(p075)
            datos.append(p076)
            datos.append(p077)
            datos.append(p078)
            datos.append(p079)
            datos.append(p080)
            datos.append(p081)
            datos.append(p082)
            datos.append(p083)
            datos.append(p084)
            datos.append(p085)
            datos.append(p086)
            datos.append(p087)
            datos.append(p088)
            datos.append(p089)
        
        print(datos)
        #print("Tamaño de vector de respuestas",datos.count())
        #print("Tamaño de ids de preguntas", preguntas_encuesta_ids.count())

        xx = []
        for a,b in zip(datos,preguntas_encuesta_ids):
            v={
                'respuesta': a,
                'pregunta': b,
                'proveedor_id': proveedor_id,
            }
            xx.append(v)
        
        print(xx)
        for x in xx:
            r = proveedor_respuestas(
                respuesta=x['respuesta'],
                pregunta=x['pregunta'],
                proveedor_id=x['proveedor_id'],
            )
            r.save()
        

    #ATENCION: Sección para grabación de respuestas de constatación de documentos y preguntas para Calidad.
    
    #Graba respuestas de segmento de Catalogo
        #Productos y Servicios Varios
        if proveedor_id.categoria.id == 19:
        
            pregunta_catalogo = proveedor_encuesta.objects.get(id=145)
            # pregunta_cod_eti = proveedor_encuesta.objects.get(id=274)

            documento_catalogo = proveedor_documentos.objects.get(id=208)
            # documento_cod_eti = proveedor_documentos.objects.get(id=109)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_catalogo,
                    proveedor_id = proveedor_id,
                    documento = documento_catalogo,
                    )
                r.save()
                
                

        #Gestores Ambientales de Desecho
        elif proveedor_id.categoria.id == 5:
            pregunta_certificacion_ambiental = proveedor_encuesta.objects.get(id=201)
            pregunta_luae = proveedor_encuesta.objects.get(id=205)
            pregunta_certificado_destruccion = proveedor_encuesta.objects.get(id=203)
            pregunta_contrato = proveedor_encuesta.objects.get(id=290)
            
            documento_certificacion_ambiental  = proveedor_documentos.objects.get(id=85)
            documento_luae = proveedor_documentos.objects.get(id=86)
            documento_contrato = proveedor_documentos.objects.get(id=200)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificacion_ambiental,
                    proveedor_id = proveedor_id,
                    documento = documento_certificacion_ambiental ,
                    )
                r.save()
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_luae,
                    proveedor_id = proveedor_id,
                    documento = documento_luae ,
                    )
                r.save()
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_destruccion,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_contrato,
                    proveedor_id = proveedor_id,
                    documento = documento_contrato ,
                    )
                r.save()
            

        #Fundas/Rollos

        elif proveedor_id.categoria.id == 4:
            pregunta_certificado_migra = proveedor_encuesta.objects.get(id=102)
            pregunta_aceptacion_especificaciones = proveedor_encuesta.objects.get(id=104)
            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=106)
            pregunta_aprobaciones_contacto = proveedor_encuesta.objects.get(id=108)
            pregunta_basc = proveedor_encuesta.objects.get(id=109)
            pregunta_gfsi = proveedor_encuesta.objects.get(id=114)
            pregunta_calidad_lote = proveedor_encuesta.objects.get(id=117)
            pregunta_formato_informacion = proveedor_encuesta.objects.get(id=123)
            pregunta_contrato = proveedor_encuesta.objects.get(id=124)

            documento_aprobaciones_contacto  = proveedor_documentos.objects.get(id=7)
            documento_basc = proveedor_documentos.objects.get(id=8)
            documento_gfsi = proveedor_documentos.objects.get(id=9)
            documento_formato_informacion = proveedor_documentos.objects.get(id=12)
            documento_contrato = proveedor_documentos.objects.get(id=174)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_migra,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_aceptacion_especificaciones,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_aprobaciones_contacto,
                    proveedor_id = proveedor_id,
                    documento = documento_aprobaciones_contacto,
                    )
                r.save()
                
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_basc,
                    proveedor_id = proveedor_id,
                    documento = documento_basc,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_gfsi,
                    proveedor_id = proveedor_id,
                    documento = documento_gfsi,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_calidad_lote,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_formato_informacion,
                    proveedor_id = proveedor_id,
                    documento = documento_formato_informacion,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_contrato,
                    proveedor_id = proveedor_id,
                    documento = documento_contrato,
                    )
                r.save()

        #CAJAS

        elif proveedor_id.categoria.id == 18:
            pregunta_info_contacto = proveedor_encuesta.objects.get(id=176)
            pregunta_aceptacion_especificaciones = proveedor_encuesta.objects.get(id=161)
            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=163)
            pregunta_gfsi = proveedor_encuesta.objects.get(id=164)
            pregunta_contrato = proveedor_encuesta.objects.get(id=167)
            pregunta_basc = proveedor_encuesta.objects.get(id=169)
            pregunta_calidad_lote = proveedor_encuesta.objects.get(id=170)

            documento_contacto_emergencia  = proveedor_documentos.objects.get(id=67)
            documento_basc = proveedor_documentos.objects.get(id=64)
            documento_gfsi = proveedor_documentos.objects.get(id=63)
            documento_contrato = proveedor_documentos.objects.get(id=177)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_info_contacto,
                    proveedor_id = proveedor_id,
                    documento = documento_contacto_emergencia,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_aceptacion_especificaciones,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_basc,
                    proveedor_id = proveedor_id,
                    documento = documento_basc,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_gfsi,
                    proveedor_id = proveedor_id,
                    documento = documento_gfsi,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_calidad_lote,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_contrato,
                    proveedor_id = proveedor_id,
                    documento = documento_contrato,
                    )
                r.save()



        #Productos Agricolas
        elif proveedor_id.categoria.id == 20:

            pregunta_ingredientes = proveedor_encuesta.objects.get(id=192)
            pregunta_revision_total = proveedor_encuesta.objects.get(id=291)
            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=329)
            pregunta_documentacion_agrocalidad = proveedor_encuesta.objects.get(id=330)
            pregunta_hojas_msds = proveedor_encuesta.objects.get(id=331)
    
            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_ingredientes,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_revision_total,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_documentacion_agrocalidad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_hojas_msds,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()


        #Laboratorios de Análisis
        elif proveedor_id.categoria.id == 9:

            pregunta_iso = proveedor_encuesta.objects.get(id=195)
            documento_iso = proveedor_documentos.objects.get(id=74)

            pregunta_carta_garantia = proveedor_encuesta.objects.get(id=197)
            documento_carta_garantia = proveedor_documentos.objects.get(id=75)

            pregunta_alcance_certificacion = proveedor_encuesta.objects.get(id=334)
            documento_alcance = proveedor_documentos.objects.get(id=233)



            for x in range(1):

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_iso,
                        proveedor_id = proveedor_id,
                        documento = documento_iso,
                        )
                r.save()

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_carta_garantia,
                        proveedor_id = proveedor_id,
                        documento = documento_carta_garantia,
                        )
                r.save()

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_alcance_certificacion,
                        proveedor_id = proveedor_id,
                        documento = documento_alcance,
                        )
                r.save()



        #Insumos de Laboratorio
        elif proveedor_id.categoria.id == 11:

            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=244)
            documento_fichas_tecnicas = proveedor_documentos.objects.get(id=197)

            pregunta_hoja_msds = proveedor_encuesta.objects.get(id=245)
            documento_hoja_msds = proveedor_documentos.objects.get(id=78)

            pregunta_certificado_calidad = proveedor_encuesta.objects.get(id=247)
            documento_certificado_calidad = proveedor_documentos.objects.get(id=80)

            pregunta_certificado_entidadoficial = proveedor_encuesta.objects.get(id=248)
            documento_certificado_entidadoficial = proveedor_documentos.objects.get(id=81)

            for x in range(1):

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_fichas_tecnicas,
                        proveedor_id = proveedor_id,
                        documento = documento_fichas_tecnicas,
                        )
                r.save()
                
                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_hoja_msds,
                        proveedor_id = proveedor_id,
                        documento = documento_hoja_msds,
                        )
                r.save()
                
                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_certificado_calidad,
                        proveedor_id = proveedor_id,
                        documento = documento_certificado_calidad,
                        )
                r.save()

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta = pregunta_certificado_entidadoficial,
                        proveedor_id = proveedor_id,
                        documento = documento_certificado_entidadoficial,
                        )
                r.save()

        elif proveedor_id.categoria.id == 3:

            pregunta_certificado_npma = proveedor_encuesta.objects.get(id=130)
            documento_npma = proveedor_documentos.objects.get(id=39)

            pregunta_certificado_aecpu = proveedor_encuesta.objects.get(id=132)
            documento_aecpu = proveedor_documentos.objects.get(id=40)

            pregunta_certificado_mip = proveedor_encuesta.objects.get(id=134)
            #documento_mip = proveedor_documentos.objects.get(id=41)

            pregunta_poliza_responsabilidad = proveedor_encuesta.objects.get(id=138)
            documento_poliza_responsabilidad = proveedor_documentos.objects.get(id=43)

            pregunta_formato_contacto_emergencia = proveedor_encuesta.objects.get(id=141)
            documento_formato_contacto_emergencia = proveedor_documentos.objects.get(id=44)

            pregunta_certificado_arcsa = proveedor_encuesta.objects.get(id=143)
            documento_arcsa = proveedor_documentos.objects.get(id=45)

            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=280)
            #documento_fichas_tecnicas = proveedor_documentos.objects.get(id=179)

            pregunta_permiso_funcionamiento = proveedor_encuesta.objects.get(id=281)
            documento_permiso_func = proveedor_documentos.objects.get(id=183)

            pregunta_contrato_firmado = proveedor_encuesta.objects.get(id=283)
            documento_contrato = proveedor_documentos.objects.get(id=185)

            pregunta_hojas_de_seguridad = proveedor_encuesta.objects.get(id=282)

            pregunta_certificado_calidad = proveedor_encuesta.objects.get(id=284)

            pregunta_etiquetas_de_productos = proveedor_encuesta.objects.get(id=321)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_npma,
                    proveedor_id = proveedor_id,
                    documento = documento_npma,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_aecpu,
                    proveedor_id = proveedor_id,
                    documento = documento_aecpu,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_mip,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_poliza_responsabilidad,
                    proveedor_id = proveedor_id,
                    documento = documento_poliza_responsabilidad,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_formato_contacto_emergencia,
                    proveedor_id = proveedor_id,
                    documento = documento_formato_contacto_emergencia,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_arcsa,
                    proveedor_id = proveedor_id,
                    documento = documento_arcsa,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_permiso_funcionamiento,
                    proveedor_id = proveedor_id,
                    documento = documento_permiso_func,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_contrato_firmado,
                    proveedor_id = proveedor_id,
                    documento = documento_contrato,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_hojas_de_seguridad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_calidad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_etiquetas_de_productos,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

        #QUIMICOS PARA EL CALDERO
        elif proveedor_id.categoria.id == 14:
           
            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=159)

            pregunta_analisis_cumplimiento_fda = proveedor_encuesta.objects.get(id=255)
            documento_analisis_cumplimiento_fda = proveedor_documentos.objects.get(id=95)

            pregunta_registro_sanitario = proveedor_encuesta.objects.get(id=326)
            documento_registro_sanitario = proveedor_documentos.objects.get(id=231)

            pregunta_hojas_de_seguridad = proveedor_encuesta.objects.get(id=254)

            pregunta_certificado_calidad = proveedor_encuesta.objects.get(id=328)

            pregunta_etiquetas_de_productos = proveedor_encuesta.objects.get(id=327)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_hojas_de_seguridad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_analisis_cumplimiento_fda,
                    proveedor_id = proveedor_id,
                    documento = documento_analisis_cumplimiento_fda,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_registro_sanitario,
                    proveedor_id = proveedor_id,
                    documento = documento_registro_sanitario,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_calidad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_etiquetas_de_productos,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

        #QUIMICOS DESINFECTANTES
        elif proveedor_id.categoria.id == 12:
           
            pregunta_fichas_tecnicas = proveedor_encuesta.objects.get(id=178)
            pregunta_hojas_de_seguridad = proveedor_encuesta.objects.get(id=229)
            pregunta_ingredientes_activos = proveedor_encuesta.objects.get(id=230) 
            pregunta_certificado_reduccion = proveedor_encuesta.objects.get(id=231)
            pregunta_certificado_calidad = proveedor_encuesta.objects.get(id=233)
            pregunta_etiquetas = proveedor_encuesta.objects.get(id=324)
            
            pregunta_registro_sanitario = proveedor_encuesta.objects.get(id=325)
            documento_registro_sanitario = proveedor_documentos.objects.get(id=230)

            pregunta_contacto_emergencia = proveedor_encuesta.objects.get(id=234)
            documento_contacto_emergencia = proveedor_documentos.objects.get(id=60)

            
            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_fichas_tecnicas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_hojas_de_seguridad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_ingredientes_activos,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_registro_sanitario,
                    proveedor_id = proveedor_id,
                    documento = documento_registro_sanitario,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_calidad,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_certificado_reduccion,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_etiquetas,
                    proveedor_id = proveedor_id,
                    documento = None,
                    )
                r.save() 

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_contacto_emergencia,
                    proveedor_id = proveedor_id,
                    documento = documento_contacto_emergencia,
                    )
                r.save() 


            
        #TRANSPORTE (OTROS)    
        elif proveedor_id.categoria.id == 17:

            pregunta_documento_operacion = proveedor_encuesta.objects.get(id=270)
            documento_operacion = proveedor_documentos.objects.get(id=104)

            pregunta_matricula = proveedor_encuesta.objects.get(id=271)
            documento_matricula = proveedor_documentos.objects.get(id=105)

            pregunta_poliza_vehiculo = proveedor_encuesta.objects.get(id=272)
            documento_poliza_vehiculo = proveedor_documentos.objects.get(id=106)

            pregunta_poliza_ocupantes = proveedor_encuesta.objects.get(id=273)
            documento_poliza_ocupantes = proveedor_documentos.objects.get(id=107)

            pregunta_lista_choferes = proveedor_encuesta.objects.get(id=294)
            documento_lista_choferes = proveedor_documentos.objects.get(id=209)

            pregunta_contrato_firmado = proveedor_encuesta.objects.get(id=295)
            documento_contrato_firmado = proveedor_documentos.objects.get(id=211)

            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_documento_operacion,
                    proveedor_id = proveedor_id,
                    documento = documento_operacion,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_matricula,
                    proveedor_id = proveedor_id,
                    documento = documento_matricula,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_poliza_vehiculo,
                    proveedor_id = proveedor_id,
                    documento = documento_poliza_vehiculo,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_poliza_ocupantes,
                    proveedor_id = proveedor_id,
                    documento = documento_poliza_ocupantes,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_lista_choferes,
                    proveedor_id = proveedor_id,
                    documento = documento_lista_choferes,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta =  pregunta_contrato_firmado,
                    proveedor_id = proveedor_id,
                    documento = documento_contrato_firmado,
                    )
                r.save()

        


    #Graba respuestas vacías para documentos de Existencia y Legalidad (Seccion Existencia y Legalidad) y Codigo de Etica
        
        tipo = proveedor_det.objects.get(codigo_id=proveedor_id)
        
        tip = tipo.empresa_tipo
        print(tip)

        # for i in tipo:
        #     tip = i.empresa_tipo
        
        if tip != 'Persona Natural':
            print("Entra a la grabacion de preguntas")

        
            preg213 = proveedor_encuesta.objects.get(id=213)
            docu_ruc = proveedor_documentos.objects.get(id=25)
            
            preg99 = proveedor_encuesta.objects.get(id=99)
            docu_nombra = proveedor_documentos.objects.get(id=26)
            
            preg215 = proveedor_encuesta.objects.get(id=215)
            docu_ref = proveedor_documentos.objects.get(id=27)
            
            preg216 = proveedor_encuesta.objects.get(id=216)
            docu_iess = proveedor_documentos.objects.get(id=30)
            
            preg218 = proveedor_encuesta.objects.get(id=218)
            docu_basico = proveedor_documentos.objects.get(id=31)
            
            preg219 = proveedor_encuesta.objects.get(id=219)
            docu_cliente = proveedor_documentos.objects.get(id=32)
            
            preg125 = proveedor_encuesta.objects.get(id=125)
            docu_accionistas = proveedor_documentos.objects.get(id=38)
            
            preg217 = proveedor_encuesta.objects.get(id=217)
            docu_proveedor = proveedor_documentos.objects.get(id=226)

            pregunta_codigo_etica = proveedor_encuesta.objects.get(id=274)
            documento_codigo_etica = proveedor_documentos.objects.get(id=109)
            
            for x in range(1):
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg213,
                    proveedor_id = proveedor_id,
                    documento = docu_ruc,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg99,
                    proveedor_id = proveedor_id,
                    documento = docu_nombra,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg215,
                    proveedor_id = proveedor_id,
                    documento = docu_ref,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg216,
                    proveedor_id = proveedor_id,
                    documento = docu_iess,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg218,
                    proveedor_id = proveedor_id,
                    documento = docu_basico,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg219,
                    proveedor_id = proveedor_id,
                    documento = docu_cliente,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg125,
                    proveedor_id = proveedor_id,
                    documento = docu_accionistas,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg217,
                    proveedor_id = proveedor_id,
                    documento = docu_proveedor,
                    )
                r.save()

                r = proveedor_respuestas(
                        respuesta = None,
                        pregunta =  pregunta_codigo_etica,
                        proveedor_id = proveedor_id,
                        documento = documento_codigo_etica,
                        )
                r.save()
       
        else:


            preg27 = proveedor_encuesta.objects.get(id=27)
            docu_ruc = proveedor_documentos.objects.get(id=13)

            preg29 = proveedor_encuesta.objects.get(id=29)
            docu_bancaria = proveedor_documentos.objects.get(id=14)

            preg31 = proveedor_encuesta.objects.get(id=31)
            docu_iess = proveedor_documentos.objects.get(id=17)

            preg33 = proveedor_encuesta.objects.get(id=33)
            docu_basico = proveedor_documentos.objects.get(id=18)

            preg35 = proveedor_encuesta.objects.get(id=35)
            docu_cliente = proveedor_documentos.objects.get(id=19)

            pregunta_codigo_etica = proveedor_encuesta.objects.get(id=274)
            documento_codigo_etica = proveedor_documentos.objects.get(id=108)
        
            for x in range(1):

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg27,
                    proveedor_id = proveedor_id,
                    documento = docu_ruc,
                    )
                r.save()
            
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg29,
                    proveedor_id = proveedor_id,
                    documento = docu_bancaria,
                    )
                r.save()
            
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg31,
                    proveedor_id = proveedor_id,
                    documento = docu_iess,
                    )
                r.save()
            
                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg33,
                    proveedor_id = proveedor_id,
                    documento = docu_basico,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = preg35,
                    proveedor_id = proveedor_id,
                    documento = docu_cliente,
                    )
                r.save()

                r = proveedor_respuestas(
                    respuesta = None,
                    pregunta = pregunta_codigo_etica,
                    proveedor_id = proveedor_id,
                    documento = documento_codigo_etica,
                    )
                r.save()



        #Actualiza los datos de información general del proveedor.
        updateHead(empresa)

        
        
        return render(request,'proveedores/confirma_success_carga_respuestas.html',{'codigo_id':empresa})

    else:
        nombre_empresa = request.GET.get("nombre_empresa")
        formdet = NombreEmpresaForm(request.GET or None)

        #Sección para determinar si el proveedor ya ha contestado la encuesta de proveedores

        try:
            proveedor_id = proveedor.objects.get(id=request.GET.get('nombre_empresa'))
            evalua_responde_encuesta = proveedor_respuestas.objects.get(pregunta=1,proveedor_id=proveedor_id.id)

            fecha_responde_encuesta = evalua_responde_encuesta.respuesta

        except:
            proveedor_id = None
            evalua_responde_encuesta = None
            fecha_responde_encuesta = None

        

    return render(request,'proveedores/importar.html',{'form':formdet,'nombre_empresa':nombre_empresa,'fecha_responde_encuesta':fecha_responde_encuesta})



#Carga de Fichas Técnicas

class VerFichas(View):
    model = proveedor_fichas
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        valor = True
        fichas_pendientes = proveedor_fichas.objects.filter(proveedor=proveedor_id).order_by('nombre_ficha')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,es_ficha=True).order_by('nombre_corto')
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 
        
        categoria = proveedor.objects.all().filter(id=proveedor_id)
        
        cate = 0
        for i in categoria:
            cate = i.categoria.id
        
        tipo = proveedor_det.objects.all().filter(codigo_id=proveedor_id)
        
        tip = 0
        for i in tipo:
            tip = i.empresa_tipo

        #Gran Bifurcacion para tratar customizacion de Categoria Productos y Servicios de Uso Agrícola

        if cate != 20:

            query3 = proveedor_fichas.objects.filter(proveedor=prov_id).order_by('nombre_ficha')

            contador = 0
            cargados =[]
            lista2 =[]
            
            #Copiamos el queryset query3 de documentos solicitados en una lista
            for i in query3:
                lista2.append(str(i))
    
            dimen = len(lista2)

            #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
            for i in documentos_list:
                for j,k in zip(lista2,range(dimen)):
                    if str(j) in str(i.nombre_corto):
                        contador+=1
                        cargados.append(str(j))
                        i.nombre_corto=str(j)
                        i.save()
                        lista2.pop(k)
            
            print("Los No encontrados",lista2)
            print("Los Cargados",cargados)


            q = proveedor_fichas.objects.filter(nombre_ficha__in=cargados)
            
            try:
                r = documentos_list.first()
                fecha_carga = r.fecha_documento.strftime('%Y-%m-%d')
            except:
                r = None
                fecha_carga = None    
            
            #listfocsids es la lista que tiene los IDs de todos los documentos cargados
            listdocids = []
            for i in q:
                iddoc = i.id
                listdocids.append(iddoc)

            #listatotal es la lista que tiene los IDS de todos los documentos solicitados
            listatotal = []
            for i in query3:
                idtotal = i.id
                listatotal.append(idtotal)
            
            
            #print("LISTAENCONTRADOS",listdocids)
            #print("LISTATOTAL",listatotal)

            #noencontrados es la lista que tiene los IDS de los documentos no encontrados
            noencontrados =[]
            noencontrados = list(set(listatotal) - set(listdocids))

            print("LISTA IDS NO ENCONTRADOS",noencontrados)

            
            #Continua rutina para mostrar documentos cargados VS solicitados
            
            lista2.sort()
            print("Elementos eliminados",contador)
            
            #Comprueba si los documentos cargados se encuentran completos
            dimenfinal = len(lista2)
            
            #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

            q3 = proveedor_fichas.objects.filter(id__in=noencontrados).order_by('nombre_ficha')
            
           

            return render(self.request, 'proveedores/ver_fichas.html', {'documentos_list': documentos_list,'proveedor':empresa,'valor':valor,'fichas_pendientes':q3,'fecha_cargados':fecha_carga,'record':r})

        else:
            print("Categoria Agricola")         


       
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        fecha_carga = datetime.now()
        form = Fichas(self.request.POST, self.request.FILES)

        
        if form.is_valid():
            doc = form.save(commit=False)
            doc.proveedor_id = prove_id
            doc.es_ficha=True
            doc.fecha_documento = fecha_carga
            doc.actualiza = request.user.username
            # doc.save() 
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()

        else:
            data = {'is_valid': False}
        
        return JsonResponse(data)

def actualizaMasivoFechaFicha(request):
    proveedor_id = request.GET.get("cod")
    fecha = request.GET.get("fecha_actualiza")

    try:
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,es_ficha=True).update(
            fecha_documento=fecha,actualiza=request.user.username)
        
        messages.info(request, 'Fecha Actualizada!')
    except:
        messages.info(request, 'Ninguna Fecha Actualizada!')

    return HttpResponseRedirect(reverse('proveedores:ver_fichas', kwargs={'pk':proveedor_id}))


class VerEtiquetasProductos(View):
    model = proveedor_etiquetas_productos
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        valor = True
        etiquetas_pendientes = proveedor_etiquetas_productos.objects.filter(proveedor=proveedor_id).order_by('nombre_etiqueta_producto')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,es_etiqueta_producto=True).order_by('nombre_corto')
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 
        
        categoria = proveedor.objects.get(id=proveedor_id)
        cate = categoria.id

        tipo = proveedor_det.objects.get(codigo_id=proveedor_id)
        tip = tipo.empresa_tipo


        #Bifurcacion para tratar customizacion de Categoria Productos y Servicios de Uso Agrícola

        if cate != 20:

            query3 = proveedor_etiquetas_productos.objects.filter(proveedor=prov_id).order_by('nombre_etiqueta_producto')

            contador = 0
            cargados =[]
            lista2 =[]
            
            #Copiamos el queryset query3 de documentos solicitados en una lista
            for i in query3:
                lista2.append(str(i))
    
            dimen = len(lista2)

            #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
            for i in documentos_list:
                for j,k in zip(lista2,range(dimen)):
                    if str(j) in str(i.nombre_corto):
                        contador+=1
                        cargados.append(str(j))
                        i.nombre_corto=str(j)
                        i.save()
                        lista2.pop(k)
            
            print("Los No encontrados",lista2)
            print("Los Cargados",cargados)


            q = proveedor_etiquetas_productos.objects.filter(nombre_etiqueta_producto__in=cargados)
            
            
            #listfocsids es la lista que tiene los IDs de todos los documentos cargados
            listdocids = []
            for i in q:
                iddoc = i.id
                listdocids.append(iddoc)

            #listatotal es la lista que tiene los IDS de todos los documentos solicitados
            listatotal = []
            for i in query3:
                idtotal = i.id
                listatotal.append(idtotal)
            
            
            #print("LISTAENCONTRADOS",listdocids)
            #print("LISTATOTAL",listatotal)

            #noencontrados es la lista que tiene los IDS de los documentos no encontrados
            noencontrados =[]
            noencontrados = list(set(listatotal) - set(listdocids))

            print("LISTA IDS NO ENCONTRADOS",noencontrados)

            
            #Continua rutina para mostrar documentos cargados VS solicitados
            
            lista2.sort()
            print("Elementos eliminados",contador)
            
            #Comprueba si los documentos cargados se encuentran completos
            dimenfinal = len(lista2)
            
            #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

            q3 = proveedor_etiquetas_productos.objects.filter(id__in=noencontrados).order_by('nombre_etiqueta_producto')
            

            return render(self.request, 'proveedores/ver_etiquetas_productos.html', {'documentos_list': documentos_list,'proveedor':empresa,'valor':valor,'etiquetas_productos_pendientes':q3})

        else:
            print("Categoria Agricola")         


       
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        form = EtiquetasProductos(self.request.POST, self.request.FILES)

        
        if form.is_valid():
            doc = form.save(commit=False)
            doc.proveedor_id = prove_id
            doc.es_etiqueta_producto=True
            # doc.save() 
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()

        else:
            data = {'is_valid': False}
        
        return JsonResponse(data)





class VerHojasMSDS(View):
    model = proveedor_hojas_msds
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        valor = True
        hojas_pendientes = proveedor_hojas_msds.objects.filter(proveedor=proveedor_id).order_by('nombre_hoja_msds')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,es_hoja_msds=True).order_by('nombre_corto')
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 
        
        categoria = proveedor.objects.get(id=proveedor_id)
        cate = categoria.id

        tipo = proveedor_det.objects.get(codigo_id=proveedor_id)
        tip = tipo.empresa_tipo


        #Control para tratar customizacion de Categoria Productos y Servicios de Uso Agrícola

        if cate != 20:

            query3 = proveedor_hojas_msds.objects.filter(proveedor=prov_id).order_by('nombre_hoja_msds')

            contador = 0
            cargados =[]
            lista2 =[]
            
            #Copiamos el queryset query3 de documentos solicitados en una lista
            for i in query3:
                lista2.append(str(i))
    
            dimen = len(lista2)

            #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
            for i in documentos_list:
                for j,k in zip(lista2,range(dimen)):
                    if str(j) in str(i.nombre_corto):
                        contador+=1
                        cargados.append(str(j))
                        i.nombre_corto=str(j)
                        i.save()
                        lista2.pop(k)
            
            print("Los No encontrados",lista2)
            print("Los Cargados",cargados)


            q = proveedor_hojas_msds.objects.filter(nombre_hoja_msds__in=cargados)
            
            
            #listfocsids es la lista que tiene los IDs de todos los documentos cargados
            listdocids = []
            for i in q:
                iddoc = i.id
                listdocids.append(iddoc)

            #listatotal es la lista que tiene los IDS de todos los documentos solicitados
            listatotal = []
            for i in query3:
                idtotal = i.id
                listatotal.append(idtotal)
            
            
            #print("LISTAENCONTRADOS",listdocids)
            #print("LISTATOTAL",listatotal)

            #noencontrados es la lista que tiene los IDS de los documentos no encontrados
            noencontrados =[]
            noencontrados = list(set(listatotal) - set(listdocids))

            print("LISTA IDS NO ENCONTRADOS",noencontrados)

            
            #Continua rutina para mostrar documentos cargados VS solicitados
            
            lista2.sort()
            print("Elementos eliminados",contador)
            
            #Comprueba si los documentos cargados se encuentran completos
            dimenfinal = len(lista2)
            
            #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

            q3 = proveedor_hojas_msds.objects.filter(id__in=noencontrados).order_by('nombre_hoja_msds')
            

            return render(self.request, 'proveedores/ver_hojas_msds.html', {'documentos_list': documentos_list,'proveedor':empresa,'valor':valor,'hojas_msds_pendientes':q3})

        else:
            print("Categoria Agricola")         


       
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        form = HojasMSDS(self.request.POST, self.request.FILES)

        
        if form.is_valid():
            doc = form.save(commit=False)
            doc.proveedor_id = prove_id
            doc.es_hoja_msds=True
            # doc.save() 
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()

        else:
            data = {'is_valid': False}
        
        return JsonResponse(data)


#Carga de Fichas, Hojas de Seguridad y Documentos de Agrocalidad

class VerFichasPes(View):
    model = proveedor_fichas
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        valor = True
        fichas_pendientes = proveedor_fichas.objects.filter(proveedor=proveedor_id).order_by('nombre_ficha')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,es_ficha_pes=True).order_by('nombre_corto')
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 
        
        categoria = proveedor.objects.get(id=proveedor_id)
        cate = categoria.id

        tipo = proveedor_det.objects.get(codigo_id=proveedor_id)
        tip = tipo.empresa_tipo


        #Control para tratar customizacion de Categoria Productos y Servicios de Uso Agrícola

        #Desindentar Ctrl + ?

        query3 = proveedor_fichas.objects.filter(proveedor=prov_id,subcategoria=1).order_by('nombre_ficha')

        contador = 0
        cargados =[]
        lista2 =[]
        
        #Copiamos el queryset query3 de documentos solicitados en una lista
        for i in query3:
            lista2.append(str(i))

        dimen = len(lista2)

        #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
        for i in documentos_list:
            for j,k in zip(lista2,range(dimen)):
                if str(j) in str(i.nombre_corto):
                    contador+=1
                    cargados.append(str(j))
                    i.nombre_corto=str(j)
                    i.save()
                    lista2.pop(k)
        
        print("Los No encontrados",lista2)
        print("Los Cargados",cargados)


        q = proveedor_fichas.objects.filter(nombre_ficha__in=cargados,categoria=1)

            
            
        #listfocsids es la lista que tiene los IDs de todos los documentos cargados
        listdocids = []
        for i in q:
            iddoc = i.id
            listdocids.append(iddoc)

        #listatotal es la lista que tiene los IDS de todos los documentos solicitados
        listatotal = []
        for i in query3:
            idtotal = i.id
            listatotal.append(idtotal)
        
        
        #print("LISTAENCONTRADOS",listdocids)
        #print("LISTATOTAL",listatotal)

        #noencontrados es la lista que tiene los IDS de los documentos no encontrados
        noencontrados =[]
        noencontrados = list(set(listatotal) - set(listdocids))

        print("LISTA IDS NO ENCONTRADOS",noencontrados)

        
        #Continua rutina para mostrar documentos cargados VS solicitados
        
        lista2.sort()
        print("Elementos eliminados",contador)
        
        #Comprueba si los documentos cargados se encuentran completos
        dimenfinal = len(lista2)
        
        #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

        q3 = proveedor_fichas.objects.filter(id__in=noencontrados).order_by('nombre_ficha')
        

        return render(self.request, 'proveedores/ver_fichas_pes.html', {'documentos_list': documentos_list,'proveedor':empresa,'valor':valor,'fichas_pendientes':q3,'subcate':q})

         
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        form = Fichas(self.request.POST, self.request.FILES)

        
        if form.is_valid():
            doc = form.save(commit=False)
            doc.proveedor_id = prove_id
            doc.es_ficha_pes=True
            # doc.save() 
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()

        else:
            data = {'is_valid': False}
        
        return JsonResponse(data)

    


class VerFichasAgricolas(View):
    model = proveedor_fichas
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        subcate = kwargs['subcate_id']

        valor = True
        fichas_pendientes = proveedor_fichas.objects.filter(proveedor=proveedor_id).order_by('nombre_ficha')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id,subcategoria=subcate).order_by('nombre_corto')
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 

        categoria = proveedor.objects.all().filter(id=proveedor_id)
        
        cate = 0
        for i in categoria:
            cate = i.categoria.id
        
        tipo = proveedor_det.objects.all().filter(codigo_id=proveedor_id)
        
        tip = 0
        for i in tipo:
            tip = i.empresa_tipo

        
        query3 = proveedor_fichas.objects.filter(proveedor=prov_id,subcategoria=subcate).order_by('nombre_ficha')

        contador = 0
        cargados =[]
        lista2 =[]
        
        #Copiamos el queryset query3 de documentos solicitados en una lista
        for i in query3:
            lista2.append(str(i))

        dimen = len(lista2)

        #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
        for i in documentos_list:
            for j,k in zip(lista2,range(dimen)):
                if str(j) in str(i.nombre_corto):
                    contador+=1
                    cargados.append(str(j))
                    i.nombre_corto=str(j)
                    i.save()
                    lista2.pop(k)
        
        print("Los No encontrados",lista2)
        print("Los Cargados",cargados)


        q = proveedor_fichas.objects.filter(nombre_ficha__in=cargados)
        
        
        #listfocsids es la lista que tiene los IDs de todos los documentos cargados
        listdocids = []
        for i in q:
            iddoc = i.id
            listdocids.append(iddoc)

        #listatotal es la lista que tiene los IDS de todos los documentos solicitados
        listatotal = []
        for i in query3:
            idtotal = i.id
            listatotal.append(idtotal)
        
        
        #print("LISTAENCONTRADOS",listdocids)
        #print("LISTATOTAL",listatotal)

        #noencontrados es la lista que tiene los IDS de los documentos no encontrados
        noencontrados =[]
        noencontrados = list(set(listatotal) - set(listdocids))

        print("LISTA IDS NO ENCONTRADOS",noencontrados)

        
        #Continua rutina para mostrar documentos cargados VS solicitados
        
        lista2.sort()
        print("Elementos eliminados",contador)
        print(type(subcate))
        
        #Comprueba si los documentos cargados se encuentran completos
        dimenfinal = len(lista2)
        
        #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

        q3 = proveedor_fichas.objects.filter(id__in=noencontrados).order_by('nombre_ficha')
        
        #Q4 es un diccionario que incluye la empresa y la subcategoria
        q4 = []
        d = {
            'empresa':prov_id,
            'subcategoria':subcate,
            }
        q4.append(d) 


        return render(self.request, 'proveedores/ver_fichas_agricolas.html', {'documentos_list': documentos_list,'proveedor':empresa,'valor':valor,'fichas_pendientes':q3,'fichadocumento':q4})

#SUBCATEGORIAS AGRICOLAS (NO BORRAR): SE USAN CON LA LLAMADA A "{% url 'proveedores:ver_fichas_agricolas' nombre.codigo_id.id 1 %}"
#    1 FICHAS TECNICAS Pesticidas
#    2 FICHAS TECNICAS Foliares
#    3. FICHAS TECNICAS Fertilizantes
#    4. FICHAS TECNICAS MATERIA ORGANICA

#    13. FICHAS TECNICAS SEMILLAS

#    5. REGISTROS MAGAP PESTICIDAS
#    6. REGISTROS MAGAP Foliares
#    7. REGISTROS MAGAP Fertilizantes
#    8. REGISTROS MAGAP MAteria Orgánica
#    14. REGISTROS MAGAP SEMILLAS

#    9. MSDS PESTICIDAS
#    10. MSDS Foliares
#    11. MSDS Fertilizantes
#    12. MSDS Materia Orgánica

#    15. MSDS Semillas
#    16. Etiquetas de productos Foliares 

    # 17. Etiquetas de productos Pesticidas
    # 18. Etiquetas de productos Fertilizantes
    # 19. Etiquetas de productos MAteria Orgánica
    # 20. Etiquetas de productis Semillas

    # 21. Análisis Pesticidas
    # 22. Análsis Foliares
    # 23. Análisis Fertilizantes
    # 24. Analsiis Materia Orgánica
    # 25. Análisis Semillas

    # 26. Certificados Pesticidas
    # 27. Certificados Foliares
    # 28. Certificados Fertilizantes
    # 29. Certificados MAO
    # 30. Certificados Semillas
  
       
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        form = Fichas(self.request.POST, self.request.FILES)
        subcate = kwargs['subcate_id']

        print(type(subcate))
        #if int(subcate)==1:
        
        if subcate == '1':
        
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_ficha_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '2':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_ficha_fol=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        elif subcate == '3':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_ficha_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        

        elif subcate == '4':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_ficha_mao=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        elif subcate == '5':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_agro_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '6':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_agro_fol=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        elif subcate == '7':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_agro_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        elif subcate == '8':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_agro_mao=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '9':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_msds_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '10':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_msds_fol=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '11':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_msds_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        elif subcate == '12':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_msds_mao=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '13':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_ficha_sem=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '14':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_agro_sem=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}
        
        elif subcate == '15':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_msds_sem=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '16':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_etiqueta_foliar=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '17':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_etiqueta_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '18':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_etiqueta_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '19':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_etiqueta_mao =True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '20':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_etiqueta_sem =True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '21':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_analisis_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '22':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_analisis_fol=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '23':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_analisis_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '24':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_analisis_mao=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '25':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_analisis_sem=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '26':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_certificado_pes=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '27':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_certificado_fol=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '28':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_certificado_fer=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()
        
        elif subcate == '29':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_certificado_mao=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

        elif subcate == '30':
            if form.is_valid():
                doc = form.save(commit=False)
                doc.proveedor_id = prove_id
                doc.subcategoria = subcate
                doc.es_certificado_sem=True
                data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
                nombre_corto = doc.archivos.name
                nueva_cadena = nombre_corto.replace("_"," ")
                doc.nombre_corto=nueva_cadena
                doc.save()

            else:
                data = {'is_valid': False}

        return JsonResponse(data)


        


def verFichasCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verMSDSCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_hoja_msds=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 
    

def verEtiquetasCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_producto=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_etiquetas_single_calidad:

        return render(request,'proveedores/ver_etiquetas_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_etiquetas_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 
    

def verFichasFerCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fer=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 


def verFichasPesCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_pes=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verFichasFolCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fol=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 


def verFichasMaoCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_mao=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verFichasSemCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_sem=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_fichas_single_calidad:

        return render(request,'proveedores/ver_fichas_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_fichas_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 



def verMSDSFerCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fer=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verMSDSPesCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_pes=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verMSDSFolCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fol=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verMSDSMaoCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_mao=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verMSDSSemCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_sem=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_msds_single_calidad:

        return render(request,'proveedores/ver_msds_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_msds_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

   

def verAgrocalidadFerCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fer=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_agrocalidad_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_agrocalidad_fer_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 
    
def verAgrocalidadPesCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_pes=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_agrocalidad_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_agrocalidad_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verAgrocalidadFolCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fol=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_agrocalidad_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_agrocalidad_fol_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 


def verAgrocalidadMaoCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_mao=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_agrocalidad_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_agrocalidad_mao_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 

def verAgrocalidadSemCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_sem=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_agrocalidad_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_agrocalidad_sem_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 


def verEtiquetasPesCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_eti_pes=True).order_by('nombre_corto')
    prove_det = proveedor_det.objects.get(codigo_id=pk)

    if prove_det.observaciones_agrocalidad_single_calidad:

        return render(request,'proveedores/ver_etiquetas_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_single_calidad}) 
    
    else:
        return render(request,'proveedores/ver_etiquetas_pes_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':''}) 


#Función General para ver documentos de subcategorias Agricolas en revisión de Calidad

def verDocuSubcatesCalidad(request,pk,id):
    prove = proveedor.objects.get(id=pk)
    
   #Referencia para ID
    # 1 FICHAS TECNICAS Pesticidas
    # 2 FICHAS TECNICAS Foliares
    # 3. FICHAS TECNICAS Fertilizantes
    # 4. FICHAS TECNICAS MATERIA ORGANICA

    # 13. FICHAS TECNICAS SEMILLAS

    # 5. REGISTROS AGROCALIDAD PESTICIDAS
    # 6. REGISTROS AGROCALIDAD Foliares
    # 7. REGISTROS AGROCALIDAD Fertilizantes
    # 8. REGISTROS AGROCALIDAD MAteria Orgánica

    # 14. REGISTROS AGROCALIDAD SEMILLAS

    # 9. MSDS PESTICIDAS
    # 10. MSDS Foliares
    # 11. MSDS Fertilizantes
    # 12. MSDS Materia Orgánica

    # 15. MSDS Semillas
    # 16.Etiquetas de productos Foliares
    # 17. Etiquetas de productos Pesticidas
    # 18. Etiquetas de productos Fertilizantes
    # 19. Etiquetas de productos MAteria Orgánica
    # 20. Etiquetas de productis Semillas

    # 21. Análisis Pesticidas
    # 22. Análsis Foliares
    # 23. Análisis Fertilizantes
    # 24. Analsiis Materia Orgánica
    # 25. Análisis Semillas

    # 26. Certificados Pesticidas
    # 27. Certificados Foliares
    # 28. Certificados Fertilizantes
    # 29. Certificados MAO
    # 30. Certificados Semillas

    if id == 1:
        etiqueta = "Fichas Técnicas Pesticidas"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_pes=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_fichas_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    elif id == 2:
        etiqueta = "Fichas Técnicas Foliares"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fol=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_fichas_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 3:
        etiqueta = "Fichas Técnicas Fertilizantes"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_fichas_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 4:
        etiqueta = "Fichas Técnicas Materia Orgánica"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_fichas_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 5:
        etiqueta = "Agrocalidad Pesticidas"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_agro_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_agro_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 6:
        etiqueta = "Agrocalidad Foliares"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fol=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_agro_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_agro_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 7:
        etiqueta = "Agrocalidad Fertilizantes"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_agro_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_agro_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 8:
        etiqueta = "Agrocalidad Materia Orgánica"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_agro_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 9:
        etiqueta = "MSDS Pesticidas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_pes=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_msds_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 10:
        etiqueta = "MSDS Foliares"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fol=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_msds_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 


    elif id == 11:
        etiqueta = "MSDS Fertilizantes"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_msds_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 12:
        etiqueta = "MSDS Materia Orgánica"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_msds_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    

    elif id == 13:
        etiqueta = "Fichas Técnicas Semillas"

        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_fichas_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_fichas_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 14:
        etiqueta = "Agrocalidad Semillas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_agro_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_agro_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 15:
        etiqueta = "MSDS Semillas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_msds_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_msds_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 16:
        etiqueta = "Etiquetas de productos Foliares"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_foliar=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_etiquetas_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 17:
        etiqueta = "Etiquetas de productos Pesticidas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_pes=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_etiquetas_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 18:
        etiqueta = "Etiquetas de productos Fertilizantes"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_etiquetas_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 19:
        etiqueta = "Etiquetas de productos Materia Orgánica"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_etiquetas_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 20:
        etiqueta = "Etiquetas de productos Semillas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_etiquetas_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_etiquetas_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 21:
        etiqueta = "Análisis Pesticidas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_pes=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_analisis_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_analisis_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 22:
        etiqueta = "Análisis Foliares"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_fol=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_analisis_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_analisis_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 23:
        etiqueta = "Análisis Fertilizantes"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_analisis_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_analisis_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 24:
        etiqueta = "Análisis Materia Orgánica"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_analisis_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_analisis_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 25:
        etiqueta = "Análisis Semillas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_analisis_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_analisis_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 26:
        etiqueta = "Certificados Pesticidas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_pes=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_certificados_pes_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_certificados_pes_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 27:
        etiqueta = "Certificados Foliares"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_fol=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_certificados_fol_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_certificados_fol_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 28:
        etiqueta = "Certificados Fertilizantes"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_fer=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_certificados_fer_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_certificados_fer_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 29:
        etiqueta = "Certificados Materia Orgánica"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_mao=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_certificados_mao_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_certificados_mao_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 

    
    elif id == 30:
        etiqueta = "Certificados Semillas"
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_sem=True).order_by('nombre_corto')
        prove_det = proveedor_det.objects.get(codigo_id=pk)

        if prove_det.observaciones_certificados_sem_calidad:

            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':prove_det.observaciones_certificados_sem_calidad,'etiqueta':etiqueta,'subcate':id}) 
    
        else:
            return render(request,'proveedores/ver_docuall_subcate_agricola_calidad.html',{'form':prove,'documentos_list':documentos_list,'prove':prove.id,'registro_cal':'','etiqueta':etiqueta,'subcate':id}) 



def confirma_revision_ficha(request,id,pk):
    prove = proveedor.objects.get(id=pk)
    valor_revisa = request.POST.get("revisado")
    valor_marca = request.POST.get("marca")
    print(valor_revisa)
    print(valor_marca)

    if request.method == 'POST':
        valor_revisa = request.POST.get("revisado")
        print(valor_revisa)
        if valor_revisa:
            if valor_revisa and valor_marca:
                actualiza = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                revisado_cal=valor_revisa)
                actualiza2 = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                marca_cal=valor_marca)
            else:
                actualiza = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                    revisado_cal=valor_revisa)

            return HttpResponseRedirect(reverse('proveedores:ver_fichas_calidad',kwargs={'pk':prove.id}))

        elif valor_marca:
            if valor_marca and valor_revisa:
                actualiza = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                revisado_cal=valor_revisa)
                actualiza2 = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                marca_cal=valor_marca)
            else:
                actualiza = documentos_prove.objects.filter(id=id,proveedor=pk).update(
                    marca_cal=valor_marca)

            return HttpResponseRedirect(reverse('proveedores:ver_fichas_calidad',kwargs={'pk':prove.id}))
        



def observaciones_revision_ficha(request,pk):
    prove = proveedor.objects.get(id=pk)
    
    if request.method == 'POST':
        valor_revisa = request.POST.get("registro")   
       
        actualiza = proveedor_det.objects.filter(codigo_id=pk).update(
            observaciones_fichas_single_calidad=valor_revisa)
        
                    
        return HttpResponseRedirect(reverse('proveedores:ver_fichas_calidad',kwargs={'pk':prove.id}))

def observaciones_revision_msds(request,pk):
    prove = proveedor.objects.get(id=pk)
    
    if request.method == 'POST':
        valor_revisa = request.POST.get("registro")   
       
        actualiza = proveedor_det.objects.filter(codigo_id=pk).update(
            observaciones_msds_single_calidad=valor_revisa)
        
                    
        return HttpResponseRedirect(reverse('proveedores:ver_msds_calidad',kwargs={'pk':prove.id}))

def observaciones_revision_eti(request,pk):
    prove = proveedor.objects.get(id=pk)
    
    if request.method == 'POST':
        valor_revisa = request.POST.get("registro")   
       
        actualiza = proveedor_det.objects.filter(codigo_id=pk).update(
            observaciones_eti_single_calidad=valor_revisa)
        
                    
        return HttpResponseRedirect(reverse('proveedores:ver_etiquetas_calidad',kwargs={'pk':prove.id}))




#Función para determinar si la documentación base de un proveedor se encuentra completa y contedo de documentos cargados UP: 4-NOV-2021
def verifica_documentacion_categoria(prov_id, documentos_list, empresa, cate, tip):
    
    lista_total = proveedor_documentos.objects.filter(categoria=cate)
    # #codigo_de_etica
    # codigo_etica = proveedor_documentos.objects.filter(nombre_documento="CODIGO_DE_ETICA",empresa_tipo=tip).first()

    # lista_total = lista_docs_categoria | codigo_etica

    contador = 0
    cargados =[]
    lista =[]
        
    for i in lista_total:
        lista.append(str(i))
 
    dimen = len(lista)

    for i in documentos_list:
        for j,k in zip(lista,range(dimen)):
            if str(j) in str(i.archivos):
                contador+=1
                cargados.append(str(j))
                lista.pop(k)
        
    print("La lista final",lista)
    #print("Los Cargados",cargados)

    #Guarda marca en campo para identificar cumplimiento de documentos enviados
    if len(lista) == 0:
        provecumple = proveedor_det.objects.get(codigo_id=prov_id)
        provecumple.cumple_entrega_docs = 'SI'
        provecumple.save()
          
    else:
        provecumple = proveedor_det.objects.get(codigo_id=prov_id)
        provecumple.cumple_entrega_docs = 'NO'
        provecumple.save()

    #Consulta y guarda el numero de documentos cargados
    conteoq = documentos_prove.objects.all().filter(proveedor=prov_id)
    numcargados = conteoq.count()
    proveedor_det.objects.filter(codigo_id=prov_id).update(num_documentos_cargados=numcargados)


    return None


class BasicUploadView(View):
    model = documentos_prove
    def get(self, request, **kwargs):
        proveedor_id=kwargs['pk']
        prov_id = kwargs['pk']
        

        #documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id).filter(Q(es_ficha=None) | Q(es_ficha=False) | Q(es_hoja_msds=None) | Q(es_hoja_msds=False )).order_by('archivos')
        documentos_list = documentos_prove.objects.filter(proveedor=proveedor_id).exclude(Q(es_ficha=True)
         | Q(es_hoja_msds=True)| Q(es_etiqueta_producto=True)| Q(es_ficha_pes=True) | Q(es_ficha_fol=True)
         | Q(es_ficha_fer=True) | Q(es_ficha_mao=True) | Q(es_ficha_sem=True) | Q(es_agro_pes=True) 
         | Q(es_agro_fol=True) | Q(es_agro_fer=True) | Q(es_agro_mao=True) | Q(es_agro_sem=True) 
         | Q(es_msds_pes=True) | Q(es_msds_fol=True) | Q(es_msds_fer=True) | Q(es_msds_mao=True)
         | Q(es_msds_sem=True) | Q(es_etiqueta_foliar=True) | Q(es_etiqueta_pes=True ) | Q(es_etiqueta_fer=True) 
         | Q(es_etiqueta_mao=True) | Q(es_etiqueta_sem=True) | Q(es_certificado_pes=True) | Q(es_certificado_fol=True)
         | Q(es_certificado_fer=True) | Q(es_certificado_mao=True) | Q(es_certificado_sem=True)
         | Q(es_analisis_pes=True) | Q(es_analisis_fol=True) | Q(es_analisis_fer=True) | Q(es_analisis_mao=True) | Q(es_analisis_sem=True)).order_by('archivos')
        

        cabecera = documentos_prove.objects.filter(proveedor=proveedor_id).exclude(Q(es_ficha=True) | Q(es_hoja_msds=True)| Q(es_etiqueta_producto=True)).filter(
            Q(archivos__icontains='RUC') | Q(archivos__icontains='IESS') | Q(
                archivos__icontains='NOMBRAMIENTO_REP')).order_by('-archivos')
        
        empresa = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id=proveedor_id) 
        
        categoria = proveedor.objects.all().filter(id=proveedor_id)
        
        cate = 0
        for i in categoria:
            cate = i.categoria.id
        
        tipo = proveedor_det.objects.all().filter(codigo_id=proveedor_id)
        
        tip = 0
        for i in tipo:
            tip = i.empresa_tipo

        #Control para tratar customizacion de Categoria Productos y Servicios de Uso Agrícola

        if cate != 20:


            query = proveedor_documentos.objects.filter(categoria=cate).exclude(proveedor__isnull=False)
            query2 = proveedor_documentos.objects.filter(empresa_tipo=tip)
            query4 = proveedor_documentos.objects.filter(proveedor=prov_id,nombre_documento='FICHAS_TECNICAS')
            

            #query3 tiene la lista de documentos completa solicitada
            query3 = query | query2 | query4
            query3 = query3.order_by('empresa_tipo')

            print("Query3",query3)

            contador = 0
            cargados =[]
            lista2 =[]
            
            #Copiamos el queryset query3 de documentos solicitados en una lista
            for i in query3:
                lista2.append(str(i))
    
            dimen = len(lista2)

            #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
            for i in documentos_list:
                for j,k in zip(lista2,range(dimen)):
                    if str(j) in str(i.archivos):
                        contador+=1
                        cargados.append(str(j))
                        i.nombre_corto=str(j)
                        i.save()
                        lista2.pop(k)
            
            print("La lista final Lista2",lista2)
            print("Los Cargados",cargados)


            #Marca con "SI" en tabla de respuestas si los documentos han sido cargados

            q1 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,categoria=cate)
            
            #Incluir el Codigo de Etica
            if tip =='Persona Natural':
                # q2 = proveedor_documentos.objects.filter(id=108)
                q2 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,empresa_tipo='Persona Natural')
            else:
                q2 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,empresa_tipo='Persona Jurídica')

            #q es el query de todos los documentos solicitados
            #q1 es el query de los documentos solicitados de la categoria 
            #q2 es el query de los documentos solicitados de persona natural o jurìdica
            q = q1 | q2

            # print("Hola Q1",q1)
            print("Hola Q2",q2)
            
            #listfocsids es la lista que tiene los IDs de todos los documentos cargados
            listdocids = []
            for i in q:
                iddoc = i.id
                listdocids.append(iddoc)

            #listatotal es la lista que tiene los IDS de todos los documentos solicitados
            listatotal = []
            for i in query3:
                idtotal = i.id
                listatotal.append(idtotal)
            
            
            print("LISTAENCONTRADOS",listdocids)
            #print("LISTATOTAL",listatotal)

            #noencontrados es la lista que tiene los IDS de los documentos no encontrados
            noencontrados =[]
            noencontrados = list(set(listatotal) - set(listdocids))

            print("LISTA IDS NO ENCONTRADOS",noencontrados)

            
            #Con esta linea se actualizan las respuestas de los documentos encontrados

            curioso = proveedor_respuestas.objects.filter(proveedor_id=prov_id,documento__in=listdocids)
            print("Curioso",curioso)
            queryup = proveedor_respuestas.objects.filter(proveedor_id=prov_id,documento__in=listdocids).update(respuesta='SI')
        
            #Con esta linea se actualizan las respuestas de los documentos encontrados
            queryup = proveedor_respuestas.objects.filter(proveedor_id=prov_id,documento__in=noencontrados).update(respuesta='NO')
            
            
            #FIN MArca "SI"
            
            #Continua rutina para mostrar documentos cargados VS solicitados
            
            lista2.sort()
            print("Elementos eliminados",contador)
            
            #Comprueba si los documentos cargados se encuentran completos
            dimenfinal = len(lista2)
            
            #Esta seccion sirve para marcar como Verdadero en el modelo documentos_prove a aquellos documentos cargados que son de la categoria

            q4 = proveedor_documentos.objects.filter(categoria=cate).exclude(nombre_documento='FICHAS_TECNICAS')
            contador = 0
            cargados_cate =[]
            lista_categoria =[]
            
            #Copiamos el queryset query que tiene la lista de documentos de la categoria en una lista
            for i in q4:
                lista_categoria.append(str(i))
    
            dimen_categoria = len(lista_categoria)

            for i in documentos_list:
                for j,k in zip(lista_categoria,range(dimen_categoria)):
                    if str(j) in str(i.archivos):
                        contador+=1
                        cargados_cate.append(str(j))
                        i.es_doc_categoria=True
                        i.save()
                        
            #Llama a función para determinar si la documentación de la categoría está completa

            verifica_documentacion_categoria(prov_id, documentos_list, empresa, cate, tip)


            #Esta secciòn sirve para controlar los documentos como FICHAS que se pueden marcar como completos
            #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

            q3 = proveedor_documentos.objects.filter(id__in=noencontrados).order_by('nombre_documento')
            
            q4 = []
            for i in q3:
                prov_id
                d = {
                    'id':i.id,
                    'nombre_documento':i.nombre_documento,
                    'categoria':i.categoria,
                    'completo':i.completo,
                    'id_prov':prov_id,               
                }
                q4.append(d) 
        

            #Saca la cabecera de 3 documentos RUC, Nobramiento y Cumplimiento Obligaciones IESS

            documentos_listf = documentos_list.difference(cabecera).order_by('archivos')

            #Saca la nota 
            v = proveedor.objects.filter(id=proveedor_id)

            for nota in v:
                califica_num = nota.calificacion
                califica = nota.grade
                catego = nota.categoria


            return render(self.request, 'proveedores/ver_documentos.html', {'documentos_list': documentos_listf,'cabecera':cabecera,'proveedor':empresa,'lista_solicitada':q4,'cate':cate,'catego':catego,'nota':califica_num,'grade':califica})

        else:
            try:

                respuesta_pes = proveedor_respuestas.objects.filter(proveedor_id=prov_id).get(pregunta=297)
                respuesta_fol = proveedor_respuestas.objects.filter(proveedor_id=prov_id).get(pregunta=298)
                respuesta_fer = proveedor_respuestas.objects.filter(proveedor_id=prov_id).get(pregunta=299)
                respuesta_mao = proveedor_respuestas.objects.filter(proveedor_id=prov_id).get(pregunta=300)
                respuesta_sem = proveedor_respuestas.objects.filter(proveedor_id=prov_id).get(pregunta=319)

            except:
                respuesta_pes='NO'
                respuesta_fol='NO'
                respuesta_fer='NO'
                respuesta_mao='NO'
                respuesta_sem='NO'

            #print(respuesta_pes,type(respuesta_pes))
            if str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria=1)

            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,2])

            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,2,3])

            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,2,3,4])
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria=2)
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria=3)
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria=4)
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria=5)

            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,4])
        
            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,3])
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[2,4])
        
            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'NO' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,2,4])
        
            elif str(respuesta_pes) == 'SI' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[1,3,4])
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'NO':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[2,3])
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'NO' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[3,4])
        
            elif str(respuesta_pes) == 'NO' and str(respuesta_fol) == 'SI' and str(respuesta_fer) == 'SI' and str(respuesta_mao) == 'SI':
                query = proveedor_documentos.objects.filter(categoria=cate).filter(sub_categoria__in=[2,3,4])
        

            query2 = proveedor_documentos.objects.filter(empresa_tipo=tip)
            #query4 = proveedor_documentos.objects.filter(proveedor=prov_id,nombre_documento='FICHAS_TECNICAS')
            

            #query3 tiene la lista de documentos completa solicitada
            query3 = query | query2 
            query3 = query3.order_by('empresa_tipo')

            contador = 0
            cargados =[]
            lista2 =[]
            
            #Copiamos el queryset query3 de documentos solicitados en una lista
            for i in query3:
                lista2.append(str(i))
    
            dimen = len(lista2)

            #Eliminamos las ocurrencias de documentos ya cargados en lista2. Lista 2 tendrpa al final la lista de documentos no encontrados
            for i in documentos_list:
                for j,k in zip(lista2,range(dimen)):
                    if str(j) in str(i.archivos):
                        contador+=1
                        cargados.append(str(j))
                        i.nombre_corto=str(j)
                        i.save()
                        lista2.pop(k)
            
            print("La lista final Lista2",lista2)
            print("Los Cargados",cargados)


            #Marca con "SI" en tabla de respuestas si los documentos han sido cargados

            q1 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,categoria=cate)
            
            #Incluir el Codigo de Etica
            if tip =='Persona Natural':
                # q2 = proveedor_documentos.objects.filter(id=108)
                q2 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,empresa_tipo='Persona Natural')
            else:
                q2 = proveedor_documentos.objects.filter(nombre_documento__in=cargados,empresa_tipo='Persona Jurídica')

            #q es el query de todos los documentos solicitados
            #q1 es el query de los documentos solicitados de la categoria 
            #q2 es el query de los documentos solicitados de persona natural o jurìdica
            q = q1 | q2

            # print("Hola Q1",q1)
            print("Hola Q2",q2)
            
            #listfocsids es la lista que tiene los IDs de todos los documentos cargados
            listdocids = []
            for i in q:
                iddoc = i.id
                listdocids.append(iddoc)

            #listatoral es la lista que tiene los IDS de todos los documentos solicitados
            listatotal = []
            for i in query3:
                idtotal = i.id
                listatotal.append(idtotal)
            
            
            #print("LISTAENCONTRADOS",listdocids)
            #print("LISTATOTAL",listatotal)

            #noencontrados es la lista que tiene los IDS de los documentos no encontrados
            noencontrados =[]
            noencontrados = list(set(listatotal) - set(listdocids))

            print("LISTA IDS NO ENCONTRADOS",noencontrados)

            
            #Con esta linea se actualizan las respuestas de los documentos encontrados
            queryup = proveedor_respuestas.objects.filter(proveedor_id=prov_id,documento__in=listdocids).update(respuesta='SI')
        
            #Con esta linea se actualizan las respuestas de los documentos encontrados
            queryup = proveedor_respuestas.objects.filter(proveedor_id=prov_id,documento__in=noencontrados).update(respuesta='NO')
            
            
            #FIN MArca "SI"
            
            #Continua rutina para mostrar documentos cargados VS solicitados
            
            lista2.sort()
            print("Elementos eliminados",contador)
            
            #Comprueba si los documentos cargados se encuentran completos
            dimenfinal = len(lista2)
            
            #Esta seccion sirve para marcar como Verdadero en el modelo documentos_prove a aquellos documentos cargados que son de la categoria

            q4 = proveedor_documentos.objects.filter(categoria=cate).exclude(nombre_documento='FICHAS_TECNICAS')
            contador = 0
            cargados_cate =[]
            lista_categoria =[]
            
            #Copiamos el queryset query que tiene la lista de documentos de la categoria en una lista
            for i in q4:
                lista_categoria.append(str(i))
    
            dimen_categoria = len(lista_categoria)

            for i in documentos_list:
                for j,k in zip(lista_categoria,range(dimen_categoria)):
                    if str(j) in str(i.archivos):
                        contador+=1
                        cargados_cate.append(str(j))
                        i.es_doc_categoria=True
                        i.save()
                        
            #Llama a función para determinar si la documentación de la categoría está completa

            verifica_documentacion_categoria(prov_id, documentos_list, empresa, cate, tip)

            #Esta secciòn sirve para controlar los documentos como FICHAS que se pueden marcar como completos
            #Se transforma a query lo que queda en lista2 que son los documntos pendientes de entrega

            q3 = proveedor_documentos.objects.filter(id__in=noencontrados).order_by('sub_categoria','nombre_documento')
            
            q4 = []
            for i in q3:
                prov_id
                d = {
                    'id':i.id,
                    'nombre_documento':i.nombre_documento,
                    'categoria':i.categoria,
                    'completo':i.completo,
                    'id_prov':prov_id,               
                }
                q4.append(d) 
        

            #Saca la cabecera de 3 documentos RUC, Nobramiento y Cumplimiento Obligaciones IESS

            documentos_listf = documentos_list.difference(cabecera).order_by('archivos')

            #Saca la nota 
            v = proveedor.objects.filter(id=proveedor_id)

            for nota in v:
                califica_num = nota.calificacion
                califica = nota.grade
                catego = nota.categoria

            return render(self.request, 'proveedores/ver_documentos.html', {'documentos_list': documentos_listf,'cabecera':cabecera,'proveedor':empresa,'lista_solicitada':q4,'cate':cate,'pes':respuesta_pes,'fol':respuesta_fol,'fer':respuesta_fer,'mao':respuesta_mao,'sem':respuesta_sem,'catego':catego,'nota':califica_num,'grade':califica})




   
    def post(self, request, **kwargs):

        prove_id = kwargs['pk']  
        form = Documentos(self.request.POST, self.request.FILES)
        
        print(form)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.proveedor_id = prove_id
            #doc.save() 
            data = {'is_valid': True, 'name': doc.archivos.name, 'url': doc.archivos.url}
            nombre_corto = doc.archivos.name
            nueva_cadena = nombre_corto.replace("_"," ")
            doc.nombre_corto=nueva_cadena
            doc.save()
        
        else:
            data = {'is_valid': False}
        return JsonResponse(data)
            


def render_template_x_tipo_proveedor(request):
    
    form = RegistroFormP()
    formdet = RegistroFormDetP(request.GET or None)
    parametros = RegistroFormDetP(request.GET or None) 

    return render(request,'proveedores/nuevo_registro_rollos_fundas.html',{'form':form,'form2':formdet,'form3':parametros})


def render_template_x_proveedor(request):
    proveedor = request.GET.get("nombre_empresa")
    
    formdet = NombreEmpresaForm(request.GET or None)
    
    return render(request,'proveedores/importar.html',{'form':formdet})


@login_required
def listar_proveedores(request):
     queryset = request.GET.get("buscar")
     print(queryset) 
     if queryset:

        proveedoresq = proveedor_det.objects.all().select_related("codigo_id").filter(
            Q(codigo_id__nombre_empresa__icontains = queryset) |
            Q(codigo_id__nombre_comercial__icontains = queryset) |
            Q(codigo_id__ruc = queryset) |
            Q(giro_negocio__icontains = queryset)).order_by('-codigo_id__fecha_modifica')
            

        paginator = Paginator(proveedoresq, 25)
        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
            
     else:

        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        proveedoresq = proveedor_det.objects.all().select_related("codigo_id").order_by('-codigo_id__fecha_modifica')
       
        paginator = Paginator(proveedoresq, 25)

        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
    
     return render(request,'proveedores/busqueda_proveedores.html',{'form':proveedores})


@login_required
def NotificaSuccess(request):
    return render(request,'proveedores/notifica_success.html')

@login_required
def ConfirmaSuccessCarga(request):
    return render(request,'proveedores/confirma_success_carga_respuestas.html')

@login_required
def EnviarACalidad(request,pk):
    
    prove_nombre = proveedor.objects.get(id=pk)
    categoria = prove_nombre.categoria

    if request.method == 'GET':
        actualiza = proveedor_det.objects.get(codigo_id__id=pk)
        if actualiza.enviado_a_calidad == True:
            print("Si reconoce")
            mensaje = True
            print(mensaje)
        
            return render(request,'proveedores/enviar_a_calidad.html',{'prove_nombre':prove_nombre,'mensaje':mensaje})

        return render(request,'proveedores/enviar_a_calidad.html',{'prove_nombre':prove_nombre})

    elif request.method == 'POST':
        msj_reenvio = request.POST.get("comentarios-admin")
        actualiza = proveedor_det.objects.get(codigo_id__id=pk)
        if actualiza.enviado_a_calidad == True:
            actualiza = proveedor_det.objects.filter(codigo_id__id=pk).update(observaciones_administrativas=msj_reenvio)
            

        fecha = datetime.now()

        actualiza = proveedor_det.objects.filter(codigo_id__id=pk).update(enviado_a_calidad=True,fecha_enviado_a_calidad=fecha)

        #Envío de notificación a Calidad
            
        envioMailParaRevisionCalidad('Solicitud de Revisión de Documentación de Proveedor', 'auditoria@ecofroz.com,gerencia.administrativa@ecofroz.com,dmencias@ecofroz.com,asistente.administrativa@ecofroz.com', 'email_revision_proveedor_calidad.html', prove_nombre, categoria, msj_reenvio,'')
        #envioMailParaRevisionCalidad('Solicitud de Revisión de Documentación de Proveedor', 'dmencias@ecofroz.com', 'email_revision_proveedor_calidad.html', prove_nombre, categoria, msj_reenvio,'')


        return redirect('proveedores:notifica_success')


def listar_proveedores_para_adqui(request):
     queryset = request.GET.get("buscar")
     print(queryset) 
     if queryset:

        proveedoresq = proveedor_det.objects.all().select_related("codigo_id").filter(
            Q(codigo_id__nombre_empresa__icontains = queryset) |
            Q(codigo_id__nombre_comercial__icontains = queryset) |
            Q(codigo_id__ruc = queryset) |
            Q(giro_negocio__icontains = queryset)).order_by('-codigo_id__fecha_modifica')
            

        
        paginator = Paginator(proveedoresq, 25)
        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
            
     else:

        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        proveedoresq = proveedor_det.objects.all().select_related("codigo_id").order_by('-codigo_id__fecha_modifica')
       
        paginator = Paginator(proveedoresq, 25)

        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
    
     return render(request,'proveedores/para_adqui.html',{'form':proveedores})

@login_required
def listar_proveedores_para_calidad(request):
     queryset = request.GET.get("buscar")
     print(queryset) 
     if queryset:

        proveedoresq = proveedor_det.objects.all().select_related("codigo_id").filter(enviado_a_calidad=True).filter(
            Q(codigo_id__nombre_empresa__icontains = queryset) |
            Q(codigo_id__nombre_comercial__icontains = queryset) |
            Q(codigo_id__ruc = queryset) |
            Q(giro_negocio__icontains = queryset)).order_by('-codigo_id__fecha_modifica')
            

        paginator = Paginator(proveedoresq, 25)
        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
            
     else:

        # activosx = detalle_desc_activo.objects.all().select_related("desc_activo_codigo").filter(id__lte=50)
        proveedoresq = proveedor_det.objects.filter(enviado_a_calidad=True).select_related(
            "codigo_id").order_by('-fecha_enviado_a_calidad')
       
        paginator = Paginator(proveedoresq, 25)

        page = request.GET.get('page')
        proveedores = paginator.get_page(page)
    
     return render(request,'proveedores/listar_proveedores_para_calidad.html',{'form':proveedores})

@login_required
def nuevo_registro(request):

    if request.method == 'POST':
        form = RegistroFormP(request.POST)
        formdet = RegistroFormDetP(request.POST,request.FILES)

        if form.is_valid() and formdet.is_valid():
        # if form.is_valid():
            registro = formdet.save(commit=False)
            registro.codigo_id = form.save()
            # registro = form.save()
            registro.save()
            
            cod = proveedor.objects.all().order_by('-fecha_registro').first()

            califica_ecofroz(cod.id,None,None,None,None)

            # return HttpResponse('registro_guardado')
            return redirect('proveedores:listar_proveedores')

        else:
            return HttpResponse("No se pudo guardar. RUC ya existe")

    else:
        form = RegistroFormP()
        formdet = RegistroFormDetP()
        
        return render(request,'proveedores/nuevo_registroV2.html', {'form':form, 'form2':formdet})


def califica_ecofroz(pk,respuesta1,respuesta2,respuesta3,respuesta4):
    print("Llegó")
    print(pk)
    print(respuesta1)
    print(respuesta2)
    print(respuesta3)
    print(respuesta4)


    #    #IDS de preguntas de retroalimentaciòn Ecofroz
    #    #ID 206 ¿El proveedor encuestado cumple muy frecuentemente con el procedimiento de facturacion?
    #    #ID 207 ¿El proveedor encuestado cumple muy frecuentemente con el procedimiento de entrega?
    #    #ID 208 ¿El proveedor encuestado cumple muy frecuentemente con el procedimiento de compra?
    #    #ID 209 De acuerdo con la normativa BASC,  se ha verificado antecedentes del  Representante Legal de la Empresa ?

        
    q206 = proveedor_encuesta.objects.get(id=206)
    q207 = proveedor_encuesta.objects.get(id=207)
    q208 = proveedor_encuesta.objects.get(id=208)
    q209 = proveedor_encuesta.objects.get(id=209)

    id_prove = proveedor.objects.get(id=pk)

    respuesta206 = proveedor_respuestas.objects.filter(pregunta=206,proveedor_id=pk)
    respuesta207 = proveedor_respuestas.objects.filter(pregunta=207,proveedor_id=pk)
    respuesta208 = proveedor_respuestas.objects.filter(pregunta=208,proveedor_id=pk)
    respuesta209 = proveedor_respuestas.objects.filter(pregunta=209,proveedor_id=pk)
    
    
    if respuesta206 or respuesta207 or respuesta208 or respuesta209:
        respuesta206 = proveedor_respuestas.objects.filter(pregunta=206,proveedor_id=pk).update(respuesta=respuesta1)
        respuesta207 = proveedor_respuestas.objects.filter(pregunta=207,proveedor_id=pk).update(respuesta=respuesta2)
        respuesta208 = proveedor_respuestas.objects.filter(pregunta=208,proveedor_id=pk).update(respuesta=respuesta3)
        respuesta209 = proveedor_respuestas.objects.filter(pregunta=209,proveedor_id=pk).update(respuesta=respuesta4)

    else:
        r = proveedor_respuestas(
            respuesta = respuesta1,
            proveedor_id = id_prove,
            pregunta = q206,
            )
        r.save()
    
        r = proveedor_respuestas(
            respuesta = respuesta2,
            proveedor_id = id_prove,
            pregunta = q207,
            )
        r.save()
        
        r = proveedor_respuestas(
            respuesta = respuesta3,
            proveedor_id = id_prove,
            pregunta = q208,
            )
        r.save()
        
        r = proveedor_respuestas(
            respuesta = respuesta4,
            proveedor_id = id_prove,
            pregunta = q209,
            )
        r.save()


    #dentifica del Rango de calificacion A, B, C, D ó F si no tiene calificacion

    query1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))
    

    calculaCalificacion(query1)

    query = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))

    
    suma_notausointerno = 1
    for calif in query:
        if calif.pregunta.id == 206:
            print("1zzzz")
            if calif.calificacion == 2:
                print("2zzz")
                suma_notausointerno+=2
            else:
                None
        
        elif calif.pregunta.id == 207:
            print("1yyyy")
            if calif.calificacion == 2:
                print("syyy")
                suma_notausointerno+=2
            else:
                None
        
        elif calif.pregunta.id == 208:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None

        elif calif.pregunta.id == 209:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None
        else:
            print("No hay la pregunta")

    print('Aqui la nota final',suma_notausointerno)

    
    if suma_notausointerno == 9:
        gradev = 'A'
    if suma_notausointerno >6 and suma_notausointerno <9:
        gradev = 'B'
    if suma_notausointerno >1 and suma_notausointerno <7:
        gradev = 'C'
    if suma_notausointerno == 1:
        gradev = 'F'
    
   
    consulta2 = proveedor.objects.get(id=pk)
    consulta2.grade = gradev
    consulta2.save()
    
    return None


#Vista para ver respuestas a preguntas
@login_required
def ver_proceso_calificacion(request):
    categoriaf = request.GET.get("categoria_proveedor")
    nombre_proveedor = request.GET.get("nombre_empresa")
    calf_minima = request.GET.get("calificacionminima")
    calf_maxima = request.GET.get("calificacionmaxima")
    encuesta = request.GET.get("tiene_encuesta")
    busqueda = request.GET.get("buscar") 
    categoriacomp = request.GET.get("categoria")
    criticos = request.GET.get("criticos")

    if categoriaf:
            
        proveedores = proveedor_det.objects.all().filter(categoria_proveedor=categoriaf).select_related("codigo_id")
        preguntas = proveedor_encuesta.objects.all().order_by('id')
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__categoria=categoriaf).order_by('-codigo_id__calificacion','-num_trabajadores_fijos')

        conteo = query.count()

        print(conteo)

        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        
   
        return render(request,'proveedores/ver_proceso_calificacion.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query}) 


    elif nombre_proveedor:

        preguntas = proveedor_encuesta.objects.all().order_by('id')
           
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__id=nombre_proveedor).order_by('-codigo_id__calificacion')


        conteo = query.count()
       
      
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = ParaExportarExcelForm(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        

        return render(request,'proveedores/ver_proceso_calificacion.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'form5':parametros3,'preguntas':preguntas,'count':conteo,'query':query}) 

    elif criticos:

        preguntas = proveedor_encuesta.objects.all().order_by('id')
           
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__proveedor_critico=True).order_by('-codigo_id__calificacion')


        conteo = query.count()
       
      
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = ParaExportarExcelForm(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        

        return render(request,'proveedores/ver_proceso_calificacion.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'form5':parametros3,'preguntas':preguntas,'count':conteo,'query':query}) 



    formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None)

    preguntas = proveedor_encuesta.objects.all().order_by('id')

    query = proveedor_det.objects.all().select_related(
        'codigo_id').order_by('-codigo_id__calificacion','-num_trabajadores_fijos')

        
    conteo = query.count()

    parametros = FiltrarForm(request.GET or None) 
    parametros2 = FiltrarFormDet(request.GET or None) 
    parametros3 = ParaExportarExcelForm(request.GET or None) 
    parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)

    return render(request,'proveedores/ver_proceso_calificacion.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'form4':parametros3,'preguntas':preguntas,'count':conteo,'query':query,'busqueda':busqueda,'encuesta':encuesta}) 


#Vista para ver respuestas a preguntas
@login_required
def ver_preguntas(request):
    categoriaf = request.GET.get("categoria_proveedor")
    nombre_proveedor = request.GET.get("nombre_empresa")
    calf_minima = request.GET.get("calificacionminima")
    calf_maxima = request.GET.get("calificacionmaxima")
    encuesta = request.GET.get("tiene_encuesta")
    busqueda = request.GET.get("buscar") 
    categoriacomp = request.GET.get("categoria")
    criticos = request.GET.get("criticos")

    print(criticos)
      
    if encuesta:

            if encuesta and busqueda:
            
                
                preguntas = 1
                
                query = proveedor_det.objects.all().select_related(
                        'codigo_id').filter(
                                giro_negocio__icontains=busqueda).order_by(
                                    '-codigo_id__calificacion')

                conteo = query.count()

                parametros = FiltrarForm(request.GET or None) 
                parametros2 = FiltrarFormDet(request.GET or None) 
                parametros3 = FiltrarFormSegmento(request.GET or None) 
            
                return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'form4':parametros3,'preguntas':preguntas,'count':conteo,'query':query,'busqueda':busqueda,'encuesta':encuesta}) 

    

            elif encuesta and calf_minima:
                if encuesta and calf_maxima and calf_minima:
                    calf_max = int(calf_maxima)
                    calf_min = int(calf_minima)
                    
                    preguntas = 1
                
                    query = proveedor_det.objects.all().select_related(
                        'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                            codigo_id__calificacion__lte=calf_max).filter(
                                codigo_id__calificacion__gte=calf_min).order_by(
                                    '-codigo_id__calificacion')

                    conteo = query.count()

                    parametros = FiltrarForm(request.GET or None) 
                    parametros2 = FiltrarFormDet(request.GET or None) 
            
                    return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'calf_max':calf_max,'calf_min':calf_min}) 

                
                calf_min = int(calf_minima)
                preguntas = 1
            
                query = proveedor_det.objects.all().select_related(
                    'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                        codigo_id__calificacion__gte=calf_min).order_by('-codigo_id__calificacion')

                conteo = query.count()

                parametros = FiltrarForm(request.GET or None) 
                parametros2 = FiltrarFormDet(request.GET or None) 
        
                return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'calf_min':calf_min}) 

            if encuesta and calf_maxima:
                if encuesta and calf_maxima and calf_minima:
                    calf_max = int(calf_maxima)
                    calf_min = int(calf_minima)
                    
                    preguntas = 1
                
                    query = proveedor_det.objects.all().select_related(
                        'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                            codigo_id__calificacion__lte=calf_max).filter(
                                codigo_id__calificacion__gte=calf_min).order_by(
                                    '-codigo_id__calificacion')

                    conteo = query.count()

                    parametros = FiltrarForm(request.GET or None) 
                    parametros2 = FiltrarFormDet(request.GET or None) 
            
                    return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'calf_max':calf_max,'calf_min':calf_min}) 

            if encuesta and categoriaf:
                if encuesta and categoriaf and calf_minima:
                    calf_min = int(calf_minima)
                    
                    preguntas = 1
                
                    query = proveedor_det.objects.all().select_related(
                        'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                            codigo_id__categoria=categoriaf).filter(
                                codigo_id__calificacion__gte=calf_min).order_by(
                                    '-codigo_id__calificacion')

                    conteo = query.count()

                    parametros = FiltrarForm(request.GET or None) 
                    parametros2 = FiltrarFormDet(request.GET or None) 
            
                    return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'categoria':categoriaf,'calf_min':calf_min}) 


                
                preguntas = 1
            
                query = proveedor_det.objects.all().select_related(
                    'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                        codigo_id__categoria=categoriaf).order_by('-codigo_id__calificacion')

                conteo = query.count()

                parametros = FiltrarForm(request.GET or None) 
                parametros2 = FiltrarFormDet(request.GET or None) 

                return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'categoria':categoriaf}) 

        
            else:

                preguntas = 1
                
                query = proveedor_det.objects.all().select_related(
                    'codigo_id').filter(codigo_id__respondio_encuesta=True ).filter(
                        codigo_id__categoria=categoriaf).order_by('-codigo_id__calificacion')

                conteo = query.count()

                parametros = FiltrarForm(request.GET or None) 
                parametros2 = FiltrarFormDet(request.GET or None) 
        
                return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'categoria':categoriaf}) 


    elif busqueda:
        preguntas = 1
                
        query = proveedor_det.objects.all().select_related(
            'codigo_id').filter(
                giro_negocio__icontains=busqueda).order_by(
                    '-codigo_id__calificacion','num_trabajadores_fijos')


        conteo = query.count()

        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
            
        return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'busqueda':busqueda}) 


    
    elif categoriaf:
        
        print(categoriaf)
        print("Entro por aqui DMP")
            
        proveedores = proveedor_det.objects.all().filter(categoria_proveedor=categoriaf).select_related("codigo_id")
        preguntas = proveedor_encuesta.objects.all().order_by('id')
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__categoria=categoriaf).order_by('-codigo_id__calificacion','-num_trabajadores_fijos')

        conteo = query.count()

        print(conteo)

        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        
   
        return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query}) 

    elif categoriacomp:
        
        print(categoriacomp)
        print("!!!!!!!!!!!!!!!!!!!!")

        proveedor1 = proveedor.objects.filter(categoria=categoriacomp)
        print(proveedor1)

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedor1,proveedor2=proveedor1)

   
        return render(request,'proveedores/ver_preguntas.html',{'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp,'proveedor1':proveedor1}) 


    elif nombre_proveedor:

        preguntas = proveedor_encuesta.objects.all().order_by('id')
           
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__id=nombre_proveedor).order_by('-codigo_id__calificacion')


        conteo = query.count()
       
      
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = ParaExportarExcelForm(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        

        return render(request,'proveedores/ver_preguntas.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'form5':parametros3,'preguntas':preguntas,'count':conteo,'query':query}) 


    elif calf_minima:
        print("Entrò en Calf mìnima")

       
        if calf_minima and calf_maxima:

            calf_min = int(calf_minima)
            calf_max = int(calf_maxima)

            preguntas = 1
            

            query = proveedor_det.objects.all().select_related(
                'codigo_id').filter(codigo_id__calificacion__gte=calf_min ).filter(
                    codigo_id__calificacion__lte=calf_max).order_by('-codigo_id__calificacion')


            conteo = query.count()

            parametros = FiltrarForm(request.GET or None) 
            parametros2 = FiltrarFormDet(request.GET or None)
            parametros4 = NombreEmpresaConRespuestasForm(request.GET or None) 

            return render(request,'proveedores/ver_preguntas.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'calf_min':calf_min,'calf_max':calf_max}) 
           
        else: 

            calf_min = int(calf_minima)

            preguntas = 1
            

            query = proveedor_det.objects.all().select_related(
                'codigo_id').filter(codigo_id__calificacion__gte=calf_min ).order_by('-codigo_id__calificacion')

            
            conteo = query.count()

            parametros = FiltrarForm(request.GET or None) 
            parametros2 = FiltrarFormDet(request.GET or None) 

            return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query}) 

    elif calf_maxima:
        if calf_maxima and encuesta:
            calf_max = int(calf_maxima)

            preguntas = 1
            

            query = proveedor_det.objects.all().select_related(
                'codigo_id').filter(codigo_id__calificacion__lte=calf_max ).filter(
                    codigo_id__respondio_encuesta=True).order_by('-codigo_id__calificacion')


            conteo = query.count()

            parametros = FiltrarForm(request.GET or None) 
            parametros2 = FiltrarFormDet(request.GET or None) 

            return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query,'encuesta':encuesta,'calf_max':calf_max}) 
  
        else:

            calf_max = int(calf_maxima)

            preguntas = 1
            
            query = proveedor_det.objects.all().select_related(
                'codigo_id').filter(codigo_id__calificacion__lte=calf_max ).order_by('-codigo_id__calificacion')


            conteo = query.count()

            parametros = FiltrarForm(request.GET or None) 
            parametros2 = FiltrarFormDet(request.GET or None) 
        
            return render(request,'proveedores/ver_preguntas.html',{'form2':parametros, 'form3':parametros2,'preguntas':preguntas,'count':conteo,'query':query}) 


    elif criticos:

        preguntas = proveedor_encuesta.objects.all().order_by('id')
           
        
        query = proveedor_det.objects.all().select_related('codigo_id').filter(codigo_id__proveedor_critico=True).order_by('-codigo_id__calificacion')


        conteo = query.count()
       
      
        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = ParaExportarExcelForm(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        

        return render(request,'proveedores/ver_preguntas.html',{'form':parametros4,'form2':parametros, 'form3':parametros2,'form5':parametros3,'preguntas':preguntas,'count':conteo,'query':query}) 



    else:
        
        print("Enntró DM")

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None)


        preguntas = proveedor_encuesta.objects.all().order_by('id')



        query = proveedor_det.objects.all().select_related(
            'codigo_id').order_by('-codigo_id__calificacion','-num_trabajadores_fijos')

        
       
        conteo = query.count()

        parametros = FiltrarForm(request.GET or None) 
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = ParaExportarExcelForm(request.GET or None) 
        parametros4 = NombreEmpresaConRespuestasForm(request.GET or None)
        
        return render(request,'proveedores/ver_preguntas.html',{'form':parametros4, 'form3':parametros2,'form5':parametros3,'preguntas':preguntas,'query':query,'count':conteo}) 


#Funcion para Actualizar contadores

def actualiza_contadores(request):
    
       
    query = proveedor_det.objects.all().select_related(
            'codigo_id').order_by('-codigo_id__calificacion','-num_trabajadores_fijos')

        
    # Sección que actuaiza el número de documentos cargados al sistema y el número de documentos solicitados 
    # por categoria de todos los proveedores

    listids = []
    for i in query:
        idprove = i.codigo_id.id
        listids.append(idprove)
        #print(listids)
                
    #Consulta y Guarda el número de documentos cargados al sistema
                
    for i in listids:
        conteoq = documentos_prove.objects.all().filter(proveedor=i)
        consulta = proveedor_det.objects.get(codigo_id=i)
        numcargados = conteoq.count()

        categoria = proveedor.objects.all().filter(id=i)
        cate = 0
        for j in categoria:
            cate = j.categoria.id
        
        tipo = proveedor_det.objects.all().filter(codigo_id=i)
        
        tip = 0
        for k in tipo:
            tip = k.empresa_tipo
        
        qnew1 = proveedor_documentos.objects.all().filter(categoria=cate)
        qnew2 = proveedor_documentos.objects.all().filter(empresa_tipo=tip)
        #queryficha = proveedor_documentos.objects.all().filter(proveedor=i)
        qnew3 = qnew1 | qnew2 
        cuentadocs = qnew3.count()

        

        #proveedor_det.objects.filter(codigo_id=i).update(num_documentos_cargados=numcargados,numero_de_fichas=cuentadocs)
        proveedor_det.objects.filter(codigo_id=i).update(num_documentos_cargados=numcargados)


         #Busca documentos cargados en qnew3 y los excluye de la lista a ser mostrada en template
        contador = 0
        cargados =[]
        lista2 =[]
        
        #Copiamos el queryset de documentos solicitados en una lista
        for p in qnew3:
            lista2.append(str(p))
        
        dimen = len(lista2)
       
        #Eliminamos las ocurrencias de documentos ya cargados en lista2
        for l in conteoq:
            for m,n in zip(lista2,range(dimen)):
                if str(m) in str(l.archivos):
                    contador+=1
                    cargados.append(str(m))
                    lista2.pop(n)
        
    
        #Comprueba si los documentos cargados se encuentran completos
        dimenfinal = len(lista2)
    
        if dimenfinal == 0:
            #Guarda marca en campo para identificar cumplimiento de documentos enviados
            #provecumple = proveedor_det.objects.get(codigo_id=i)

            # consulta = proveedor_det.objects.get(codigo_id=i)
            # consulta.cumple_entrega_docs = 'SI'
            # consulta.save()

            proveedor_det.objects.filter(codigo_id=i).update(cumple_entrega_docs = 'SI')

        else:
            #provecumple = proveedor_det.objects.get(codigo_id=i)
            # consulta.cumple_entrega_docs = 'NO'
            # consulta.save()
            proveedor_det.objects.filter(codigo_id=i).update(cumple_entrega_docs = 'NO')
            

        
    return redirect('proveedores:ver_preguntas')

    




#Funcion para verificar estado de documentos de proveedores cargados

def control_subida_documentos():
    #FUNDAS ROLLOS CAT ID 4

    categoria = proveedor.objects.filter(categoria_id=4)

    listids = []
    for i in categoria:
        idprove = i.id
        listids.append(idprove)
    
    print(listids)
   
    for i in listids:
        documentos = documentos_prove.objects.filter(proveedor_id=i)
        print(i)
        print(documentos)
        if documentos:
            cuenta = 0
            for j in documentos:
                print("Hola",j.archivos)

                if 'RUC' in str(j.archivos):
                    cuenta+=1
                    print("Si entró",cuenta)
                    
                elif "NOMBRAMIENTO REP" in str(j.archivos):
                    cuenta+=1
                
                elif "BANCARIA" in str(j.archivos):
                    cuenta+=1
                
                elif "IESS" in str(j.archivos):
                    cuenta+=1

                elif "SERVICIO BASICO" in str(j.archivos):
                    cuenta+=1
                
                elif "REFERENCIA CLIENTE" in str(j.archivos):
                    cuenta+=1

                elif "REFERENCIA PROVEEDOR" in str(j.archivos):
                    cuenta+=1
                
                elif "NOMINA ACCIONISTAS" in str(j.archivos):
                    cuenta+=1

                elif "CERTIFICADO DE MIGRACION" in str(j.archivos):
                    cuenta+=1

                elif "ACEPTACION DE ESPECIFICACIONES" in str(j.archivos):
                    cuenta+=1

                elif "FICHA TECNICA1" or "FICHA TECNICA2" or "FICHA TECNICA3" or "FICHA TECNICA4" in str(j.archivos):
                    cuenta+=1

                elif "CONTACTO ALIMENTOS" in str(j.archivos):
                    cuenta+=1 
                
                elif "CERTIFICADO BASC" or "AUDITORIA BASC ECOFROZ" in str(j.archivos):
                    cuenta+=1
                
                elif "CERTIFICADO GFSI" or "AUDITORIA GFSI ECOFROZ" in str(j.archivos):
                    cuenta+=1
                
                elif "TRANSPORTE EXCLUSIVO" in str(j.archivos):
                    cuenta+=1
                
                elif "TRANSPORTE SELLADO" in str(j.archivos):
                    cuenta+=1
                elif "CONTACTO EMERGENCIA" in str(j.archivos):
                    cuenta+=1
                
                

              
        else:
            print("No hay documentos")

    return None




#Vista para filtrar proveedores

def filtrar_proveedores(request):
    actividad = request.GET.get("actividad")
    categoria = request.GET.get("categoria_proveedor")
    estado = request.GET.get("proveedor_estado")
    nombre_proveedor = request.GET.get("codigo_id")
    encuesta = request.GET.get("tiene_encuesta")

    if encuesta:
        print("Hola mundo") 

        return render(request,'proveedores/filtrar_proveedores.html',{}) 

    
      
    if actividad:
            
        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None) 

        proveedorx = proveedor_det.objects.all().select_related("codigo_id").filter(actividad__id = actividad)
       
    
        proveedorcount = proveedorx.count()

        paginator = Paginator(proveedorx, 25)
        page = request.GET.get('page')
        proveedor = paginator.get_page(page)

        return render(request,'proveedores/filtrar_proveedores.html',{'form':proveedor,'form2':parametros,'form3':parametros2,'var':proveedorcount,'proveedorx':proveedorx}) 

    elif categoria:
        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None) 

        proveedorx = proveedor_det.objects.all().select_related("codigo_id").filter(categoria_proveedor__id = categoria)
        print(proveedorx)
        print(categoria)
      
    
        proveedorcount = proveedorx.count()

        paginator = Paginator(proveedorx, 25)
        page = request.GET.get('page')
        proveedor = paginator.get_page(page)

        return render(request,'proveedores/filtrar_proveedores.html',{'form':proveedor,'form2':parametros,'form3':parametros2,'var':proveedorcount,'proveedorx':proveedorx}) 

    elif nombre_proveedor:
        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None) 
        print("paso por aqui")
        print("parametros2")

        proveedorx = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id = nombre_proveedor)
        print(proveedorx)
        print(categoria)
      
    
        proveedorcount = proveedorx.count()

        paginator = Paginator(proveedorx, 25)
        page = request.GET.get('page')
        proveedor = paginator.get_page(page)

        return render(request,'proveedores/filtrar_proveedores.html',{'form':proveedor,'form2':parametros,'form3':parametros2,'var':proveedorcount,'proveedorx':proveedorx}) 


    
    else:
        print("paso por aqui")
      
        parametros = FiltrarForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None) 
        print("paso por aqui")
        print("parametros2")

        proveedorx = proveedor_det.objects.all().select_related("codigo_id").filter(codigo_id__id = nombre_proveedor)
      
    
        proveedorcount = proveedorx.count()

        paginator = Paginator(proveedorx, 25)
        page = request.GET.get('page')
        proveedor = paginator.get_page(page)

        return render(request,'proveedores/filtrar_proveedores.html',{'form':proveedor,'form2':parametros,'form3':parametros2,'var':proveedorcount,'proveedorx':proveedorx}) 


#Clases

class calificaEcofroz(UpdateView):
    model = proveedor
    second_model = proveedor_det
    template_name = 'proveedores/califica_ecofroz.html'
    form_class = RegistroFormP
    second_form_class = RegistroFormDetP
    success_url = reverse_lazy('proveedores:listar_proveedores_para_adqui')

    def get_context_data(self, **kwargs):
        context = super(calificaEcofroz, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        proveedor = self.model.objects.get(id=pk)
        detalle = self.second_model.objects.get(codigo_id=proveedor.id)

        #Consulta para mostrr el documento de antecedentes penales
       
        try:
            antecedentes = documentos_prove.objects.filter(proveedor=pk,archivos__icontains='ANTECEDENTES')
    
        except antecedentes.DoesNotExist:
            antecedentes = None

        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'antecedentes' not in context:
            context['antecedentes'] = antecedentes
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_proveedor = kwargs['pk']
        proveedorm = self.model.objects.get(id=id_proveedor)
        detalle = self.second_model.objects.get(codigo_id=proveedorm.id)
        form = self.form_class(request.POST, instance=proveedorm)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        print('aqui')
        if form.is_valid() and form2.is_valid():
            print("Entró bien")
            respuesta1 = request.POST.get('cumple_procedimiento_facturacion')
            respuesta2 = request.POST.get('cumple_procedimiento_entrega')
            respuesta3 = request.POST.get('cumple_procedimiento_compra')
            respuesta4 = request.POST.get('antecedentes_verificados_rep_legal')

            form.save()
            form2.save()
            print('entro aqui')
            califica_ecofroz(kwargs['pk'],respuesta1,respuesta2,respuesta3,respuesta4)
            return HttpResponseRedirect(self.get_success_url())
            # return HttpResponse("Exito!!")

        else:
            print(form.errors)
            print(form2.errors)
            return HttpResponse("No se puede gurdar")
            
@login_required
def calificaEcofrozCalidad(request,pk):
    prove = proveedor.objects.get(id=pk)
    cate = prove.categoria.id
    documentos_list = documentos_prove.objects.all().filter(proveedor=pk).filter(es_doc_categoria=True).filter(Q(es_ficha__isnull=True) | Q(es_ficha=False)).order_by('archivos')
    detalle_proveedor = proveedor_det.objects.get(codigo_id=pk)
    observaciones_administrativas = detalle_proveedor.observaciones_administrativas

    prove_nombre = prove.nombre_empresa
    categoria = prove.categoria.nombre_categoria

    if request.method == 'POST':
        retro_proc_facturacion = request.POST.get("proc-facturacion")
        retro_proc_entrega = request.POST.get("proc-entrega")
        retro_proc_compra = request.POST.get("proc-compra")
        contesta_reclamos = request.POST.get("reclamos")
        valor_aprueba = request.POST.get("combo")
        observaciones = request.POST.get("comentarios")

        if cate == 4: #FUNDAS/ROLLOS
            valor_migracion = request.POST.get("102")
            valor_especificaciones = request.POST.get("104")
            valor_fichas_tecnicas = request.POST.get("106")
            valor_cert_calidad_lote = request.POST.get("117")

            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=102).update(
                respuesta = valor_migracion)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=104).update(
                respuesta = valor_especificaciones)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=106).update(
                respuesta = valor_fichas_tecnicas)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=117).update(
                respuesta = valor_cert_calidad_lote)
                
        if cate == 18: #CAJAS
            
            valor_especificaciones = request.POST.get("161")
            valor_fichas_tecnicas = request.POST.get("163")
            valor_cert_calidad_lote = request.POST.get("170")

            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=161).update(
                respuesta = valor_especificaciones)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=163).update(
                respuesta = valor_fichas_tecnicas)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=170).update(
                respuesta = valor_cert_calidad_lote)

        if cate == 8: #LABORATORIO DE CALIBRACION
            
            valor_patrones = request.POST.get("250")
            valor_procedimientos = request.POST.get("289")

            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=250).update(
                respuesta = valor_patrones)
            actualiza_respuestas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=289).update(
                respuesta = valor_procedimientos)
        
        if cate == 11: #INSUMOS DE LABORATORIO

            valor_fichas = request.POST.get("244")
            valor_hojas_msds = request.POST.get("245")
            valor_cert_calidad_producto = request.POST.get("247")
            valor_entidad_oficial = request.POST.get("248")

            valor_fichas_técnicas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=244).update(
                respuesta = valor_fichas)
            
            valor_msds = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=245).update(
                respuesta = valor_hojas_msds)

            valor_calidad_producto = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=247).update(
                respuesta = valor_cert_calidad_producto)
            
            valor_entidad = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=248).update(
                respuesta = valor_entidad_oficial)

        if cate == 3: #CONTROL DE PLAGAS
            valor_certificado_mip = request.POST.get("134")
            valor_fichas_tecnicas = request.POST.get("280")
            valor_hojas_msds = request.POST.get("282")
            valor_certificado_calidad_lote = request.POST.get("284")
            valor_etiquetas_productos = request.POST.get("321")

            valor_mip = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=134).update(
                respuesta = valor_certificado_mip)
            
            valor_fichas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=280).update(
                respuesta = valor_fichas_tecnicas)

            valor_hojas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=282).update(
                respuesta = valor_hojas_msds)
            
            valor_cert_calidad = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=284).update(
                respuesta = valor_certificado_calidad_lote)
            
            valor_etiquetas = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=321).update(
                respuesta = valor_etiquetas_productos)

        if cate == 7: #KORES
            valor_fichas = request.POST.get("279")
            
            valor_fic = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=279).update(
                respuesta = valor_fichas)
        
        if cate == 13: #DETERGENTES
            valor_hojas = request.POST.get("180")
            valor_calidad_lote = request.POST.get("185")
            valor_fichas = request.POST.get("228")
            valor_etiquetas = request.POST.get("322")
            
            valor_h = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=180).update(
                respuesta = valor_hojas)

            valor_cl = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=185).update(
                respuesta = valor_calidad_lote)
            
            valor_fic = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=228).update(
                respuesta = valor_fichas)
            
            valor_eti = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=322).update(
                respuesta = valor_etiquetas)

        if cate == 12: #DESINFECTANTES
            valor_hojas = request.POST.get("229")
            valor_calidad_lote = request.POST.get("233")
            valor_fichas = request.POST.get("178")
            valor_etiquetas = request.POST.get("324")
            valor_ingredientes = request.POST.get("230")
            valor_reduccion = request.POST.get("231")
            
            valor_h = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=229).update(
                respuesta = valor_hojas)
            valor_c = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=233).update(
                respuesta = valor_calidad_lote)
            valor_f = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=178).update(
                respuesta = valor_fichas)
            valor_et = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=324).update(
                respuesta = valor_etiquetas)
            valor_in = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=230).update(
                respuesta = valor_ingredientes)
            valor_red = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=231).update(
                respuesta = valor_reduccion)

        if cate == 14: #CALDERO
            valor_hojas = request.POST.get("254")
            valor_calidad_lote = request.POST.get("328")
            valor_fichas = request.POST.get("159")
            valor_etiquetas = request.POST.get("327")
            
            valor_h = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=254).update(
                respuesta = valor_hojas)
            valor_c = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=328).update(
                respuesta = valor_calidad_lote)
            valor_fic = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=159).update(
                respuesta = valor_fichas)
            valor_eti = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=327).update(
                respuesta = valor_etiquetas)

        if cate == 10: #MATERIA PRIMA
            valor_formato = request.POST.get("153")
            
            valor_f = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=153).update(
                respuesta = valor_formato)
        
        if cate == 5: #GESTORES DESECHO
            valor_destruccion = request.POST.get("203")
            
            valor_d = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=203).update(
                respuesta = valor_destruccion)

        if cate == 1: #CATERING
            valor_bpm = request.POST.get("238")
            valor_seguimiento_medico = request.POST.get("239")
            valor_afiliacion = request.POST.get("242")
            
            valor_bp = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=238).update(
                respuesta = valor_bpm)
            valor_s = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=239).update(
                respuesta = valor_seguimiento_medico)
            valor_a = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=242).update(
                respuesta = valor_afiliacion)

        if cate == 20:
            valor_ingredientes_activos = request.POST.get("192")
            valor_fichas = request.POST.get("329")
            valor_agrocalidad = request.POST.get("330")
            valor_msds = request.POST.get("331")
            

            valor_ingre = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=192).update(
                respuesta = valor_ingredientes_activos)
            valor_fic = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=329).update(
                respuesta = valor_fichas)
            valor_agro = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=330).update(
                respuesta = valor_agrocalidad)
            valor_m = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=331).update(
                respuesta=valor_msds)
            #Con la respuesta afirmativa de Calidad se obtienen 26 puntos
            if valor_aprueba == 'True':
                valor_final = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=291).update(
                    respuesta = 'SI')
            elif valor_aprueba == 'False':
                valor_final = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=291).update(
                    respuesta = 'NO')
            else:
                valor_final = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta=291).update(
                    respuesta = '--')



        actualiza = proveedor_det.objects.filter(codigo_id=pk).update(
            revisado_por_calidad=valor_aprueba,observaciones_calidad=observaciones,
            cumple_procedimiento_facturacion_calidad=retro_proc_facturacion,
            cumple_procedimiento_entrega_calidad=retro_proc_entrega,
            cumple_procedimiento_compra_calidad=retro_proc_compra,
            contesta_reclamos_calidad=contesta_reclamos)

        actualizaCalificacionIndividualsinRetorno(pk)

    

        envioMailRespuestaCalidad('Nueva respuesta de Calidad en calificación de Proveedores', 'gerencia.administrativa@ecofroz.com,asistente.administrativa@ecofroz.com,dmencias@ecofroz.com', 'email_confirma_revisado_calidad.html', prove_nombre, categoria,observaciones,valor_aprueba,'')
 
        return redirect('proveedores:listar_proveedores_para_calidad')
    
    else: #GET
        detalle=proveedor_det.objects.filter(codigo_id=pk)

        print(cate,type(cate))
        if cate == 4: #Fundas Rollos
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[102,104,106,117]).select_related('pregunta').order_by('pregunta')

        elif cate == 18: #CAJAS

            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[161,163,170]).select_related('pregunta').order_by('pregunta')

        elif cate == 8: #LABORATORIO DE CALIBRACION
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[250,289]).select_related('pregunta').order_by('pregunta')

        elif cate == 9: #LABORATORIO DE ANALISIS
            preguntas_validacion_categoria = []
        
        elif cate == 11: #MATERIALES DE LABORATORIO (INSUMOS)
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[244,245,247,248]).select_related('pregunta').order_by('pregunta')

        elif cate == 3: #CONTROL DE PLAGAS
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[134,280,282,284,321]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 7: #KORES
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[279]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 13: #DETERGENTES
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[228,180,185,322]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 12: #DESINFECTANTES
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[178,229,230,231,233,324]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 14: #CALDERO
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[159,254,327,328]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 10: #MATERIA PRIMA
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[153]).select_related('pregunta').order_by('pregunta')
        
        elif cate == 2: #ACOPIO
            preguntas_validacion_categoria = []

        elif cate == 5: #GESTORES DESECHO
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[203]).select_related('pregunta').order_by('pregunta')
      
        elif cate == 1: #CATERING
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[238,239,242]).select_related('pregunta').order_by('pregunta')
      
        elif cate == 20: #PRODUCTOS Y SERVICIOS DE USO AGRICOLA
            respuesta_pes=''
            respuesta_fol=''
            respuesta_fer=''
            respuesta_mao=''
            respuesta_sem=''

            respuesta_pes = proveedor_respuestas.objects.filter(proveedor_id=pk).get(pregunta=297)
            respuesta_fol = proveedor_respuestas.objects.filter(proveedor_id=pk).get(pregunta=298)
            respuesta_fer = proveedor_respuestas.objects.filter(proveedor_id=pk).get(pregunta=299)
            respuesta_mao = proveedor_respuestas.objects.filter(proveedor_id=pk).get(pregunta=300)
            respuesta_sem = proveedor_respuestas.objects.filter(proveedor_id=pk).get(pregunta=319)
            
            
            preguntas_validacion_categoria = proveedor_respuestas.objects.filter(proveedor_id=pk,pregunta__in=[192,329,330,331]).select_related('pregunta').order_by('pregunta')


            #Consultas para saber si las carpetas de subcategorias tienen o no documentos

            ficha_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_pes=True)
            ficha_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fol=True)
            ficha_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_fer=True)
            ficha_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_mao=True)
            ficha_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_ficha_sem=True)

            agro_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_pes=True)
            agro_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fol=True)
            agro_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_fer=True)
            agro_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_mao=True)
            agro_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_agro_sem=True)

            msds_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_pes=True)
            msds_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fol=True)
            msds_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_fer=True)
            msds_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_mao=True)
            msds_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_msds_sem=True)

            eti_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_pes=True)
            eti_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_foliar=True)
            eti_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_fer=True)
            eti_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_mao=True)
            eti_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_etiqueta_sem=True)

            ana_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_pes=True)
            ana_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_fol=True)
            ana_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_fer=True)
            ana_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_mao=True)
            ana_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_analisis_sem=True)

            certi_pes = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_pes=True)
            certi_fol = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_fol=True)
            certi_fer = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_fer=True)
            certi_mao = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_mao=True)
            certi_sem = documentos_prove.objects.all().filter(proveedor=pk).filter(es_certificado_sem=True)


            return render(request,'proveedores/califica_ecofroz_calidad.html',{'form':prove,'documentos_list':documentos_list,'detalle':detalle,'preguntas_validacion_categoria':preguntas_validacion_categoria,'pes':respuesta_pes,'fol':respuesta_fol,'fer':respuesta_fer,'mao':respuesta_mao,'sem':respuesta_sem,
            'ficha_pes':ficha_pes,'ficha_fol':ficha_fol,'ficha_fer':ficha_fer,'ficha_mao':ficha_mao,'ficha_sem':ficha_sem,
            'agro_pes':agro_pes,'agro_fol':agro_fol,'agro_fer':agro_fer,'agro_mao':agro_mao,'agro_sem':agro_sem,
            'msds_pes':msds_pes,'msds_fol':msds_fol,'msds_fer':msds_fer,'msds_mao':msds_mao,'msds_sem':msds_sem,
            'eti_pes':eti_pes,'eti_fol':eti_fol,'eti_fer':eti_fer,'eti_mao':eti_mao,'eti_sem':eti_sem,
            'ana_pes':ana_pes,'ana_fol':ana_fol,'ana_fer':ana_fer,'ana_mao':ana_mao,'ana_sem':ana_sem,
            'certi_pes':certi_pes,'certi_fol':certi_fol,'certi_fer':certi_fer,'certi_mao':certi_mao,'certi_sem':certi_sem,'observaciones_administrativas':observaciones_administrativas})
       

        else:
            return HttpResponse("Ooops!! Categoria de proveedor no configurada!!")

        return render(request,'proveedores/califica_ecofroz_calidad.html',{'form':prove,'documentos_list':documentos_list,'detalle':detalle,'preguntas_validacion_categoria':preguntas_validacion_categoria,'observaciones_administrativas':observaciones_administrativas})
        
        
        
class ProveedorDelete(DeleteView):
    model = proveedor
    template_name = 'proveedores/eliminar_proveedor.html'
    success_url = reverse_lazy('proveedores:listar_proveedores')


class ProveedorUpdate(UpdateView):
    model = proveedor
    second_model = proveedor_det
    template_name = 'proveedores/editar_proveedor.html'
    form_class = RegistroFormP
    second_form_class = RegistroFormDetP
    success_url = reverse_lazy('proveedores:listar_proveedores')

    def get_context_data(self, **kwargs):
        context = super(ProveedorUpdate, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        proveedor = self.model.objects.get(id=pk)
        detalle = self.second_model.objects.get(codigo_id=proveedor.id)
        categoria = self.model.objects.filter(id=pk)
        documentos_list = documentos_prove.objects.all().filter(proveedor=pk).order_by('archivos')
        fichas_list = proveedor_fichas.objects.all().filter(proveedor=pk).order_by('nombre_ficha')
        trazabilidad = proveedor_det.objects.all().filter(codigo_id=pk).select_related('codigo_id')


        cate = 0
        for i in categoria:
            cate = i.categoria.id
        
        print("Categoria",cate)
        
        tipo = self.second_model.objects.filter(codigo_id=pk)
        
        tip = 0
        for i in tipo:
            tip = i.empresa_tipo
        
        listaso = 0
        for i in documentos_list:
            listaso = str(i.archivos)

        asientocount = 0
        for i in tipo:
            asientocount = i.asientos
        
        color = ""
        for i in tipo:
            color = i.color
        
        
        query = proveedor_documentos.objects.all().filter(categoria=cate)
        query2 = proveedor_documentos.objects.all().filter(empresa_tipo=tip)

        #Union de las consultas de fichas y documntos por tipo de empresa (Persona o Juridica)
        query3 = query | query2
        query3 = query3.order_by('empresa_tipo')    

        #self.model.objects.get(id=pk).update(num_documentos_solicitados_categoria=query3.count())


        #Crea una lista con un diccionario con la lista de documentos solicitados y el estado del documento en el sistema (Cargado SI/NO)
      
        lista2 = []
        for i in query3:
            lista2.append(str(i))
         
        dimen = len(lista2)
        print(dimen)
        print(lista2)
      
        vec = []
       
        contador = 0
       
        for i in documentos_list:
            for j,k in zip(lista2,range(dimen)):
                if str(j) in str(i.archivos):
                    vector = {
                        'documento_sol':str(j),
                        'cargado':'SI',
                    }
                    vec.append(vector)
                    lista2.pop(k)
                
        for i in lista2:
            vector2 = {
                'documento_sol':str(i),
                'cargado':'NO',
            }
            vec.append(vector2)
        

        vec_count = len(vec)
        entregados_count = vec_count - len(lista2)



        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        if 'numfichas' not in context:
            context['numfichas'] = detalle.numero_de_fichas
        if 'proveedor' not in context:
            context['proveedor'] = proveedor
        if 'query3' not in context:
            context['query3'] = query3  
        if 'documentos_list' not in context:
            context['documentos_list'] = documentos_list
        if 'vec' not in context:
            context['vec'] = vec
        if 'vec_count' not in context:
            context['vec_count'] = vec_count
        if 'entregados_count' not in context:
            context['entregados_count'] = entregados_count
        if 'trazabilidad' not in context:
            context['trazabilidad'] = trazabilidad
        if 'asientocount' not in context:
            context['asientocount'] = asientocount
        if 'color' not in context:
            context['color'] = color

        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_proveedor = kwargs['pk']
        proveedorm = self.model.objects.get(id=id_proveedor)
        detalle = self.second_model.objects.get(codigo_id=proveedorm.id)
        categoria = self.model.objects.filter(id=id_proveedor)
        form = self.form_class(request.POST, instance=proveedorm)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
        # fichas_list = proveedor_fichas.objects.all().filter(proveedor=id_proveedor).order_by('nombre_ficha')

        if form.is_valid() and form2.is_valid():
        #if form.is_valid():
            form.save()
            form2.save()
            # detalle.tiene_fichas = request.POST.get('tiene_fichas')
            detalle.asientos = request.POST.get('asientos')
            detalle.color = request.POST.get('color')
            detalle.save()
            
            proveedorm.usuario_modifica = request.user.username
            proveedorm.save()

            return HttpResponseRedirect(self.get_success_url())
            # return HttpResponse("Exito!!")
        else:
            return HttpResponse("Error No se pudo guardar")
            #return HttpResponseRedirect(self.get_success_url())

class VerDocumentos(UpdateView):
    model = proveedor
    second_model = proveedor_det
    template_name = 'proveedores/ver_documentos.html'
    form_class = RegistroFormP
    second_form_class = RegistroFormDetP
    success_url = reverse_lazy('proveedores:listar_proveedores')

    def get_context_data(self, **kwargs):
        context = super(VerDocumentos, self).get_context_data(**kwargs)
        pk = self.kwargs.get('pk',0)
        proveedor = self.model.objects.get(id=pk)
        detalle = self.second_model.objects.get(codigo_id=proveedor.id)
        if 'form' not in context:
            context['form'] = self.form_class()
        if 'form2' not in context:
            context['form2'] = self.second_form_class(instance=detalle)
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object
        id_proveedor = kwargs['pk']
        proveedorm = self.model.objects.get(id=id_proveedor)
        detalle = self.second_model.objects.get(codigo_id=proveedorm.id)
        form = self.form_class(request.POST, instance=proveedorm)
        form2 = self.second_form_class(request.POST, request.FILES, instance=detalle)
      
        if form.is_valid() and form2.is_valid():
        #if form.is_valid():
            form.save()
            form2.save()
            return HttpResponseRedirect(self.get_success_url())
            # return HttpResponse("Exito!!")
        else:
            return HttpResponse("Error No se pudo guardar")
            #return HttpResponseRedirect(self.get_success_url())


#FUNCION GLOBAL QUE CALCULA NOTAS DE PROVEEDORES
def calculaCalificacion(prove):
    
    bandera = 0
    bandera1 = 0
    bandera2 = 0
    bandera3 = 0

    for idprove in prove:
        pk = idprove.proveedor_id.id
        cate = idprove.proveedor_id.categoria.id

    print(pk)
    print(cate)

    aprobacion_calidad = proveedor_det.objects.get(codigo_id=pk)

    if aprobacion_calidad.revisado_por_calidad == True:
        puntua = 1
    else:
        puntua = 0
    
    
    for calif in prove:
        
        
        #PREGUNTAS ESPECIFICAS
        #LABORATORIOS DE CALIBRACION
        if calif.pregunta.id == 249:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 250:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 251:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 288:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 289:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        #INSUMOS DE LABORATORIO 
        if calif.pregunta.id == 244:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 245:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 246:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 247:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 248:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #LABORATORIO DE ENSAYO
        if calif.pregunta.id == 195:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 197:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 199:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        
        
        #KOREX SLIP SHEET
        if calif.pregunta.id == 268:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 277:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 279:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 280:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 281:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 282:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 283:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 284:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #TRANSPORTE (OTROS)
        if calif.pregunta.id == 269:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
 
        if calif.pregunta.id == 270:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 271:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
            
        if calif.pregunta.id == 272:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 273:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        
        if calif.pregunta.id == 294:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 295:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        

        #CENTROS DE ACOPIO
        if calif.pregunta.id == 224:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 225:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 172:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 252:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 278:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 285:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 286:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 296:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #QUIMICOS DE LIMPIEZA DETERGENTES
        if calif.pregunta.id == 258:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 322:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 228:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 180:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 182:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 184:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 185:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 187:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta =='':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 189:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        

        #QUIMICOS DE LIMPIEZA DESINFECTANTES
        if calif.pregunta.id == 178:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 324:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 325:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 229:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 230:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 231:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 232:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 233:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 234:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #CONTROL DE PLAGAS
        if calif.pregunta.id == 321:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 130:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 132:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 134:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 136:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 138:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 139:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 141:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 143:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #PRODUCTOS Y SERVICIOS VARIOS #19
        if calif.pregunta.id == 144: 
            print("vino")   
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                print("vino vino")
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 145:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
                
        if calif.pregunta.id == 146:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NO APLICA':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #FIN DE PRODUCTOS Y SERVICIOS VARIOS

        #CATERING
        if calif.pregunta.id == 237:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
                
        if calif.pregunta.id == 238:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 239:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 240:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 241:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 242:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 243:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #MATERIA PRIMA CATEGORIA #10
        if calif.pregunta.id == 147:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera2 = 0
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 148:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 149:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 153:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 154:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta ==  '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 156:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 157:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        #FIN DE MATERIA PRIMA
        
        #QUIMICOS PARA EL CALDERO # 14
        if calif.pregunta.id == 328:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 327:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 326:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 254:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        if calif.pregunta.id == 159:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 255:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()


        #FIN DE QUIMICOS PARA EL CALDERO

        #GESTORES AMBIENTALES DE DESECHOS CATEGORIA # 5
        if calif.pregunta.id == 205:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 201:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        if calif.pregunta.id == 203:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()


        if calif.pregunta.id == 290:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        #FIN DE GESTORES AMBIENTALES DE DESECHOS

        # if calif.pregunta.pregunta == '¿ Envió el certificado de migración ?':
        #FUNDAS/ROLLOS
        if calif.pregunta.id == 102:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
           
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        # if calif.pregunta.pregunta == '¿Envió la aceptación de las especificaciones de Ecofroz S.A?':
        if calif.pregunta.id == 104:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        # if calif.pregunta.pregunta == '¿Envió las especificación del material de empaque (vigencia 1 año)?':
        if calif.pregunta.id == 106:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        
        # if calif.pregunta.pregunta == '¿Envió las aprobaciones de contacto con los alimentos (materia prima)?':
        if calif.pregunta.id == 108:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()



  #TRANSPORTE CONTENEDORES      
  # if calif.pregunta.pregunta == ¿ Envió el listado de choferes autorizados ?:
        if calif.pregunta.id == 128:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 260:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()


        # if calif.pregunta.pregunta == '¿Tiene certificación de seguridad BASC?':
        if calif.pregunta.id == 109:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
               

        # if calif.pregunta.pregunta == 'En caso de que NO tenga certificación BASC. ¿Tiene homologación y/o Ecofroz ha realizado una auditoria a su empresa?':
        if calif.pregunta.id == 111: 
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0

            else:
                # if bandera1 == 1:
                calif.calificacion = 0
                calif.save()
                

                
        # if calif.pregunta.pregunta == 'Confirmación de envío de certificación BASC u homologación o Auditoria de Ecofroz por:':
        if calif.pregunta.id == 113:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
        
            else:
                calif.calificacion = 0
                calif.save()
        
        # if calif.pregunta.pregunta == '¿Tiene certificado de homologación GFSI?':
        if calif.pregunta.id == 114: 
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0

            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
                
        if calif.pregunta.id == 115:
            if calif.respuesta == 'NO':  
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
                bandera1 = 0
            else:
                if bandera1 == 1:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == 'En caso de tener certificado GFSI o de auditoria Ecofroz lo envió por:':
        if calif.pregunta.id == 116:  
            if calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        # if calif.pregunta.pregunta == '¿Entrega certificado de calidad por lote?':
        if calif.pregunta.id == 117:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == '¿Envió el certificado de que el transporte es exclusivo para este material?':
        if calif.pregunta.id == 119:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        # if calif.pregunta.pregunta == '¿ Envió el certificado de que el transporte debe llegar sellado a Ecofroz S.A?':
        if calif.pregunta.id == 121:
            if calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        # if calif.pregunta.pregunta == '¿Envió el formato de información del contacto de emergencia debidamente completado?':
        if calif.pregunta.id == 123:  
            if calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == 'Contrato vigente firmado con Ecofroz S.A':
        if calif.pregunta.id == 124:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        #CAJAS

        if calif.pregunta.id == 320:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 4
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 161:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()


        # if calif.pregunta.id == '¿ Envió el documento con las especificaciones/fichas técnicas de todos los productos vendidos a Ecofroz S.A (vigencia 1 año) ?':
        if calif.pregunta.id == 163:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == '¿Tiene certificado de homologación GFSI o auditoria realizada por Ecofroz S.A?':
        if calif.pregunta.id == 164:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == 'Confirmó el envió del certificado de Homologacion GFSI o auditoria realizada por Ecofroz por:':
        

        if calif.pregunta.id == 166:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        # if calif.pregunta.pregunta == '¿Tiene certificación de seguridad BASC o certificado de homologacion o auditoria de Ecofroz?':
        if calif.pregunta.id == 167:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        # Certificado BASC o de homologación y/o auditoria realizada por Ecofroz?':
        if calif.pregunta.id == 169:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == '¿Tiene el certificado de calidad por lote?':
        if calif.pregunta.id == 170:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()
        
        # if calif.pregunta.pregunta == 'Tiene el certificado de que el transporte debe llegar sellado a Ecofroz S.A':
        if calif.pregunta.id == 171:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        # if calif.pregunta.pregunta == '¿Tiene el certificado de que el transporte es exclusivo para este material?':
        if calif.pregunta.id == 174:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                # calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        # if calif.pregunta.pregunta == 'Envió la información del contacto de emergencia por:':
        if calif.pregunta.id == 176:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if puntua == 1:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #CODIGO DE ETICA
        if calif.pregunta.id == 274:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        #PRODUCTOS Y SERVICIOS DE USO AGRICOLA
        
        if calif.pregunta.id == 319:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 190:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
                
        if calif.pregunta.id == 191:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save() 
                
        if calif.pregunta.id == 192:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
                
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save() 

                
        if calif.pregunta.id == 193:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save() 

        if calif.pregunta.id == 291:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()      
        
        if calif.pregunta.id == 292:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 293:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 329:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 330:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 331:  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'XXX':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #CONSTRUCTORES
        
        if calif.pregunta.id == 340:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        if calif.pregunta.id == 341:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        if calif.pregunta.id == 342:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        if calif.pregunta.id == 343:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        if calif.pregunta.id == 339:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()


        if calif.pregunta.id == 336:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()


        if calif.pregunta.id == 337:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        if calif.pregunta.id == 338:
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()    
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                calif.save()

        #FIN DE CALCULO DE CALIFICACION PREGUNTAS ESPECIFICAS


        #COMPROMISOS
        if calif.pregunta.id == 1:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 2:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 97:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 3:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 4:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()


        #EXISTENCIA Y LEGALIDAD
        if calif.pregunta.id == 276:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 275:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 27:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 29:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 31:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 33:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 35:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'xxx':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()


        if calif.pregunta.id == 213:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 99:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 215:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 216:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 100:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 217:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 218:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 219:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 125:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        if calif.pregunta.id == 126:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        #BPC
        if calif.pregunta.id == 42:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 43:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 44:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 45:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 46:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 47:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #BPS
        if calif.proveedor_id.categoria.id == 21:
            if calif.pregunta.id == 53:
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                    calif.save()

            if calif.pregunta.id == 54:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                    calif.save()

            if calif.pregunta.id == 55:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                    calif.save()


            if calif.pregunta.id == 56:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica_constructores
                    calif.save()

            if calif.pregunta.id == 57:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

            if calif.pregunta.id == 58:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

            if calif.pregunta.id == 59:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        
        else: 
            if calif.pregunta.id == 48:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 49:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 50:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 51:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 52:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 53:
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 54:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 55:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()


            if calif.pregunta.id == 56:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

            if calif.pregunta.id == 57:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

            if calif.pregunta.id == 58:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

            if calif.pregunta.id == 59:    
                if calif.respuesta == 'NO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'no':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'nO':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta == 'No lo tengo':
                    calif.calificacion = 0
                    calif.save()
                elif calif.respuesta is None:
                    calif.calificacion = 0
                    calif.save()
                else:
                    calif.calificacion = 0
                    calif.save()

        #BPA
        if calif.pregunta.id == 60:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 64:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 65:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        if calif.pregunta.id == 66:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 67:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #BPRS
        if calif.pregunta.id == 68:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 69:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'NA':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'N/A':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 70:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 71:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 72:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
            
        if calif.pregunta.id == 73:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 74:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        #BPL
        if calif.pregunta.id == 75:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 76:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 77:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 78:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 79:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 80:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 81:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 82:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        #BPPC
        if calif.pregunta.id == 83:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 84:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 85:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        #RETROALIMENTACION ECOFROZ
        if calif.pregunta.id == 86:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 87:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 88:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 89:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 90:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 91:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 92:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()
        
        if calif.pregunta.id == 93:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()

        if calif.pregunta.id == 94:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        if calif.pregunta.id == 95:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
        
        if calif.pregunta.id == 96:    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'None':
                calif.calificacion = 0
                calif.save()
            else:
                calif.calificacion = 0
                calif.save()

        #CALIFICACION ECOFROZ
        if calif.pregunta.id == 206:
            print("Llega 206")    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                print("Reconoce el None en 206")
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if calif.respuesta == 'SI':
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()
                

        
        if calif.pregunta.id == 207:
            print("Llega 207")    
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if calif.respuesta == 'SI':
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

        if calif.pregunta.id == 208:  
            print("Llega 208")  
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if calif.respuesta == 'SI':
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

        if calif.pregunta.id == 209: 
            print("Llega 209")   
            if calif.respuesta == 'NO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'no':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'nO':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == 'No lo tengo':
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta is None:
                calif.calificacion = 0
                calif.save()
            elif calif.respuesta == '':
                calif.calificacion = 0
                calif.save()
            else:
                if calif.respuesta == 'SI':
                    calif.calificacion = calif.pregunta.puntaje_pregunta_especifica
                    calif.save()

    return None


def actualizaCalificacionIndividual(request,pk):
    id_prove=pk
    codigo_id = request.GET.get("nombre_empresa")

    q = proveedor.objects.filter(id=pk)

    for cat in q:
        cate = cat.categoria.id
    
    prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

    calculaCalificacion(prove)

    
    calif_total = proveedor_respuestas.objects.filter(proveedor_id__id=pk).exclude(
        respuesta='xxx').exclude(respuesta='XXX').exclude(
            pregunta_id__in=[222,223,206,207,208,209])
        

    nombre = proveedor.objects.all().get(id=pk)
    
    suma_prove1 = 0
    for nota in prove:
        suma_prove1+=nota.calificacion
        
    #Encera el recuento parcial de puntos obtenidos si el puntaje excede a los 30 de categoria que es el máximo
    if suma_prove1 > 30:
        suma_prove1 = 0
        
    suma_total = 0
    for nota in calif_total:
        suma_total+=nota.calificacion
        
    print("Esta es la calificacion numerica",suma_total)
        
    if suma_total > 100:
        suma_total = 100

        #Guarda el calculo de nota individual en el campo calificacion de la cabecera proveedor
    consulta = proveedor.objects.get(id=pk)
    consulta.calificacion = suma_total
    consulta.save()

        
    query = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))
        

    suma_notausointerno = 1
    
    for calif in query:
        if calif.pregunta.id == 206:
            print("1zzzzXXX")
            if calif.calificacion == 2:
                print("2zzzXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 207:
            print("1yyyyXXX")
            if calif.calificacion == 2:
                print("syyyXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 208:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None

        elif calif.pregunta.id == 209:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None
        else:
            print("No hay la pregunta")

        print('Aqui la nota final',suma_notausointerno)

    
    if suma_notausointerno == 9:
        gradev = 'A'
    if suma_notausointerno >6 and suma_notausointerno <9:
        gradev = 'B'
    if suma_notausointerno >1 and suma_notausointerno <7:
        gradev = 'C'
    if suma_notausointerno == 1:
        gradev = 'F'
    
   
    consulta2 = proveedor.objects.get(id=pk)
    consulta2.grade = gradev
    consulta2.save()

    messages.info(request, 'Calificación Actualizada!')
    #mensaje = 'Calificación Actualizada!'
    #return render(request,'proveedores/confirma_success_actualiza_nota_indivi.html',{'codigo_id':pk,'mensaje':mensaje})
    #return render(request,'proveedores/ver_documentos.html',{'codigo_id':pk,'mensaje':mensaje})
    return HttpResponseRedirect(reverse('proveedores:ver_documentos',kwargs={'pk':pk}))


def actualizaCalificacionIndividualTest(request,pk):
    id_prove=pk
    codigo_id = request.GET.get("nombre_empresa")

    q = proveedor.objects.filter(id=pk)

    for cat in q:
        cate = cat.categoria.id
    
    prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

    calculaCalificacion(prove)

    
    calif_total = proveedor_respuestas.objects.filter(proveedor_id__id=pk).exclude(
        respuesta='xxx').exclude(respuesta='XXX').exclude(
            pregunta_id__in=[222,223,206,207,208,209])
        

    nombre = proveedor.objects.all().get(id=pk)
    
    suma_prove1 = 0
    for nota in prove:
        suma_prove1+=nota.calificacion
        
    #Encera el recuento parcial de puntos obtenidos si el puntaje excede a los 30 de categoria que es el máximo
    if suma_prove1 > 30:
        suma_prove1 = 0
        
    suma_total = 0
    for nota in calif_total:
        suma_total+=nota.calificacion
        
    print("Esta es la calificacion numerica",suma_total)
        
    if suma_total > 100:
        suma_total = 100

        #Guarda el calculo de nota individual en el campo calificacion de la cabecera proveedor
    consulta = proveedor.objects.get(id=pk)
    consulta.calificacion = suma_total
    consulta.save()

        
    query = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))
        

    suma_notausointerno = 1
    
    for calif in query:
        if calif.pregunta.id == 206:
            print("1zzzzXXX")
            if calif.calificacion == 2:
                print("2zzzXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 207:
            print("1yyyyXXX")
            if calif.calificacion == 2:
                print("syyyXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 208:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None

        elif calif.pregunta.id == 209:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None
        else:
            print("No hay la pregunta")

        print('Aqui la nota final',suma_notausointerno)

    
    if suma_notausointerno == 9:
        gradev = 'A'
    if suma_notausointerno >6 and suma_notausointerno <9:
        gradev = 'B'
    if suma_notausointerno >1 and suma_notausointerno <7:
        gradev = 'C'
    if suma_notausointerno == 1:
        gradev = 'F'
    
   
    consulta2 = proveedor.objects.get(id=pk)
    consulta2.grade = gradev
    consulta2.save()

    proveedoresq = proveedor_det.objects.all().select_related("codigo_id").order_by('-codigo_id__fecha_modifica')
       
    paginator = Paginator(proveedoresq, 25)

    page = request.GET.get('page')
    proveedores = paginator.get_page(page)
    
    mensaje = 'Calificación Actualizada!'
    return render(request,'proveedores/busqueda_proveedores.html',{'form':proveedores,'mensaje':mensaje})


    


def actualizaCalificacionIndividualsinRetorno(pk):
    id_prove=pk
    #codigo_id = request.GET.get("nombre_empresa")

    q = proveedor.objects.filter(id=pk)

    for cat in q:
        cate = cat.categoria.id
    
    prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

    calculaCalificacion(prove)

    
    calif_total = proveedor_respuestas.objects.filter(proveedor_id__id=pk).exclude(
        respuesta='xxx').exclude(respuesta='XXX').exclude(
            pregunta_id__in=[222,223,206,207,208,209])
        

    nombre = proveedor.objects.all().get(id=pk)
    
    suma_prove1 = 0
    for nota in prove:
        suma_prove1+=nota.calificacion
        
    #Encera el recuento parcial de puntos obtenidos si el puntaje excede a los 30 de categoria que es el máximo
    if suma_prove1 > 30:
        suma_prove1 = 0
        
    suma_total = 0
    for nota in calif_total:
        suma_total+=nota.calificacion
        
    print("Esta es la calificacion numerica",suma_total)
        
    if suma_total > 100:
        suma_total = 100

        #Guarda el calculo de nota individual en el campo calificacion de la cabecera proveedor
    consulta = proveedor.objects.get(id=pk)
    consulta.calificacion = suma_total
    consulta.save()

        
    query = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))
        

    suma_notausointerno = 1
    
    for calif in query:
        if calif.pregunta.id == 206:
            print("1zzzzXXX")
            if calif.calificacion == 2:
                print("2zzzXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 207:
            print("1yyyyXXX")
            if calif.calificacion == 2:
                print("syyyXXX")
                suma_notausointerno+=2
            else:
                None
            
        elif calif.pregunta.id == 208:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None

        elif calif.pregunta.id == 209:
            if calif.calificacion == 2:
                suma_notausointerno+=2
            else:
                None
        else:
            print("No hay la pregunta")

        print('Aqui la nota final',suma_notausointerno)

    
    if suma_notausointerno == 9:
        gradev = 'A'
    if suma_notausointerno >6 and suma_notausointerno <9:
        gradev = 'B'
    if suma_notausointerno >1 and suma_notausointerno <7:
        gradev = 'C'
    if suma_notausointerno == 1:
        gradev = 'F'
    
   
    consulta2 = proveedor.objects.get(id=pk)
    consulta2.grade = gradev
    consulta2.save()

    return None



def verRespuestasEncuesta(request,pk):

    id_prove=pk
    codigo_id = request.GET.get("nombre_empresa")
    segmento_escogido = request.GET.get("nombre_segmento")

    
    q = proveedor.objects.filter(id=pk)

    for cat in q:
        cate = cat.categoria.id
    

    if segmento_escogido:
        if codigo_id:
            pk = codigo_id
            print("Segmento con cambio de proveedor entró por aqui DM")
            q = proveedor.objects.filter(id=pk)

            for cat in q:
                cate = cat.categoria.id

            prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                pregunta__segmento__id=segmento_escogido).filter(
                    proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX')

            # calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            #     'proveedor_id').filter(proveedor_id__id=codigo_id).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

            calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223).exclude(
                pregunta_id=206).exclude(pregunta_id=207).exclude(pregunta_id=208).exclude(pregunta_id=209)
        
            print(cate)
            print(prove)
            print(segmento_escogido)
            print(codigo_id)
            print(pk)

            nombre = proveedor.objects.all().get(id=pk)
            parametros2 = FiltrarFormDet(request.GET or None) 
            parametros = NombreEmpresaConRespuestasForm(request.GET or None)
            parametros3 = SegmentoForm(request.GET or None)

        
            
            calculaCalificacion(prove)

            empresa_tip = proveedor.objects.all().filter(id=pk)
            for i in empresa_tip:
                tipo = i.tipo_empresa
            
            print("Hola Tipo",tipo)
            
            if tipo == 'Persona Jurídica':
                print("Entra papà2")

                prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                    pregunta__segmento__id=segmento_escogido).filter(
                        proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(
                            pregunta__pregunta__icontains='*')
            
            else:
                print("Por este otro lado papá")
                prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                        pregunta__segmento__id=segmento_escogido).filter(
                            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX')
                            



            suma_prove1 = 0
            for nota in prove:
                suma_prove1+=nota.calificacion
            
            suma_total = 0
            for nota in calif_total:
                suma_total+=nota.calificacion
            
            if suma_total > 100:
                suma_total = 100

            v = proveedor.objects.filter(id=codigo_id)

            for nota in v:
                califica = nota.grade 
                catego = nota.categoria
            
            print(califica)
            
            return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form2':prove,'form3':parametros2,'form4':parametros3,'nombre':nombre,'id_prove':id_prove,'segmento_escogido':segmento_escogido,'suma_prove1':suma_prove1,'suma_total':suma_total,'califica':califica,'categoria':catego,'grade':califica})

        else:

            print("Segmento por primera vez")
            prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                pregunta__segmento__id=segmento_escogido).filter(
                    proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX')

           

            nombre = proveedor.objects.all().get(id=pk)
            parametros2 = FiltrarFormDet(request.GET or None) 
            parametros = NombreEmpresaConRespuestasForm(request.GET or None)
            parametros3 = SegmentoForm(request.GET or None)

        
            calculaCalificacion(prove)

            empresa_tip = proveedor.objects.all().filter(id=pk)
            for i in empresa_tip:
                tipo = i.tipo_empresa
            
            print("Hola Tipo",tipo)
            
            if tipo == 'Persona Jurídica':
                print("Entra papà1")

                prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                    pregunta__segmento__id=segmento_escogido).filter(
                        proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(
                            pregunta__pregunta__icontains='*')
            
            else:
                print("Entra else papà1")

                prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                        pregunta__segmento__id=segmento_escogido).filter(
                            proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX')
                            #.filter(Q(pregunta__pregunta__icontains="*"))
                            

            # calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            #     'proveedor_id').filter(proveedor_id__id=pk).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

            calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223).exclude(
                pregunta_id=206).exclude(pregunta_id=207).exclude(pregunta_id=208).exclude(pregunta_id=209)
        


            # calculaCalificacion(calif_total)


            suma_prove1 = 0
            for nota in prove:
                suma_prove1+=nota.calificacion
            
            suma_total = 0
            for nota in calif_total:
                suma_total+=nota.calificacion
            
            if suma_total > 100:
                suma_total = 100

            v = proveedor.objects.filter(id=pk)

            for nota in v:
                califica = nota.grade 
                catego = nota.categoria
            
            
            return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form2':prove,'form3':parametros2,'form4':parametros3,'nombre':nombre,'id_prove':id_prove,'segmento_escogido':segmento_escogido,'suma_prove1':suma_prove1,'suma_total':suma_total,'califica':califica,'categoria':catego,'grade':califica})


    else:
        
        if codigo_id:
            if segmento_escogido:
                print("SSSSSSSSSSSSSSSSSSSSS")
                prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=codigo_id).order_by('pregunta').filter(pregunta__segmento__nombre_segmento__icontains=segmento_escogido)

                nombre = proveedor.objects.all().get(id=codigo_id)
                parametros2 = FiltrarFormDet(request.GET or None) 



                return render(request,'proveedores/ver_respuestas_individual.html',{'form':prove,'form3':parametros2,'nombre':nombre,'id_prove':codigo_id,'segmento_escogido':segmento_escogido,'grade':califica})

            print("Otro proveedor escogido para buscar segmentos DM")
            print(codigo_id)
            pk = codigo_id

            prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__id').exclude(respuesta='xxx')
            
            # calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                # 'proveedor_id').filter(proveedor_id__id=codigo_id)
            
            calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223).exclude(
                pregunta_id=206).exclude(pregunta_id=207).exclude(pregunta_id=208).exclude(pregunta_id=209)
        
            

            nombre = proveedor.objects.all().get(id=pk)
            parametros = NombreEmpresaConRespuestasForm(request.GET or None)
            parametros3 = SegmentoForm(request.GET or None) 

            calculaCalificacion(prove)
            calculaCalificacion(calif_total)


            suma_prove1 = 0
            for nota in prove:
                suma_prove1+=nota.calificacion
            
            if suma_prove1 > 40:
                suma_prove1 = 0
            
            suma_total = 0
            for nota in calif_total:
                suma_total+=nota.calificacion
                
            form = []
                
            for x in prove:
                vec = {
                        'pregunta':x.pregunta,
                        'respuesta':x.respuesta,
                        'calificacion':x.calificacion,
                        }
                form.append(vec)
            
            v = proveedor.objects.filter(id=pk)

            for nota in v:
                califica = nota.grade
                catego = nota.categoria
                    

            return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form4':parametros3,'nombre':nombre,'id_prove':id_prove,'segmento_escogido':segmento_escogido,'suma_prove1':suma_prove1,'suma_total':suma_total,'califica':califica,'categoria':catego,'grade':califica})
           
           
        print("BBBBBBBBBBBBBBBBbbbb")
        prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta__pregunta_campo_modelo_id').filter(
                    proveedor_id__categoria__id=cate).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(pregunta_id=222).exclude(pregunta_id=223)

        calculaCalificacion(prove)


        #calif_total = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        #    'proveedor_id').filter(proveedor_id__id=pk).exclude(respuesta='xxx').exclude(respuesta='XXX').exclude(
        #        pregunta_id__in=[222,223,206,207,208,209])
        
        calif_total = proveedor_respuestas.objects.filter(proveedor_id__id=pk).exclude(
            respuesta='xxx').exclude(respuesta='XXX').exclude(
                pregunta_id__in=[222,223,206,207,208,209])
        

        nombre = proveedor.objects.all().get(id=pk)
        parametros = NombreEmpresaConRespuestasForm(request.GET or None)
        parametros2 = FiltrarFormDet(request.GET or None) 
        parametros3 = SegmentoForm(request.GET or None)

        suma_prove1 = 0
        for nota in prove:
            suma_prove1+=nota.calificacion
        
        #Encera el recuento parcial de puntos obtenidos si el puntaje excede a los 30 de categoria que es el máximo
        if suma_prove1 > 30:
            suma_prove1 = 0
        
        suma_total = 0
        for nota in calif_total:
            suma_total+=nota.calificacion
        
        print("Esta es la calificacion numerica",suma_total)
        
        if suma_total > 100:
            suma_total = 100

        #Guarda el calculo de nota individual en el campo calificacion de la cabecera proveedor
        consulta = proveedor.objects.get(id=pk)
        consulta.calificacion = suma_total
        consulta.save()

        

        query = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=pk).filter(pregunta_id__in = (206,207,208,209))
        

        suma_notausointerno = 1
    
        for calif in query:
            if calif.pregunta.id == 206:
                print("1zzzzXXX")
                if calif.calificacion == 2:
                    print("2zzzXXX")
                    suma_notausointerno+=2
                else:
                    None
            
            elif calif.pregunta.id == 207:
                print("1yyyyXXX")
                if calif.calificacion == 2:
                    print("syyyXXX")
                    suma_notausointerno+=2
                else:
                    None
            
            elif calif.pregunta.id == 208:
                if calif.calificacion == 2:
                    suma_notausointerno+=2
                else:
                    None

            elif calif.pregunta.id == 209:
                if calif.calificacion == 2:
                    suma_notausointerno+=2
                else:
                    None
            else:
                print("No hay la pregunta")

        print('Aqui la nota final',suma_notausointerno)

    
        if suma_notausointerno == 9:
            gradev = 'A'
        if suma_notausointerno >6 and suma_notausointerno <9:
            gradev = 'B'
        if suma_notausointerno >1 and suma_notausointerno <7:
            gradev = 'C'
        if suma_notausointerno == 1:
            gradev = 'F'
    
   
        consulta2 = proveedor.objects.get(id=pk)
        consulta2.grade = gradev
        consulta2.save()

        v = proveedor.objects.filter(id=pk)

        for nota in v:
            califica = nota.grade
            catego = nota.categoria

        return render(request,'proveedores/ver_respuestas_individual.html',{'form':parametros,'form3':parametros2,'form4':parametros3,'nombre':nombre,'id_prove':id_prove,'segmento_escogido':segmento_escogido,'suma_total':suma_total,'suma_prove1':suma_prove1,'califica':califica,'categoria':catego})


def verRespuestasEspecificas(request):
    
    categoriacomp = request.GET.get("categoria")
    proveedor1 = request.GET.get("proveedor1")
    proveedor2  = request.GET.get("proveedor2")
    proveedor3  = request.GET.get("proveedor2")
    segmento = request.GET.get("segmento")
    
    if categoriacomp:
        if categoriacomp and proveedor1 and proveedor2:
            if categoriacomp and proveedor1 and proveedor2 and segmento:
                print(segmento)
                print("SSSSSSSSSSSSSSSSssYY")

                proveedorx = proveedor.objects.filter(categoria=categoriacomp)
            

                formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedorx,proveedor2=proveedorx)

                formulario_segmento = FiltrarFormSegmento(request.GET or None)


                prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=proveedor1).order_by('pregunta').filter(pregunta__segmento=segmento)
            
                prove2 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=proveedor2).order_by('pregunta').filter(pregunta__segmento=segmento)
            
                # vec=prove1.union(prove2)
                
                form = []
                
                for x,y in zip(prove1,prove2):
                    vec = {
                        'pregunta':x.pregunta,
                        'proveedor1':x.respuesta,
                        'calificacion1':x.calificacion,
                        'pregunta2':y.pregunta,
                        'proveedor2':y.respuesta,
                        'califiacion2':y.calificacion
                    }
                    form.append(vec)
                    
                nombre1 = proveedor.objects.get(id=proveedor1)
                nombre2 = proveedor.objects.get(id=proveedor2)


                return render(request,'proveedores/ver_respuestas_especificas.html',{'form5':formulario_segmento,'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp,'form':form,'nombre1':nombre1,'nombre2':nombre2}) 

            print(categoriacomp)
            print("?????????????DM")

            proveedorx = proveedor.objects.filter(categoria=categoriacomp)
        

            formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedorx,proveedor2=proveedorx)

            formulario_segmento = FiltrarFormSegmento(request.GET or None)


            prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=proveedor1).filter(pregunta__categoria_proveedor_id=categoriacomp).exclude(respuesta='xxx').order_by('pregunta')
        
            prove2 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=proveedor2).filter(pregunta__categoria_proveedor_id=categoriacomp).exclude(respuesta='xxx').order_by('pregunta')
        
            #Las siguientes líneas seirven para realizar el cálculo de la calificación en las preguntas específicas

            calculaCalificacion(prove1)
            calculaCalificacion(prove2)

            
            suma_prove1 = 0
            for nota in prove1:
                suma_prove1+=nota.calificacion
            
            suma_prove2 = 0
            for nota in prove2:
                suma_prove2+=nota.calificacion

                
            form = []
            
            for x,y in zip(prove1,prove2):
                vec = {
                        'pregunta':x.pregunta,
                        'proveedor1':x.respuesta,
                        'calificacion1':x.calificacion,
                        'pregunta2':y.pregunta,
                        'proveedor2':y.respuesta,
                        'calificacion2':y.calificacion
                    }
                form.append(vec)
                  
            nombre1 = proveedor.objects.get(id=proveedor1)
            nombre2 = proveedor.objects.get(id=proveedor2)
            
            
            

            return render(request,'proveedores/ver_respuestas_especificas.html',{'form5':formulario_segmento,'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp,'form':form,'nombre1':nombre1,'nombre2':nombre2,'suma_prove1':suma_prove1,'suma_prove2':suma_prove2}) 

        
        print(categoriacomp)
        print("!!!!!!!!!!!!!!!!!!!!")

        proveedor1 = proveedor.objects.filter(categoria=categoriacomp)
        

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedor1,proveedor2=proveedor1)

        
        return render(request,'proveedores/ver_respuestas_especificas.html',{'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp}) 

    else:
        
        print("Enntró")

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None)

       
        return render(request,'proveedores/ver_respuestas_especificas.html',{'form4':formulario_resultados_consulta}) 



def verPreguntasEspecificas(request):
    
    categoriacomp = request.GET.get("categoria_proveedor")
    
    if categoriacomp:
        print("Entra con categoria")
        cod_etica = proveedor_encuesta.objects.filter(id=274)
        categoria = proveedor_encuesta.objects.filter(
            segmento__id=14).filter(categoria_proveedor__id=categoriacomp).order_by('pregunta_campo_modelo_id')

        preguntas = categoria | cod_etica
        
        formulario_resultados_consulta = FiltrarFormDet(request.GET or None)
        
        return render(request,'proveedores/ver_preguntas_especificas.html',{'form4':formulario_resultados_consulta,'form':preguntas}) 

    else:
        print("Entra sin categoria")

        preguntas = proveedor_encuesta.objects.filter(
            segmento__id=14).order_by('pregunta_campo_modelo_id')
            
        formulario_resultados_consulta = FiltrarFormDet(request.GET or None)

        return render(request,'proveedores/ver_preguntas_especificas.html',{'form4':formulario_resultados_consulta,'form':preguntas}) 


def verCalifSegmento(request):
    
    categoriacomp = request.GET.get("tipo")
    segmento = request.GET.get("segmento")
    proveedor_nombre  = request.GET.get("proveedor")

    #prove = proveedor.objects.all().order_by("nombre_empresa").filter(respondio_encuesta=True).filter(tipo_empresa="Persona Natural")

    prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(pregunta__segmento=1).order_by(
                        'proveedor_id__nombre_empresa','pregunta_id').filter(
                            proveedor_id__respondio_encuesta=True).filter(
                                proveedor_id__tipo_empresa="Persona Natural").distinct('proveedor_id__nombre_empresa')
    
    print(prove)

    pregu_compromisos = proveedor_encuesta.objects.all().filter(segmento=1).order_by('pregunta_campo_modelo_id')




    
    return render(request,'proveedores/ver_calif_x_tipo_x_segmento.html',{'form':prove,'pregu_compromisos':pregu_compromisos}) 


def verPlantillas(request):
    
    categoria = request.GET.get("categoria_proveedor")
    form = FiltrarFormDet(request.GET or None)

    print(categoria)

    if categoria:
        print(categoria,type(categoria))

        
        if categoria == '3':  #PLAGAS
            plantilla = proveedor_encuesta_control_plagas.objects.all().select_related('pregunta_id').order_by('ordenamiento')

        elif categoria == '14':  #QUIMICOS PARA EL CALDERO
            plantilla = proveedor_encuesta_quimicos_caldero.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        

        elif categoria == '20':  #PRODUCTOS Y SERVICIOS De USO AGRICOLA
            plantilla = proveedor_encuesta_control_plagas.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        
        elif categoria == '19':  #PRODUCTOS Y SERVICIOS VARIOS
            plantilla = proveedor_encuesta_prod_serv_varios.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        
        elif categoria == '11':  #INSUMOS
            plantilla = proveedor_encuesta_laboratorio_insumos.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        
        elif categoria == '17':  #TRANSPORTE OTROS
            plantilla = proveedor_encuesta_transporte_otros.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        
        elif categoria == '9':  #LABORATORIO ANALISIS
            plantilla = proveedor_encuesta_laboratorio_analisis.objects.all().select_related('pregunta_id').order_by('ordenamiento') 
        
        try:
            if plantilla:
                return render(request,'proveedores/ver_plantillas.html',{'plantilla':plantilla,'form':form,'categoria':categoria}) 

        except:
            return HttpResponse("Categoria de proveedor no configurada")   
            

            
    return render(request,'proveedores/ver_plantillas.html',{'form':form}) 








def verRespuestasCompara(request):
    
    categoriacomp = request.GET.get("categoria")
    proveedor1 = request.GET.get("proveedor1")
    proveedor2  = request.GET.get("proveedor2")
    segmento = request.GET.get("segmento")
    
    if categoriacomp:
        if categoriacomp and proveedor1 and proveedor2:
            if categoriacomp and proveedor1 and proveedor2 and segmento:
                print(type(segmento))
                print("SSSSSSSSSSSSSSSSss")

                proveedorx = proveedor.objects.filter(categoria=categoriacomp)
            

                formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedorx,proveedor2=proveedorx)

                formulario_segmento = FiltrarFormSegmento(request.GET or None)


                prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=proveedor1).order_by('pregunta__pregunta_campo_modelo_id').filter(pregunta__segmento=segmento).exclude(respuesta='xxx')
            
                prove2 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                'proveedor_id').filter(proveedor_id__id=proveedor2).order_by('pregunta__pregunta_campo_modelo_id').filter(pregunta__segmento=segmento).exclude(respuesta='xxx')

                if segmento == '3':
                    tipo_empresa1 = str(prove1[0])
                    tipo_empresa2 = str(prove2[0])
                
                    print(type(tipo_empresa1))
                    print(tipo_empresa2)
                    
                    if tipo_empresa1 == 'Persona Jurídica' and tipo_empresa2 == 'Persona Natural':
                        print("Entra DMP")
                        prove2 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                            'proveedor_id').filter(proveedor_id__id=proveedor2).order_by('pregunta__segundo_ordenamiento').filter(pregunta__segmento=segmento).exclude(respuesta='xxx')

                    elif tipo_empresa1 == 'Persona Natural' and tipo_empresa2 == 'Persona Jurídica':
                        print("Entra 2DMP")
                        prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                            'proveedor_id').filter(proveedor_id__id=proveedor1).order_by('pregunta__segundo_ordenamiento').filter(pregunta__segmento=segmento).exclude(respuesta='xxx')

                  
                # vec=prove1.union(prove2)
                
                calculaCalificacion(prove1) 
                calculaCalificacion(prove2) 

                suma_prove1 = 0
                for nota in prove1:
                    suma_prove1+=nota.calificacion
                
                if suma_prove1 > 100:
                    suma_prove1 = 100
                    
                suma_prove2 = 0
                for nota in prove2:
                    suma_prove2+=nota.calificacion
                
                if suma_prove2 > 100:
                    suma_prove2 = 100


                form = []
                
                for x,y in zip(prove1,prove2):
                    vec = {
                        'pregunta':x.pregunta,
                        'proveedor1':x.respuesta,
                        'calificacion1':x.calificacion,
                        'pregunta2':y.pregunta,
                        'proveedor2':y.respuesta,
                        'calificacion2':y.calificacion,
                    }
                    form.append(vec)
                    
                nombre1 = proveedor.objects.get(id=proveedor1)
                nombre2 = proveedor.objects.get(id=proveedor2)
                print(nombre1)

                
                return render(request,'proveedores/ver_respuestas_comp.html',{'form5':formulario_segmento,'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp,'form':form,'nombre1':nombre1,'nombre2':nombre2,'suma_prove1':suma_prove1,'suma_prove2':suma_prove2}) 

            print(categoriacomp)
            print("?????????????XX")

            proveedorx = proveedor.objects.filter(categoria=categoriacomp)
        

            formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedorx,proveedor2=proveedorx)

            formulario_segmento = FiltrarFormSegmento(request.GET or None)


            prove1 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=proveedor1).exclude(respuesta='xxx').exclude(respuesta='XXX').order_by('pregunta')
        
            prove2 = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
            'proveedor_id').filter(proveedor_id__id=proveedor2).exclude(respuesta='xxx').exclude(respuesta='XXX').order_by('pregunta')
        
            # vec=prove1.union(prove2)
            calculaCalificacion(prove1) 
            calculaCalificacion(prove2) 

            suma_prove1 = 0
            for nota in prove1:
                suma_prove1+=nota.calificacion
            
            if suma_prove1 > 100:
                suma_prove1 = 100
                
            suma_prove2 = 0
            for nota in prove2:
                suma_prove2+=nota.calificacion
            
            if suma_prove2 > 100:
                suma_prove2 = 100
            

            form = []
            
            for x,y in zip(prove1,prove2):
                vec = {
                    'pregunta':x.pregunta,
                    'proveedor1':x.respuesta,
                    'calificacion1':x.calificacion,
                    'pregunta2':y.pregunta,
                    'proveedor2':y.respuesta,
                    'calificacion2':y.calificacion,
                }
                form.append(vec)
                  
            nombre1 = proveedor.objects.get(id=proveedor1)
            nombre2 = proveedor.objects.get(id=proveedor2)
            print(nombre1)
            print(suma_prove1)
            print(suma_prove2)
            
            
            return render(request,'proveedores/ver_respuestas_comp.html',{'form5':formulario_segmento,'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp,'form':form,'nombre1':nombre1,'nombre2':nombre2,'suma_prove1':suma_prove1,'suma_prove2':suma_prove2}) 

        

        
        print(categoriacomp)
        print("!!!!!!!!!!!!!!!!!!!!")

        proveedor1 = proveedor.objects.filter(categoria=categoriacomp)
        

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None,categoria=categoriacomp, proveedor1=proveedor1,proveedor2=proveedor1)

        
        return render(request,'proveedores/ver_respuestas_comp.html',{'form4':formulario_resultados_consulta,'categoriacomp':categoriacomp}) 

    else:
        
        print("Enntró")

        formulario_resultados_consulta = ResultadoConsultaForm(request.GET or None)

       
        return render(request,'proveedores/ver_respuestas_comp.html',{'form4':formulario_resultados_consulta}) 

def confirma_completo(request,pk,prove):
    
    print(pk)
    updocu = proveedor_documentos.objects.filter(id=pk,proveedor=prove).update(completo=True)
    prove_nombre = proveedor.objects.get(id=prove)

    try:
        docu_cargado = documentos_prove.objects.get(proveedor=prove,archivos='FICHAS_TECNICAS')
    
    except documentos_prove.DoesNotExist:
        docu_cargado = None
    
    if docu_cargado:

        print(docu_cargado)

        actualiza = documentos_prove.objects.filter(id=docu_cargado.id).delete()
        print(pk,prove)
        upd2= proveedor_documentos.objects.filter(proveedor=prove_nombre,nombre_documento='FICHAS_TECNICAS').update(completo=False)
        
    else:
        r = documentos_prove(
            archivos = 'FICHAS_TECNICAS',
            proveedor = prove_nombre,
            nombre_corto = 'FICHAS_TECNICAS',
            es_ficha = None,
            )
        r.save()
    

    return redirect ('proveedores:ver_documentos', prove )


def confirma_completo_agricola(request,pk,prove,tipo):
    
    print(pk)
    print(tipo)

    updocu = proveedor_documentos.objects.filter(id=pk,proveedor=prove).update(completo=True)
    prove_nombre = proveedor.objects.get(id=prove)

    try:
        if tipo == 'FICHAS_TECNICAS_PESTICIDAS':
            docu_cargado = documentos_prove.objects.get(proveedor=prove,archivos='FICHAS_TECNICAS_PESTICIDAS')
        elif tipo == 'FICHAS_TECNICAS_FOLIARES':
            docu_cargado = documentos_prove.objects.get(proveedor=prove,archivos='FICHAS_TECNICAS_FOLIARES')
        elif tipo == 'FICHAS_TECNICAS_FERTILIZANTES':
            docu_cargado = documentos_prove.objects.get(proveedor=prove,archivos='FICHAS_TECNICAS_FERTILIZANTES')
        elif tipo == 'FICHAS_TECNICAS_MATERIA_ORGANICA':
            docu_cargado = documentos_prove.objects.get(proveedor=prove,archivos='FICHAS_TECNICAS_MATERIA_ORGANICA')
    


    except documentos_prove.DoesNotExist:
        docu_cargado = None
    
    if docu_cargado:

        actualiza = documentos_prove.objects.filter(id=docu_cargado.id).delete()
    
        upd2= proveedor_documentos.objects.filter(proveedor=prove_nombre,nombre_documento=tipo).update(completo=False)
        
    else:
        
        if tipo == 'FICHAS_TECNICAS_PESTICIDAS':
        
            r = documentos_prove(
                archivos = 'FICHAS_TECNICAS_PESTICIDAS',
                proveedor = prove_nombre,
                nombre_corto = 'FICHAS_TECNICAS_PESTICIDAS',
                es_ficha = None,
                )
            r.save()
        
        elif tipo == 'FICHAS_TECNICAS_FOLIARES':
            r = documentos_prove(
                archivos = 'FICHAS_TECNICAS_FOLIARES',
                proveedor = prove_nombre,
                nombre_corto = 'FICHAS_TECNICAS_FOLIARES',
                es_ficha = None,
                )
            r.save()
        
        elif tipo == 'FICHAS_TECNICAS_FERTILIZANTES':
            r = documentos_prove(
                archivos = 'FICHAS_TECNICAS_FERTILIZANTES',
                proveedor = prove_nombre,
                nombre_corto = 'FICHAS_TECNICAS_FERTILIZANTES',
                es_ficha = None,
                )
            r.save()
    
        elif tipo == 'FICHAS_TECNICAS_MATERIA_ORGANICA':
            r = documentos_prove(
                archivos = 'FICHAS_TECNICAS_MATERIA_ORGANICA',
                proveedor = prove_nombre,
                nombre_corto = 'FICHAS_TECNICAS_MATERIA_ORGANICA',
                es_ficha = None,
                )
            r.save()

    return redirect ('proveedores:ver_documentos', prove )

#Funciòn para recoger Registros Agrocalidad, Certificados y Fichas Tecnicas Cat. Agricola

@login_required
def accede_menu_subcategorias_agricolas_pes(request,pk,respuesta):
    prove=proveedor.objects.get(id=pk)
    subcate='PESTICIDAS'

    prove = proveedor.objects.get(id=pk)
    subcate_id=1
    
    return render(request,'proveedores/ver_menu_subcategorias.html',{'prove':prove,'subcate':subcate,'subcate_id':subcate_id}) 

def accede_menu_subcategorias_agricolas_fol(request,pk,respuesta):
    prove=proveedor.objects.get(id=pk)
    subcate='FOLIARES'

    prove = proveedor.objects.get(id=pk)
    
    return render(request,'proveedores/ver_menu_subcategorias.html',{'prove':prove,'subcate':subcate}) 

def accede_menu_subcategorias_agricolas_fer(request,pk,respuesta):
    prove=proveedor.objects.get(id=pk)
    subcate='FERTILIZANTES'

    prove = proveedor.objects.get(id=pk)
    
    return render(request,'proveedores/ver_menu_subcategorias.html',{'prove':prove,'subcate':subcate}) 

def accede_menu_subcategorias_agricolas_mao(request,pk,respuesta):
    prove=proveedor.objects.get(id=pk)
    subcate='MATERIA ORGANICA'

    prove = proveedor.objects.get(id=pk)
    
    return render(request,'proveedores/ver_menu_subcategorias.html',{'prove':prove,'subcate':subcate}) 





def updateHead(pk):
    # pk = 59
    r6 = proveedor_respuestas.objects.get(pregunta=6,proveedor_id= pk)  #Dirección
    r7 = proveedor_respuestas.objects.get(pregunta=7,proveedor_id= pk) #Horario de Trabajo
    r8 = proveedor_respuestas.objects.get(pregunta=8,proveedor_id= pk)  #Rep. Legal
    r10 = proveedor_respuestas.objects.get(pregunta=10,proveedor_id= pk) # Persona que se pone en contacto con Ecofroz
    r11 = proveedor_respuestas.objects.get(pregunta=11,proveedor_id= pk) #Teléfono Fijo
    r12 = proveedor_respuestas.objects.get(pregunta=12,proveedor_id= pk) #Celular
    r14 = proveedor_respuestas.objects.get(pregunta=14,proveedor_id= pk) #Giro de Negocio
    r314 = proveedor_respuestas.objects.get(pregunta=314,proveedor_id= pk) #Ventas (USD) de productos y/o servicios en el 2020
    r16 = proveedor_respuestas.objects.get(pregunta=16,proveedor_id= pk) #Ventas (USD) de productos y/o servicios en el 2019
    r17 = proveedor_respuestas.objects.get(pregunta=17,proveedor_id= pk) #N° de años en el mercado/antigüedad
    r18 = proveedor_respuestas.objects.get(pregunta=18,proveedor_id= pk) #N° de trabajadores fijos
    r19 = proveedor_respuestas.objects.get(pregunta=19,proveedor_id= pk) #Numero de clientes
    r20 = proveedor_respuestas.objects.get(pregunta=20,proveedor_id= pk) #Numero de proveedores
    r21 = proveedor_respuestas.objects.get(pregunta=21,proveedor_id= pk) #3 Valores
    r23 = proveedor_respuestas.objects.get(pregunta=23,proveedor_id= pk) #Direccion Web
    r24 = proveedor_respuestas.objects.get(pregunta=24,proveedor_id= pk) # De acuerdo a la siguiente clasificación de proveedores, escoja la actividad que su empresa realiza 
    r22 = proveedor_respuestas.objects.get(pregunta=22,proveedor_id= pk) #Tiene página web
    r25 = proveedor_respuestas.objects.get(pregunta=25,proveedor_id= pk)  #Persona Natural o Jurídica

    #r37 = proveedor_respuestas.objects.get(pregunta=37,proveedor_id= pk) #Ventas a Ecofroz 2018
    r38 = proveedor_respuestas.objects.get(pregunta=38,proveedor_id= pk) #Ventas (USD) de productos y/o servicios a ECOFROZ S.A en el 2019
    r315 = proveedor_respuestas.objects.get(pregunta=315,proveedor_id= pk) #Ventas (USD) de productos y/o servicios a ECOFROZ S.A en el 2020
    
    print(r6,type(r6))
    
    consulta = proveedor.objects.get(id=pk)
    #proveedor.objects.filter(id=pk).update(respuesta=r6)
    consulta.respondio_encuesta = True
    consulta.direccion_matriz = str(r6)
    consulta.save()
    consulta.horario_trabajo = str(r7)
    consulta.representante_legal = str(r8)
    consulta.nombre_contacto_ecofroz = str(r10)
    consulta.telefono = str(r11)
    consulta.celular = str(r12)
    consulta.tipo_empresa = r25.respuesta
    consulta.save()

    consulta2 = proveedor_det.objects.get(codigo_id=pk)
    consulta2.giro_negocio = str(r14)
    #consulta2.monto_ventas_2018 = str(r15)
    consulta2.monto_ventas_2019 = str(r16)
    consulta2.monto_ventas_2020 = str(r314)
    consulta2.antiguedad = r17.respuesta
    consulta2.num_trabajadores_fijos = r18.respuesta
    consulta2.num_clientes = r19.respuesta
    consulta2.num_proveedores = r20.respuesta
    consulta2.valores_empresa = str(r21)
    consulta2.web = str(r23)
    consulta2.empresa_tipo = r25.respuesta
    #consulta2.ventas_ecofroz_2018 = r37.respuesta
    consulta2.ventas_ecofroz_2019 = r38.respuesta
    consulta2.ventas_ecofroz_2020 = r315.respuesta
    consulta2.tieneweb = r22.respuesta
   
    consulta2.save()
    

    return None


def exportExcelGeneral(request):

    query = proveedor_det.objects.all().select_related('codigo_id').order_by('codigo_id__nombre_empresa','codigo_id__respondio_encuesta')

    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'Reporte General de Proveedores Ecofroz S.A'

    ws.merge_cells('B1:K1')
    ws['A3'] = 'No'
    ws['B3'] = 'PROVEEDOR'
    ws['C3'] = 'TIPO'
    ws['D3'] = 'RUC'
    ws['E3'] = 'CRITICO'
    ws['F3'] = 'CATEGORIA'
    ws['G3'] = 'CALIFICACION PROVEEDOR'
    ws['H3'] = 'CALIFICACION ADQUISICIONES'
    ws['I3'] = 'GIRO NEGOCIO'
    ws['J3'] = 'REPRESENTANTE LEGAL'
    ws['K3'] = 'PERSONA SE CONTACTA CON ECOFROZ'
    ws['L3'] = 'CELULAR'
    ws['M3'] = 'TELEFONO FIJO'
    ws['N3'] = 'NUMERO TRABAJADORES FIJOS'
    ws['O3'] = 'VENTAS TOTALES 2018'
    ws['P3'] = 'VENTAS TOTALES 2019'
    ws['Q3'] = 'VENTAS TOTALES 2020'
    ws['R3'] = 'VENTAS TOTALES 2021'
    ws['S3'] = 'VENTAS A ECOFROZ 2018'
    ws['T3'] = 'VENTAS A ECOFROZ 2019'
    ws['U3'] = 'VENTAS A ECOFROZ 2020'
    ws['V3'] = 'VENTAS A ECOFROZ 2021'
    ws['W3'] = 'NUMERO DOCUMENTOS CARGADOS'
    ws['X3'] = 'RESPONDIO ENCUESTA'
    ws['Y3'] = 'DIRECCION'
    ws['Z3'] = 'FECHA RESPUESTA ENCUESTA'
    #ws['AA3'] = 'ID'
    

    count = 4
    rowcount = 1

    for registro in query:
        ws.cell(row = count, column = 1).value =  rowcount
        ws.cell(row = count, column = 2).value =  registro.codigo_id.nombre_empresa
        ws.cell(row = count, column = 3).value =  registro.empresa_tipo
        ws.cell(row = count, column = 4).value =  registro.codigo_id.ruc
        ws.cell(row = count, column = 5).value =  registro.codigo_id.proveedor_critico
        ws.cell(row = count, column = 6).value =  str(registro.codigo_id.categoria)
        ws.cell(row = count, column = 7).value =  registro.codigo_id.calificacion
        ws.cell(row = count, column = 8).value =  registro.codigo_id.grade
        ws.cell(row = count, column = 9).value =  registro.giro_negocio
        ws.cell(row = count, column = 10).value =  registro.codigo_id.representante_legal
        ws.cell(row = count, column = 11).value =  registro.codigo_id.nombre_contacto_ecofroz
        ws.cell(row = count, column = 12).value =  registro.codigo_id.celular 
        ws.cell(row = count, column = 13).value =  registro.codigo_id.telefono 
        ws.cell(row = count, column = 14).value =  registro.num_trabajadores_fijos       
        ws.cell(row = count, column = 15).value =  round(float(registro.monto_ventas_2018),0)
        ws.cell(row = count, column = 16).value =  round(float(registro.monto_ventas_2019),0)
        ws.cell(row = count, column = 17).value =  round(float(registro.monto_ventas_2020),0)
        ws.cell(row = count, column = 18).value =  round(float(registro.monto_ventas_2021),0)
        ws.cell(row = count, column = 19).value =  round(float(registro.ventas_ecofroz_2018),0)
        ws.cell(row = count, column = 20).value =  round(float(registro.ventas_ecofroz_2019),0)
        ws.cell(row = count, column = 21).value =  round(float(registro.ventas_ecofroz_2020),0)
        ws.cell(row = count, column = 22).value =  round(float(registro.ventas_ecofroz_2021),0)
        
        ws.cell(row = count, column = 23).value =  registro.num_documentos_cargados
        ws.cell(row = count, column = 24).value =  registro.codigo_id.respondio_encuesta
        ws.cell(row = count, column = 25).value =  registro.codigo_id.direccion_matriz
        ws.cell(row = count, column = 26).value =  str(registro.codigo_id.fecha_eleboracion)
        #ws.cell(row = count, column = 25).value =  registro.codigo_id.id

        count+=1
        rowcount+=1
    
    
    nombre_archivo = "Reporte_General_de_Proveedores_Ecofroz.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response




def exportExcelRespuestasXCategoria(request):
    categoria = request.GET.get("categoria")
    # segmento = request.GET.get("segmento")
    segmento = request.GET.get("segmento2")

    print(categoria)
    print(segmento)

    if categoria:

    
        prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(proveedor_id__categoria=categoria).order_by('proveedor_id__id','pregunta_id')
        
    
        
        pregu = proveedor_respuestas.objects.all().select_related('pregunta').filter(proveedor_id__categoria=categoria).order_by(
            'pregunta')
        
        form = []
                    
        for x in prove:
            vec = {
                'pregunta':x.pregunta,
                'respuesta':x.respuesta,
            }
            form.append(vec)
                
        print(form)
        
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Preguntas y Respuestas Formulario de Proveedores Ecofroz S.A'

        ws.merge_cells('B1:K1')
        ws['A3'] = 'No'
        ws['B3'] = 'PREGUNTA'
        ws['C3'] = 'PROVEEDOR1'
        ws['D3'] = 'PROVEEDOR2'
        ws['E3'] = 'PROVEEDOR3'
        ws['F3'] = 'PROVEEDOR4'
        ws['G3'] = 'PROVEEDOR5'
        ws['H3'] = 'PROVEEDOR6'
        ws['I3'] = 'PROVEEDOR7'
        ws['J3'] = 'PROVEEDOR8'
        ws['K3'] = 'PROVEEDOR9'
        ws['L3'] = 'PROVEEDOR10'
        ws['M3'] = 'PROVEEDOR11'
        ws['N3'] = 'PROVEEDOR12'
        ws['O3'] = 'PROVEEDOR13'
        ws['P3'] = 'PROVEEDOR14'
        ws['Q3'] = 'PROVEEDOR15'
        ws['R3'] = 'PROVEEDOR16'
        ws['S3'] = 'PROVEEDOR17'
        ws['T3'] = 'PROVEEDOR18'
        ws['U3'] = 'PROVEEDOR19'
        ws['V3'] = 'PROVEEDOR20'
        ws['W3'] = 'PROVEEDOR21'
        ws['X3'] = 'PROVEEDOR22'
        ws['Y3'] = 'PROVEEDOR23'
        ws['Z3'] = 'PROVEEDOR24'
        ws['AA3'] = 'PROVEEDOR25'
        ws['AB3'] = 'PROVEEDOR26'
        ws['AC3'] = 'PROVEEDOR27'
        ws['AD3'] = 'PROVEEDOR28'
        ws['AE3'] = 'PROVEEDOR29'
        ws['AF3'] = 'PROVEEDOR30'
        ws['AG3'] = 'PROVEEDOR31'
        ws['AH3'] = 'PROVEEDOR32'
        ws['AI3'] = 'PROVEEDOR33'
        ws['AJ3'] = 'PROVEEDOR34'
        ws['AK3'] = 'PROVEEDOR35'
        ws['AL3'] = 'PROVEEDOR36'
        ws['AM3'] = 'PROVEEDOR37'
        ws['AN3'] = 'PROVEEDOR38'
        ws['AO3'] = 'PROVEEDOR39'
        ws['AP3'] = 'PROVEEDOR40'

            
        count = 4
        column_num = 3
        column_preg = 2
        contador = 0
        rowcount = 1  

        for registro in form:
            if str(registro['pregunta']) == 'Fecha de elaboración del cuestionario (El formulario debe ser enviado máximo en los próximos 8 días laborables)' and contador >0:
                contador+=1
                count=4
                column_num+=1
                column_preg+=2
                rowcount=1 

            ws.cell(row = count, column = 1).value =  rowcount
            ws.cell(row = count, column = 2).value =  str(registro['pregunta'])
            ws.cell(row = count, column = column_num).value =  registro['respuesta']
                            
            count+=1
            contador+=1
            rowcount+=1

            print(registro['pregunta'])
        
            
            

        nombre_archivo = "Reporte_de_Respuestas_Encuesta_Proveedores_Ecofroz_Proveedor_.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

    elif segmento:
        
        prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
                    'proveedor_id').filter(pregunta__segmento=segmento).filter(
                        proveedor_id__respondio_encuesta=True).exclude(pregunta__id=13).exclude(
                            pregunta__id=99).exclude(pregunta__id=100).exclude(pregunta__id=125).exclude(
                                pregunta__id=126).exclude(pregunta__id=213).exclude(pregunta__id=215).exclude(pregunta__id=216).exclude(
                                    pregunta__id=218).exclude(pregunta__id=219).exclude(respuesta='xxx').exclude(
                            respuesta='XXX').order_by('proveedor_id__nombre_empresa','pregunta__id','pregunta__segmento')
       

        proveedor_list = proveedor.objects.all().filter(respondio_encuesta=True).order_by('nombre_empresa')

        final_prove_list = []
        for i in proveedor_list:
            if i not in final_prove_list:
                final_prove_list.append(i)


        pregu = []
        resp =  []
        criti =  []
        tipo_emp = []

        final_pregu = []
        prove_nombre = []

        for x in final_prove_list:
            vec = {
                'critico':x.proveedor_critico,
                #'critico':x.nombre_empresa,
            }
            criti.append(vec)
       
        for x in final_prove_list:
            vec = {
                'tipo_empre':x.tipo_empresa,
            }
            tipo_emp.append(vec)

        for x in prove:
            vec = {
                'respuesta':x.respuesta,
            }
            resp.append(vec)
                    
       
        for x in prove:
            vec = {
                'pregunta':str(x.pregunta),
            }
            pregu.append(vec)
        
        print("Tamaño de pregu",len(pregu))
        
        #Descartar duplicados
        final_pregu = []
        for i in pregu:
            if i not in final_pregu:
                final_pregu.append(i)
        
        print("Tamaño de final pregu",len(final_pregu))
        
        for x in prove:
            vec2 = {
                'nombre':x.proveedor_id.nombre_empresa,
            }
            prove_nombre.append(vec2)
        
        print("Numero total de registros: ",len(prove_nombre))
        
        #Descartar duplicados
        final = []
        for i in prove_nombre:
            if i not in final:
                final.append(i)
        
        print("Numero total de proveedores: ",len(final))
        # print(final)
        
             
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Respuestas de proveedores por Segmento de preguntas'

        posicionx = ['K3','L3','M3','N3','O3','P3','Q3','R3','S3','T3','U3','V3','W3','X3','Y3','Z3',
                        'AA3','AB3','AC3','AD3','AE3','AF3','AG3']
        # for x in range(len(final_pregu)):
        #     x = str('I'+str(x+3))
        #     posicionx.append(x)

        
        posiciony = []
        for x in range(len(final)):
            x = str('B'+str(x+4))
            posiciony.append(x)
        
        posicionycriti = []
        for x in range(len(final)):
            x = str('I'+str(x+4))
            posicionycriti.append(x)
        
        posicionytipo = []
        for x in range(len(final)):
            x = str('J'+str(x+4))
            posicionytipo.append(x)

        
        ws.merge_cells('B1:K1')
        ws['A3'] = 'No'
        ws['B3'] = 'PROVEEDOR'
        ws['I3'] = 'CRITICO'
        ws['J3'] = 'TIPO EMPRESA'

        for x,y in zip(final_pregu,posicionx):
            ws[y] = x['pregunta']
        
        
        for x,y in zip(final,posiciony):
            ws[y] = x['nombre']  

        for x,y in zip(criti,posicionycriti):
            ws[y] = x['critico']  

        for x,y in zip(tipo_emp,posicionytipo):
            ws[y] = x['tipo_empre'] 


        sizepregu = len(final_pregu)

        print(sizepregu)

        count = 0
        column_num = 11
        rowcount = 4  

        
        for registro in prove:
            if count < sizepregu:
                ws.cell(row = rowcount, column = column_num).value =  registro.respuesta
                column_num+=1
                count+=1
            else:
                count=0
                column_num = 11
                rowcount+=1
                ws.cell(row = rowcount, column = column_num).value =  registro.respuesta
                column_num+=1
                count+=1


        #     print(registro['pregunta'])
        
            
            

        nombre_archivo = "Reporte_de_Respuestas_Encuesta_Proveedores_por_Segmento.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
        


    
def exportExcelEncuesta(request,pk):

    
    prove = proveedor_respuestas.objects.all().select_related('pregunta').select_related(
        'proveedor_id').filter(proveedor_id__id=pk).order_by('pregunta')
    
    nombre = proveedor.objects.all().get(id=pk)

      
    wb = Workbook()
    ws = wb.active
    ws['B1'] = 'Reporte de Preguntas y Respuestas Formulario de Proveedores Ecofroz S.A, Proveedor: '+nombre.nombre_empresa

    ws.merge_cells('B1:K1')
    ws['B3'] = 'No'
    ws['C3'] = 'PREGUNTA'
    ws['D3'] = 'PROVEEDOR1'
    ws['E3'] = 'PROVEEDOR2'
    ws['F3'] = 'PROVEEDOR3'
    ws['G3'] = 'PROVEEDOR4'
    ws['H3'] = 'PROVEEDOR5'
    ws['I3'] = 'PROVEEDOR6'
    ws['J3'] = 'PROVEEDOR7'
    ws['K3'] = 'PROVEEDOR8'
    ws['L3'] = 'PROVEEDOR9'
    ws['M3'] = 'PROVEEDOR10'
    ws['N3'] = 'PROVEEDOR11'
    ws['O3'] = 'PROVEEDOR12'
    ws['P3'] = 'PROVEEDOR13'
    ws['Q3'] = 'PROVEEDOR14'
    ws['R3'] = 'PROVEEDOR15'
    ws['S3'] = 'PROVEEDOR16'
    ws['T3'] = 'PROVEEDOR17'
    ws['U3'] = 'PROVEEDOR18'
    ws['V3'] = 'PROVEEDOR19'
    ws['W3'] = 'PROVEEDOR20'
    ws['X3'] = 'PROVEEDOR21'
    ws['Y3'] = 'PROVEEDOR22'
    ws['Z3'] = 'PROVEEDOR23'
    ws['AA3'] = 'PROVEEDOR24'
    ws['AB3'] = 'PROVEEDOR25'
    ws['AC3'] = 'PROVEEDOR26'
    ws['AD3'] = 'PROVEEDOR27'
    ws['AE3'] = 'PROVEEDOR28'
    ws['AF3'] = 'PROVEEDOR29'
    ws['AG3'] = 'PROVEEDOR30'
    ws['AH3'] = 'PROVEEDOR31'
    ws['AI3'] = 'PROVEEDOR32'
    ws['AJ3'] = 'PROVEEDOR33'
    ws['AK3'] = 'PROVEEDOR34'
    ws['AL3'] = 'PROVEEDOR35'
    ws['AM3'] = 'PROVEEDOR36'
    ws['AN3'] = 'PROVEEDOR37'
    ws['AO3'] = 'PROVEEDOR38'
    ws['AP3'] = 'PROVEEDOR39'
    ws['AQ3'] = 'PROVEEDOR40'
    
        
    count = 4
    rowcount = 1  

    for registro in prove:
        ws.cell(row = count, column = 2).value =  rowcount
        ws.cell(row = count, column = 3).value =  registro.pregunta.pregunta
        ws.cell(row = count, column = 4).value =  registro.respuesta
            
                         
        count+=1
        rowcount+=1

    nombre_archivo = "Reporte_de_Respuestas_Encuesta_Proveedores_Ecofroz_Proveedor_.xlsx"
    response = HttpResponse(content_type = "application/ms-excel")
    content = "attachment; filename = {0}".format(nombre_archivo)
    response['Content-Disposition'] = content
    wb.save(response)
    return response



class to_excelclass(TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        proveedor_id = 107

        proveedor = proveedor_encuesta.objects.all().select_related('codigo_id').filter(id=proveedor_id)

        registros = proveedor_encuesta.objects.all().order_by('id')

        columns = proveedor.objects.all().order_by('id')
            
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de Preguntas y Respuestas Formulario de Proveedores Ecofroz S.A'

        ws.merge_cells('B1:I1')
        ws['B3'] = 'No'
        ws['C3'] = 'PREGUNTA'
        ws['D3'] = 'PROVEEDOR1'
        ws['E3'] = 'PROVEEDOR2'
        ws['F3'] = 'PROVEEDOR3'
        ws['G3'] = 'PROVEEDOR4'
        ws['H3'] = 'PROVEEDOR5'
        ws['I3'] = 'PROVEEDOR6'
        ws['J3'] = 'PROVEEDOR7'
        ws['K3'] = 'PROVEEDOR8'
        ws['L3'] = 'PROVEEDOR9'
        ws['M3'] = 'PROVEEDOR10'
        ws['N3'] = 'PROVEEDOR11'
        ws['O3'] = 'PROVEEDOR12'
        ws['P3'] = 'PROVEEDOR13'
        ws['Q3'] = 'PROVEEDOR14'
        ws['R3'] = 'PROVEEDOR15'
        ws['S3'] = 'PROVEEDOR16'
        ws['T3'] = 'PROVEEDOR17'
        ws['U3'] = 'PROVEEDOR18'
        ws['V3'] = 'PROVEEDOR19'
        ws['W3'] = 'PROVEEDOR20'
        ws['X3'] = 'PROVEEDOR21'
        ws['Y3'] = 'PROVEEDOR22'
        ws['Z3'] = 'PROVEEDOR23'
        ws['AA3'] = 'PROVEEDOR24'
        ws['AB3'] = 'PROVEEDOR25'
        ws['AC3'] = 'PROVEEDOR26'
        ws['AD3'] = 'PROVEEDOR27'
        ws['AE3'] = 'PROVEEDOR28'
        ws['AF3'] = 'PROVEEDOR29'
        ws['AG3'] = 'PROVEEDOR30'
        ws['AH3'] = 'PROVEEDOR31'
        ws['AI3'] = 'PROVEEDOR32'
        ws['AJ3'] = 'PROVEEDOR33'
        ws['AK3'] = 'PROVEEDOR34'
        ws['AL3'] = 'PROVEEDOR35'
        ws['AM3'] = 'PROVEEDOR36'

        cuenta = columns.count()
        count = 4
        rowcount = 1  

        for registro in registros:
            ws.cell(row = count, column = 2).value =  rowcount
            ws.cell(row = count, column = 3).value =  registro.pregunta
            ws.cell(row = count, column = 4).value =  registro.consulta
            ws.cell(row = count, column = 5).value =  registro.consulta2
            ws.cell(row = count, column = 6).value =  registro.consulta3
            ws.cell(row = count, column = 7).value =  registro.consulta4
            ws.cell(row = count, column = 8).value =  registro.consulta5
            ws.cell(row = count, column = 9).value =  registro.consulta6
            ws.cell(row = count, column = 10).value =  registro.consulta7
            ws.cell(row = count, column = 11).value =  registro.consulta8
            ws.cell(row = count, column = 12).value =  registro.consulta9
            ws.cell(row = count, column = 13).value =  registro.consulta10
            # ws.cell(row = count, column = 14).value =  registro.consulta11
            ws.cell(row = count, column = 14).value =  registro.consulta12
            ws.cell(row = count, column = 15).value =  registro.consulta13
            ws.cell(row = count, column = 16).value =  registro.consulta14
            ws.cell(row = count, column = 17).value =  registro.consulta15
            ws.cell(row = count, column = 18).value =  registro.consulta16
            ws.cell(row = count, column = 19).value =  registro.consulta17
            ws.cell(row = count, column = 20).value =  registro.consulta18
            ws.cell(row = count, column = 21).value =  registro.consulta19         
            ws.cell(row = count, column = 22).value =  registro.consulta20
            ws.cell(row = count, column = 23).value =  registro.consulta21
            ws.cell(row = count, column = 24).value =  registro.consulta22
            ws.cell(row = count, column = 25).value =  registro.consulta23
            ws.cell(row = count, column = 26).value =  registro.consulta25
            ws.cell(row = count, column = 27).value =  registro.consulta26
            ws.cell(row = count, column = 28).value =  registro.consulta27
            ws.cell(row = count, column = 39).value =  registro.consulta28
            ws.cell(row = count, column = 30).value =  registro.consulta29
            ws.cell(row = count, column = 31).value =  registro.consulta30
            ws.cell(row = count, column = 32).value =  registro.consulta31
            ws.cell(row = count, column = 33).value =  registro.consulta32
            ws.cell(row = count, column = 34).value =  registro.consulta33
            ws.cell(row = count, column = 35).value =  registro.consulta34
            ws.cell(row = count, column = 36).value =  registro.consulta35
            ws.cell(row = count, column = 37).value =  registro.consulta36
            ws.cell(row = count, column = 38).value =  registro.consulta37
            
                         
            count+=1
            rowcount+=1

        nombre_archivo = "Reporte_de_Respuestas_Encuesta_Proveedores_Ecofroz.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

def verExpiracionDocs(request):
    form = ExpiraDocumentos(request.GET or None)
    empresa = request.GET.get("nombre_empresa")
    docum = request.GET.get("documento")
    tdocu = request.GET.get("todos-docu")
    inicio = str(request.GET.get("inicio"))
    fin = str(request.GET.get("fin"))
    
    
    if empresa:
        empre = proveedor.objects.get(id=int(empresa))
        sq1 = documentos_prove.objects.all().select_related('proveedor').filter(
        nombre_corto__in=['RUC','NOMBRAMIENTO_REP_LEGAL','CUMPLIMIENTO_IESS']).filter(proveedor=int(empresa)).order_by(
        'fecha_documento')
        dc = documentos_prove.objects.filter(es_doc_categoria=True)
        qc1 = sq1 | dc 
        sq = qc1.filter(proveedor=int(empresa)).order_by('fecha_documento')
        # var = sq.count()

        query = var = None

    elif docum:
        
        q = documentos_prove.objects.all().select_related('proveedor').filter(
        nombre_corto=str(docum)).order_by(
        'fecha_documento')
        sq = empre = None

        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE proveedores.proveedor_estados_select")
        record = proveedor_estados_select()
        record.docu = str(docum)  
        record.save()
        
        paginator = Paginator(q, 150)
        page = request.GET.get('page')
        query = paginator.get_page(page)
        var = q.count()

    elif tdocu:

        q1 = documentos_prove.objects.all().select_related('proveedor').filter(
        nombre_corto__in=['RUC','NOMBRAMIENTO_REP_LEGAL','CUMPLIMIENTO_IESS'])
    
        q2 = documentos_prove.objects.filter(es_doc_categoria=True)
        q3 = q1 | q2
        q = q3.order_by('fecha_documento')
    
        sq = empre = None
        cursor = connection.cursor()
        cursor.execute("TRUNCATE TABLE proveedores.proveedor_estados_select")
        record = proveedor_estados_select()
        record.tdocu = True  
        record.save()
        
        paginator = Paginator(q, 150)
        page = request.GET.get('page')
        query = paginator.get_page(page)
        var = q.count()

    elif inicio != 'None':
        if inicio and fin:
            iniciodate = datetime.strptime(inicio, '%Y-%m-%d').date()
            findate = datetime.strptime(fin, '%Y-%m-%d' ).date()
            if inicio == fin:
                print(inicio,fin)
                print(iniciodate)
                
                q1 = documentos_prove.objects.filter(fecha_documento=iniciodate)
                query = q1.order_by('fecha_documento')
                var = q1.count()
                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE proveedores.proveedor_estados_select")
                record = proveedor_estados_select()
                record.fecini = iniciodate  
                record.fecfin = findate
                record.save()
            else:
                q1 = documentos_prove.objects.filter(fecha_documento__range=[iniciodate,findate])
                query = q1.order_by('fecha_documento')
                var = q1.count()
                cursor = connection.cursor()
                cursor.execute("TRUNCATE TABLE proveedores.proveedor_estados_select")
                record = proveedor_estados_select()
                record.fecini = iniciodate  
                record.fecfin = findate
                record.save()

            
            sq = empre = None
            
    else:
        q = sq = empre = query = tdocu = var = None
       
      
    return render(request,'proveedores/ver_documentos_expira.html',{'form':form,'query':query,'squery':sq,'empre':empre,'tdocu':tdocu, 'var':var}) 

class to_excelclass(LoginRequiredMixin,TemplateView):
    
    def get(self,request,*args,**kwargs):
        
        registro = proveedor_estados_select.objects.all().first()
        
        codigo = registro.codigo
        docu = registro.docu
        tdocu = registro.tdocu
        fecini = registro.fecini
        fecfin = registro.fecfin
        
        if codigo:
            
            sq1 = documentos_prove.objects.all().select_related('proveedor').filter(
                nombre_corto__in=['RUC','NOMBRAMIENTO_REP_LEGAL','CUMPLIMIENTO_IESS']).filter(proveedor=int(codigo)).order_by('fecha_documento')
            dc = documentos_prove.objects.filter(es_doc_categoria=True)
            qc1 = sq1 | dc 
            sq = qc1.filter(proveedor=int(codigo)).order_by('fecha_documento')

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte de Vencimiento de Documentos Proveedores'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'NOMBRE DOCUMENTO'
            ws['D3'] = 'FECHA EMISION/VIGENCIA'
            
            count = 4
            rowcount = 1  

            for registro in sq:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  registro.nombre_corto
                ws.cell(row = count, column = 4).value =  registro.fecha_documento
                
                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Expiracion_Documentos.xlsx"
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        elif docu:
            q = documentos_prove.objects.all().select_related('proveedor').filter(nombre_corto=str(docu)).order_by(
                'fecha_documento')
            
            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte de Vencimiento de Documentos Proveedores'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'PROVEEDOR'
            ws['D3'] = 'CATEGORIA'
            ws['E3'] = 'NOMBRE DOCUMENTO'
            ws['F3'] = 'FECHA EMISION/VIGENCIA'
            
            count = 4
            rowcount = 1  

            for registro in q:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  registro.proveedor.nombre_empresa
                ws.cell(row = count, column = 4).value =  registro.proveedor.categoria.nombre_categoria
                ws.cell(row = count, column = 5).value =  registro.nombre_corto
                ws.cell(row = count, column = 6).value =  registro.fecha_documento
                
                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Vencimiento_Documentos_Proveedores_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        elif tdocu:
            q1 = documentos_prove.objects.all().select_related('proveedor').filter(
                nombre_corto__in=['RUC','NOMBRAMIENTO_REP_LEGAL','CUMPLIMIENTO_IESS'])
    
            q2 = documentos_prove.objects.filter(es_doc_categoria=True)
            q3 = q1 | q2
            q = q3.order_by('fecha_documento')

            wb = Workbook()
            ws = wb.active
            ws['B1'] = 'Reporte de Vencimiento de Documentos Proveedores'

            ws.merge_cells('B1:E1')
            ws['B3'] = 'No'
            ws['C3'] = 'PROVEEDOR'
            ws['D3'] = 'CATEGORIA'
            ws['E3'] = 'NOMBRE DOCUMENTO'
            ws['F3'] = 'FECHA EMISION/VIGENCIA'
            
            count = 4
            rowcount = 1  

            for registro in q:
                ws.cell(row = count, column = 2).value =  rowcount
                ws.cell(row = count, column = 3).value =  registro.proveedor.nombre_empresa
                ws.cell(row = count, column = 4).value =  registro.proveedor.categoria.nombre_categoria
                ws.cell(row = count, column = 5).value =  registro.nombre_corto
                ws.cell(row = count, column = 6).value =  registro.fecha_documento
                
                count+=1
                rowcount+=1

            nombre_archivo = "Reporte_de_Vencimiento_Documentos_Proveedores_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
            response = HttpResponse(content_type = "application/ms-excel")
            content = "attachment; filename = {0}".format(nombre_archivo)
            response['Content-Disposition'] = content
            wb.save(response)
            return response

        elif fecini and fecfin:
            
            if fecini == fecfin:    
                q1 = documentos_prove.objects.filter(fecha_documento=fecini)
                query = q1.order_by('fecha_documento')
                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte de Vencimiento de Documentos Proveedores'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'PROVEEDOR'
                ws['D3'] = 'CATEGORIA'
                ws['E3'] = 'NOMBRE DOCUMENTO'
                ws['F3'] = 'FECHA EMISION/VIGENCIA'
                
                count = 4
                rowcount = 1  

                for registro in query:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  registro.proveedor.nombre_empresa
                    ws.cell(row = count, column = 4).value =  registro.proveedor.categoria.nombre_categoria
                    ws.cell(row = count, column = 5).value =  registro.nombre_corto
                    ws.cell(row = count, column = 6).value =  registro.fecha_documento
                    
                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Vencimiento_Documentos_Proveedores_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response

                
            else:
                q1 = documentos_prove.objects.filter(fecha_documento__range=[fecini,fecfin])
                query = q1.order_by('fecha_documento')
                wb = Workbook()
                ws = wb.active
                ws['B1'] = 'Reporte de Vencimiento de Documentos Proveedores'

                ws.merge_cells('B1:E1')
                ws['B3'] = 'No'
                ws['C3'] = 'PROVEEDOR'
                ws['D3'] = 'CATEGORIA'
                ws['E3'] = 'NOMBRE DOCUMENTO'
                ws['F3'] = 'FECHA EMISION/VIGENCIA'
                
                count = 4
                rowcount = 1  

                for registro in query:
                    ws.cell(row = count, column = 2).value =  rowcount
                    ws.cell(row = count, column = 3).value =  registro.proveedor.nombre_empresa
                    ws.cell(row = count, column = 4).value =  registro.proveedor.categoria.nombre_categoria
                    ws.cell(row = count, column = 5).value =  registro.nombre_corto
                    ws.cell(row = count, column = 6).value =  registro.fecha_documento
                    
                    count+=1
                    rowcount+=1

                nombre_archivo = "Reporte_de_Vencimiento_Documentos_Proveedores_Ecofroz-{date}.xlsx".format(date=datetime.now().strftime('%Y-%m-%d'))
                response = HttpResponse(content_type = "application/ms-excel")
                content = "attachment; filename = {0}".format(nombre_archivo)
                response['Content-Disposition'] = content
                wb.save(response)
                return response

                


